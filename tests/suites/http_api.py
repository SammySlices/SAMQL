"""Live HTTP API suite."""
import json
import os
import shutil
import sys
import tempfile
import threading
import time
import urllib.error
import urllib.request

from fixtures import ROWS
from harness import _api_urlopen, eq, need, run, skip
from paths import BACKEND, ROOT

DATADIR = None
CSV = None
JSONF = None

class Client:
    def __init__(self, base, token=None):
        self.base = base
        self.token = token

    def call(self, method, path, payload=None, timeout=60, headers=None):
        data = json.dumps(payload).encode() if payload is not None else None
        rq = urllib.request.Request(self.base + path, data=data,
                                    method=method)
        if data:
            rq.add_header("Content-Type", "application/json")
        supplied = {str(k).lower() for k in (headers or {})}
        if (self.token and path.startswith("/api/")
                and "x-samql-token" not in supplied):
            rq.add_header("X-SamQL-Token", self.token)
        for key, value in (headers or {}).items():
            rq.add_header(key, value)
        try:
            r = urllib.request.urlopen(rq, timeout=timeout)
            raw = r.read()
            return r.status, raw
        except urllib.error.HTTPError as e:
            return e.code, e.read()

    def js(self, method, path, payload=None, timeout=60, headers=None):
        st, raw = self.call(method, path, payload, timeout, headers)
        try:
            return st, json.loads(raw)
        except Exception:
            return st, {"_raw": raw[:200]}


def http_tests(datadir, csv_path, json_path, host, port, online):
    global DATADIR, CSV, JSONF
    DATADIR, CSV, JSONF = datadir, csv_path, json_path
    import server  # noqa: ensure module import works (contract test reuses)
    httpd, port = server.make_server(host, port)
    server._HTTPD = httpd
    th = threading.Thread(target=httpd.serve_forever, daemon=True)
    th.start()
    c = Client(f"http://{host}:{port}", httpd.samql_api_token)
    feats = {}
    state = {"rid": None}

    def t_health():
        st, d = c.js("GET", "/api/health")
        eq(st, 200, "status")
        need(d.get("version"), "no version")
        need(d.get("build"), "no build id in /api/health")

    def t_request_boundary_security():
        # Same-origin/no-Origin clients continue to work.
        st, _ = c.js("GET", "/api/health")
        eq(st, 200, "ordinary local client allowed")
        # A hostile browser origin cannot preflight or call a sensitive route.
        st, d = c.js("POST", "/api/nuke", {}, headers={
            "Origin": "https://evil.example",
        })
        eq(st, 403, "hostile Origin rejected")
        st, raw = c.call("OPTIONS", "/api/nuke", headers={
            "Origin": "https://evil.example",
            "Access-Control-Request-Method": "POST",
        })
        eq(st, 403, "hostile preflight rejected")
        # DNS-rebinding style Host names are rejected before route execution.
        st, d = c.js("GET", "/api/health", headers={"Host": "evil.example"})
        eq(st, 403, "unapproved Host rejected")

    def t_api_token_and_json_limit():
        # The HTML shell sets an HttpOnly capability cookie. Unauthenticated
        # local processes still cannot use sensitive endpoints without it.
        import http.client
        conn = http.client.HTTPConnection(host, port, timeout=10)
        conn.request("GET", "/")
        resp = conn.getresponse()
        body = resp.read()
        eq(resp.status, 200, "HTML shell status")
        cookies = resp.getheader("Set-Cookie") or ""
        need("samql_api_token=" in cookies and "HttpOnly" in cookies,
             "HTML shell sets an HttpOnly API capability cookie")
        need(httpd.samql_api_token.encode() not in body,
             "HTML body must not embed the process API token")
        conn.close()

        no_token = Client(c.base)
        st, d = no_token.js("GET", "/api/tables")
        eq(st, 403, "missing token rejected")
        need("token" in (d.get("error") or "").lower(),
             "missing-token response is explicit")
        st, _ = c.js("GET", "/api/tables", headers={
            "X-SamQL-Token": "definitely-wrong",
        })
        eq(st, 403, "wrong token rejected")
        st, d = no_token.js("GET", "/api/tables", headers={
            "Cookie": "samql_api_token=%s" % httpd.samql_api_token,
        })
        eq(st, 200, "HttpOnly cookie authenticates API calls")

        old = os.environ.get("SAMQL_JSON_BODY_MB")
        os.environ["SAMQL_JSON_BODY_MB"] = "1"
        try:
            st, d = c.js("POST", "/api/query", {"sql": "x" * (1024 * 1024)})
            eq(st, 413, "oversized JSON rejected before buffering")
            need("SAMQL_JSON_BODY_MB" in (d.get("error") or ""),
                 "JSON ceiling names its setting")
        finally:
            if old is None:
                os.environ.pop("SAMQL_JSON_BODY_MB", None)
            else:
                os.environ["SAMQL_JSON_BODY_MB"] = old

    def t_static_path_containment():
        import http.client
        import tempfile as _tf
        base = _tf.mkdtemp(prefix="samql_static_guard_")
        served = os.path.join(base, "dist")
        sibling = os.path.join(base, "dist_evil")
        os.makedirs(served)
        os.makedirs(sibling)
        open(os.path.join(served, "index.html"), "w", encoding="utf-8").write("OK")
        open(os.path.join(sibling, "secret.txt"), "w", encoding="utf-8").write("SECRET")
        old = server._frontend_dir_live
        server._frontend_dir_live = lambda: served
        try:
            conn = http.client.HTTPConnection(host, port, timeout=10)
            conn.request("GET", "/../dist_evil/secret.txt")
            resp = conn.getresponse()
            resp.read()
            eq(resp.status, 403, "sibling-prefix traversal rejected")
            conn.close()
            if hasattr(os, "symlink"):
                link = os.path.join(served, "escape")
                try:
                    os.symlink(sibling, link, target_is_directory=True)
                except (OSError, NotImplementedError):
                    link = None
                if link:
                    st, raw = c.call("GET", "/escape/secret.txt")
                    eq(st, 403, "symlink escape rejected")
        finally:
            server._frontend_dir_live = old
            shutil.rmtree(base, ignore_errors=True)

    def t_unmatched_multipart_cleanup():
        from samql_core import tmputil
        boundary = "----samqlCleanupBoundary"
        payload = ("--" + boundary + "\r\n"
                   "Content-Disposition: form-data; name=\"files\"; "
                   "filename=\"orphan.bin\"\r\n"
                   "Content-Type: application/octet-stream\r\n\r\n"
                   "orphan-data\r\n--" + boundary + "--\r\n").encode()
        before = {n for n in os.listdir(tmputil.instance_dir())
                  if n.startswith("up_")}
        rq = urllib.request.Request(c.base + "/api/not-a-route", data=payload,
                                    method="POST", headers={
            "Content-Type": "multipart/form-data; boundary=" + boundary,
            "Content-Length": str(len(payload)),
            "X-SamQL-Token": c.token,
        })
        try:
            urllib.request.urlopen(rq, timeout=10)
            need(False, "unknown route should 404")
        except urllib.error.HTTPError as e:
            eq(e.code, 404, "unknown multipart route returns 404")
            e.read()
        after = {n for n in os.listdir(tmputil.instance_dir())
                 if n.startswith("up_")}
        eq(after, before, "unmatched upload leaves no spool file")


    def t_features():
        st, d = c.js("GET", "/api/features")
        eq(st, 200, "status")
        feats.update(d)

    def t_fs_list():
        st, d = c.js("GET", "/api/fs/list?path=" + DATADIR)
        eq(st, 200, "status")
        need("entries" in d or "items" in d or isinstance(d, dict),
             "no listing payload")
        need(d.get("home"), "fs/list did not report a home dir")
        sc = d.get("shortcuts")
        need(isinstance(sc, list) and sc, "fs/list returned no shortcuts")
        need("Home" in [x.get("label") for x in sc], "no Home shortcut")
        need(all(os.path.isdir(x["path"]) for x in sc),
             "a shortcut points to a non-directory")

    def t_wave1_stateful_server_rail():
        # Wave 1 stabilization: one stateful cross-route journey proves that a
        # real load can feed query -> filtered page -> export, and that a bad
        # query does not poison the next request. Individual routes already
        # have focused tests; this guards their lifecycle glue and cleanup.
        # (Also covers POST /api/load/start + progress polling.)
        d = tempfile.mkdtemp(prefix="samql_wave1_http_")
        csv_path = os.path.join(d, "wave1_stateful.csv")
        table = None
        try:
            with open(csv_path, "w", encoding="utf-8", newline="") as f:
                f.write("id,name,category,amount\n")
                f.write("1,alpha,A,12.5\n2,beta,B,30\n")
                f.write("3,gamma,A,7.25\n4,delta,B,42\n")
            st, start = c.js("POST", "/api/load/start", {
                "path": csv_path, "destination": "sqlite",
                "mode": "materialize",
            })
            eq(st, 200, "wave1 load start")
            job = start.get("job_id")
            need(job, "wave1 load returned no job id: %r" % start)
            progress = {}
            for _ in range(600):
                st, progress = c.js("GET", "/api/load/progress/%s" % job)
                if progress.get("state") in ("done", "error", "cancelled"):
                    break
                time.sleep(0.02)
            eq(progress.get("state"), "done",
               "wave1 load did not finish: %r" % progress)
            loaded = progress.get("loaded") or []
            need(isinstance(loaded, list) and loaded,
                 "wave1 load returned no table: %r" % progress)
            table = loaded[0].get("name") or loaded[0].get("table")
            need(table, "wave1 loaded-table name missing: %r" % loaded[0])

            qtable = '"%s"' % str(table).replace('"', '""')
            st, result = c.js("POST", "/api/query", {
                "sql": "SELECT id,name,category,amount FROM %s" % qtable,
                "target": "__local__", "read_only": True,
            })
            eq(st, 200, "wave1 query status")
            rid = result.get("result_id")
            need(rid and not result.get("error"),
                 "wave1 query failed: %r" % result)

            st, page = c.js("POST", "/api/result/%s/page" % rid, {
                "limit": 20,
                "filters": [{"column": "category", "op": "equals",
                             "value": "A"}],
                "sort_col": "amount", "descending": False,
            })
            eq(st, 200, "wave1 filtered page status")
            eq(page.get("total_rows"), 2, "wave1 filtered row count")
            cat_i = page.get("columns", []).index("category")
            need(all(r[cat_i] == "A" for r in page.get("rows", [])),
                 "wave1 filtered page leaked another category")

            st, raw = c.call("POST", "/api/result/%s/export" % rid,
                             {"fmt": "csv"})
            eq(st, 200, "wave1 export status")
            need(b"alpha" in raw and b"delta" in raw and b"category" in raw,
                 "wave1 CSV export lost expected rows/header")

            st, bad = c.js("POST", "/api/query", {
                "sql": "SELECT * FROM definitely_missing_wave1_table",
                "target": "__local__",
            })
            need(st in (200, 400) and bad.get("error"),
                 "wave1 invalid query must fail cleanly: %r / %r" % (st, bad))
            st, recovered = c.js("POST", "/api/query", {
                "sql": "SELECT 1 AS recovered", "target": "__local__",
            })
            eq(st, 200, "wave1 recovery query status")
            need(recovered.get("result_id") and not recovered.get("error"),
                 "server did not recover after an invalid query: %r" % recovered)
        finally:
            if table:
                safe = '"%s"' % str(table).replace('"', '""')
                try:
                    c.js("POST", "/api/query", {
                        "sql": "DROP TABLE IF EXISTS %s" % safe,
                        "target": "__local__",
                    })
                except Exception:
                    pass
            shutil.rmtree(d, ignore_errors=True)

    def t_load_files_start_cancel():
        # drag-and-drop's background loader: multipart upload -> a job that can
        # be polled + cancelled. A tiny file finishes near-instantly (so we
        # can't race a mid-load cancel deterministically), but we can prove the
        # route plumbing: the job reaches "done", and cancelling a finished or
        # unknown job is a safe no-op.
        import urllib.request as _u
        csv = b"id,name\n1,alpha\n2,beta\n3,gamma\n"
        boundary = "----samqlfilesBOUNDARY"
        destfield = ("--%s\r\n"
                     'Content-Disposition: form-data; name="destination"\r\n'
                     "\r\nsqlite\r\n") % boundary
        pre = ("--%s\r\n"
               'Content-Disposition: form-data; name="files"; '
               'filename="drop.csv"\r\n'
               "Content-Type: text/csv\r\n\r\n") % boundary
        body = (destfield.encode("latin-1") + pre.encode("latin-1") + csv
                + ("\r\n--%s--\r\n" % boundary).encode("latin-1"))
        rq = _u.Request(f"http://{host}:{port}/api/load/files-start",
                        data=body, method="POST")
        rq.add_header("Content-Type",
                      "multipart/form-data; boundary=" + boundary)
        rq.add_header("X-SamQL-Token", c.token)
        r = _u.urlopen(rq, timeout=30)
        d = json.loads(r.read())
        eq(r.status, 200, "files-start status")
        job = d.get("job_id")
        need(job, f"no job id: {d}")
        p = {}
        for _ in range(600):
            st, p = c.js("GET", f"/api/load/progress/{job}")
            if p.get("state") in ("done", "error", "cancelled"):
                break
            time.sleep(0.02)
        eq(p.get("state"), "done", f"files-start load did not finish: {p}")
        need((p.get("rows") or 0) >= 3, f"expected >=3 rows loaded: {p}")
        # cancelling an already-finished job reports nothing to cancel
        st, cc = c.js("POST", f"/api/load/cancel/{job}")
        eq(st, 200, "cancel status")
        eq(cc.get("cancelled"), False, "cancel on a finished job is a no-op")
        # cancelling an unknown job id is safe
        st, cc = c.js("POST", "/api/load/cancel/not-a-real-job-xyz")
        need(cc.get("cancelled") is False, "bogus cancel -> cancelled False")

    def t_malformed_upload_no_connection_drop():
        # A malformed/oversized upload must never drop the socket: an unhandled
        # exception during multipart parsing would close the connection with no
        # response, which the UI reports as the opaque "Failed to fetch". The
        # dispatcher wraps the parse and returns a clean 400, and the server
        # stays responsive afterward.
        import urllib.request as _u
        import urllib.error as _ue
        # static: the guard is wired (parse_multipart wrapped + 400 on failure)
        srv_src = open(os.path.join(BACKEND, "server.py"),
                       encoding="utf-8").read()
        di = srv_src.find("def _dispatch_api")
        seg = srv_src[di:di + 3500]
        pi = seg.find("stream_multipart(")
        need(pi != -1, "dispatcher calls stream_multipart")
        guard = seg[max(0, pi - 120):pi + 1600]
        need("try:" in guard and "_send_json(400" in guard,
             "stream_multipart wrapped so a bad upload returns 400, not a reset")
        # behavioural: force the parser to raise; the request must still get a
        # clean 400 and the connection must NOT drop.
        orig = server.stream_multipart

        def _boom(*a, **k):
            raise ValueError("simulated malformed multipart")

        server.stream_multipart = _boom
        dropped = False
        status = None
        try:
            bb = b"----samqlbadBOUNDARY"
            data = b"--" + bb + b"\r\nbad\r\n--" + bb + b"--\r\n"
            rq = _u.Request(f"http://{host}:{port}/api/load/files-start",
                            data=data, method="POST")
            rq.add_header("Content-Type",
                          "multipart/form-data; boundary=----samqlbadBOUNDARY")
            rq.add_header("X-SamQL-Token", c.token)
            try:
                r = _u.urlopen(rq, timeout=30)
                status = r.status
                r.read()
            except _ue.HTTPError as e:
                status = e.code
                e.read()
            except Exception:
                dropped = True  # connection reset == the "failed to fetch" bug
        finally:
            server.stream_multipart = orig
        need(not dropped, "a parse failure must not drop the connection")
        # the drain-then-retire contract (the Windows RST fix) is wired
        need("_drain_body(length)" in srv_src
             and srv_src.count("self._drain_body(") >= 2
             and "def _drain_body" in srv_src,
             "both failure paths drain the unread body before responding")
        eq(status, 400, "a malformed upload returns a clean 400")
        # the server survived and still serves requests
        st, _h = c.js("GET", "/api/health")
        eq(st, 200, "server still responsive after a malformed upload")

    def t_flatten_progress():
        import tempfile as _tf
        outd = _tf.mkdtemp(prefix="t_http_ft_")
        st, d = c.js("POST", "/api/flatten/start",
                     {"json_path": JSONF, "out_dir": outd})
        eq(st, 200, "flatten start status")
        job = d.get("job_id")
        need(job, f"no flatten job id: {d}")
        p = {}
        for _ in range(600):
            st, p = c.js("GET", f"/api/flatten/progress/{job}")
            if p.get("state") in ("done", "error"):
                break
            time.sleep(0.02)
        eq(p.get("state"), "done", f"flatten did not finish: {p}")
        eq(p.get("stage"), "done", "final stage is done")
        need(p.get("result") and p["result"].get("ok"),
             f"flatten result missing/not ok: {p.get('result')}")
        need((p["result"].get("table_count") or 0) >= 2,
             "flatten produced the root + a nested-array child table")

    def t_load_jobs_list():
        # /api/load/jobs lists recent/active load jobs so the UI can reattach
        # after a reload. Shape: {jobs: [{id, state, name, bytes_*, ...}]}. Load
        # jobs ran earlier in this suite, so the contract is exercised with real
        # entries.
        st, d = c.js("GET", "/api/load/jobs")
        eq(st, 200, "load/jobs status")
        jobs = d.get("jobs")
        need(isinstance(jobs, list), "jobs is a list: %r" % d)
        for j in jobs:
            for k in ("id", "state", "name", "bytes_done", "bytes_total"):
                need(k in j, "job entry missing %s: %r" % (k, j))

    def t_tables():
        st, d = c.js("GET", "/api/tables")
        eq(st, 200, "status")
        names = [t["name"] for t in d.get("tables", [])]
        need(names, f"no tables: {d}")

    def t_query():
        st, d = c.js("POST", "/api/query",
                     {"sql": "SELECT id, name, score, category FROM data"})
        eq(st, 200, "status")
        eq(d["total_rows"], ROWS, "total")
        state["rid"] = d["result_id"]

    def t_query_preview():
        st, d = c.js("POST", "/api/query", {
            "sql": "SELECT 1 AS id, 'a' AS name UNION ALL "
                   "SELECT 2, 'b' UNION ALL SELECT 3, 'c' ORDER BY id",
            "preview_limit": 2,
        })
        eq(st, 200, "preview query status")
        eq(d.get("total_rows"), 2, "preview query is bounded")
        need(d.get("preview") is True and d.get("preview_limit") == 2,
             "preview query metadata returned over HTTP")

    def t_gzip():
        import gzip as _gz
        import urllib.request as _u

        pad = "x" * 300  # make the page exceed the gzip threshold
        body = json.dumps(
            {"sql": f"SELECT id, name, '{pad}' AS pad FROM data"}
        ).encode()
        rq = _u.Request(
            f"http://{host}:{port}/api/query", data=body, method="POST"
        )
        rq.add_header("Content-Type", "application/json")
        rq.add_header("Accept-Encoding", "gzip")
        rq.add_header("X-SamQL-Token", c.token)
        r = _u.urlopen(rq, timeout=30)
        enc = r.headers.get("Content-Encoding")
        raw = r.read()
        need(enc == "gzip", f"expected gzip encoding, got {enc!r}")
        d = json.loads(_gz.decompress(raw))
        need(d.get("total_rows", 0) >= 1, "gzip body did not decode")

    def t_page():
        st, d = c.js("POST", f"/api/result/{state['rid']}/page",
                     {"offset": 0, "limit": 4, "sort_col": "score",
                      "descending": True})
        eq(st, 200, "status")
        eq(float(d["rows"][0][2]), 88.5, "sorted page (score desc)")

    def t_filter():
        st, d = c.js("POST", f"/api/result/{state['rid']}/page",
                     {"limit": 5,
                      "filters": [{"column": "category", "op": "equals",
                                   "value": "alpha"}]})
        eq(st, 200, "status")
        need(d.get("filtered") is True, "filtered flag not set")
        eq(d["total_rows"], ROWS // 3, "filtered total (one category)")
        need(all(r[3] == "alpha" for r in d["rows"]), "row leaked filter")
        st, d = c.js("POST", f"/api/result/{state['rid']}/page",
                     {"filters": [{"column": "id", "op": "is_null"}]})
        eq(d["total_rows"], 0, "is_null over a non-null column should be empty")

    def t_cancel_route():
        # route + handler wiring (behavioural cancel is covered in-process);
        # an unknown id is a valid 200 with ok:false.
        st, d = c.js("POST", "/api/query/not-a-real-id/cancel")
        eq(st, 200, "status")
        need("ok" in d, f"no ok field: {d}")

    def t_reconcile():
        # Compare the loaded table against a copy whose names all differ.
        # The report carries totals + a per-field row; the underlying rows
        # are fetched via the drilldown / profile endpoints.
        st, _ = c.js("POST", "/api/query",
                     {"sql": "CREATE TABLE data2 AS "
                      "SELECT id, name || '_x' AS name FROM data"})
        eq(st, 200, "create data2")
        st, d = c.js("POST", "/api/reconcile",
                     {"left": "data", "right": "data2", "keys": ["id"],
                      "compare": ["name"]})
        eq(st, 200, "reconcile status")
        need("totals" in d and "fields" in d, f"bad shape: {d}")
        t = d["totals"]
        eq(t["a_only"], 0, "same keys -> only-left 0")
        eq(t["b_only"], 0, "same keys -> only-right 0")
        eq(t["non_matching"], ROWS, "every name differs")
        eq(t["matching"], 0, "nothing matches")
        need("rows" not in d, "reconcile must not inline rows")
        eq(len(d["fields"]), 1, "one compared field")
        eq(d["fields"][0]["non_matching"], ROWS, "per-field non-matching")
        # drill down the name break and page it
        st, dd = c.js("POST", "/api/reconcile/drilldown",
                      {"left": "data", "right": "data2", "keys": ["id"],
                       "bucket": "non_matching", "field": "name"})
        eq(st, 200, "drilldown status")
        eq(dd["count"], ROWS, "drilldown non-matching count")
        rid = dd["result_id"]
        need(rid, "expected a pageable drilldown result id")
        st, pg = c.js("POST", f"/api/result/{rid}/page",
                      {"offset": 0, "limit": 5})
        eq(st, 200, "page drilldown result")
        eq(pg["total_rows"], ROWS, "drilldown result pages to true count")
        # profile the underlying rows
        st, pr = c.js("POST", "/api/reconcile/profile",
                      {"left": "data", "right": "data2", "keys": ["id"],
                       "bucket": "non_matching", "field": "name"})
        eq(st, 200, "reconcile profile status")
        eq(pr["total_rows"], ROWS, "profile row count")
        # materialise a SELECT and confirm the endpoint returns its columns and
        # that the materialised table is reconcilable (here against data2)
        st, mt = c.js("POST", "/api/materialize",
                      {"name": "__nb_http", "sql": "SELECT id, name FROM data",
                       "target": "__local__"})
        eq(st, 200, "materialize status")
        need(mt.get("columns") == ["id", "name"] and not mt.get("error"),
             f"materialize shape: {mt}")
        # M1: the staging table must not show up in /api/tables
        st, tl = c.js("GET", "/api/tables")
        need(
            st == 200
            and all(
                not t["name"].startswith("__")
                for t in (tl.get("tables") or [])
            ),
            "staging __ tables leaked into /api/tables",
        )
        st, rc = c.js("POST", "/api/reconcile",
                      {"left": "__nb_http", "right": "data2", "keys": ["id"],
                       "compare": ["name"]})
        eq(st, 200, "reconcile over materialised status")
        eq(rc["totals"]["non_matching"], ROWS,
           "materialised table reconciles against data2")
        st, bad = c.js("POST", "/api/materialize",
                       {"name": "bad name", "sql": "SELECT 1"})
        need(bad.get("error"), f"bad materialise name should error: {bad}")

    def t_page_projection():
        st, q = c.js("POST", "/api/query", {"sql": "SELECT * FROM data"})
        eq(st, 200, "query status")
        rid = q.get("result_id")
        need(rid, "no result_id")
        st, pg = c.js("POST", f"/api/result/{rid}/page",
                      {"offset": 0, "limit": 3, "columns": ["name", "id"]})
        eq(st, 200, "page status")
        eq(pg["columns"], ["name", "id"],
           f"projection columns: {pg.get('columns')}")
        eq(len(pg["rows"][0]), 2, "projected row width")

    def t_export():
        st, raw = c.call("POST", f"/api/result/{state['rid']}/export",
                         {"fmt": "csv"})
        eq(st, 200, "status")
        need(len(raw) > 0, "empty export")
        need(b"," in raw or b"\n" in raw, "not csv-ish")

    def t_profile():
        st, d = c.js("POST", "/api/table/data/profile", {"engine": "sqlite"})
        eq(st, 200, "status")
        need(d.get("columns"), "no profile columns")

    def t_chart():
        st, d = c.js("POST", "/api/chart/data",
                     {"table": "data", "engine": "sqlite",
                      "chart_type": "bar", "x": "category", "y": "score",
                      "agg": "sum"})
        eq(st, 200, "status")
        need("error" not in d, f"chart error: {d.get('error')}")

    def t_pivot():
        st, d = c.js("POST", "/api/pivot",
                     {"table": "data", "engine": "sqlite",
                      "rows": ["category"], "value": "score", "agg": "sum"})
        eq(st, 200, "status")
        need("error" not in d, f"pivot error: {d.get('error')}")

    def t_sql_format():
        st, d = c.js("POST", "/api/sql/format", {"sql": "select 1"})
        eq(st, 200, "status")
        need("result" in d, "no format result")

    def t_statement_at():
        sql = "SELECT 1;\nSELECT 2 FROM data;"
        st, d = c.js("POST", "/api/sql/statement-at",
                     {"sql": sql, "pos": sql.index("FROM")})
        eq(st, 200, "status")
        need("SELECT 2" in d.get("sql", ""), "wrong statement")

    def t_saved():
        st, d = c.js("POST", "/api/saved",
                     {"name": "h1", "sql": "SELECT 1", "tags": []})
        eq(st, 200, "save status")
        st, d = c.js("GET", "/api/saved")
        need(any(x.get("name") == "h1" for x in d.get("saved", [])),
             "saved not listed")
        st, d = c.js("DELETE", "/api/saved", {"name": "h1"})
        eq(st, 200, "delete status")

    def t_workflows():
        g = {"nodes": [{"id": "i", "type": "input", "x": 10, "y": 20,
                        "config": {"table": "data"}}], "edges": []}
        st, d = c.js("POST", "/api/workflows", {"name": "wf-http", "graph": g})
        eq(st, 200, "workflow save status")
        need(d.get("ok"), "workflow save ok")
        st, d = c.js("GET", "/api/workflows")
        need(any(x.get("name") == "wf-http" for x in d.get("workflows", [])),
             "saved workflow listed")
        st, d = c.js("POST", "/api/workflows/load", {"name": "wf-http"})
        eq(st, 200, "workflow load status")
        need((d.get("graph") or {}).get("nodes", [{}])[0].get("x") == 10,
             "workflow load returns the full graph with positions")
        st, d = c.js("DELETE", "/api/workflows", {"name": "wf-http"})
        eq(st, 200, "workflow delete status")
        st, d = c.js("GET", "/api/workflows")
        need(not any(x.get("name") == "wf-http"
                     for x in d.get("workflows", [])),
             "deleted workflow no longer listed")
        # kind-aware: the same name can exist once per kind (IDE/Journal/Node/Dashboard)
        c.js("POST", "/api/workflows",
             {"name": "shared", "kind": "node",
              "graph": {"nodes": [{}], "edges": []}})
        c.js("POST", "/api/workflows",
             {"name": "shared", "kind": "ide", "graph": {"sql": "SELECT 42"}})
        c.js("POST", "/api/workflows",
             {"name": "shared", "kind": "dashboard",
              "graph": {
                  "version": 2,
                  "activeId": "d1",
                  "dashboards": [{
                      "id": "d1",
                      "name": "Main",
                      "widgets": [{"id": "w1"}, {"id": "w2"}],
                  }],
              }})
        st, d = c.js("GET", "/api/workflows")
        shared = [x for x in d.get("workflows", []) if x.get("name") == "shared"]
        eq(sorted(x.get("kind") for x in shared), ["dashboard", "ide", "node"],
           "same name coexists across kinds")
        st, d = c.js("POST", "/api/workflows/load",
                     {"name": "shared", "kind": "ide"})
        eq((d.get("graph") or {}).get("sql"), "SELECT 42",
           "load by kind returns the matching payload")
        st, d = c.js("POST", "/api/workflows/load",
                     {"name": "shared", "kind": "dashboard"})
        eq((d.get("graph") or {}).get("version"), 2,
           "dashboard kind load returns the workspace")
        c.js("DELETE", "/api/workflows", {"name": "shared", "kind": "ide"})
        st, d = c.js("GET", "/api/workflows")
        shared2 = sorted(x.get("kind") for x in d.get("workflows", [])
                         if x.get("name") == "shared")
        eq(shared2, ["dashboard", "node"], "deleting one kind leaves the others")
        c.js("DELETE", "/api/workflows", {"name": "shared", "kind": "node"})
        c.js("DELETE", "/api/workflows", {"name": "shared", "kind": "dashboard"})

    def t_workspace_file_http():
        import tempfile, os as _osf, shutil as _sh
        d = tempfile.mkdtemp(prefix="samql_wf_")
        try:
            p = _osf.path.join(d, "demo.samql.json")
            content = ('{"samql":"workflow","kind":"ide","name":"Demo",'
                       '"payload":{"sql":"SELECT 1"}}')
            st, r = c.js("POST", "/api/workspace/save-file",
                         {"path": p, "content": content})
            eq(st, 200, "save-file status")
            need(r.get("ok") and r.get("name") == "demo.samql.json",
                 "save-file returns the file name")
            need(_osf.path.isfile(p), "file is actually written to disk")
            st, r = c.js("POST", "/api/workspace/open-file", {"path": p})
            eq(st, 200, "open-file status")
            eq(r.get("content"), content,
               "open-file round-trips the content exactly")
            st, r = c.js("POST", "/api/workspace/save-file",
                         {"path": _osf.path.join(d, "nope", "x.json"),
                          "content": "x"})
            eq(st, 400, "saving into a missing folder is rejected")
            st, r = c.js("POST", "/api/workspace/open-file",
                         {"path": _osf.path.join(d, "ghost.json")})
            eq(st, 400, "opening a missing file is rejected")
        finally:
            _sh.rmtree(d, ignore_errors=True)

    def t_table_create_http():
        st, d = c.js("POST", "/api/table/create", {
            "name": "grid_http", "columns": ["a", "b"],
            "rows": [["1", "x"], ["2", "y"]], "destination": "sqlite"})
        eq(st, 200, "table create status")
        need(d.get("ok"), "table create ok")
        tname = d.get("table", "grid_http")
        st, d = c.js("GET", "/api/tables")
        rows = d if isinstance(d, list) else d.get("tables", d.get("all", []))
        need(any(t.get("name") == tname for t in rows),
             "the created table appears in the tables list")

    def t_catalog_endpoints_http():
        # No catalog is loaded in the test server (and importing needs a live
        # SQL Server), so we just confirm the endpoints route and return a
        # graceful error rather than 404 / 500.
        st, d = c.js("GET", "/api/catalog/columns?name=NoSuchTable")
        eq(st, 200, "catalog columns route responds")
        need(d.get("error"), "unknown catalog table returns an error")
        st, d = c.js("POST", "/api/catalog/import", {"name": "NoSuchTable"})
        eq(st, 200, "catalog import route responds")
        need(d.get("error"), "importing an unknown table returns an error")

    def t_nodeflow_write_http():
        g = {"nodes": [
            {"id": "i", "type": "input", "config": {"table": "data"}},
            {"id": "w", "type": "write", "config": {"name": "http_written"}}],
            "edges": [{"from": {"node": "i", "port": "out"},
                       "to": {"node": "w", "port": "in"}}]}
        st, d = c.js("POST", "/api/nodeflow/write",
                     {"graph": g, "node": "w", "name": "http_written"})
        eq(st, 200, "nodeflow write status")
        need(d.get("ok"), "nodeflow write ok")
        st, d = c.js("GET", "/api/tables")
        rows = d if isinstance(d, list) else d.get("tables", d.get("all", []))
        need(any(t.get("name") == "http_written" for t in rows),
             "the written table appears in the tables list")

    def t_history():
        st, d = c.js("GET", "/api/history")
        eq(st, 200, "status")
        need("history" in d, "no history key")

    def t_nodeflow_export_many_http():
        import tempfile
        import os
        d_out = tempfile.mkdtemp()
        # the input node feeds two outputs -> it's the shared node, built once
        # for the batch. Different formats + a name de-collision are exercised.
        g = {"nodes": [
            {"id": "i", "type": "input", "config": {"table": "data"}},
            {"id": "o1", "type": "output", "config": {}},
            {"id": "o2", "type": "output", "config": {}}],
            "edges": [
            {"from": {"node": "i", "port": "out"},
             "to": {"node": "o1", "port": "in"}},
            {"from": {"node": "i", "port": "out"},
             "to": {"node": "o2", "port": "in"}}]}
        items = [
            {"node_id": "o1", "folder": d_out, "fmt": "csv",
             "base_name": "first"},
            {"node_id": "o2", "folder": d_out, "fmt": "json",
             "base_name": "second"}]
        st, d = c.js("POST", "/api/nodeflow/export-many",
                     {"graph": g, "items": items})
        eq(st, 200, "export-many status")
        need(d.get("ok"), "export-many ok: %s" % d.get("error"))
        res = d.get("results") or []
        eq(len(res), 2, "one result per output")
        need(all(r.get("ok") for r in res), "both outputs exported")
        need(os.path.isfile(os.path.join(d_out, "first.csv")),
             "csv file written")
        need(os.path.isfile(os.path.join(d_out, "second.json")),
             "json file written")
        # two outputs with the SAME base name must not clobber each other
        items2 = [
            {"node_id": "o1", "folder": d_out, "fmt": "csv",
             "base_name": "dup"},
            {"node_id": "o2", "folder": d_out, "fmt": "csv",
             "base_name": "dup"}]
        st, d2 = c.js("POST", "/api/nodeflow/export-many",
                      {"graph": g, "items": items2})
        eq(st, 200, "export-many (dup names) status")
        files = {r.get("file") for r in (d2.get("results") or []) if r.get("ok")}
        eq(len(files), 2, "same base name de-collided into two files")
        # empty folder defaults to Downloads; a non-existent path is rejected
        dl = tempfile.mkdtemp()
        old_env = os.environ.get("SAMQL_DOWNLOADS_DIR")
        os.environ["SAMQL_DOWNLOADS_DIR"] = dl
        try:
            st, d3 = c.js("POST", "/api/nodeflow/export-many",
                          {"graph": g, "items": [
                              {"node_id": "o1", "folder": "", "fmt": "csv",
                               "base_name": "dlmany"}]})
            eq(st, 200, "export-many empty folder status")
            need(d3.get("ok"),
                 "empty folder exports to Downloads: %s" % d3.get("error"))
            need(os.path.isfile(os.path.join(dl, "dlmany.csv")),
                 "Downloads file from export-many")
        finally:
            if old_env is None:
                os.environ.pop("SAMQL_DOWNLOADS_DIR", None)
            else:
                os.environ["SAMQL_DOWNLOADS_DIR"] = old_env
        st, d4 = c.js("POST", "/api/nodeflow/export-many",
                      {"graph": g, "items": [
                          {"node_id": "o1",
                           "folder": "/definitely/missing/folder",
                           "fmt": "csv"}]})
        need(d4.get("error"), "missing folder rejected")

    def t_nodeflow_columns_batch_http():
        g = {"nodes": [
            {"id": "a", "type": "sql",
             "config": {"sql": "SELECT 1 AS x, 2 AS y", "label": "a"}},
            {"id": "b", "type": "sql",
             "config": {"sql": "SELECT 3 AS p", "label": "b"}}],
            "edges": []}
        st, d = c.js("POST", "/api/nodeflow/columns-batch",
                     {"graph": g, "requests": [
                         {"node": "a", "port": "out"},
                         {"node": "b", "port": "out"}]})
        eq(st, 200, "columns-batch status")
        res = d.get("results") or []
        eq(len(res), 2, "one result per request")
        byid = {(r["node"], r["port"]): r.get("columns") for r in res}
        eq(byid[("a", "out")], ["x", "y"], "batch columns for a")
        eq(byid[("b", "out")], ["p"], "batch columns for b")

    def t_memory():
        st, d = c.js("GET", "/api/memory")
        eq(st, 200, "status")
        st, d = c.js("POST", "/api/memory/free")
        eq(st, 200, "free status")

    def t_sweep_temp():
        st, d = c.js("POST", "/api/maintenance/sweep-temp")
        eq(st, 200, "sweep-temp status")
        need("removed" in d and "instance_bytes" in d,
             "sweep-temp returns removed + instance_bytes")
        need(isinstance(d.get("instance_bytes"), int) and d["instance_bytes"] >= 0,
             "instance_bytes is a non-negative integer")

    def t_rename_change_drop():
        st, d = c.js("POST", "/api/table/rename",
                     {"engine": "sqlite", "old": "data", "new": "data2"})
        eq(st, 200, "rename status")
        st, d = c.js("POST", "/api/table/change-type",
                     {"engine": "sqlite", "table": "data2", "col": "score",
                      "new_type": "TEXT"})
        eq(st, 200, "change-type status")
        st, d = c.js("POST", "/api/table/drop",
                     {"engine": "sqlite", "name": "data2"})
        eq(st, 200, "drop status")

    def t_clear():
        st, d = c.js("POST", "/api/clear")
        eq(st, 200, "status")

    def t_mssql_drivers():
        st, d = c.js("GET", "/api/mssql/drivers")
        if st != 200:
            skip("mssql drivers endpoint unavailable without pyodbc")
        need(isinstance(d, dict), "bad drivers payload")

    def t_mssql_tables_guard():
        # Listing tables on a connection that was never opened must fail fast.
        st, d = c.js("POST", "/api/mssql/tables", {"name": "nope"})
        eq(st, 400, "tables on dead connection should 400")

    def t_mssql_import_guard():
        # Import returns a structured error (not a 500) when the named
        # connection is not active, so the UI can surface it.
        st, d = c.js("POST", "/api/mssql/import",
                     {"name": "nope", "query": "SELECT 1", "base_name": "x"})
        eq(st, 200, "import route status")
        need(isinstance(d, dict) and d.get("error"),
             "import on dead connection should return an error: %r" % d)

    def t_api_fetch():
        if not online:
            skip("network test (run with --online)")
        st, d = c.js("POST", "/api/api-fetch",
                     {"url": "https://jsonplaceholder.typicode.com/users",
                      "base_name": "users", "destination": "sqlite"}, 30)
        eq(st, 200, "status")

    def t_health_app_marker():
        # the single-instance probe identifies SamQL by the health 'app' field
        st, d = c.js("GET", "/api/health")
        eq(st, 200, "status")
        eq(str(d.get("app", "")).lower(), "samql",
           "health must mark app=SamQL so a probe can recognise it")
        need(server._probe_existing(host, port),
             "probe should see this live SamQL instance")

    def t_brand_favicon():
        # The server must return a real PNG at /favicon.ico and /app-icon.png
        # (NOT fall through to the SPA's index.html), plus a web manifest
        # pointing at the icon, so the browser tab shows the SamQL mark.
        for p in ("/favicon.ico", "/app-icon.png", "/logo.png"):
            st, raw = c.call("GET", p)
            eq(st, 200, f"{p} status")
            # .461: /favicon.ico is now a REAL multi-size ICO (the
            # taskbar-branding asset); the png paths stay png.
            want = (b"\x00\x00\x01\x00" if p.endswith(".ico")
                    else b"\x89PNG\r\n\x1a\n")
            need(raw[:len(want)] == want,
                 f"{p} must return its image type, not HTML")
        st, raw = c.call("GET", "/manifest.webmanifest")
        eq(st, 200, "manifest status")
        m = json.loads(raw)
        need(m.get("icons") and m["icons"][0]["src"] == "/app-icon.png",
             "manifest points at the app icon")

    def t_focus_endpoint():
        st, d = c.js("POST", "/api/focus", {})
        eq(st, 200, "status")
        eq(d.get("ok"), True, "focus returns ok")
        need("focused" in d, "focus reports whether a window was surfaced")
        # no native window under test, so nothing to surface
        eq(d.get("focused"), False, "no pywebview window in the harness")

    def t_window_helpers_safe():
        # with no pywebview present these must degrade quietly
        eq(server._focus_window(), False, "focus is False without a window")
        server._destroy_window()  # must not raise
        import socket as _sock
        sk = _sock.socket()
        sk.bind((host, 0))
        dead = sk.getsockname()[1]
        sk.close()
        eq(server._probe_existing(host, dead), False,
           "probe of a closed port is False")

    def t_single_instance_attaches():
        # a second launch against a running instance attaches and exits
        import webbrowser as _wb
        opened = {"url": None}
        orig_open = server.webbrowser.open
        orig_argv = sys.argv
        server.webbrowser.open = lambda u, *a, **k: opened.__setitem__("url", u)
        sys.argv = ["server.py", "--host", host, "--port", str(port)]
        try:
            raised = False
            try:
                server._maybe_attach_to_running()
            except SystemExit as e:
                raised = True
                eq(e.code, 0, "attach exits cleanly")
            need(raised, "second launch must attach + exit, not fall through")
            eq(opened["url"], "http://%s:%s/" % (host, port),
               "attach opens a browser to the running instance (no window)")
        finally:
            server.webbrowser.open = orig_open
            sys.argv = orig_argv

    def t_browser_open_mode():
        # The UI opens in a NORMAL browser window (tabs + address bar),
        # preferring Chrome, then Edge, then the system default. Detection is
        # built from install roots + registry + PATH. The launch is a plain
        # [exe, url] -- an ordinary window on the user's own profile, not a
        # process we own or wait on. The actual Popen can't run in this
        # sandbox, so detection + the launch command + the mode routing are
        # exercised here and the launcher is source-guarded.
        fake = {"PROGRAMFILES": r"C:\Program Files",
                "PROGRAMFILES(X86)": r"C:\Program Files (x86)",
                "LOCALAPPDATA": r"C:\Users\x\AppData\Local"}
        cands = server._chromium_candidates(fake)
        ch = [i for i, p in enumerate(cands) if p.lower().endswith("chrome.exe")]
        ed = [i for i, p in enumerate(cands) if p.lower().endswith("msedge.exe")]
        need(ch and ed and min(ch) < min(ed),
             "candidates must list Chrome before Edge")
        eq(server._find_chromium(exists=lambda p: True, env=fake).lower()
           .endswith("chrome.exe"), True, "find prefers Chrome when present")
        eq(server._find_chromium(exists=lambda p: "Edge" in p, env=fake).lower()
           .endswith("msedge.exe"), True, "find falls back to Edge")
        eq(server._find_chromium(exists=lambda p: False, env=fake), None,
           "find returns None when no browser exists")

        # _launch_browser_window opens a PLAIN window: argv is exactly
        # [exe, url], with none of the old chromeless app-window flags.
        orig_find = server._find_chromium
        orig_popen = server.subprocess.Popen
        captured = {"argv": None}
        server._find_chromium = lambda *a, **k: r"C:\chrome.exe"
        server.subprocess.Popen = lambda argv, *a, **k: (
            captured.__setitem__("argv", list(argv)) or object())
        try:
            eq(server._launch_browser_window("http://127.0.0.1:8765/"), True,
               "launch returns True when a browser is found")
            eq(captured["argv"], [r"C:\chrome.exe", "http://127.0.0.1:8765/"],
               "browser launch is a plain [exe, url] window")
            for flag in ("--app=", "--user-data-dir=",
                         "--hide-crash-restore-bubble"):
                need(not any(str(a).startswith(flag) for a in captured["argv"]),
                     "no chromeless app-window flag survives: %s" % flag)
        finally:
            server._find_chromium = orig_find
            server.subprocess.Popen = orig_popen

        # No Chromium browser found -> launch reports False (caller falls back).
        server._find_chromium = lambda *a, **k: None
        try:
            eq(server._launch_browser_window("http://h/"), False,
               "launch returns False when no browser is found")
        finally:
            server._find_chromium = orig_find

        # The source carries the new launcher and none of the app-window bits.
        srv = open(os.path.join(BACKEND, "server.py"), encoding="utf-8").read()
        need("def _launch_browser_window" in srv,
             "the plain browser-window launcher is present")
        for gone in ("_app_argv", "_ui_profile_dir", "_launch_app_window",
                     "--user-data-dir", "chromeapp"):
            need(gone not in srv,
                 "the chromeless app-window machinery is gone: %s" % gone)

        # Mode routing. 'browser' forces the system default (webbrowser.open)
        # and must NOT launch a Chromium window.
        orig_launch = server._launch_browser_window
        orig_open = server.webbrowser.open
        opened = {"u": None}
        server._launch_browser_window = lambda u: (_ for _ in ()).throw(
            AssertionError("browser mode must use the system default browser"))
        server.webbrowser.open = lambda u, *a, **k: opened.__setitem__("u", u)
        try:
            r = server._open_window_or_browser("http://h/", "browser")
            eq(r, None, "browser mode returns no window handle")
            eq(opened["u"], "http://h/", "browser mode opens the default browser")
        finally:
            server._launch_browser_window = orig_launch
            server.webbrowser.open = orig_open

        # 'auto' with a browser present opens it and waits on nothing.
        opened = {"u": None, "fallback": None}
        server._launch_browser_window = lambda u: (
            opened.__setitem__("u", u) or True)
        server.webbrowser.open = lambda u, *a, **k: opened.__setitem__(
            "fallback", u)
        try:
            r = server._open_window_or_browser("http://h/", "auto")
            eq(r, None, "auto mode returns no window handle")
            eq(opened["u"], "http://h/", "auto opens a Chrome/Edge window")
            eq(opened["fallback"], None,
               "auto does not also open the default browser when Chrome is up")
        finally:
            server._launch_browser_window = orig_launch
            server.webbrowser.open = orig_open

        # 'auto' with NO browser present falls back to the system default.
        opened = {"u": None}
        server._launch_browser_window = lambda u: False
        server.webbrowser.open = lambda u, *a, **k: opened.__setitem__("u", u)
        try:
            r = server._open_window_or_browser("http://h/", "auto")
            eq(r, None, "auto fallback returns no handle")
            eq(opened["u"], "http://h/", "auto falls back to the default browser")
        finally:
            server._launch_browser_window = orig_launch
            server.webbrowser.open = orig_open

        # 'window' opens a pywebview window when present, else the browser. No
        # pywebview is installed in this harness, so it must fall back cleanly.
        opened = {"u": None}
        server.webbrowser.open = lambda u, *a, **k: opened.__setitem__("u", u)
        try:
            r = server._open_window_or_browser("http://h/", "window")
            if r is None:
                eq(opened["u"], "http://h/",
                   "window mode without pywebview falls back to the browser")
            else:
                eq(r[0], "webview", "window mode returns a pywebview handle")
        finally:
            server.webbrowser.open = orig_open

    def t_chrome_and_window_launch_paths():
        # Dual start paths: Chrome/browser via server.py --browser (and the
        # default auto Chrome preference), and the native window app via
        # --window / Start-SamQL-AppWindow.ps1 / launcher_app.py. Flag parsing
        # + both entrypoints must stay wired; full pywebview UI is covered by
        # backend launcher tests with a fake webview.
        p = server._build_arg_parser()
        eq(server._ui_mode(p.parse_args(["--browser"])), "browser",
           "--browser forces browser UI mode")
        eq(server._ui_mode(p.parse_args(["--window"])), "window",
           "--window forces native window UI mode")
        eq(server._ui_mode(p.parse_args(["--no-browser"])), "none",
           "--no-browser serves headlessly")
        eq(server._ui_mode(p.parse_args([])), "auto",
           "default mode is auto (prefer Chrome/Edge window)")

        # --browser must not open a pywebview handle; --window may.
        opened = {"u": None}
        orig_open = server.webbrowser.open
        orig_launch = server._launch_browser_window
        server.webbrowser.open = lambda u, *a, **k: opened.__setitem__("u", u)
        server._launch_browser_window = lambda u: (_ for _ in ()).throw(
            AssertionError("--browser must use webbrowser.open, not Chrome Popen"))
        try:
            r = server._open_window_or_browser("http://launch-browser/", "browser")
            eq(r, None, "browser launch path returns no window handle")
            eq(opened["u"], "http://launch-browser/",
               "browser launch path opens the system browser")
        finally:
            server.webbrowser.open = orig_open
            server._launch_browser_window = orig_launch

        aw = os.path.join(ROOT, "Start-SamQL-AppWindow.ps1")
        la = os.path.join(BACKEND, "launcher_app.py")
        need(os.path.isfile(aw), "Start-SamQL-AppWindow.ps1 is the window-app entry")
        need(os.path.isfile(la), "launcher_app.py backs the window-app path")
        ps = open(aw, encoding="utf-8").read()
        lap = open(la, encoding="utf-8").read()
        srv = open(os.path.join(BACKEND, "server.py"), encoding="utf-8").read()
        need("def _ui_mode" in srv and "--browser" in srv and "--window" in srv,
             "server.py exposes --browser and --window launch flags")
        need("SamQL-AppWindow" in ps and "launcher_app.py" in ps,
             "AppWindow script targets the native launcher")
        need("def main" in lap and "webview" in lap,
             "launcher_app drives the pywebview window app")
        # Chrome preference is shared with the browser-window launcher test.
        need("chrome" in ps.lower() or "Browser" in ps,
             "AppWindow script can prefer Chrome when falling back")

    def t_errors_log_http():
        # The error log captures server-side failures with debuggable detail
        # and can be read + cleared over the API. Clear first, trigger a known
        # handled error (unknown load job -> 404), confirm it was recorded.
        c.js("DELETE", "/api/errors")
        st, d0 = c.js("GET", "/api/errors")
        eq(st, 200, "GET /api/errors status")
        need("errors" in d0 and "version" in d0, "errors payload shape")
        c.js("GET", "/api/load/progress/not-a-real-job-xyz")
        st, d = c.js("GET", "/api/errors")
        eq(st, 200, "GET /api/errors after trigger")
        errs = d.get("errors") or []
        hit = [e for e in errs
               if e.get("status") == 404
               and "/api/load/progress/" in (e.get("path") or "")]
        need(hit, "the 404 was logged: %s" % errs[:2])
        e0 = hit[0]
        need(e0.get("ts") and e0.get("error"),
             "log entry carries a timestamp + message")
        need(e0.get("method") == "GET", "log entry records the HTTP method")
        # Soft application errors (HTTP 200 + {error}) also land in the log.
        st, bad = c.js("POST", "/api/query",
                       {"sql": "SELECT * FROM __samql_no_such_table_xyz__",
                        "target": "duckdb"})
        eq(st, 200, "bad query still returns HTTP 200")
        need(bad.get("error"), "bad query payload carries error")
        st, d3 = c.js("GET", "/api/errors")
        soft = [e for e in (d3.get("errors") or [])
                if e.get("kind") == "QueryError"
                and "/api/query" in (e.get("path") or "")]
        need(soft, "soft query errors are logged: %s"
             % [(e.get("kind"), e.get("error"))
                for e in (d3.get("errors") or [])[:5]])
        need(soft[0].get("detail"), "soft query log keeps request detail")
        st, dc = c.js("DELETE", "/api/errors")
        eq(st, 200, "DELETE /api/errors status")
        st, d2 = c.js("GET", "/api/errors")
        eq(len(d2.get("errors") or []), 0, "error log cleared")

    def t_graceful_shutdown_cleans_temp():
        # The graceful-shutdown handler must (a) remove THIS instance's temp
        # directory and (b) sweep directories left by previous instances that
        # were killed before cleaning up -- so temp dirs never accumulate across
        # runs. Exercise the real handler with its destructive steps stubbed
        # (no server/session teardown, no real file removal), then restore.
        import samql_core.tmputil as _tu
        calls = {"cleanup": 0, "sweep": 0}
        orig_cleanup, orig_sweep = _tu.cleanup_instance, _tu.sweep_stale
        orig_httpd, orig_session = server._HTTPD, server.SESSION
        orig_done = server._SHUTDOWN_DONE
        try:
            _tu.cleanup_instance = lambda: calls.__setitem__(
                "cleanup", calls["cleanup"] + 1)
            _tu.sweep_stale = lambda: calls.__setitem__(
                "sweep", calls["sweep"] + 1)
            server._HTTPD = None       # skip the real server stop
            server.SESSION = None      # skip the real session shutdown
            server._SHUTDOWN_DONE = False
            server._graceful_shutdown()
            eq(calls["cleanup"], 1, "shutdown removes this instance's temp dir")
            eq(calls["sweep"], 1,
               "shutdown sweeps stale temp dirs left by prior runs")
        finally:
            _tu.cleanup_instance, _tu.sweep_stale = orig_cleanup, orig_sweep
            server._HTTPD, server.SESSION = orig_httpd, orig_session
            server._SHUTDOWN_DONE = orig_done

    def t_nodeflow_run_http():
        # the core flow-run route over HTTP -- exercises body parsing
        # (graph/node/port) and the response shape, not just the Session method.
        # Self-contained (literal SQL) so it doesn't depend on test ordering.
        g = {"nodes": [
                {"id": "src", "type": "sql", "config": {
                    "sql": "SELECT 1 AS id, 'a' AS v UNION ALL "
                           "SELECT 2,'b' UNION ALL SELECT 3,'c'"}},
                {"id": "o", "type": "output", "config": {"label": "out"}}],
             "edges": [{"from": {"node": "src", "port": "out"},
                        "to": {"node": "o", "port": "in"}}]}
        st, d = c.js("POST", "/api/nodeflow/run",
                     {"graph": g, "node": "o", "port": "out"})
        eq(st, 200, "nodeflow/run status")
        need(not d.get("error"), "nodeflow/run over HTTP: %s" % d.get("error"))
        eq(d.get("total_rows"), 3, "nodeflow/run returned rows over HTTP")
        st, d = c.js("POST", "/api/nodeflow/run", {
            "graph": g, "node": "o", "port": "out",
            "preview": True, "preview_limit": 2,
        })
        eq(st, 200, "nodeflow preview status")
        eq(d.get("total_rows"), 2, "nodeflow preview is bounded")
        need(d.get("preview") is True and d.get("preview_limit") == 2,
             "nodeflow preview metadata returned over HTTP")

    def t_nodeflow_run_batch_http():
        # Multi-target Run all route. The scheduler may group shared branches
        # or parallelise disjoint DuckDB branches, but the HTTP contract always
        # returns one ordinary result envelope per requested terminal.
        g = {"nodes": [
                {"id": "a", "type": "sql", "config": {
                    "sql": "SELECT 1 AS id UNION ALL SELECT 2"}},
                {"id": "ao", "type": "output", "config": {"label": "A"}},
                {"id": "b", "type": "sql", "config": {
                    "sql": "SELECT 10 AS id UNION ALL SELECT 20 UNION ALL SELECT 30"}},
                {"id": "bo", "type": "output", "config": {"label": "B"}}],
             "edges": [
                {"from": {"node": "a", "port": "out"},
                 "to": {"node": "ao", "port": "in"}},
                {"from": {"node": "b", "port": "out"},
                 "to": {"node": "bo", "port": "in"}}]}
        st, d = c.js("POST", "/api/nodeflow/run-batch", {
            "graph": g,
            "requests": [{"node": "ao", "port": "out"},
                         {"node": "bo", "port": "out"}],
        })
        eq(st, 200, "nodeflow/run-batch status")
        need(d.get("ok") is True and len(d.get("results") or []) == 2,
             "nodeflow/run-batch result shape: %r" % d)
        by_node = {r.get("node"): r for r in d.get("results") or []}
        eq(by_node["ao"].get("total_rows"), 2, "first batch branch rows")
        eq(by_node["bo"].get("total_rows"), 3, "second batch branch rows")

    def t_nodeflow_export_http():
        # the single-node export route over HTTP (sibling of the export-many
        # route already covered, and of the export-node bug). Writes a real file.
        import tempfile
        import os as _os
        g = {"nodes": [
                {"id": "src", "type": "sql", "config": {
                    "sql": "SELECT 1 AS id, 'a' AS v UNION ALL "
                           "SELECT 2,'b' UNION ALL SELECT 3,'c'"}},
                {"id": "o", "type": "output", "config": {"label": "out"}}],
             "edges": [{"from": {"node": "src", "port": "out"},
                        "to": {"node": "o", "port": "in"}}]}
        outd = tempfile.mkdtemp(prefix="t_http_exp_")
        st, d = c.js("POST", "/api/nodeflow/export",
                     {"graph": g, "node": "o", "out_dir": outd,
                      "fmt": "csv", "base_name": "exp"})
        eq(st, 200, "nodeflow/export status")
        need(not d.get("error"), "nodeflow/export over HTTP: %s" % d.get("error"))
        eq(d.get("rows"), 3, "export reported every row over HTTP")
        need(_os.path.isfile(_os.path.join(outd, "exp.csv")),
             "export wrote the file over HTTP")

    def t_iterator_run_http():
        # the iterator controller route over HTTP. while/run shares this exact
        # wiring ({graph, node_id, query_id} -> run_while), so this also covers
        # that route's request/response contract.
        g = {"nodes": [
                {"id": "body", "type": "sql",
                 "config": {"sql": "SELECT ${i} AS n"}},
                {"id": "it", "type": "iterator", "config": {
                    "var": "i", "table": "acc_http",
                    "driver": {"kind": "list", "values": ["1", "2", "3"]}}}],
             "edges": [{"from": {"node": "body", "port": "out"},
                        "to": {"node": "it", "port": "in"}}]}
        st, d = c.js("POST", "/api/iterator/run",
                     {"graph": g, "node_id": "it"})
        eq(st, 200, "iterator/run status")
        need(not d.get("error"), "iterator/run over HTTP: %s" % d.get("error"))
        eq(d.get("passes"), 3, "iterator/run reports passes over HTTP")
        eq(d.get("rows"), 3, "iterator accumulator rows over HTTP")

    def t_secrets_http():
        # the saved-credential store over HTTP. The store needs an OS keystore
        # (DPAPI on Windows); where that's absent, set returns ok=False but the
        # routes must still answer with the right shape. Assert that contract on
        # both, keyed off the reported availability.
        st, d = c.js("POST", "/api/secrets/available", {})
        eq(st, 200, "secrets/available status")
        need("available" in d, "available shape: %s" % d)
        avail = bool(d["available"])
        st, d = c.js("POST", "/api/secrets/status", {"keys": ["__t_http_k"]})
        eq(st, 200, "secrets/status status")
        need("available" in d and "saved" in d, "status shape: %s" % d)
        st, d = c.js("POST", "/api/secrets/set",
                     {"key": "__t_http_k", "value": "v"})
        eq(st, 200, "secrets/set status")
        eq(bool(d.get("ok")), avail, "set ok tracks store availability")
        st, d = c.js("POST", "/api/secrets/status", {"keys": ["__t_http_k"]})
        eq(bool(d["saved"].get("__t_http_k")), avail,
           "status reflects a stored key only when the store is available")
        st, d = c.js("POST", "/api/secrets/delete", {"key": "__t_http_k"})
        eq(st, 200, "secrets/delete status")
        need("ok" in d, "delete returns ok over HTTP")

    def t_connection_profiles_http():
        # Named connection profiles over HTTP: fields in the registry, password
        # only via the DPAPI secret store (never in the profile JSON).
        key = "api:__t_http_profile"
        st, d = c.js("POST", "/api/connection-profiles/list", {})
        eq(st, 200, "connection-profiles/list status")
        need("profiles" in d and "secrets_available" in d,
             "list shape: %s" % d)
        st, d = c.js("POST", "/api/connection-profiles/upsert", {
            "kind": "api",
            "name": "__t_http_profile",
            "fields": {"url": "https://example.test/data", "auth_user": "u"},
            "password": "pw-http-test",
        })
        eq(st, 200, "connection-profiles/upsert status")
        need(d.get("ok") is True and (d.get("profile") or {}).get("key") == key,
             "upsert returns profile key: %s" % d)
        avail = bool(d.get("secrets_available"))
        if avail:
            need(d.get("has_secret") is True,
                 "password stored when secrets available")
        st, d = c.js("POST", "/api/connection-profiles/get", {"key": key})
        eq(st, 200, "connection-profiles/get status")
        fields = (d.get("profile") or {}).get("fields") or {}
        eq(fields.get("url"), "https://example.test/data", "get returns fields")
        need("pw-http-test" not in str(d),
             "get payload never includes the password")
        st, d = c.js("POST", "/api/connection-profiles/list", {})
        names = [p.get("key") for p in (d.get("profiles") or [])]
        need(key in names, "list includes the upserted profile")
        st, d = c.js("POST", "/api/connection-profiles/delete", {"key": key})
        eq(st, 200, "connection-profiles/delete status")
        need(d.get("ok") is True, "delete ok")
        st, d = c.js("POST", "/api/connection-profiles/get", {"key": key})
        eq(st, 404, "get after delete is 404")

    def t_nodeflow_validate_http():
        g = {"nodes": [
                {"id": "s", "type": "sql", "config": {
                    "sql": "SELECT 1 AS id, 5 AS v UNION ALL SELECT 2, 9"}},
                {"id": "v", "type": "validate", "config": {}}],
             "edges": [{"from": {"node": "s", "port": "out"},
                        "to": {"node": "v", "port": "in"}}]}
        st, d = c.js("POST", "/api/nodeflow/validate",
                     {"graph": g, "node": "v",
                      "checks": [{"type": "rows_min", "n": 2},
                                 {"type": "rows_max", "n": 2}]})
        eq(st, 200, "nodeflow/validate status")
        need(not d.get("error"), "validate over HTTP: %s" % d.get("error"))
        eq(d.get("total_rows"), 2, "validate counted rows over HTTP")
        by = {(r["type"], r["target"]): r["pass"] for r in d["results"]}
        eq(by[("rows_min", "2")], True, "rows_min passes over HTTP")
        eq(by[("rows_max", "2")], True, "rows_max passes (2 <= 2) over HTTP")

    def t_nodeflow_chart_http():
        g = {"nodes": [
                {"id": "s", "type": "sql", "config": {
                    "sql": "SELECT 'east' AS region, 10 AS amount UNION ALL "
                           "SELECT 'west', 20 UNION ALL SELECT 'east', 5"}},
                {"id": "ch", "type": "chart", "config": {"label": "ch"}}],
             "edges": [{"from": {"node": "s", "port": "out"},
                        "to": {"node": "ch", "port": "in"}}]}
        st, d = c.js("POST", "/api/nodeflow/chart",
                     {"graph": g, "node": "ch",
                      "spec": {"chart_type": "bar", "x": "region",
                               "y": "amount", "agg": "sum"}})
        eq(st, 200, "nodeflow/chart status")
        need(not d.get("error"), "chart over HTTP: %s" % d.get("error"))
        need(any(k in d for k in ("rows", "data", "series", "labels", "spec")),
             "chart returned a payload over HTTP: %s" % list(d)[:6])

    def t_nodeflow_reconcile_http():
        g = {"nodes": [
                {"id": "l", "type": "sql", "config": {
                    "sql": "SELECT 'a' AS k, 10 AS amt UNION ALL SELECT 'b', 20"}},
                {"id": "r", "type": "sql", "config": {
                    "sql": "SELECT 'a' AS k, 10 AS amt UNION ALL SELECT 'b', 25"}},
                {"id": "rc", "type": "reconcile", "config": {"label": "rc"}}],
             "edges": [{"from": {"node": "l", "port": "out"},
                        "to": {"node": "rc", "port": "left"}},
                       {"from": {"node": "r", "port": "out"},
                        "to": {"node": "rc", "port": "right"}}]}
        st, d = c.js("POST", "/api/nodeflow/reconcile",
                     {"graph": g, "node": "rc", "keys": ["k"],
                      "compare": ["amt"]})
        eq(st, 200, "nodeflow/reconcile status")
        need(not d.get("error"), "reconcile over HTTP: %s" % d.get("error"))
        need(any(k in d for k in ("totals", "summary", "rows", "result")),
             "reconcile returned a result over HTTP: %s" % list(d)[:6])

    def t_nodeflow_browse_http():
        g = {"nodes": [
                {"id": "s", "type": "sql", "config": {
                    "sql": "SELECT 1 AS id, 'x' AS v UNION ALL SELECT 2, 'y'"}},
                {"id": "b", "type": "browse", "config": {"label": "b"}}],
             "edges": [{"from": {"node": "s", "port": "out"},
                        "to": {"node": "b", "port": "in"}}]}
        st, d = c.js("POST", "/api/nodeflow/browse", {"graph": g, "node": "b"})
        eq(st, 200, "nodeflow/browse status")
        need(not d.get("error"), "browse over HTTP: %s" % d.get("error"))
        need(any(k in d for k in ("columns", "rows", "profile", "fields")),
             "browse returned a column profile over HTTP: %s" % list(d)[:6])

    def t_api_preview_http():
        if not online:
            skip("network test (run with --online)")
        st, d = c.js("POST", "/api/api-preview",
                     {"url": "https://jsonplaceholder.typicode.com/users"}, 30)
        eq(st, 200, "api-preview status")
        need(not d.get("error"), "api-preview over HTTP: %s" % d.get("error"))

    def t_excel_sheets_http():
        if not feats.get("openpyxl"):
            skip("openpyxl not installed")
        import openpyxl as _xl
        import tempfile
        import os as _os
        d = tempfile.mkdtemp(prefix="t_http_xls_")
        p = _os.path.join(d, "book.xlsx")
        wb = _xl.Workbook()
        wb.active.title = "Alpha"
        wb.create_sheet("Beta")
        wb.save(p)
        wb.close()
        st, dd = c.js("POST", "/api/excel/sheets", {"path": p})
        eq(st, 200, "excel/sheets status")
        eq(dd.get("sheets"), ["Alpha", "Beta"],
           "excel/sheets lists a disk workbook's sheets over HTTP")

    def t_excel_peek_http():
        # multipart upload -> sheet list, the drag-and-drop sheet picker's route.
        # The JSON client can't do multipart, so build the body by hand.
        if not feats.get("openpyxl"):
            skip("openpyxl not installed")
        import openpyxl as _xl
        import io as _io
        import urllib.request as _u
        wb = _xl.Workbook()
        wb.active.title = "Sheet1"
        wb.create_sheet("Sheet2")
        buf = _io.BytesIO()
        wb.save(buf)
        wb.close()
        boundary = "----samqlpeekBOUNDARY"
        pre = ("--%s\r\n"
               'Content-Disposition: form-data; name="file"; '
               'filename="b.xlsx"\r\n'
               "Content-Type: application/octet-stream\r\n\r\n") % boundary
        body = (pre.encode("latin-1") + buf.getvalue()
                + ("\r\n--%s--\r\n" % boundary).encode("latin-1"))
        rq = _u.Request(f"http://{host}:{port}/api/excel/peek",
                        data=body, method="POST")
        rq.add_header("Content-Type",
                      "multipart/form-data; boundary=" + boundary)
        rq.add_header("X-SamQL-Token", c.token)
        r = _u.urlopen(rq, timeout=30)
        dd = json.loads(r.read())
        eq(r.status, 200, "excel/peek status")
        eq(dd.get("sheets"), ["Sheet1", "Sheet2"],
           "excel/peek lists the uploaded workbook's sheets over HTTP")

    def t_directory_read_http():
        import tempfile
        import os as _os
        d = tempfile.mkdtemp(prefix="t_http_dir_")
        p = _os.path.join(d, "one.csv")
        open(p, "w", newline="").write("id,name\n1,ada\n2,bob\n")
        st, dd = c.js("POST", "/api/directory/read", {"path": p})
        eq(st, 200, "directory/read status")
        need(dd.get("ok"), "directory/read ok over HTTP: %s" % dd.get("error"))
        eq(dd.get("rows"), 2, "directory/read loaded the file over HTTP")

    def t_folder_read_http():
        import tempfile
        import os as _os
        d = tempfile.mkdtemp(prefix="t_http_fld_")
        open(_os.path.join(d, "a.csv"), "w", newline="").write(
            "id,amt\n1,10\n2,20\n")
        open(_os.path.join(d, "b.csv"), "w", newline="").write(
            "id,amt\n3,30\n")
        st, dd = c.js("POST", "/api/folder/read", {"folder": d})
        eq(st, 200, "folder/read status")
        need(dd.get("ok"), "folder/read ok over HTTP: %s" % dd.get("error"))
        eq(dd.get("files"), 2, "folder/read stacked both files over HTTP")
        eq(dd.get("rows"), 3, "folder/read stacked all rows over HTTP")

    def t_status_endpoint():
        st, d = c.js("GET", "/api/status")
        eq(st, 200, "status code")
        for k in ("operations", "engines", "threads", "restoring",
                  "concurrent_reads"):
            need(k in d, "/api/status missing %s" % k)
        need(isinstance(d["operations"], list), "operations must be a list")
        need("sqlite" in d["engines"] and "duckdb" in d["engines"],
             "engines must list sqlite + duckdb")
        need(isinstance(d["threads"], int) and d["threads"] >= 1,
             "threads must be a positive int")
        need(d["engines"]["sqlite"]["busy"] is False, "idle sqlite not busy")
        # busy must flip while another thread holds the engine connection lock
        grab, rel = threading.Event(), threading.Event()

        def holder():
            with server.SESSION.db.write_lock:
                grab.set()
                rel.wait(2.0)
        t = threading.Thread(target=holder)
        t.start()
        need(grab.wait(2.0), "holder never acquired the lock")
        _, d2 = c.js("GET", "/api/status")
        rel.set()
        t.join()
        need(d2["engines"]["sqlite"]["busy"] is True,
             "status must report sqlite busy while the lock is held")

    def t_status_last_error():
        # The heartbeat monitor turns red on a recent failure, so /api/status
        # must surface the most recent server error (with an age the UI uses to
        # decide "recent") and a readable route.
        server.log_server_error("POST", "/api/query", 500, "binder",
                                "BinderException: monitor test failure")
        _, d = c.js("GET", "/api/status")
        le = d.get("last_error")
        need(le and "monitor test failure" in le.get("error", ""),
             "status must surface the most recent server error")
        need(isinstance(le.get("age_s"), (int, float)),
             "last_error needs an age so the UI can flag recency")
        eq(le.get("route"), "POST /api/query", "route is recorded readably")

    def t_concurrent_reads_endpoint():
        # POST {} reports state; POST {on: ...} toggles it; /api/status reflects
        # it. Reset to off at the end so the shared session stays clean.
        try:
            st, d = c.js("POST", "/api/settings/concurrent-reads", {})
            eq(st, 200, "status code")
            need("concurrent_reads" in d, "endpoint reports the flag")
            _, d = c.js("POST", "/api/settings/concurrent-reads", {"on": True})
            need(d.get("concurrent_reads") is True, "flag turns on")
            _, sd = c.js("GET", "/api/status")
            need(sd.get("concurrent_reads") is True,
                 "/api/status reflects the flag on")
            _, d = c.js("POST", "/api/settings/concurrent-reads",
                        {"on": False})
            need(d.get("concurrent_reads") is True,
                 ".426: the opt-out is accepted and CLAMPED -- always on")
            _, sd = c.js("GET", "/api/status")
            need(sd.get("concurrent_reads") is True,
                 "/api/status stays on after an off attempt")
        finally:
            try:
                c.js("POST", "/api/settings/concurrent-reads", {"on": False})
            except Exception:
                pass

    def t_flow_cache_settings_endpoint():
        sess = server.SESSION
        old_enabled = bool(sess.flow_cache)
        old_entries = int(sess.flow_cache_max)
        old_mb = int(sess.flow_cache_bytes_max // (1024 * 1024))
        try:
            st, info = c.js("GET", "/api/settings/flow-cache")
            eq(st, 200, "flow-cache settings GET status")
            for key in ("enabled", "size", "max", "bytes", "bytes_max",
                        "hits", "misses", "evictions", "largest"):
                need(key in info, "flow-cache telemetry missing " + key)
            st, changed = c.js("POST", "/api/settings/flow-cache", {
                "enabled": True, "max_entries": 7, "max_mb": 16,
            })
            eq(st, 200, "flow-cache settings POST status")
            need(changed.get("enabled") is True, "cache remains enabled")
            eq(changed.get("max"), 7, "entry limit applies live")
            eq(changed.get("mb_max"), 16.0, "byte limit applies live")
            st, bad = c.js("POST", "/api/settings/flow-cache",
                           {"max_entries": "many"})
            eq(st, 400, "junk cache limit is a clean client error")
            need(bad.get("error"), "junk cache limit explains the error")
            st, reset = c.js("POST", "/api/settings/flow-cache", {
                "clear": True, "reset_stats": True,
            })
            eq(st, 200, "cache clear/reset status")
            eq(reset.get("size"), 0, "cache clear is immediate")
            eq(reset.get("hits"), 0, "counter reset is immediate")
        finally:
            sess.configure_flow_cache(enabled=old_enabled,
                                      max_entries=old_entries, max_mb=old_mb)

    def t_load_thresholds_settings_endpoint():
        sess = server.SESSION
        prev = dict(sess.config.get("load_thresholds") or {})
        try:
            st, info = c.js("GET", "/api/settings/load-thresholds")
            eq(st, 200, "load-thresholds GET status")
            need(isinstance(info.get("thresholds"), dict),
                 "thresholds map present")
            need("ondisk_mb" in info["thresholds"], "ondisk_mb field")
            st, changed = c.js("POST", "/api/settings/load-thresholds", {
                "thresholds": {"ondisk_mb": 77, "filecache_gb": 12},
            })
            eq(st, 200, "load-thresholds POST status")
            eq(changed["thresholds"]["ondisk_mb"]["value"], 77,
               "ondisk applied via API")
            eq(changed["thresholds"]["ondisk_mb"]["source"], "override",
               "source is override")
            st, reset = c.js("POST", "/api/settings/load-thresholds",
                             {"reset": True})
            eq(st, 200, "load-thresholds reset status")
            need(reset["thresholds"]["ondisk_mb"]["source"] in
                 ("default", "env"),
                 "reset drops override source")
        finally:
            if prev:
                sess.configure_load_thresholds(updates=prev)
            else:
                sess.configure_load_thresholds(reset=True)

    def t_load_preflight_endpoint():
        st, miss = c.js("POST", "/api/load/preflight", {"path": ""})
        eq(st, 200, "preflight empty path status")
        need(not miss.get("ok") and miss.get("blockers"),
             "empty path blocked: %r" % miss)
        # Real tiny file — should be ok with optional warnings only.
        p = os.path.join(tempfile.gettempdir(), "samql_preflight_api.csv")
        with open(p, "w", encoding="utf-8") as f:
            f.write("a,b\n1,2\n")
        try:
            st, ok = c.js("POST", "/api/load/preflight", {"path": p})
            eq(st, 200, "preflight existing file status")
            need(ok.get("exists") and ok.get("ok"),
                 "tiny csv preflight ok: %r" % ok)
            need("warnings" in ok and "blockers" in ok,
                 "shape includes warnings/blockers")
        finally:
            try:
                os.unlink(p)
            except OSError:
                pass

    def t_engine_reset_endpoint():
        # Point the server at a throwaway session so the reset (which wipes +
        # rebuilds engines) can't disturb the shared suite session. Verifies the
        # endpoint contract AND an end-to-end rebuild from the manifest.
        import time as _t
        from samql_core import Session
        orig = server.SESSION
        s = Session()
        try:
            server.SESSION = s
            s.load_file(CSV, destination="sqlite", base_name="rtbl")
            s.record_load("file", CSV, destination="sqlite", base_name="rtbl")
            need(any("rtbl" in n for n in s.db.table_columns),
                 "precondition: table loaded")
            old_db = s.db
            st, d = c.js("POST", "/api/engine/reset", {})
            eq(st, 200, "reset status")
            need(d.get("ok") is True, "reset must report ok")
            need("sqlite" in (d.get("reset") or []),
                 "reset must list refreshed engines")
            need(s.db is not old_db, "engine must be a fresh instance")
            for _ in range(80):
                if not s.restoring and list(s.db.table_columns):
                    break
                _t.sleep(0.05)
            need(any("rtbl" in n for n in s.db.table_columns),
                 "reset over HTTP must rebuild the table from the manifest")
            st2, _ = c.js("GET", "/api/health")
            eq(st2, 200, "server must stay alive after an engine reset")
        finally:
            server.SESSION = orig
            try:
                s.duckdb = None
                s.shutdown()
            except Exception:
                pass

    def t_concurrent_loads():
        # two background loads started together both complete -- proves the
        # engine slot serializes them through the real load rail without
        # deadlocking (one runs, the other queues, then advances).
        ids = []
        for _ in range(2):
            st, d = c.js("POST", "/api/load/start",
                         {"path": CSV, "destination": "sqlite"})
            eq(st, 200, "concurrent load start ok")
            ids.append(d.get("job_id") or d.get("job"))
        need(all(ids), "both loads got job ids")
        for jb in ids:
            fin = None
            for _ in range(600):
                st, pr = c.js("GET", "/api/load/progress/%s" % jb)
                if pr.get("state") in ("done", "error"):
                    fin = pr.get("state")
                    break
                time.sleep(0.02)
            eq(fin, "done", "load %s finished without deadlock" % jb)

    def t_tasks_feed():
        # the unified activity feed: a started load shows up as a normalized
        # card (kind + coarse state + honest progress mode). This is the same
        # list the tray polls in place of the per-task progress modals.
        st, d = c.js("GET", "/api/tasks")
        eq(st, 200, "tasks feed status")
        need(isinstance(d.get("tasks"), list), "tasks is a list")
        st, d = c.js("POST", "/api/load/start",
                     {"path": CSV, "destination": "sqlite"})
        eq(st, 200, "load start status")
        job = d.get("job_id") or d.get("job")
        need(job, "no job id from load start")
        st, d = c.js("GET", "/api/tasks")
        eq(st, 200, "tasks feed status after start")
        crd = next((t for t in d["tasks"] if t.get("id") == job), None)
        need(crd, "the started load appears as a card")
        eq(crd["kind"], "load", "card is tagged kind=load")
        need(crd["state"] in ("running", "done"), "card has a coarse state")
        need(crd["progress"]["mode"] in ("bytes", "spinner"),
             "a sized load shows a bytes bar once it advances (spinner before)")
        eq(crd["cancellable"], crd["state"] != "done",
           "running cards cancellable, finished ones not")
        need("title" in crd and "phase" in crd, "card carries title + phase")

    def t_optimize_no_duckdb():
        # the sandbox server has no DuckDB, so converting a table to Parquet
        # must fail cleanly with a 400 (never a 500) and name the engine. On a
        # DuckDB box the named table just doesn't exist -> still a clean refusal.
        st, d = c.js("POST", "/api/table/optimize-start", {"name": "whatever"})
        if feats.get("duckdb"):
            need(st in (200, 400), "optimize-start responds cleanly")
        else:
            eq(st, 400, "no DuckDB -> a clean 400 (not a 500)")
            need("DuckDB" in (d.get("error") or ""),
                 "the 400 names the DuckDB requirement")
        # the server is still responsive afterwards
        st2, _h = c.js("GET", "/api/health")
        eq(st2, 200, "server still responsive after optimize-start")

    try:
        cases = [
            ("GET /api/health", t_health),
            ("request boundary rejects hostile Origin + Host",
             t_request_boundary_security),
            ("API token + JSON request ceiling",
             t_api_token_and_json_limit),
            ("static serving rejects traversal + symlink escapes",
             t_static_path_containment),
            ("unknown multipart route reclaims its spool",
             t_unmatched_multipart_cleanup),
            ("GET /api/health app marker", t_health_app_marker),
            ("GET /favicon.ico + /app-icon.png + manifest (browser tab icon)",
             t_brand_favicon),
            ("POST /api/focus", t_focus_endpoint),
            ("single-instance window helpers", t_window_helpers_safe),
            ("single-instance second launch attaches", t_single_instance_attaches),
            ("browser-window launcher (detect + launch + mode routing)",
             t_browser_open_mode),
            ("Chrome/browser and window-app launch paths",
             t_chrome_and_window_launch_paths),
            ("GET /api/features", t_features),
            ("GET /api/fs/list", t_fs_list),
            ("Wave 1 stateful load/query/filter/export/recovery rail",
             t_wave1_stateful_server_rail),
            ("POST /api/load/files-start + cancel (drag-drop background load)", t_load_files_start_cancel),
            ("malformed upload returns 400, never drops the connection", t_malformed_upload_no_connection_drop),
            ("GET /api/status (activity dashboard: shape + busy flag)", t_status_endpoint),
            ("GET /api/status (surfaces last error for the monitor)", t_status_last_error),
            ("POST /api/engine/reset (recovery: payload + server stays up)", t_engine_reset_endpoint),
            ("POST /api/settings/concurrent-reads (toggle + status reflects it)", t_concurrent_reads_endpoint),
            ("GET/POST /api/settings/flow-cache (telemetry + live limits)",
             t_flow_cache_settings_endpoint),
            ("GET/POST /api/settings/load-thresholds (load size gates)",
             t_load_thresholds_settings_endpoint),
            ("POST /api/load/preflight (large-file checklist)",
             t_load_preflight_endpoint),
            ("POST /api/flatten/start + progress", t_flatten_progress),
            ("GET /api/tables", t_tables),
            ("GET /api/load/jobs (reattach list shape)", t_load_jobs_list),
            ("POST /api/query", t_query),
            ("POST /api/query (fast preview)", t_query_preview),
            ("gzip large response", t_gzip),
            ("POST /api/result/{id}/page", t_page),
            ("POST /api/result/{id}/page (filters)", t_filter),
            ("POST /api/result/{id}/page (projection)", t_page_projection),
            ("POST /api/query/{id}/cancel", t_cancel_route),
            ("POST /api/result/{id}/export", t_export),
            ("POST /api/table/{name}/profile", t_profile),
            ("POST /api/chart/data", t_chart),
            ("POST /api/pivot", t_pivot),
            ("POST /api/reconcile", t_reconcile),
            ("POST /api/sql/format", t_sql_format),
            ("POST /api/sql/statement-at", t_statement_at),
            ("GET/POST/DELETE /api/saved", t_saved),
            ("GET/POST/DELETE /api/workflows (+load)", t_workflows),
            ("save / open a workflow file on disk", t_workspace_file_http),
            ("POST /api/table/create", t_table_create_http),
            ("POST /api/table/optimize-start (no DuckDB -> 400)",
             t_optimize_no_duckdb),
            ("GET /api/tasks (unified activity feed)", t_tasks_feed),
            ("two concurrent loads both finish (engine slot, no deadlock)",
             t_concurrent_loads),
            ("GET/POST /api/catalog/columns + import", t_catalog_endpoints_http),
            ("POST /api/nodeflow/write", t_nodeflow_write_http),
            ("POST /api/nodeflow/export-many (shared pass)",
             t_nodeflow_export_many_http),
            ("POST /api/nodeflow/columns-batch", t_nodeflow_columns_batch_http),
            ("GET /api/history", t_history),
            ("GET/DELETE /api/errors (error log)", t_errors_log_http),
            ("graceful shutdown cleans + sweeps temp dirs", t_graceful_shutdown_cleans_temp),
            ("GET/POST /api/memory", t_memory),
            ("POST /api/maintenance/sweep-temp", t_sweep_temp),
            ("rename / change-type / drop", t_rename_change_drop),
            ("POST /api/clear", t_clear),
            ("GET /api/mssql/drivers", t_mssql_drivers),
            ("POST /api/mssql/tables (guard)", t_mssql_tables_guard),
            ("POST /api/mssql/import (guard)", t_mssql_import_guard),
            ("POST /api/nodeflow/run (flow over HTTP)", t_nodeflow_run_http),
            ("POST /api/nodeflow/run-batch (Run all over HTTP)",
             t_nodeflow_run_batch_http),
            ("POST /api/nodeflow/export (single, over HTTP)",
             t_nodeflow_export_http),
            ("POST /api/iterator/run (over HTTP)", t_iterator_run_http),
            ("POST /api/secrets/* (store over HTTP)", t_secrets_http),
            ("POST /api/connection-profiles/* (named profiles over HTTP)",
             t_connection_profiles_http),
            ("POST /api/nodeflow/validate (over HTTP)", t_nodeflow_validate_http),
            ("POST /api/nodeflow/chart (over HTTP)", t_nodeflow_chart_http),
            ("POST /api/nodeflow/reconcile (over HTTP)",
             t_nodeflow_reconcile_http),
            ("POST /api/nodeflow/browse (over HTTP)", t_nodeflow_browse_http),
            ("POST /api/api-preview (over HTTP)", t_api_preview_http),
            ("POST /api/excel/sheets (over HTTP)", t_excel_sheets_http),
            ("POST /api/excel/peek (multipart, over HTTP)", t_excel_peek_http),
            ("POST /api/directory/read (over HTTP)", t_directory_read_http),
            ("POST /api/folder/read (over HTTP)", t_folder_read_http),
            ("POST /api/api-fetch", t_api_fetch),
        ]
        from test_wave3_stabilization import http_cases as _wave3_http_cases
        cases.extend(_wave3_http_cases(ROOT, CSV, c.base,
                                       httpd.samql_api_token, skip))
        _only = os.environ.get("SAMQL_ONLY")
        _skip = os.environ.get("SAMQL_SKIP")
        if os.environ.get("SAMQL_GROUP") not in (None, "", "http"):
            return
        for name, fn in cases:
            if _only and _only not in name:
                continue
            if _skip and _skip in name:
                continue
            run("http", name, fn)
    finally:
        try:
            httpd.shutdown()
        except Exception:
            pass
        if getattr(server, "SESSION", None) is not None:
            try:
                server.SESSION.shutdown()
            except Exception:
                pass
