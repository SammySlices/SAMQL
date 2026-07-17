"""Column-level data-flow lineage for node-graph workflows.

Given a node graph, trace each *output* (terminal) column back through the
intermediate steps to the input-file columns that feed it. The result is two
collections:

* ``passthrough`` -- output columns that arrive from an input file unchanged
  (same name, no transform anywhere on the path).
* ``derived``     -- one row per *hop* for every output column that has at
  least one upstream transformation. A hop is ``(source field -> used in)``;
  the chain recurses from the final element down to the input columns, exactly
  like the reference layout: e.g. for ``MaturityDate``::

      Input File   Source Field         Used in              DataElement
      <file>       UniqueID             MaturityDate         MaturityDate
      <file>       AsOfDate             MaturityDate         MaturityDate
      <file>       SourceCurrency       ConvertedCurrency    MaturityDate
      Derived      ConvertedCurrency    MaturityDate         MaturityDate
      <file>       EncumberedFlag       IssueDate            MaturityDate
      Derived      IssueDate            SettlementMechanism  MaturityDate
      Derived      SettlementMechanism  MaturityDate         MaturityDate

``Input File`` is the source file name when the field comes straight off an
input node, or the literal ``"Derived"`` when the field is itself produced by
an upstream step. The derived rows also carry the step/node that did the
derivation, and any element whose path crosses a ``sql`` node is flagged (we
treat ``sql`` as an opaque boundary -- we never decompose hand-written SQL).

The per-node column semantics mirror ``nodeflow.compile``:

* sources (input/directory/appendfolder/createtable) -- every column is a
  source column tagged with the file name;
* ``formula`` -- input columns the formulas do not overwrite pass through; each
  formula column derives from the input columns its expression references;
* ``summarize`` -- group keys pass through, aggregates derive from their col;
* ``select`` -- each kept field passes through unless it is renamed or cast;
* ``renamecols`` -- 1:1 rename (pass-through when the name is unchanged);
* transparent row-ops (filter/sort/sample/unique/dedupe/topn/limit/fill) pass
  every column straight through;
* ``sql`` -- opaque boundary (flagged, not decomposed);
* everything else (joins, unions, explode, pivot, window, ...) is handled
  generically: a column that matches an upstream name passes through from that
  side, an entirely new column is attributed (best-effort) to every upstream
  column with the node's type as the step.
"""

import os
import tempfile

from .nodeflow import upstream, _ident_cols, _snake


_SOURCE_TYPES = {"input", "directory", "appendfolder", "createtable"}
# row operations that never add, drop or rename a column
_TRANSPARENT = {"filter", "sort", "sample", "unique", "dedupe", "topn",
                "limit", "fill"}
_AGG_FUNCS = {"sum", "avg", "min", "max", "count", "countd"}
# leaf node types that are sinks/inspectors, not data the lineage should treat
# as a terminal when no explicit output exists
_NON_TERMINAL = {"output", "samqldash", "write", "chart", "dashboard", "browse",
                 "text", "validate", "reconcile", "filebrowser"}

SQL_NOTE = "This output field may have SQL on top of it."


def _norm(s):
    return str(s or "").strip().lower()


def _nodes_by_id(graph):
    return {n.get("id"): n for n in (graph.get("nodes") or []) if n.get("id")}


def _lookup(info, name):
    """Case-insensitive column lookup into a node-info dict."""
    if name in info:
        return info[name]
    nl = _norm(name)
    for k, v in info.items():
        if _norm(k) == nl:
            return v
    return None


def _step(typ, node):
    """A human label for the step that did a derivation -- the node type plus
    its label when the user named the node."""
    cfg = node.get("config") or {}
    label = (cfg.get("label") or "").strip()
    if label and _norm(label) != _norm(typ):
        return "%s \u2014 %s" % (typ, label)
    return typ


def _resolve_columns(session, graph, node_id, port):
    """Output columns of (node_id, port), via the cheap zero-row probe the
    inspectors already use. Best-effort: returns [] if it cannot be resolved
    (e.g. an engine the sandbox lacks) so the rest of the trace degrades
    gracefully rather than raising."""
    try:
        r = session.nodeflow_columns(graph, node_id, port or "out")
        return list((r or {}).get("columns") or [])
    except Exception:
        return []


def _renamecols_map(cfg, cols):
    """Replicate the renamecols node's old->new mapping for ``cols`` (same
    rules as nodeflow.compile: explicit mappings win, else case/find/replace +
    prefix/suffix, with de-duplication of collisions)."""
    prefix = cfg.get("prefix") or ""
    suffix = cfg.get("suffix") or ""
    case = _norm(cfg.get("case"))
    find = cfg.get("find") or ""
    repl = cfg.get("replace") or ""
    mp = {m.get("from"): m.get("to")
          for m in (cfg.get("mappings") or [])
          if m.get("from") and m.get("to")}
    seen = {}
    out = {}
    for c in cols:
        if c in mp:
            new = mp[c]
        else:
            new = c
            if case == "snake":
                new = _snake(new)
            elif case == "lower":
                new = new.lower()
            elif case == "upper":
                new = new.upper()
            if find:
                new = new.replace(find, repl)
            new = prefix + new + suffix
        new = (new or "").strip() or c
        base, k = new, 2
        while new in seen:
            new = "%s_%d" % (base, k)
            k += 1
        seen[new] = True
        out[c] = new
    return out


def _node_columns_info(session, graph, node_id, port, memo):
    """For (node_id, port) return an ordered dict ``out_col -> rec`` where rec
    describes how that output column is produced from the node's *direct*
    inputs::

        {"kind": "source"|"passthrough"|"derived"|"sql",
         "srcs": [(up_node, up_port, up_col), ...],   # direct inputs
         "step": "<label>",                           # derived only
         "file": "<input file>"}                      # source only

    Memoised on (node_id, port); a placeholder is stored before recursing so a
    cyclic graph cannot loop forever."""
    key = (node_id, port or "out")
    if key in memo:
        return memo[key]
    memo[key] = {}  # cycle guard

    by_id = _nodes_by_id(graph)
    node = by_id.get(node_id)
    if node is None:
        return {}
    typ = node.get("type")
    cfg = node.get("config") or {}
    info = {}

    def up_info(in_port="in"):
        un, up = upstream(graph, node_id, in_port)
        if not un:
            return ({}, None, None)
        return (_node_columns_info(session, graph, un, up or "out", memo),
                un, up or "out")

    if typ in _SOURCE_TYPES:
        fname = (cfg.get("table") or cfg.get("label") or node_id or "").strip()
        for c in _resolve_columns(session, graph, node_id, port):
            info[c] = {"kind": "source", "srcs": [], "file": fname}

    elif typ == "sql":
        for c in _resolve_columns(session, graph, node_id, port):
            info[c] = {"kind": "sql", "srcs": [], "step": _step(typ, node)}

    elif typ == "python":
        for c in _resolve_columns(session, graph, node_id, port):
            info[c] = {"kind": "sql", "srcs": [], "step": _step(typ, node)}

    elif typ == "formula":
        upi, un, up = up_info("in")
        valid = [((f.get("name") or "").strip(), (f.get("expr") or "").strip())
                 for f in (cfg.get("formulas") or [])
                 if (f.get("name") or "").strip()
                 and (f.get("expr") or "").strip()]
        fnames = {_norm(n) for n, _ in valid}
        in_lower = {_norm(c): c for c in upi.keys()}
        for c in upi.keys():
            if _norm(c) not in fnames:
                info[c] = {"kind": "passthrough", "srcs": [(un, up, c)]}
        step = _step("formula", node)
        for name, expr in valid:
            refs = {_norm(r) for r in _ident_cols(expr)}
            # keep the columns the expression references, in upstream column
            # order, so the emitted rows read top-down like the source data
            srcs = [(un, up, c) for c in upi.keys() if _norm(c) in refs]
            info[name] = {"kind": "derived", "srcs": srcs, "step": step}

    elif typ == "summarize":
        upi, un, up = up_info("in")
        in_lower = {_norm(c): c for c in upi.keys()}
        for g in (cfg.get("group_by") or []):
            if not g:
                continue
            real = in_lower.get(_norm(g), g)
            info[real] = {"kind": "passthrough", "srcs": [(un, up, real)]}
        step = _step("summarize", node)
        for a in (cfg.get("aggs") or []):
            col = (a.get("col") or "").strip()
            func = _norm(a.get("func"))
            if not col or func not in _AGG_FUNCS:
                continue
            name = (a.get("name") or "").strip() or ("%s_%s" % (func, col))
            real = in_lower.get(_norm(col), col)
            info[name] = {"kind": "derived", "srcs": [(un, up, real)],
                          "step": step}

    elif typ == "select":
        upi, un, up = up_info("in")
        in_lower = {_norm(c): c for c in upi.keys()}
        for f in (cfg.get("fields") or []):
            if f.get("keep") is False:
                continue
            col = (f.get("name") or "").strip()
            if not col:
                continue
            ty = (f.get("type") or "").strip()
            alias = (f.get("rename") or col).strip() or col
            real = in_lower.get(_norm(col), col)
            renamed = _norm(alias) != _norm(col)
            if renamed:
                # A rename is a real change worth tracing -- the user asked to
                # keep seeing these. A cast alongside it doesn't alter the
                # values, so the hop is labelled just "rename".
                info[alias] = {"kind": "derived", "srcs": [(un, up, real)],
                               "step": _step("select (rename)", node)}
            else:
                # Cast-only (or an untouched field): the values are unchanged --
                # a type change is not a data change for lineage purposes -- so
                # it passes through rather than showing as a transformation.
                info[alias] = {"kind": "passthrough", "srcs": [(un, up, real)]}

    elif typ == "renamecols":
        upi, un, up = up_info("in")
        for old, new in _renamecols_map(cfg, list(upi.keys())).items():
            if _norm(new) == _norm(old):
                info[new] = {"kind": "passthrough", "srcs": [(un, up, old)]}
            else:
                info[new] = {"kind": "derived", "srcs": [(un, up, old)],
                             "step": _step("rename", node)}

    elif typ in _TRANSPARENT:
        upi, un, up = up_info("in")
        for c in upi.keys():
            info[c] = {"kind": "passthrough", "srcs": [(un, up, c)]}

    else:
        # generic / multi-input: gather every connected upstream, then classify
        # each output column by which upstream owns its name.
        ups = []  # (info, un, up)
        for e in (graph.get("edges") or []):
            t = e.get("to") or {}
            if t.get("node") != node_id:
                continue
            f = e.get("from") or {}
            fn, fp = f.get("node"), f.get("port") or "out"
            if not fn:
                continue
            ups.append((_node_columns_info(session, graph, fn, fp, memo),
                        fn, fp))
        outcols = _resolve_columns(session, graph, node_id, port)
        if not outcols and ups:
            # nothing resolved (e.g. engine missing): fall back to the union of
            # upstream columns so a pass-through-shaped node still traces
            seen = set()
            for ui, _fn, _fp in ups:
                for c in ui.keys():
                    if _norm(c) not in seen:
                        seen.add(_norm(c))
                        outcols.append(c)
        allsrcs = []
        for ui, fn, fp in ups:
            for c in ui.keys():
                allsrcs.append((fn, fp, c))
        step = _step(typ or "node", node)
        for c in outcols:
            owner = None
            for ui, fn, fp in ups:
                real = _lookup_key(ui, c)
                if real is not None:
                    owner = (fn, fp, real)
                    break
            if owner is not None:
                info[c] = {"kind": "passthrough", "srcs": [owner]}
            elif allsrcs:
                info[c] = {"kind": "derived", "srcs": list(allsrcs),
                           "step": step}
            else:
                info[c] = {"kind": "source", "srcs": [],
                           "file": (cfg.get("table") or cfg.get("label")
                                    or node_id or "").strip()}

    memo[key] = info
    return info


def _lookup_key(info, name):
    """Like _lookup but returns the matching *key* (real column name)."""
    if name in info:
        return name
    nl = _norm(name)
    for k in info.keys():
        if _norm(k) == nl:
            return k
    return None


def _origin(session, graph, node_id, port, col, memo, depth=0):
    """Classify where ``col`` ultimately comes from, following pass-throughs
    upstream: ``("input", file)``, ``("derived", node)``, ``("sql", node)`` or
    ``("unknown",)``."""
    if depth > 256:
        return ("unknown",)
    info = _node_columns_info(session, graph, node_id, port, memo)
    rec = _lookup(info, col)
    if rec is None:
        return ("unknown",)
    kind = rec.get("kind")
    if kind == "source":
        return ("input", rec.get("file") or node_id)
    if kind == "sql":
        return ("sql", node_id)
    if kind == "passthrough":
        srcs = rec.get("srcs") or []
        if not srcs:
            return ("unknown",)
        un, up, uc = srcs[0]
        return _origin(session, graph, un, up, uc, memo, depth + 1)
    return ("derived", node_id)


def _emit_chain(session, graph, node_id, port, field, data_element,
                memo, rows, sql_set, seen, depth=0):
    """Emit the hop-rows that PRODUCE ``field`` at (node_id, port). Sources
    that are themselves derived are expanded first (post-order), so the deepest
    hop is listed before the hop that consumes it -- matching the layout."""
    if depth > 256:
        return
    info = _node_columns_info(session, graph, node_id, port, memo)
    rec = _lookup(info, field)
    if rec is None:
        return
    kind = rec.get("kind")
    if kind == "source":
        return
    if kind == "sql":
        sql_set.add(data_element)
        return
    if kind == "passthrough":
        for (un, up, uc) in (rec.get("srcs") or []):
            _emit_chain(session, graph, un, up, uc, data_element,
                        memo, rows, sql_set, seen, depth + 1)
        return
    # derived
    step = rec.get("step") or "derived"
    for (un, up, uc) in (rec.get("srcs") or []):
        o = _origin(session, graph, un, up, uc, memo)
        kind0 = o[0]
        if kind0 == "input":
            input_file = o[1]
        elif kind0 == "sql":
            input_file = "Derived"
            sql_set.add(data_element)
        else:
            input_file = "Derived"
        if kind0 in ("derived", "unknown"):
            _emit_chain(session, graph, un, up, uc, data_element,
                        memo, rows, sql_set, seen, depth + 1)
        rk = (input_file, _norm(uc), _norm(field), step, _norm(data_element))
        if rk not in seen:
            seen.add(rk)
            rows.append({"input_file": input_file, "source": uc,
                         "used_in": field, "step": step,
                         "data_element": data_element,
                         "node_id": node_id})


def _terminals(graph):
    """The (node, port) pairs whose columns are the workflow's outputs: the
    node feeding each output/write node, or -- failing that -- the leaf nodes
    that produce data."""
    nodes = graph.get("nodes") or []
    terms, seen = [], set()
    for n in nodes:
        if n.get("type") in ("output", "write") and n.get("id"):
            un, up = upstream(graph, n["id"], "in")
            if un and (un, up or "out") not in seen:
                seen.add((un, up or "out"))
                terms.append((un, up or "out"))
    if terms:
        return terms
    has_out = set()
    for e in (graph.get("edges") or []):
        f = e.get("from") or {}
        if f.get("node"):
            has_out.add(f["node"])
    for n in nodes:
        nid = n.get("id")
        if (nid and nid not in has_out
                and n.get("type") not in _NON_TERMINAL):
            if (nid, "out") not in seen:
                seen.add((nid, "out"))
                terms.append((nid, "out"))
    return terms


def build_field_lineage(session, graph):
    """Trace every terminal column. Returns
    ``{"passthrough": [...], "derived": [...], "sql_flagged": set()}``."""
    memo = {}
    passthrough, derived = [], []
    sql_set, seen_pt, elements = set(), set(), []
    for (tn, tport) in _terminals(graph):
        info = _node_columns_info(session, graph, tn, tport, memo)
        for dc in list(info.keys()):
            o = _origin(session, graph, tn, tport, dc, memo)
            if o[0] == "input":
                pk = _norm(dc)
                if pk not in seen_pt:
                    seen_pt.add(pk)
                    passthrough.append({"input_file": o[1], "source": dc,
                                        "data_element": dc})
            else:
                if dc not in elements:
                    elements.append(dc)
                seen = set()
                before = len(derived)
                _emit_chain(session, graph, tn, tport, dc, dc,
                            memo, derived, sql_set, seen)
                if o[0] == "sql":
                    sql_set.add(dc)
                    if len(derived) == before:
                        derived.append({"input_file": "Derived", "source": "",
                                        "used_in": dc, "step": "sql",
                                        "data_element": dc})
    return {"passthrough": passthrough, "derived": derived,
            "sql_flagged": sql_set}


def _transform_line(inputs, op, output):
    """Compact ``inputs → op → output`` line for diagrams and detail panels."""
    left = ", ".join(str(x) for x in (inputs or []) if x) or "—"
    mid = (op or "passthrough").strip() or "passthrough"
    right = (output or "").strip() or "—"
    return "%s → %s → %s" % (left, mid, right)


def _describe_change(node, rec, col):
    """Structural before→after description for how ``col`` is produced at
    ``node``. Discovery-only: reads node config + the lineage record, never
    materializes full tables.

    Each change carries a richer transform summary::

        inputs → expression/op → output

    plus optional type / filter / join / summarize specifics from node cfg.
    """
    typ = node.get("type") or ""
    cfg = node.get("config") or {}
    kind = rec.get("kind") or "unknown"
    srcs = [s[2] for s in (rec.get("srcs") or [])]
    label = (cfg.get("label") or "").strip() or typ
    change = {
        "summary": "",
        "detail": "",
        "inputs": list(srcs),
        "output": col,
        "expression": None,
        "mapping": None,
        "predicate": None,
        "unchanged": kind == "passthrough",
        # Richer per-stage transform (additive; older clients ignore these).
        "op": None,
        "transform": None,
        "type_from": None,
        "type_to": None,
        "group_by": None,
        "join_how": None,
    }

    def _finish(op, transform=None, **extra):
        change["op"] = op
        for k, v in extra.items():
            change[k] = v
        change["transform"] = transform or _transform_line(
            change.get("inputs"), op, change.get("output") or col)
        return change

    if kind == "source":
        fname = rec.get("file") or cfg.get("table") or label
        change["summary"] = "Loaded from %s" % fname
        change["detail"] = "Source column from input file / table"
        change["unchanged"] = False
        change["inputs"] = []
        return _finish("load", "%s → load → %s" % (fname, col))

    if kind == "sql":
        change["summary"] = "Opaque %s boundary" % (
            "SQL" if typ == "sql" else (typ or "code"))
        change["detail"] = SQL_NOTE
        expr = (cfg.get("sql") or cfg.get("code") or "").strip()
        change["expression"] = (expr[:500] + ("…" if len(expr) > 500 else "")
                                ) if expr else None
        change["unchanged"] = False
        op = "SQL" if typ == "sql" else (typ or "code")
        return _finish(op, _transform_line(srcs or ["…"], op, col))

    if typ == "formula":
        for f in (cfg.get("formulas") or []):
            name = (f.get("name") or "").strip()
            if _norm(name) != _norm(col):
                continue
            expr = (f.get("expr") or "").strip()
            change["summary"] = "%s = %s" % (col, expr or "?")
            change["expression"] = expr or None
            change["detail"] = ("Inputs: %s" % ", ".join(srcs)) if srcs else ""
            change["unchanged"] = False
            op_txt = expr or "formula"
            return _finish("formula", _transform_line(srcs, op_txt, col))

    if typ == "summarize":
        groups = [g for g in (cfg.get("group_by") or []) if g]
        for a in (cfg.get("aggs") or []):
            acol = (a.get("col") or "").strip()
            func = _norm(a.get("func"))
            name = ((a.get("name") or "").strip()
                    or ("%s_%s" % (func, acol)))
            if _norm(name) != _norm(col):
                continue
            expr = "%s(%s)" % (func or "agg", acol or "?")
            change["summary"] = "%s → %s" % (expr, col)
            change["expression"] = expr
            change["detail"] = ("Group by %s" % ", ".join(groups)
                                if groups else "No group keys")
            change["unchanged"] = False
            change["inputs"] = [acol] if acol else list(srcs)
            return _finish(
                func or "agg",
                _transform_line(change["inputs"], expr, col),
                group_by=list(groups) or None)
        change["summary"] = "Group key (value unchanged)"
        change["detail"] = ("Group by %s" % ", ".join(groups)
                            if groups else "Group key")
        change["unchanged"] = True
        return _finish(
            "group key",
            _transform_line([col], "group key", col),
            group_by=list(groups) or None)

    if typ == "select":
        for f in (cfg.get("fields") or []):
            if f.get("keep") is False:
                continue
            name = (f.get("name") or "").strip()
            alias = (f.get("rename") or name).strip() or name
            if _norm(alias) != _norm(col) and _norm(name) != _norm(col):
                continue
            ty = (f.get("type") or "").strip()
            if _norm(alias) != _norm(name):
                change["summary"] = "Rename %s → %s" % (name, alias)
                change["mapping"] = {"from": name, "to": alias}
                change["unchanged"] = False
                op = "rename"
                if ty:
                    change["type_to"] = ty
                    change["expression"] = "CAST(%s AS %s)" % (name, ty)
                    change["detail"] = "Rename with cast to %s" % ty
                    op = "rename+cast"
                return _finish(
                    op,
                    _transform_line([name], op, alias),
                    type_to=ty or None)
            if ty:
                change["summary"] = "Cast %s as %s" % (alias, ty)
                change["detail"] = "Values unchanged; type only"
                change["unchanged"] = True
                change["expression"] = "CAST(%s AS %s)" % (name or alias, ty)
                change["type_to"] = ty
                return _finish(
                    "cast",
                    _transform_line([name or alias], "CAST AS %s" % ty, alias),
                    type_to=ty)
            change["summary"] = "Selected (unchanged)"
            change["unchanged"] = True
            return _finish("select", _transform_line([name], "select", alias))

    if typ == "renamecols":
        if srcs and _norm(srcs[0]) != _norm(col):
            change["summary"] = "Rename %s → %s" % (srcs[0], col)
            change["mapping"] = {"from": srcs[0], "to": col}
            change["unchanged"] = False
            return _finish("rename", _transform_line([srcs[0]], "rename", col))
        change["summary"] = "Rename pass-through"
        change["unchanged"] = True
        return _finish("rename", _transform_line([col], "rename", col))

    if typ in ("join", "antijoin", "crossjoin", "multijoin"):
        pairs = [(k.get("left"), k.get("right"))
                 for k in (cfg.get("keys") or [])
                 if k.get("left") and k.get("right")]
        if typ == "antijoin" and not pairs:
            lk, rk = cfg.get("left_key"), cfg.get("right_key")
            if lk and rk:
                pairs = [(lk, rk)]
        how = (cfg.get("how") or cfg.get("type") or typ).strip()
        key_txt = ", ".join("%s = %s" % (l, r) for l, r in pairs) or "(no keys)"
        side = ""
        if col.lower().startswith("r_"):
            side = " from right (collision rename)"
        change["summary"] = "%s join%s" % (how, side)
        change["detail"] = "On %s" % key_txt
        change["predicate"] = key_txt if pairs else None
        change["unchanged"] = kind == "passthrough"
        if kind == "derived":
            change["detail"] = ((change["detail"] + "; ") if change["detail"]
                                else "") + "New column from join inputs"
        op = "%s join" % how
        return _finish(
            op,
            _transform_line(srcs or [col], "%s on %s" % (op, key_txt), col),
            join_how=how)

    if typ == "filter":
        pred = (cfg.get("condition") or cfg.get("expr") or "").strip()
        change["summary"] = "Filter rows"
        change["predicate"] = pred or None
        change["detail"] = ("Keep where %s" % pred) if pred else "Row filter"
        change["unchanged"] = True
        op = "filter"
        mid = ("filter(%s)" % pred) if pred else "filter"
        return _finish(op, _transform_line([col], mid, col))

    if typ == "sort":
        keys = cfg.get("keys") or cfg.get("order") or []
        bits = []
        if isinstance(keys, list) and keys:
            for k in keys:
                if isinstance(k, dict):
                    bits.append("%s %s" % (
                        k.get("col") or k.get("column") or "?",
                        (k.get("dir") or k.get("order") or "asc")))
                else:
                    bits.append(str(k))
            change["detail"] = "Order by %s" % ", ".join(bits)
        change["summary"] = "Sort rows (values unchanged)"
        change["unchanged"] = True
        mid = ("sort(%s)" % ", ".join(bits)) if bits else "sort"
        return _finish("sort", _transform_line([col], mid, col))

    if typ in ("sample", "topn", "limit"):
        n = cfg.get("n") or cfg.get("limit") or cfg.get("count")
        change["summary"] = "Keep a subset of rows"
        change["detail"] = ("Keep %s rows" % n) if n else ("via %s" % typ)
        change["unchanged"] = True
        mid = ("%s(%s)" % (typ, n)) if n else typ
        return _finish(typ, _transform_line([col], mid, col))

    if typ in _TRANSPARENT:
        change["summary"] = "Unchanged through %s" % (label or typ)
        change["unchanged"] = True
        return _finish(typ, _transform_line([col], typ, col))

    if kind == "passthrough":
        if srcs and _norm(srcs[0]) != _norm(col):
            change["summary"] = "Pass-through rename %s → %s" % (srcs[0], col)
            change["mapping"] = {"from": srcs[0], "to": col}
            change["unchanged"] = False
            return _finish(
                "rename", _transform_line([srcs[0]], "rename", col))
        change["summary"] = "Unchanged through %s" % (label or typ)
        change["unchanged"] = True
        return _finish(
            typ or "passthrough",
            _transform_line([col], typ or "passthrough", col))

    step = rec.get("step") or typ or "derived"
    change["summary"] = "Derived by %s" % step
    if srcs:
        change["detail"] = "From %s" % ", ".join(srcs)
    change["unchanged"] = False
    return _finish(step, _transform_line(srcs, step, col))


def _should_emit_stage(typ, rec, col):
    """Skip noise passthrough hops (e.g. a column riding unused through
    formula nodes) so the diagram stays readable. Always keep sources,
    real transforms, transparent row-ops, and joins."""
    kind = rec.get("kind")
    if kind in ("source", "derived", "sql"):
        return True
    if typ in _TRANSPARENT:
        return True
    if typ in ("join", "antijoin", "crossjoin", "multijoin", "union"):
        return True
    srcs = rec.get("srcs") or []
    if srcs and _norm(srcs[0][2]) != _norm(col):
        return True
    return False


def _collect_column_stages(session, graph, node_id, port, col, memo, seen,
                           stages, depth=0):
    """Walk upstream first, then append this hop — source → … → result."""
    if depth > 256:
        return
    key = (node_id, port or "out", _norm(col))
    if key in seen:
        return
    seen.add(key)

    by_id = _nodes_by_id(graph)
    node = by_id.get(node_id)
    if node is None:
        return
    info = _node_columns_info(session, graph, node_id, port, memo)
    real = _lookup_key(info, col)
    if real is None:
        return
    rec = info[real]
    for (un, up, uc) in (rec.get("srcs") or []):
        _collect_column_stages(session, graph, un, up or "out", uc, memo,
                               seen, stages, depth + 1)

    typ = node.get("type") or ""
    if not _should_emit_stage(typ, rec, real):
        return
    cfg = node.get("config") or {}
    kind = rec.get("kind") or "unknown"
    stages.append({
        "id": "%s:%s:%s" % (node_id, port or "out", real),
        "kind": kind,
        "column": real,
        "node_id": node_id,
        "node_type": typ,
        "node_label": (cfg.get("label") or "").strip() or typ,
        "step": rec.get("step") or _step(typ, node),
        "change": _describe_change(node, rec, real),
    })


def build_column_lineage(session, graph, column, node_id=None, port=None,
                         row_index=None, cell_value=None):
    """Diagram-ready lineage for one output column, reusing the same column
    semantics as ``build_field_lineage``. Returns stages ordered source →
    result, each with a structural change summary (no full-table materialize).

    When ``row_index`` is provided (cell right-click), each stage also gets a
    best-effort ``value`` sample for that row via cheap LIMIT/OFFSET probes.
    """
    column = (column or "").strip()
    graph = graph or {}
    if not column:
        return {"ok": True, "available": False, "column": "",
                "stages": [], "sql_flagged": False,
                "reason": "No column specified."}
    if not (graph.get("nodes") or []):
        return {"ok": True, "available": False, "column": column,
                "stages": [], "sql_flagged": False,
                "reason": "Lineage is only available for NodeFlow results."}

    memo = {}
    terms = ([(node_id, port or "out")] if node_id
             else _terminals(graph))
    target = None
    for tn, tport in terms:
        if not tn:
            continue
        info = _node_columns_info(session, graph, tn, tport, memo)
        if _lookup(info, column) is not None:
            target = (tn, tport or "out")
            break
    if target is None and node_id:
        target = (node_id, port or "out")
    if target is None:
        return {"ok": True, "available": False, "column": column,
                "stages": [], "sql_flagged": False,
                "reason": "Column not found in the workflow outputs."}

    stages = []
    _collect_column_stages(session, graph, target[0], target[1], column,
                           memo, set(), stages)
    if not stages:
        return {"ok": True, "available": False, "column": column,
                "stages": [], "sql_flagged": False,
                "reason": "Could not trace lineage for this column."}

    o = _origin(session, graph, target[0], target[1], column, memo)
    sql_flagged = (o[0] == "sql"
                   or any(s.get("kind") == "sql" for s in stages))

    if row_index is not None:
        _attach_stage_values(session, graph, stages, row_index, cell_value,
                             column)

    return {
        "ok": True,
        "available": True,
        "column": column,
        "terminal_node": target[0],
        "terminal_port": target[1],
        "sql_flagged": bool(sql_flagged),
        "stages": stages,
        "reason": None,
        "row_index": row_index,
    }


def _jsonish_cell(v):
    if v is None or isinstance(v, (bool, int, float, str)):
        if isinstance(v, float) and (v != v):  # NaN
            return None
        return v
    if isinstance(v, bytes):
        try:
            return v.decode("utf-8")
        except Exception:
            return repr(v)
    return str(v)


def _attach_stage_values(session, graph, stages, row_index, cell_value,
                         target_column):
    """Best-effort per-stage value samples for one result row.

    Uses compile + LIMIT 1 OFFSET — never materializes multi-GB tables for
    this path. Failures become ``available: false`` on that stage only.
    """
    try:
        ri = int(row_index)
    except (TypeError, ValueError):
        return
    if ri < 0 or ri > 100000:
        return

    probe = getattr(session, "nodeflow_probe_row_values", None)
    if not callable(probe):
        for st in stages:
            st["value"] = {
                "available": False,
                "value": None,
                "inputs": [],
                "reason": "value unavailable",
            }
        return

    for i, st in enumerate(stages):
        cols = []
        out_col = st.get("column")
        if out_col:
            cols.append(out_col)
        ch = st.get("change") or {}
        for ic in (ch.get("inputs") or []):
            if ic and ic not in cols:
                cols.append(ic)
        try:
            got = probe(graph, st.get("node_id"), "out", cols, ri) or {}
        except Exception:
            got = {"ok": False, "error": "probe failed"}
        vals = got.get("values") if isinstance(got, dict) else None
        if not isinstance(vals, dict):
            vals = {}
        available = bool(got.get("ok")) and out_col in vals
        out_val = vals.get(out_col) if out_col else None
        # Prefer the cell the user clicked for the final stage.
        if (i == len(stages) - 1 and cell_value is not None
                and _norm(out_col) == _norm(target_column)):
            out_val = _jsonish_cell(cell_value)
            available = True
        inputs = []
        for ic in (ch.get("inputs") or []):
            if ic in vals:
                inputs.append({
                    "column": ic,
                    "value": _jsonish_cell(vals[ic]),
                    "available": True,
                })
            else:
                inputs.append({
                    "column": ic,
                    "value": None,
                    "available": False,
                })
        st["value"] = {
            "available": available,
            "value": _jsonish_cell(out_val) if available else None,
            "inputs": inputs,
            "reason": None if available else (
                got.get("error") or "value unavailable"),
            "expression": ch.get("expression"),
        }


def _write_sheet(ws, headers, rows):
    from openpyxl.styles import Font, PatternFill, Alignment
    bold = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="305496")
    ws.append(headers)
    for ci, _h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=ci)
        cell.font = bold
        cell.fill = fill
        cell.alignment = Alignment(vertical="center")
    for r in rows:
        ws.append(r)
    widths = [0] * len(headers)
    for ci, h in enumerate(headers):
        widths[ci] = len(str(h))
    for r in rows:
        for ci, v in enumerate(r):
            widths[ci] = max(widths[ci], len(str(v if v is not None else "")))
    for ci, w in enumerate(widths, start=1):
        ws.column_dimensions[
            ws.cell(row=1, column=ci).column_letter].width = min(max(w + 2, 10),
                                                                  60)
    ws.freeze_panes = "A2"


def write_lineage_xlsx(lineage, path):
    """Write the two-tab workbook (Pass-through / Derived) to ``path``."""
    from openpyxl import Workbook
    sql_norm = {_norm(x) for x in (lineage.get("sql_flagged") or set())}

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Pass-through"
    _write_sheet(ws1, ["Input File", "Source Field", "DataElement"],
                 [[r["input_file"], r["source"], r["data_element"]]
                  for r in (lineage.get("passthrough") or [])])

    ws2 = wb.create_sheet("Derived")
    drows = []
    for r in (lineage.get("derived") or []):
        note = SQL_NOTE if _norm(r["data_element"]) in sql_norm else ""
        drows.append([r["input_file"], r["source"], r["used_in"],
                      r["step"], r["data_element"], note])
    _write_sheet(ws2, ["Input File", "Source Field", "Used in", "Derived By",
                       "DataElement", "Notes"], drows)
    wb.save(path)
    return path


def export_lineage_xlsx(session, graph):
    """Build the lineage and write it to a temp .xlsx. Returns
    ``{"ok": True, "path": ...}`` or ``{"ok": False, "error": ...}``."""
    try:
        import openpyxl  # noqa: F401
    except Exception:
        return {"ok": False,
                "error": "openpyxl is required to export the lineage workbook."}
    data = build_field_lineage(session, graph)
    try:
        from . import tmputil
        fd, path = tempfile.mkstemp(suffix=".xlsx", prefix="samql_lineage_",
                                    dir=tmputil.instance_dir())
    except Exception:
        fd, path = tempfile.mkstemp(suffix=".xlsx", prefix="samql_lineage_",
                                    dir=None)  # last resort: system temp
    os.close(fd)
    write_lineage_xlsx(data, path)
    return {"ok": True, "path": path,
            "passthrough": len(data["passthrough"]),
            "derived": len(data["derived"]),
            "sql_flagged": sorted(data["sql_flagged"])}
