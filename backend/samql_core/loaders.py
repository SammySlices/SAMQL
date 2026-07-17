"""File ingestion into a target engine.

CSV: delimiter sniffing, multi-encoding fallback, duplicate-header
handling, and ragged-row tolerance, streamed row-by-row into SQLite.
JSON: parsed by the tolerant reader, flattened to relational tables.
Parquet / large CSV: registered as zero-copy DuckDB views when DuckDB is
available.

Distilled from the original single-file application's loader methods.
GUI-free; progress is reported through an optional callback.
"""
import csv
import sys
import io
import os
import shutil
import tempfile

from .flatten import JSONFlattener, stream_json_records
from .nodeflow import sanitize_ident
from .sqlutil import sanitize_column_header

# Per-table row buffer kept in memory before the JSON flattener spills it
# to disk. Small/medium files never reach this, so they stay fully in-memory;
# only genuinely large JSON spills, bounding peak RAM. Spilling (pickle to disk
# + re-read) is a real cost on a deep flatten that explodes into millions of
# rows, so on a box with ample RAM we buffer more before spilling (fewer, larger
# writes). Overridable with SAMQL_JSON_SPILL_ROWS for manual tuning.
def _json_spill_rows():
    env = os.environ.get("SAMQL_JSON_SPILL_ROWS")
    if env:
        try:
            v = int(env)
            if v > 0:
                return v
        except ValueError:
            pass
    try:
        from .engines import total_physical_ram_bytes
        gb = total_physical_ram_bytes() / (1024 ** 3)
    except Exception:
        gb = 0
    return 150_000 if gb >= 16 else 50_000


JSON_SPILL_ROWS = _json_spill_rows()


class LoadCancelled(Exception):
    """Raised through a load's progress callback to abort an in-flight load.

    The progress callbacks below are wrapped in ``try/except Exception`` so a
    flaky progress reporter can never break a load -- but this one exception is
    deliberately re-raised so the caller can signal "stop now". It propagates
    out through the streaming insert (which rolls back and drops the partial
    table on the way) and unwinds the load cleanly."""
    pass


class _CountingRaw(io.RawIOBase):
    """Wraps a raw binary file and counts the bytes read, so a CSV load
    can report real progress (bytes consumed vs. file size) while still
    letting csv.reader handle quoting/multiline fields correctly."""

    def __init__(self, fileobj):
        self._f = fileobj
        self.count = 0

    def readable(self):
        return True

    def readinto(self, b):
        n = self._f.readinto(b)
        if n:
            self.count += n
        return n

    def close(self):
        try:
            self._f.close()
        except Exception:
            pass
        super().close()

# Allow very wide CSV cells (the original raised this cap too).
try:
    csv.field_size_limit(2 ** 31 - 1)
except Exception:
    try:
        csv.field_size_limit(2 ** 27)
    except Exception:
        pass

_sanitize = JSONFlattener._sanitize

# Extensions load_file knows how to peel off a filename. Anything else
# (including a raw upload stem like ``test.sam``) is kept intact so a second
# splitext does not silently drop middle segments.
_TABLE_NAME_EXTS = {
    "csv", "tsv", "txt", "json", "ndjson", "jsonl",
    "parquet", "pq", "xlsx", "xlsm", "xls",
}


def base_name_for(path_or_stem):
    """Turn a file path, upload filename, or raw stem into a SQL-safe table name.

    Always sanitizes -- uploads used to pass ``os.path.splitext(filename)[0]``
    straight through, so names like ``Usinvestments_Monthly (1) .csv`` (trailing
    space before the extension), ``test.sam.csv`` (embedded dots), or
    ``my file.csv`` landed as fragile identifiers. NodeFlow then ``.strip()``s
    the configured table name, so a trailing-space catalog name no longer
    matched and nodes returned empty / missing-table errors. One choke point
    collapses every weird convention to the same safe identifier used by
    createtable / iterator (``sanitize_ident``).

    Only peels a *known* data extension, so an upload stem of ``test.sam``
    stays ``test.sam`` (→ ``test_sam``) instead of being truncated to ``test``.

    A leading ``__`` (directory / append-from-folder / API hidden tables) is
    preserved after sanitisation -- ``sanitize_ident`` strips edge underscores,
    which would otherwise turn ``__nbfile_…`` into a visible ``nbfile_…`` and
    leak temporary loads into the tables tree.
    """
    base = os.path.basename(str(path_or_stem or "")).strip()
    root, ext = os.path.splitext(base)
    stem = root if ext.lower().lstrip(".") in _TABLE_NAME_EXTS else base
    stem = stem.strip()
    keep_hidden = stem.startswith("__")
    cleaned = sanitize_ident(stem, "table")
    if keep_hidden and cleaned and not cleaned.startswith("__"):
        cleaned = "__" + cleaned
    return cleaned


def detect_delimiter(sample):
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        return dialect.delimiter
    except csv.Error:
        pass
    lines = [ln for ln in sample.splitlines() if ln.strip()][:10]
    if not lines:
        return None
    best, best_score = None, -1.0
    for d in (",", ";", "\t", "|"):
        counts = [ln.count(d) for ln in lines]
        mean = sum(counts) / len(counts)
        if mean < 1:
            continue
        var = sum((c - mean) ** 2 for c in counts) / len(counts)
        score = mean / (1.0 + var)
        if score > best_score:
            best_score, best = score, d
    return best


def load_csv(db, path, base_name=None, sep=None, source=None, progress=None):
    """Stream a CSV into ``db`` (a DBManager). Returns
    {name, rows, columns, multiline, excluded}. If ``progress`` is given
    it is called periodically as progress(bytes_done, bytes_total)."""
    base_name = base_name_for(base_name or path)
    source = source or os.path.basename(path)
    try:
        total_size = os.path.getsize(path)
    except Exception:
        total_size = 0
    last_err = None
    # .445 [PLAN PASS 2] fix: a UTF-16 file sailed past this chain --
    # utf-8 fails on the null bytes, latin-1 "succeeds" and loads
    # mojibake (\x00-riddled headers and values, silently). Sniff the
    # BOM first and put the right codec at the FRONT of the chain.
    _encs = ("utf-8", "utf-8-sig", "latin-1")
    try:
        with open(path, "rb") as _bf:
            _head = _bf.read(2)
        if _head == b"\xff\xfe":
            _encs = ("utf-16",) + _encs
        elif _head == b"\xfe\xff":
            _encs = ("utf-16-be",) + _encs
        else:
            # BOM-less utf-16 (some Windows exports): ASCII text shows
            # a null every other byte -- odd positions for LE, even
            # for BE.
            with open(path, "rb") as _bf:
                _s64 = _bf.read(64)
            if _s64 and _s64.count(0) > len(_s64) // 3:
                odd = sum(1 for i in range(1, len(_s64), 2)
                          if _s64[i] == 0)
                even = sum(1 for i in range(0, len(_s64), 2)
                           if _s64[i] == 0)
                _encs = (("utf-16-le",) if odd >= even
                         else ("utf-16-be",)) + _encs
    except Exception:
        pass
    for enc in _encs:
        stats = {"multiline": 0, "excluded": 0}
        raw = None
        try:
            raw = open(path, "rb", buffering=0)
            counting = _CountingRaw(raw)
            buffered = io.BufferedReader(counting)
            try:
                sample_bytes = bytes(buffered.peek(65536)[:65536])
            except Exception:
                sample_bytes = b""
            sample = sample_bytes.decode(enc, errors="replace")
            f = io.TextIOWrapper(buffered, encoding=enc, newline="")
            if sep is None:
                detected = detect_delimiter(sample)
                reader = (csv.reader(f, delimiter=detected)
                          if detected else csv.reader(f))
            else:
                reader = csv.reader(f, delimiter=sep)
            raw_fields = next(reader, None) or []
            # .445 [PLAN PASS 2] fix: duplicate or blank headers LOST
            # DATA -- idx_of repointed a repeated name to its LAST
            # index (a,a,b loaded one "a" holding the second column's
            # values; ,, collapsed three columns to one). Dedupe like
            # the xlsx path always has: blanks become colN, repeats
            # get _2/_3..., and EVERY source index is read.
            seen, columns = {}, []
            for i, rf in enumerate(raw_fields):
                # Collapse whitespace first (Order Date → Order_Date, foo  bar
                # → foo_bar), then the flattener sanitizer for other non-alnum.
                nm = _sanitize(sanitize_column_header(rf)) or ("col%d" % (i + 1))
                if nm in seen:
                    seen[nm] += 1
                    nm = "%s_%d" % (nm, seen[nm])
                else:
                    seen[nm] = 0
                columns.append(nm)
            if not columns:
                columns = ["col1"]
            read_idx = list(range(len(columns)))
            n_exp = len(raw_fields)

            def _row_values():
                cnt = 0
                for rawrow in reader:
                    if rawrow is None:
                        continue
                    cnt += 1
                    if progress and (cnt % 5000 == 0):
                        try:
                            progress(counting.count, total_size)
                        except LoadCancelled:
                            raise
                        except Exception:
                            pass
                    if len(rawrow) != n_exp:
                        stats["excluded"] += 1
                        if len(rawrow) < n_exp:
                            rawrow = rawrow + [None] * (n_exp - len(rawrow))
                    yield tuple(
                        rawrow[i] if i < len(rawrow) else None
                        for i in read_idx)
                if progress:
                    try:
                        progress(total_size, total_size)
                    except LoadCancelled:
                        raise
                    except Exception:
                        pass

            name, n = db.add_table_streaming(
                base_name, columns, _row_values(), source=source)
            return {
                "name": name,
                "rows": n,
                "columns": db.table_columns.get(name, columns),
                "multiline": stats["multiline"],
                "excluded": stats["excluded"],
                "engine": "sqlite",
            }
        except UnicodeDecodeError as e:
            last_err = e
            continue
        except LoadCancelled:
            raise
        except Exception as e:
            last_err = e
            break
        finally:
            try:
                if raw is not None:
                    raw.close()
            except Exception:
                pass
    raise ValueError(f"Could not load CSV {path}: {last_err}")


def load_json(db, path, base_name=None, source=None, progress=None,
              spill_threshold=JSON_SPILL_ROWS, engine="sqlite", exclude=None):
    """Stream a JSON file into one or more tables in the given engine
    (``db`` is the engine manager; ``engine`` is its label) without holding
    the whole document in memory: records are pulled one at a time from
    ``stream_json_records`` and the flattener spills per-table row buffers
    to disk once they exceed ``spill_threshold``. ``exclude`` is an optional
    list of keys/paths to skip flattening (see JSONFlattener); skipping a heavy
    nested array removes its rows and child table entirely. Returns a list of
    {name, rows, columns, engine}."""
    base_name = base_name_for(base_name or path)
    source = source or os.path.basename(path)

    # Keep spill files under the per-instance temp dir so they're cleaned
    # up on shutdown / next-start sweep even if this load is interrupted.
    spill_dir = None
    try:
        from . import tmputil
        spill_dir = tempfile.mkdtemp(prefix="jf_", dir=tmputil.instance_dir())
    except Exception:
        spill_dir = None  # flattener makes its own temp dir if needed

    cb = None
    if progress:
        def cb(n, _p=progress):
            try:
                _p(n, 0)  # record count; byte total is unknown when streaming
            except LoadCancelled:
                raise
            except Exception:
                pass

    fl = JSONFlattener(root_name=base_name, spill_threshold=spill_threshold,
                       spill_dir=spill_dir, progress_cb=cb, exclude=exclude)

    def _pinged(gen):
        # add_table_streaming doesn't call back, so without this a Stop pressed
        # after parsing (while rows are being written) wouldn't be seen until
        # the whole table was inserted. Re-pinging cb (which re-raises
        # LoadCancelled on cancel) keeps the write phase interruptible too; the
        # parse count is already final here, so the number just holds the bar.
        if cb is None:
            yield from gen
            return
        for j, row in enumerate(gen):
            if (j % 5000) == 0:
                cb(fl._emit_total)
            yield row

    # Make the READ phase promptly interruptible. Previously the reader got no
    # cancel hook, so a Stop was only seen at the flattener's every-5000-row
    # progress checkpoint -- meaning a cancel while grinding through one big
    # record (or a slow read) could hang for a long time. Give the reader a
    # predicate so a Stop lands inside the read, even mid-record (the reader
    # polls it on every raw chunk). Two signals are OR'd: an optional predicate
    # the caller attaches to `progress`, and the engine's own cancel event (set
    # by interrupt_loads on a Stop). Clear the engine event first so a stale
    # cancel from a previous, already-unwound load can't abort this fresh one.
    _ev = getattr(db, "_cancel", None)
    if _ev is not None:
        try:
            _ev.clear()
        except Exception:
            _ev = None
    _sc_attr = getattr(progress, "should_cancel", None) if progress else None
    _sc_checks = [c for c in (_sc_attr,
                              (_ev.is_set if _ev is not None else None))
                  if callable(c)]
    should_cancel = ((lambda: any(c() for c in _sc_checks))
                     if _sc_checks else None)

    out = []
    try:
        for rec in stream_json_records(path, should_cancel=should_cancel):
            fl.add_record(rec)
        for table in fl.table_names():
            cols = fl.columns(table)
            rows = _pinged(fl.iter_rows_aligned(table))
            name, n = db.add_table_streaming(table, cols, rows,
                                             source=source)
            out.append({
                "name": name,
                "rows": n,
                "columns": db.table_columns.get(name, cols),
                "engine": engine,
            })
    finally:
        fl.close()
        if spill_dir:
            shutil.rmtree(spill_dir, ignore_errors=True)
    return out


def load_parquet(duck, path, base_name=None, source=None):
    """Register a Parquet file as a zero-copy DuckDB view. Requires
    DuckDB. Returns {name, rows, columns}."""
    base_name = base_name_for(base_name or path)
    name = duck.create_view_from_file(base_name, path, "parquet")
    cols = duck.table_columns.get(name, [])
    n = _duck_count(duck, name)
    return {"name": name, "rows": n, "columns": cols, "engine": "duckdb"}


def _excel_headers(header):
    cols, seen = [], {}
    for i, h in enumerate(header or []):
        raw = (str(h).strip() if h is not None else "")
        # Whitespace → _ only (match DuckDB load headers). Do not strip
        # punctuation here — Excel previously preserved it, and CSV already
        # applies the stronger flattener sanitizer on its own path.
        name = sanitize_column_header(raw) or ("col%d" % (i + 1))
        if name in seen:
            seen[name] += 1
            name = "%s_%d" % (name, seen[name])
        else:
            seen[name] = 0
        cols.append(name)
    return cols


def _sanitize_sheet(s):
    import re as _re
    return _re.sub(r"[^A-Za-z0-9_]+", "_", str(s or "sheet")).strip("_") or "sheet"


def _excel_sheet_stream(ws, skip, progress=None):
    """Prepare a streaming Excel sheet ingest.

    Returns ``(columns, row_iter_factory)`` or ``(None, None)`` when the sheet
    has no header row at/after ``skip``. The factory yields body row tuples
    one at a time so multi-million-row workbooks never materialise two full
    Python copies of the sheet.
    """
    it = ws.iter_rows(values_only=True)
    for _ in range(max(0, int(skip or 0))):
        if next(it, None) is None:
            return None, None
    header = next(it, None)
    if header is None:
        return None, None
    cols = _excel_headers(list(header) if header is not None else [])
    if not cols:
        return None, None
    ncol = len(cols)

    def _body():
        n = 0
        for r in it:
            cells = list(r) if r is not None else []
            vals = cells[:ncol] + [None] * (ncol - len(cells))
            if all(v is None or v == "" for v in vals):
                continue
            n += 1
            if progress is not None and (n % 5000 == 0):
                try:
                    progress(n, 0)
                except LoadCancelled:
                    raise
                except Exception:
                    pass
            yield tuple(vals)

    return cols, _body


def _excel_ingest_sheet(session, path, tbase, sname, skip, destination,
                        progress, prefer_duckdb):
    """Stream one sheet into DuckDB (preferred) or SQLite.

    openpyxl read-only iterators are one-shot, so a DuckDB failure re-opens
    the workbook for the SQLite fallback instead of buffering the sheet.
    """
    try:
        import openpyxl
    except Exception:
        raise RuntimeError("Reading Excel files needs the 'openpyxl' package, "
                           "which isn't installed.")
    try:
        from .engines import HAS_DUCKDB as _HAS_DUCKDB
    except Exception:
        _HAS_DUCKDB = False

    def _pass(wb, engine_name):
        if sname not in wb.sheetnames:
            return None
        cols, body_fn = _excel_sheet_stream(wb[sname], skip, progress)
        if cols is None:
            return None
        if engine_name == "duckdb":
            duck = session.get_duckdb()
            name, n = duck.add_table_streaming(
                tbase, cols, body_fn(), source=path)
            eng = duck
        else:
            name, n = session.db.add_table_streaming(
                tbase, cols, body_fn(), source=path)
            eng = session.db
        return {"name": name, "rows": n,
                "columns": eng.table_columns.get(name, cols),
                "engine": engine_name}

    if prefer_duckdb and destination == "duckdb" and _HAS_DUCKDB:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            return _pass(wb, "duckdb")
        except LoadCancelled:
            raise
        except Exception:
            pass
        finally:
            try:
                wb.close()
            except Exception:
                pass

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        return _pass(wb, "sqlite")
    finally:
        try:
            wb.close()
        except Exception:
            pass


def load_excel(session, path, base_name=None, destination="sqlite",
               sheet=None, header_row=1, progress=None):
    """Load an .xlsx/.xlsm workbook into one or more tables.

    ``sheet`` selects which sheet to load: ``None``/""/"__all__" loads every
    non-empty sheet (a single-sheet workbook -> one table named after the file;
    a multi-sheet workbook -> one table per sheet, ``file_sheet``); a specific
    sheet name loads just that sheet as one table named after the file.

    ``header_row`` (1-based) is the row that holds the column names; any rows
    above it are skipped (use this when a sheet has a title/banner before the
    real header). Defaults to 1 (the first row).

    Reads via openpyxl in read-only mode and streams rows straight into
    ``add_table_streaming`` — never buffers the whole sheet in Python.
    """
    try:
        import openpyxl
    except Exception:
        raise RuntimeError("Reading Excel files needs the 'openpyxl' package, "
                           "which isn't installed.")
    try:
        header_row = int(header_row)
    except Exception:
        header_row = 1
    if header_row < 1:
        header_row = 1
    skip = header_row - 1
    want = str(sheet).strip() if sheet is not None else ""
    pick_all = want in ("", "__all__", "*")
    base = base_name_for(base_name or path)
    ext = os.path.splitext(path)[1].lower()
    if ext == ".xls":
        raise RuntimeError(
            "Legacy .xls (Excel 97-2003) isn't supported. Save the workbook "
            "as .xlsx and try again, or drag-drop the .xlsx file.")
    # Peek sheet names only (no row scan) so multi-sheet naming is known up
    # front without buffering cell data.
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheetnames = list(wb.sheetnames)
    finally:
        try:
            wb.close()
        except Exception:
            pass
    if not pick_all and want not in sheetnames:
        raise RuntimeError(
            "Sheet %r wasn't found, or has no rows at/after row %d."
            % (want, header_row))
    candidates = sheetnames if pick_all else [want]
    multi = len(candidates) > 1
    prefer_duck = destination == "duckdb"
    out = []
    for sname in candidates:
        tbase = ("%s_%s" % (base, _sanitize_sheet(sname))) if multi else base
        desc = _excel_ingest_sheet(
            session, path, tbase, sname, skip, destination, progress,
            prefer_duckdb=prefer_duck)
        if desc is not None:
            out.append(desc)
    if not out:
        if not pick_all:
            raise RuntimeError(
                "Sheet %r wasn't found, or has no rows at/after row %d."
                % (want, header_row))
        raise RuntimeError("That workbook has no readable sheets.")
    return out


def excel_sheet_names(path):
    """Return the sheet names of an .xlsx/.xlsm/.xls workbook in order, without
    scanning rows, so the UI can offer a sheet picker before loading."""
    try:
        import openpyxl
    except Exception:
        raise RuntimeError("Reading Excel files needs the 'openpyxl' package, "
                           "which isn't installed.")
    if os.path.splitext(path)[1].lower() == ".xls":
        raise RuntimeError(
            "Legacy .xls (Excel 97-2003) isn't supported. Save the workbook "
            "as .xlsx and try again.")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        return [str(s) for s in wb.sheetnames]
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _norm_delim(delim):
    """Normalize a user-supplied delimiter to a single character, or ``None``
    for auto-detect. Empty / "auto" -> None; "\\t" / "tab" (or a literal tab)
    -> a real tab; otherwise the first character of the input."""
    if delim is None:
        return None
    s = str(delim)
    if s == "\t":
        return "\t"
    low = s.strip().lower()
    if low in ("", "auto"):
        return None
    if low in ("\\t", "tab"):
        return "\t"
    body = s.strip()
    return body[0] if body else None


def _duck_count(duck, name):
    try:
        _c, rows = duck.execute(f'SELECT COUNT(*) FROM "{name}"')
        return int(rows[0][0]) if rows else None
    except Exception:
        return None


def _load_into_duckdb(session, path, base_name, kind, delimiter=None,
                      json_depth=None, force_ondisk=False):
    """Load a file into DuckDB, choosing storage by size.

    A file at/above the on-disk threshold (SAMQL_ONDISK_MB) — or the hard
    floor (SAMQL_ONDISK_HARD_MB, default 256 MiB) — is streamed into a
    Parquet cache and exposed as a query-in-place view -- bounded memory for
    the whole lifecycle (the rows never live in the engine), with columnar
    column + row-group pushdown for fast repeat queries. An already-Parquet
    large file is just viewed in place (no re-copy). Smaller files
    materialise into an in-memory table, as before. Nested JSON uses a lower
    threshold (SAMQL_JSON_ONDISK_MB) because materialising structs/lists
    expands far beyond file size.

    When the on-disk Parquet route fails for JSON, retry once at
    ``json_depth=0`` (single JSON column, no STRUCT explosion). If that still
    fails, raise so ``_load_json_duckdb`` can stream-flatten instead of
    refusing the load. CSV/text may fall through to DuckDB CTAS (spill-aware)
    and are then promoted back to Parquet so multi-GB CSVs still end as an
    efficient on-disk cache. Nested JSON never uses that CTAS fallthrough
    (multi-GB STRUCT materialize risk).

    ``json_depth`` caps DuckDB nested-type expansion for JSON (flatten-off).
    ``force_ondisk`` prefers the Parquet-cache path even below the size floor
    (used when converting large nested JSON to a safer on-disk format).
    """
    from .engines import (_ondisk_min_bytes, _ondisk_hard_floor_bytes,
                          _json_ondisk_min_bytes, _is_interrupt,
                          ensure_json_engine_memory,
                          ensure_large_file_engine_memory)
    duck = session.get_duckdb()
    if kind == "json":
        ensure_json_engine_memory(duck, path)
    elif kind == "csv":
        ensure_large_file_engine_memory(duck, path)
    thresh = _ondisk_min_bytes()
    if kind == "json":
        jthresh = _json_ondisk_min_bytes()
        if jthresh is not None:
            thresh = jthresh if thresh is None else min(thresh, jthresh)
    hard = _ondisk_hard_floor_bytes()
    big = bool(force_ondisk)
    try:
        size = os.path.getsize(path)
    except OSError:
        size = 0
    if not big:
        if thresh is not None and size >= thresh:
            big = True
        elif hard is not None and size >= hard:
            big = True

    def _parquet_result(name, storage, depth):
        return [{"name": name, "rows": _duck_count(duck, name),
                 "columns": duck.table_columns.get(name, []),
                 "engine": "duckdb", "storage": storage,
                 "json_depth": depth}]

    def _promote_table_to_parquet(table_name, depth):
        """Materialize an in-engine table to a Parquet cache view (CSV
        CTAS recovery). Best-effort: on failure keep the table."""
        from . import tmputil
        origin = (getattr(duck, "table_origins", {}) or {}).get(table_name) \
            or (getattr(duck, "table_sources", {}) or {}).get(table_name) \
            or path
        cache = tmputil.new_tempfile("cache_", ".parquet")
        try:
            duck.view_to_parquet(table_name, cache)
            try:
                duck.drop_table(table_name)
            except Exception:
                pass
            name = duck.create_view_from_file(base_name, cache, "parquet")
            remember = getattr(duck, "_remember_origin", None)
            if remember:
                remember(name, origin)
            elif origin:
                getattr(duck, "table_origins", {})[name] = origin
            sys.stderr.write(
                "[samql] promoted DuckDB table %s → Parquet cache\n"
                % table_name)
            return _parquet_result(name, "parquet-cache", depth)
        except Exception as ex:
            try:
                os.unlink(cache)
            except OSError:
                pass
            sys.stderr.write(
                "[samql] Parquet promote failed (%s); keeping in-engine "
                "table\n" % ex)
            return [{"name": table_name,
                     "rows": _duck_count(duck, table_name),
                     "columns": duck.table_columns.get(table_name, []),
                     "engine": "duckdb", "storage": "table",
                     "json_depth": depth}]

    if big:
        try:
            if kind == "parquet":
                name = duck.create_view_from_file(base_name, path, "parquet")
                return _parquet_result(name, "parquet-view", json_depth)
            name = duck.load_file_to_parquet_view(
                base_name, path, kind, delimiter=delimiter,
                json_depth=json_depth)
            return _parquet_result(name, "parquet-cache", json_depth)
        except LoadCancelled:
            raise
        except Exception as e:
            if _is_interrupt(e):
                raise                       # a cancel: abort, don't fall back
            # Flatten-off shallow loads: one more Parquet attempt as a single
            # json column before handing off to the streaming flattener.
            # Flatten-on (json_depth=None) skips this — stream-flatten is the
            # better recovery than a opaque JSON-column Parquet.
            if kind == "json" and json_depth is not None and json_depth != 0:
                try:
                    sys.stderr.write(
                        "[samql] on-disk Parquet failed at depth=%s (%s); "
                        "retrying as JSON-column Parquet (depth=0)\n"
                        % (json_depth, e))
                    name = duck.load_file_to_parquet_view(
                        base_name, path, kind, delimiter=delimiter,
                        json_depth=0)
                    return _parquet_result(name, "parquet-cache", 0)
                except LoadCancelled:
                    raise
                except Exception as e2:
                    if _is_interrupt(e2):
                        raise
                    e = e2
            if kind == "json":
                # Do not CTAS nested JSON into engine RAM when size/force_ondisk
                # required the Parquet path. Surface a recoverable error so
                # _load_json_duckdb can stream-flatten (bounded memory) instead
                # of a silent in-engine materialize of multi-GB STRUCT nests.
                raise RuntimeError(
                    "On-disk Parquet load failed for large JSON %s (%s). "
                    "Streaming flatten will be tried next."
                    % (path, e)
                ) from e
            # CSV / delimited: spill-aware CTAS, then promote to Parquet so
            # multi-GB loads still end as an efficient on-disk cache.
            sys.stderr.write(
                "[samql] on-disk Parquet failed for %s (%s); "
                "materializing via DuckDB then converting to Parquet\n"
                % (path, e))
            name = duck.create_table_from_file(
                base_name, path, kind, delimiter=delimiter,
                json_depth=json_depth)
            return _promote_table_to_parquet(name, json_depth)
    name = duck.create_table_from_file(base_name, path, kind,
                                       delimiter=delimiter,
                                       json_depth=json_depth)
    return [{"name": name, "rows": _duck_count(duck, name),
             "columns": duck.table_columns.get(name, []), "engine": "duckdb",
             "json_depth": json_depth}]


def _json_file_size(path):
    try:
        return os.path.getsize(path)
    except OSError:
        return 0


def _stream_flatten_into_duckdb(session, path, base_name, progress=None,
                                spill=None, exclude=None):
    """Bounded-memory Python flatten → DuckDB tables (spill + streaming insert).

    Used for multi-GB single documents and as an OOM fallback when the native
    nested DuckDB read cannot allocate. Returns the same descriptor list as
    ``load_json``."""
    from .engines import ensure_json_engine_memory
    duck = session.get_duckdb()
    ensure_json_engine_memory(duck, path)
    return load_json(duck, path, base_name, progress=progress,
                     spill_threshold=spill if spill is not None
                     else JSON_SPILL_ROWS,
                     engine="duckdb", exclude=exclude)


def _shred_nested_json_load(session, res, root_id=None):
    """Run in-engine flatten_table on a nested DuckDB JSON load result."""
    tables = [t.get("table") or t.get("name") for t in res
              if isinstance(t, dict)]
    tbl = next((t for t in tables if t), None)
    # .501 audit: ONE on-load flatten engine. flatten_table owns the load
    # path; the UNNEST engine remains for right-click Flatten.
    fr = (session.flatten_table(tbl, base=tbl, root_id=root_id)
          if tbl else None)
    if fr and fr.get("cancelled"):
        raise LoadCancelled()
    if fr and not fr.get("error") and fr.get("created"):
        out = []
        for t in fr["created"]:
            out.append({"table": t["name"],
                        "rows": t.get("rows"),
                        "engine": "duckdb",
                        "method": "flatten-table",
                        "key": "_sk"})
        if fr.get("root_id"):
            out[0]["root_id_stats"] = fr["root_id"]
        return out
    if fr and fr.get("error"):
        sys.stderr.write(
            "[samql] on-load flatten failed (%s); the nested "
            "load stands\n" % fr["error"])
    else:
        sys.stderr.write(
            "[samql] on-load flatten reported nothing to "
            "do; the nested table stands\n")
    return res


def _load_json_duckdb(session, path, base_name, flatten=True, progress=None,
                      spill=None, exclude=None, root_id=None,
                      full_nested=False):
    """DuckDB JSON load for array / NDJSON / concat / single-object shapes.

    Every JSON format is converted before query use:

    1. Non-NDJSON shapes are rewritten to temp NDJSON (``_json_source_for_read``).
    2. NDJSON is COPY'd to an on-disk Parquet cache (file-cached when possible).
    3. Flatten-off uses shallow ``maximum_depth`` (default 2, then 0 on failure)
       so deep nesting stays JSON instead of exploding into STRUCTs.
    4. ``full_nested`` (shred post-pass): same as flatten-off routing but with
       uncapped nested types so a later ``flatten_table`` sees real STRUCT/LIST
       columns instead of depth-capped JSON scalars.
    5. Flatten-on loads full nested types from that Parquet, then runs
       in-engine ``flatten_table``. A multi-GB *single* object still uses the
       Python streaming flattener (spill) because one STRUCT cannot be
       materialised safely. Any Parquet-route failure (OOM or otherwise)
       also falls back to the streamer so large files still load.
    """
    from .engines import (
        _is_interrupt, _is_duckdb_oom, _sniff_json_format,
        _json_stream_flatten_min_bytes, _json_shallow_depth_default,
        _json_ondisk_min_bytes, _ondisk_min_bytes,
    )
    size = _json_file_size(path)
    fmt = _sniff_json_format(path)
    stream_flat_min = _json_stream_flatten_min_bytes()
    giant_object = (
        fmt == "object"
        and stream_flat_min is not None
        and size >= stream_flat_min
    )

    if flatten and giant_object:
        # One multi-hundred-MB / multi-GB document: native nested read would
        # buffer the whole STRUCT. Stream-flatten with spill instead.
        sys.stderr.write(
            "[samql] large single-object JSON (%d MiB); "
            "streaming flatten into DuckDB\n" % (size // (1024 * 1024)))
        return _stream_flatten_into_duckdb(
            session, path, base_name, progress=progress,
            spill=spill, exclude=exclude)

    def _native(json_depth=None, force_ondisk=False):
        return _load_into_duckdb(
            session, path, base_name, "json",
            json_depth=json_depth,
            force_ondisk=force_ondisk)

    def _ondisk_for_size():
        thresh = _json_ondisk_min_bytes()
        if thresh is None:
            thresh = _ondisk_min_bytes()
        return bool(thresh is not None and size >= thresh)

    if not flatten and full_nested:
        # UI Flatten/shred toggle: need full STRUCT/LIST types for the
        # relational shred pass, but do NOT shred here (session does that).
        # Prefer Parquet only above the size floor so small files stay fast.
        force_ondisk = _ondisk_for_size()
        sys.stderr.write(
            "[samql] JSON load (shred prep, %s): %s with full nested types\n"
            % (fmt,
               ("converting to Parquet" if force_ondisk
                else "materializing")))
        try:
            return _native(json_depth=None, force_ondisk=force_ondisk)
        except LoadCancelled:
            raise
        except Exception as e:
            if _is_interrupt(e):
                raise
            sys.stderr.write(
                "[samql] full-nested JSON load failed (%s); "
                "streaming flatten into DuckDB\n" % e)
            return _stream_flatten_into_duckdb(
                session, path, base_name, progress=progress,
                spill=spill, exclude=exclude)

    if not flatten:
        # Flatten-off: shallow depth. Prefer on-disk Parquet only when the file
        # is large enough -- small files CTAS + struct-expand so keys show as
        # columns without a full relational shred. Large files stay Parquet-
        # backed for bounded memory.
        force_ondisk = _ondisk_for_size()
        depths = [_json_shallow_depth_default()]
        if depths[0] != 0:
            depths.append(0)
        last_err = None
        for i, depth in enumerate(depths):
            try:
                if i > 0:
                    sys.stderr.write(
                        "[samql] nested JSON failed at maximum_depth=%s "
                        "(%s); retrying as JSON-column%s\n"
                        % (depths[i - 1], last_err,
                           " Parquet" if force_ondisk else ""))
                else:
                    sys.stderr.write(
                        "[samql] JSON load (flatten off, %s): "
                        "%s with maximum_depth=%d\n"
                        % (fmt,
                           ("converting to Parquet" if force_ondisk
                            else "materializing"),
                           depth))
                return _native(json_depth=depth, force_ondisk=force_ondisk)
            except LoadCancelled:
                raise
            except Exception as e:
                if _is_interrupt(e):
                    raise
                last_err = e
                if i + 1 >= len(depths):
                    break
                continue
        # Native path exhausted. Do NOT silently stream-flatten here: that
        # violates Flatten-off (creates many relational tables) and on a
        # multi-GB nested file can run for hours looking like a stall.
        # Flatten-on still uses the streamer as a recovery path below.
        raise RuntimeError(
            "Could not load nested JSON with Flatten off (%s). "
            "Convert the file to .ndjson / .jsonl, free disk for the temp "
            "NDJSON→Parquet rewrite, or turn Flatten on if you want "
            "relational tables." % last_err) from last_err

    # Flatten on: always convert to Parquet first (full nested types), then
    # in-engine shred. Any Parquet-route failure (OOM or otherwise) →
    # Python streamer (bounded memory) so large files still load.
    sys.stderr.write(
        "[samql] JSON load (flatten on, %s): converting to Parquet "
        "then shred\n" % fmt)
    try:
        res = _native(force_ondisk=True)
    except LoadCancelled:
        raise
    except Exception as e:
        if _is_interrupt(e):
            raise
        why = "OOM" if _is_duckdb_oom(e) else "Parquet conversion failed"
        sys.stderr.write(
            "[samql] nested JSON %s during flatten-on load (%s); "
            "streaming flatten into DuckDB\n" % (why, e))
        return _stream_flatten_into_duckdb(
            session, path, base_name, progress=progress,
            spill=spill, exclude=exclude)
    try:
        return _shred_nested_json_load(session, res, root_id=root_id)
    except LoadCancelled:
        raise
    except Exception as e:
        if _is_interrupt(e):
            raise
        sys.stderr.write(
            "[samql] on-load flatten failed (%s); the "
            "nested load stands\n" % e)
        return res


def load_file(session, path, destination="sqlite", base_name=None,
              progress=None, delimiter=None, mode="materialize",
              sheet=None, header_row=1, exclude=None, flatten=None,
              root_id=None, full_nested=False):
    """Dispatch a single file to the right loader based on extension and
    the requested destination engine. ``session`` provides .db and
    .get_duckdb(). ``progress`` (optional) is called as
    progress(bytes_done, bytes_total) during a CSV->SQLite load.
    ``delimiter`` (CSV / delimited text only) forces a single-character field
    separator instead of auto-detection; it is ignored for JSON / Parquet /
    Excel. ``mode`` is "materialize" (default; copy into a table) or "view"
    (query the file in place via a DuckDB view -- no copy). View mode needs
    DuckDB and a streamable format (CSV / JSON / Parquet); anything else
    (Excel, or DuckDB unavailable) safely falls back to a materialized load.
    ``full_nested`` (JSON + DuckDB, flatten off) loads uncapped STRUCT/LIST
    types for a subsequent shred pass without running shred in the loader.
    Returns a list of loaded-table descriptors."""
    # "auto"/"default" means "the best engine for this machine" -- DuckDB when
    # installed (far better at large files), else SQLite. The Session.load_file
    # wrapper resolves this before calling us, but the appendfolder / directory
    # readers call here directly, so resolve it here too. Without this, an
    # "auto" load fell through every `destination == "duckdb"` check below and
    # silently landed in SQLite.
    if destination in (None, "", "auto", "default"):
        try:
            destination = session.default_destination()
        except Exception:
            destination = "sqlite"
    ext = os.path.splitext(path)[1].lower().lstrip(".")
    base_name = base_name_for(base_name or path)
    if mode == "view" and ext not in ("xlsx", "xlsm", "xls"):
        # query-in-place: a DuckDB view over read_*. Needs DuckDB; if it's
        # unavailable we fall through to a normal materialized load.
        # Large CSV/JSON and non-NDJSON JSON must not use a live view: CSV/JSON
        # views re-parse every query, and top-level JSON arrays buffer in
        # DuckDB on each scan. Prefer the Parquet-cache materialize path.
        try:
            duck = session.get_duckdb()
        except Exception:
            duck = None
        if duck is not None:
            kind = ("parquet" if ext in ("parquet", "pq")
                    else "json" if ext in ("json", "ndjson", "jsonl")
                    else "csv")
            use_view = True
            if kind == "json":
                from .engines import (
                    _sniff_json_format, _is_ndjson_path,
                    _json_ondisk_min_bytes, _ondisk_min_bytes)
                if (_sniff_json_format(path) != "ndjson"
                        and not _is_ndjson_path(path)):
                    use_view = False
                else:
                    thresh = _json_ondisk_min_bytes()
                    if thresh is None:
                        thresh = _ondisk_min_bytes()
                    try:
                        if thresh is not None and os.path.getsize(path) >= thresh:
                            use_view = False
                    except OSError:
                        pass
            elif kind == "csv":
                from .engines import _ondisk_min_bytes
                thresh = _ondisk_min_bytes()
                try:
                    if thresh is not None and os.path.getsize(path) >= thresh:
                        use_view = False
                except OSError:
                    pass
            if use_view:
                name = duck.create_view_from_file(
                    base_name, path, kind, delimiter=_norm_delim(delimiter))
                cols = duck.table_columns.get(name, [])
                return [{"name": name, "rows": None, "columns": cols,
                         "engine": "duckdb", "view": True}]
        # else: fall through and materialize instead
    if ext in ("xlsx", "xlsm", "xls"):
        return load_excel(session, path, base_name, destination=destination,
                          sheet=sheet, header_row=header_row,
                          progress=progress)
    if ext in ("parquet", "pq"):
        # Small Parquet materializes into DuckDB storage (fast repeat queries);
        # a large one is viewed in place (columnar pushdown, bounded memory).
        # See _load_into_duckdb for the size split.
        return _load_into_duckdb(session, path, base_name, "parquet")
    if ext in ("json", "ndjson", "jsonl"):
        # per-load override wins (the Load modal decides now); None keeps the
        # legacy session default for callers that don't say (folder appends,
        # node loads)
        if flatten is None:
            flatten = bool(getattr(session, "flatten_on_load", False))
        flatten = bool(flatten)
        spill = (10_000 if getattr(session, "low_memory", False)
                 else JSON_SPILL_ROWS)
        if destination == "duckdb":
            return _load_json_duckdb(
                session, path, base_name, flatten=flatten,
                progress=progress, spill=spill, exclude=exclude,
                root_id=root_id, full_nested=bool(full_nested))
        # SQLite (or any non-DuckDB engine) always flattens -- it has no nested
        # column types -- so the toggle only affects the DuckDB path above.
        return load_json(session.db, path, base_name, progress=progress,
                         spill_threshold=spill, exclude=exclude)
    # default: treat as delimited text
    ndelim = _norm_delim(delimiter)
    if destination == "duckdb":
        return _load_into_duckdb(session, path, base_name, "csv",
                                 delimiter=ndelim)
    # Large CSV into SQLite thrash-queries and can OOM on analytics. When
    # DuckDB is available, auto-upgrade to the on-disk Parquet path so
    # multi-GB / multi-million-row CSVs still load. Only fail loud when
    # DuckDB is unavailable.
    from .engines import (_ondisk_min_bytes, _ondisk_hard_floor_bytes,
                          HAS_DUCKDB)
    thresh = _ondisk_min_bytes()
    hard = _ondisk_hard_floor_bytes()
    try:
        sz = os.path.getsize(path)
    except OSError:
        sz = 0
    large = ((thresh is not None and sz >= thresh)
             or (hard is not None and sz >= hard))
    if large:
        if HAS_DUCKDB:
            sys.stderr.write(
                "[samql] CSV is %.0f MiB — using DuckDB on-disk Parquet "
                "instead of SQLite\n" % (sz / (1024 * 1024.0)))
            return _load_into_duckdb(session, path, base_name, "csv",
                                     delimiter=ndelim, force_ondisk=True)
        raise RuntimeError(
            "This CSV is %.0f MiB — too large for SQLite, and DuckDB is "
            "not installed. Install duckdb (pip install duckdb) so SamQL "
            "can stream it to on-disk Parquet."
            % (sz / (1024 * 1024.0)))
    return [load_csv(session.db, path, base_name, sep=ndelim,
                     progress=progress)]
