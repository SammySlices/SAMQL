"""Session: the headless orchestration layer.

Owns the local engines (SQLite always, DuckDB lazily), routes queries
between them the way the original App did (auto-route + reactive retry on
"missing table"), caches result sets off-heap for paged/sorted browsing,
and exposes loads, profiling, exports, the tables tree, and persistence.

Everything here is GUI-free and safe to drive from an HTTP handler.
"""
import csv as _csv
import datetime as _dt
import hashlib
import json
import os
import re
import threading
import time
import sqlite3
import tempfile
import uuid

# .414: the reconcile summary as ONE full-outer-join aggregate. The
# toggle exists for tests (parity vs the legacy statements) and as an
# escape hatch; any runtime surprise already falls back silently.
RECON_SINGLE_PASS = True
# SQLite plans a null-safe join (`IS`) as a NESTED LOOP -- O(N*M), a
# hang on six-figure inputs -- while DuckDB hash-joins IS NOT DISTINCT
# FROM natively. So the one-pass is DuckDB-only in production; this
# test-only flag lets the suite execute the SQLite variant on a tiny
# fixture to keep the parity proof honest.
RECON_SINGLE_PASS_SQLITE = False

from . import loaders
from . import applyvars
from . import progress as opreg
from . import watchdog
from .engines import (DBManager, DuckDBManager, HAS_DUCKDB,
                      total_physical_ram_bytes, _ExecKeepalive)
from .nodeflow import sanitize_ident, NodeflowError
from .profiler import profile_table, _infer_type
from .rows import (DiskBackedRows, ParquetResultStore, _to_number,
                   agg_group_multi, agg_minmax, agg_histogram, agg_xy)
from . import tmputil
from .flowcache import FlowCache, PersistentFlowCache
from . import resourcebudget
from .errfmt import err_str
from . import sqlutil
from .sqlutil import (classify_sql_statement, find_statement_at,
                      split_sql_statements_spans, split_statements,
                      sqlglot_transform)
from .stores import (ConfigStore, QueryHistoryStore, SavedQueryStore,
                     LoadManifestStore, WorkflowStore)

DUCKDB_TARGET = "__duckdb__"
LOCAL_TARGET = "__local__"

# Node types that, when the incremental flow cache is on, are always
# materialised (rather than fused into a downstream query) so their output
# becomes a content-addressed cache checkpoint. These are the blocking, often
# expensive multi-input ops -- exactly the heavy upstream work you don't want
# to recompute every time you tweak a node downstream of them. (Pivots are
# always materialised regardless, for a different reason: their columns are
# only known at run time.)
_FLOW_CHECKPOINT_TYPES = frozenset({
    "join", "multijoin", "crossjoin", "antijoin", "union", "reconcile"})
DISPLAY_LIMIT = 5000
INITIAL_PAGE_ROWS = 1000   # first page returned inline; grid lazy-loads the rest

_MISSING_TABLE_PATTERNS = (
    re.compile(r"no such table:\s*([^\s]+)", re.IGNORECASE),
    re.compile(
        r"(?:Catalog Error|Binder Error)[^\n]*?"
        r"(?:Table|table) (?:with name )?[\"']?([\w.]+)[\"']?"
        r" does not exist", re.IGNORECASE),
    re.compile(
        r"Table (?:with name )?[\"']?([\w.]+)[\"']?"
        r" does not exist", re.IGNORECASE),
)


# Substrings (lower-cased) DuckDB uses when it runs out of working memory.
# Shared by the query- and flow-error enrichers and by the read-path fallback
# guard so all three agree on what "an OOM" looks like.
_OOM_MARKERS = ("outofmemoryexception", "out of memory",
                "failed to allocate", "could not allocate")


def _looks_like_oom(message):
    """True when an engine error text is a DuckDB out-of-memory failure."""
    lowered = str(message or "").lower()
    return any(m in lowered for m in _OOM_MARKERS)


def _duckdb_oom_query_message(message, engine=None):
    """Add actionable context to a DuckDB query out-of-memory error.

    DuckDB reports the allocation that *failed*, not its configured
    ``memory_limit``.  In particular, a 16 MiB request normally means the
    query has already exhausted the larger working-memory budget.
    """
    text = str(message or "")
    if not _looks_like_oom(text):
        return text
    budget = getattr(engine, "_applied_resource_memory_mb", None)
    try:
        budget_note = (" Current engine budget: ~%d MiB." % int(budget)
                       if budget else "")
    except (TypeError, ValueError):
        budget_note = ""
    return (
        "%s\n\nDuckDB exhausted its working-memory budget while executing "
        "this query.%s The failed allocation size is not the configured "
        "memory limit. Free unused memory or raise Engine tuning → memory "
        "only when RAM and temporary-disk capacity allow it. For nested JSON, "
        "extract/project the fields you need before UNNEST; expand one array "
        "at a time and avoid recursive or sibling-array cross-products."
        % (text, budget_note)
    )


def _duckdb_oom_flow_message(message, engine=None):
    """OOM guidance for a NodeFlow node/branch failure.

    Mirrors :func:`_duckdb_oom_query_message` (keeps the raw error, notes the
    failed-alloc vs limit distinction and the engine budget) and adds the same
    memory-safe next steps NodeFlow already exposes -- so a large-data flow OOM
    gets the actionable guidance IDE/Journal queries already get, instead of a
    bare ``OutOfMemoryException``. A non-OOM message is returned unchanged.
    """
    text = str(message or "")
    if not _looks_like_oom(text):
        return text
    base = _duckdb_oom_query_message(text, engine)
    return (
        "%s\n\nIn a flow, add a Filter or a row limit before the heavy node "
        "and connect only the columns you need. For a nested/opaque JSON "
        "column, turn it into real columns with a Shred node (or Flatten it "
        "upstream) instead of UNNEST-ing the whole column in one step."
        % base
    )


def json_safe(v):
    """Coerce one cell value to something json.dumps can serialize."""
    if v is None or isinstance(v, (bool, int, str)):
        return v
    if isinstance(v, float):
        if v != v or v in (float("inf"), float("-inf")):
            return str(v)
        return v
    if isinstance(v, (dict, list)):
        # Keep nested values valid JSON on the wire.  ``str(dict)`` produced a
        # Python repr with single quotes, which was only heuristically
        # displayable and made browser assertions / copy-paste inconsistent.
        # The page cap has already bounded pathological values before this
        # conversion, so this is safe for normal result pages and the dedicated
        # full-cell endpoint.
        try:
            return json.dumps(v, default=str, ensure_ascii=False,
                              separators=(",", ":"))
        except Exception:
            return str(v)
    if isinstance(v, (bytes, bytearray)):
        try:
            return v.decode("utf-8", "replace")
        except Exception:
            return repr(v)
    if isinstance(v, (_dt.datetime, _dt.date, _dt.time)):
        try:
            return v.isoformat()
        except Exception:
            return str(v)
    try:
        import decimal
        if isinstance(v, decimal.Decimal):
            f = float(v)
            return int(v) if f == int(f) else f
    except Exception:
        pass
    return str(v)


CELL_DISPLAY_CHARS = 1000        # max characters shown per cell in the grid
PAGE_DISPLAY_BUDGET = 8_000_000  # ~8 MB soft ceiling on one serialized page


def _cap_page_rows(rows, per_cell=CELL_DISPLAY_CHARS,
                   page_budget=PAGE_DISPLAY_BUDGET):
    """Bound the size of a serialized result page for display.

    A flatten-off JSON load leaves whole records in a single nested (STRUCT/
    LIST/JSON) column, so one cell can be megabytes and a 1000-row page can be
    hundreds of MB -- enough to hang or reload the browser tab on `SELECT *`.
    This truncates any oversized cell to a preview (with a marker of the true
    length): string cells directly, and STRUCT/LIST/MAP cells -- which DuckDB
    returns as Python dict/list, not str -- by serializing them to text first.
    As a hard backstop, once the page's cumulative text passes a byte budget,
    every further oversized cell is truncated hard. Scalar cells (numbers/
    bools/None) and small nested cells are untouched, so normal results are
    unaffected -- this only ever changes the DISPLAY of very large values; the
    underlying data is intact and still fully queryable (select the nested
    sub-fields, or use the structure view to see the paths). Pure -> unit
    testable. Returns (rows, capped) where `capped` is True if anything was
    shortened."""
    used = 0
    capped = False
    out = []
    for r in rows:
        rr = []
        for v in r:
            if isinstance(v, str):
                n = len(v)
                if used >= page_budget:
                    if n > 120:
                        v = v[:120] + "… [%d chars — truncated]" % n
                        capped = True
                elif n > per_cell:
                    v = v[:per_cell] + "… [%d chars — truncated]" % n
                    capped = True
                used += len(v)
            elif isinstance(v, (dict, list)):
                # .499: STRUCT / LIST / MAP columns come back from DuckDB as
                # Python dict / list, NOT str -- so the string branch above
                # never bounded them. A flatten-OFF nested column can be MBs per
                # cell, and ~1000 such rows a page of hundreds of MB, which
                # froze the browser tab receiving/parsing it -- and the server's
                # single-shot response write then wedged behind the frozen tab
                # (threads stuck in wfile.write, saturating the connection pool
                # so Stop/reset couldn't get through). Bound these too: serialize
                # to text and preview it past the per-cell / page budget. The
                # full value stays available via the cell endpoint, and the data
                # is still fully queryable (select sub-fields / use the structure
                # view). Small nested cells keep their native shape.
                try:
                    s = json.dumps(v, default=str, ensure_ascii=False)
                except Exception:
                    s = str(v)
                n = len(s)
                limit = 120 if used >= page_budget else per_cell
                if n > limit:
                    v = s[:limit] + "… [%d chars — truncated]" % n
                    capped = True
                    used += len(v)
                else:
                    used += n
            rr.append(v)
        out.append(rr)
    return out, capped


def query_hint_for_column(name, type_str):
    """A short 'how to query this column' hint for a NESTED column type, or None
    for a plain scalar (which needs no explanation). Detects DuckDB nested kinds
    from the (possibly upper-cased) type string, so a person who doesn't know
    the syntax can see how to reach into a struct / list / json / map column.

    Examples returned (name = json):
      list of records  ->  "list of records — one element: json[1].field ·
                            all rows: UNNEST(json)"
      record (struct)  ->  "record — a field: json.field"
      json blob        ->  "JSON — a field: json ->> '$.field'\""""
    import re as _re
    t = (type_str or "").strip()
    tu = t.upper()
    qn = (name if _re.match(r"^[A-Za-z_][A-Za-z0-9_]*$", name or "")
          else '"%s"' % name)
    is_list = tu.endswith("[]")
    base = (tu[:-2] if is_list else tu).strip()
    if base.startswith("STRUCT"):
        if is_list:
            return ("list of records — one element: %s[1].field · "
                    "all rows: UNNEST(%s)" % (qn, qn))
        return "record — a field: %s.field" % qn
    if base.startswith("MAP"):
        return "map — a value: %s['key']" % qn
    if base == "JSON":
        return "JSON — a field: %s ->> '$.field'" % qn
    if is_list:
        return "list — one element: %s[1] · all rows: UNNEST(%s)" % (qn, qn)
    return None


def _cols_with_hints(cols, types):
    """Build the tables-tree column entries, attaching a query hint to each
    column whose type is nested (scalars get no hint -- they need none)."""
    out = []
    # .464 pass II hardening: an engine (or a test double) answering None
    # for its type map must not take the whole tables tree down.
    types = types or {}
    for c in cols:
        ty = types.get(c, "")
        entry = {"name": c, "type": ty}
        h = query_hint_for_column(c, ty)
        if h:
            entry["hint"] = h
        out.append(entry)
    return out


def _estimate_row_bytes(row):
    """Rough in-memory size of one result row (a tuple of values). Cheap and
    approximate -- enough to keep the result cache under a memory ceiling
    rather than a pure row count, so very wide rows are accounted for."""
    n = 56  # tuple overhead
    for v in row:
        if v is None:
            n += 8
        elif isinstance(v, bool):
            n += 28
        elif isinstance(v, int):
            n += 28
        elif isinstance(v, float):
            n += 24
        elif isinstance(v, str):
            n += len(v) * 2 + 49
        elif isinstance(v, bytes):
            n += len(v) + 33
        else:
            n += 48
    return n


class _CachedResult:
    __slots__ = ("id", "cols", "store", "total", "sql", "target",
                 "engine", "created", "_sorted_cache", "capped", "cap",
                 "_fcount_cache")

    def __init__(self, rid, cols, store, total, sql, target, engine):
        self.id = rid
        self.cols = cols
        self.store = store        # list[tuple] OR DiskBackedRows
        self.total = total
        self.sql = sql
        self.target = target
        self.engine = engine
        self.created = time.time()
        self._sorted_cache = {}
        # a result whose materialization hit the safety row ceiling
        self.capped = bool(getattr(store, "capped", False))
        self.cap = getattr(store, "cap", None)

    def view(self, sort_col=None, descending=False):
        if sort_col is None or sort_col not in self.cols:
            return self.store
        ix = self.cols.index(sort_col)
        if hasattr(self.store, "sorted_view"):
            return self.store.sorted_view(ix, descending)
        key = (ix, descending)
        v = self._sorted_cache.get(key)
        if v is None:
            def _k(r):
                val = r[ix] if ix < len(r) else None
                return (val is None, _sort_key(val))
            v = sorted(self.store, key=_k, reverse=descending)
            # Keep only the most recent sorted copy so toggling sort
            # columns can't pile up multiple full copies of a big list.
            self._sorted_cache = {key: v}
        return v

    def filtered_view(self, filters, sort_col=None, descending=False):
        """Return (view, total) for a filtered (and optionally sorted)
        projection. Engine-side WHERE when the store supports it (spilled
        SQLite results); otherwise a Python predicate over any iterable
        store (small in-memory lists and Parquet-backed results)."""
        terms = []
        for f in (filters or []):
            col = (f or {}).get("column")
            if col not in self.cols:
                continue
            op = _normalize_op((f or {}).get("op"))
            terms.append((self.cols.index(col), op, (f or {}).get("value")))
        if not terms:
            return self.view(sort_col, descending), self.total
        sort_ix = (self.cols.index(sort_col)
                   if sort_col in self.cols else None)
        store = self.store
        if hasattr(store, "filtered_view") and hasattr(store, "count_view"):
            # a filtered page used to run count_view (a full filtered COUNT
            # over the store) on EVERY page fetch -- on a multi-million-row
            # parquet result that's a real scan per scroll step. Results are
            # immutable, so cache the count per canonical filter set (small
            # FIFO so hostile filter churn can't grow it).
            key = repr(terms)
            cache = getattr(self, "_fcount_cache", None)
            if cache is None:
                cache = self._fcount_cache = {}
            total = cache.get(key)
            if total is None:
                total = store.count_view(terms)
                if len(cache) >= 8:
                    cache.pop(next(iter(cache)))
                cache[key] = total
            return store.filtered_view(terms, sort_ix, descending), total
        pred = _row_predicate(terms)
        filtered = [r for r in store if pred(r)]
        if sort_ix is not None:
            def _k(r):
                val = r[sort_ix] if sort_ix < len(r) else None
                return (val is None, _sort_key(val))
            filtered.sort(key=_k, reverse=descending)
        return filtered, len(filtered)

    def close(self):
        try:
            if hasattr(self.store, "close"):
                self.store.close()
        except Exception:
            pass
        self._sorted_cache.clear()


def _sort_key(v):
    if v is None:
        return (0, 0.0)
    if isinstance(v, bool):
        return (1, float(v))
    if isinstance(v, (int, float)):
        return (1, float(v))
    return (2, str(v))


# Filter operators are word-based everywhere (gt/gte/lt/lte/equals/ne/...), but
# be forgiving if a caller (or a direct API hit) sends the symbolic form. An
# unrecognised op would otherwise fall through to "match everything", silently
# turning a filter into a no-op.
_OP_ALIASES = {
    ">": "gt", ">=": "gte", "=>": "gte", "\u2265": "gte",
    "<": "lt", "<=": "lte", "=<": "lte", "\u2264": "lte",
    "=": "equals", "==": "equals",
    "!=": "ne", "<>": "ne", "\u2260": "ne",
}


def _normalize_op(op):
    raw = str(op or "contains").strip()
    return _OP_ALIASES.get(raw, raw.lower())


def _match(v, op, val):
    """Python-side equivalent of the SQLite filter semantics, used for
    in-memory and Parquet-backed result stores."""
    if op == "is_null":
        return v is None
    if op == "not_null":
        return v is not None
    if v is None:
        return False
    if op in ("contains", "starts", "ends"):
        s = str(v).lower()
        t = str(val if val is not None else "").lower()
        if op == "contains":
            return t in s
        if op == "starts":
            return s.startswith(t)
        return s.endswith(t)
    nv = _to_number(v)
    tv = _to_number(val)
    if nv is not None and tv is not None:
        a, b = nv, tv
    else:
        a, b = str(v), str(val)
    if op == "equals":
        return a == b
    if op == "ne":
        return a != b
    if op == "gt":
        return a > b
    if op == "gte":
        return a >= b
    if op == "lt":
        return a < b
    if op == "lte":
        return a <= b
    return True


def _row_predicate(terms):
    def ok(r):
        for ix, op, val in terms:
            v = r[ix] if ix < len(r) else None
            if not _match(v, op, val):
                return False
        return True
    return ok


class _NullCtx:
    def __enter__(self):
        return self

    def __exit__(self, *exc):
        return False


def _is_interrupt(exc):
    """True if an exception came from a query being interrupted/cancelled."""
    name = exc.__class__.__name__.lower()
    s = str(exc).lower()
    return ("interrupt" in name or "interrupt" in s or "cancel" in s)


def _rename_row_vars(row, rename):
    """Turn a driver row's columns into scalar-variable bindings for one
    iterator pass, applying an optional {original_column: new_name} rename so
    the iterator's renamed fields become the ${variables} usable inside the
    loop. A column with no (or blank) rename keeps its original name; values are
    stringified (None -> "")."""
    rename = rename or {}
    out = {}
    for k, v in (row or {}).items():
        name = str(rename.get(str(k)) or "").strip() or str(k)
        out[name] = "" if v is None else str(v)
    return out


def _iter_unquoted_var_hint(msg, extra):
    """When an iterator pass fails because a ``${var}`` substituted a *text*
    value into SQL as a bare identifier -- the classic ``no such column: dog``
    from writing ``${col1}`` where ``'${col1}'`` was meant -- append a hint to
    quote it. Returns ``msg`` unchanged when it doesn't look like that case
    (numbers are valid bare, so only text values are flagged)."""
    low = (msg or "").lower()
    if not any(s in low for s in (
            "no such column", "referenced column", "column not found",
            "not found in from", 'column "', "binder error")):
        return msg
    for name, val in (extra or {}).items():
        sval = "" if val is None else str(val)
        if not sval.strip():
            continue
        try:                       # numbers are valid bare -- don't flag them
            float(sval)
            continue
        except (TypeError, ValueError):
            pass
        if sval in msg:
            return ("%s  (Tip: \"%s\" is the value of ${%s} this pass. For a "
                    "text value use {{%s}} -- it quotes automatically -- "
                    "instead of a bare ${%s}, which is read as a column name.)"
                    % (msg, sval, name, name, name))
    return msg


def _iter_missing_accum_hint(msg, graph):
    """If an error is 'no such table: X' (or a DuckDB equivalent) and X is an
    iterator/while accumulator in this graph, rewrite it to say: run the
    iterator first. The accumulator is only created once the loop runs, so
    previewing its out port (or a downstream reader) before that fails with a
    bare 'missing table' the user can't place. Returns msg unchanged otherwise."""
    import re as _re
    low = (msg or "").lower()
    if not any(s in low for s in (
            "no such table", "does not exist", "table with name")):
        return msg
    if not isinstance(graph, dict):
        return msg
    for n in (graph.get("nodes") or []):
        if not isinstance(n, dict) or n.get("type") not in (
                "iterator", "while"):
            continue
        disp = ((n.get("config") or {}).get("table") or "").strip()
        norm = sanitize_ident(disp)
        if norm and (norm in msg or (disp and disp in msg)):
            return ("The iterator's table \"%s\" doesn't exist yet -- run the "
                    "iterator first (its Run button, or Run all) to build it, "
                    "then read its output." % (disp or norm))
    return msg


def _source_is_json_file(path):
    """True iff ``path`` is a JSON/NDJSON file on disk a flatten can read
    directly. Anything else on a DuckDB table (a CSV/Parquet/text source, a
    Parquet cache from a large load, or a table with no file) is flattened by
    dumping the table to JSON first, so this only gates that direct-read
    shortcut -- it must not decide *whether* a DuckDB table can be flattened."""
    try:
        if not path or not os.path.isfile(path):
            return False
        return os.path.splitext(path)[1].lower().lstrip(".") in (
            "json", "ndjson", "jsonl")
    except Exception:
        return False


def _acquire_read_slot(slots, engine, timeout=0.35):
    """Grab a read-cursor slot: instantly when one is free; with a short
    bounded wait when the pool is saturated AND the engine's write lock
    is held (the locked fallback would stall behind a build);
    immediately False otherwise (an idle main connection makes the
    locked path cheap). Module-level so a bare test double works. .426"""
    if slots is None:
        return True
    if slots.acquire(blocking=False):
        return True
    # .438 audit fix: _DeadlineRLock has no .locked(), so the busy
    # check ALWAYS fell to False and the bounded wait never engaged --
    # the .426 enhancement was silently dead on the real engine. Probe
    # instead: a non-blocking acquire from THIS (reader) thread
    # succeeds only when the writer lock is free (RLock reentrancy
    # cannot help a different thread).
    lk = getattr(engine, "write_lock", None) or getattr(
        engine, "lock", None)
    busy = _stmt_stepping(engine)
    if lk is not None:
        try:
            got = lk.acquire(blocking=False)
            if got:
                lk.release()
            else:
                busy = True
        except TypeError:
            try:  # locks whose acquire lacks kwargs
                got = lk.acquire(False)
                if got:
                    lk.release()
                else:
                    busy = True
            except Exception:
                busy = False
        except Exception:
            busy = False
    if not busy:
        return False
    return bool(slots.acquire(timeout=timeout))


def _stmt_stepping(engine):
    """True when the engine reports a statement mid-step. Robust to
    test doubles whose attributes are not ints (.464)."""
    try:
        return int(getattr(engine, "stmt_busy", 0) or 0) > 0
    except Exception:
        return False


class _ExportCancelled(Exception):
    """Raised by _CancelIter when an export's Stop lands mid-stream."""


class _CancelIter:
    """.533: wrap an export's row view so the streaming writers become
    cancellable WITHOUT touching each writer's loop -- the check runs
    every 512 rows, cheap and prompt."""

    def __init__(self, it, cancelled):
        self._it = iter(it)
        self._c = cancelled
        self._n = 0

    def __iter__(self):
        return self

    def __next__(self):
        self._n += 1
        if (self._n & 511) == 0 and self._c():
            raise _ExportCancelled()
        return next(self._it)


class Session:
    MAX_CACHED_RESULTS = 24
    # Fetch a whole display page up front, so any result that fits on one
    # page never needs an off-heap store at all.
    _RESULT_BATCH = DISPLAY_LIMIT
    # Rows pulled per fetch when draining a larger result.
    _FETCH_BATCH = 20000

    def __init__(self, low_memory=False):
        # .501: {child_table: parent_table} for every table a flatten created,
        # recorded by BOTH flatten engines. The sidebar family tree and the
        # family-join SQL come from THIS map, not from name-prefix guessing --
        # path-only child names ("json", "legs") carry no parent in the name.
        # It also lets a re-flatten of the same source reuse its own family's
        # names instead of suffixing new copies. Session-lifetime, like the
        # DuckDB catalog it describes.
        self._table_family = {}
        # User-driven Loaded-tables sidebar order (session-scoped). List of
        # "engine:name" keys. Empty means catalog default (alphabetical).
        # Flatten/load may still append new tables; unknown keys sort after
        # ranked ones so a drag reorder never fights registration order.
        self._table_ui_order = []
        self.config = ConfigStore()
        self.history = QueryHistoryStore()
        self.saved = SavedQueryStore()
        self.workflows = WorkflowStore()
        self._migrate_saved_queries_to_workflows()
        # session restore: remember loads, replay them on the next launch
        self.manifest = LoadManifestStore()
        self.restoring = False
        self._restored_count = 0
        self._restore_started = False
        # Cache of table row counts so /api/tables (and the sidebar's column
        # expansion, which rides the same payload) doesn't re-run COUNT(*) on
        # every refresh -- that is what made expanding a freshly-loaded large
        # table feel slow. Invalidated on any table mutation.
        self._count_cache = {}
        self.low_memory = bool(low_memory or
                               self.config.get("low_memory_mode", False))
        # Materialize large DuckDB results to a temp Parquet file and page
        # from it (columnar, off-heap). Can be disabled via config.
        self.use_parquet_results = bool(
            self.config.get("parquet_results", True))
        # Fuse runs of SQL-only, single-consumer nodes into one nested query
        # instead of one temp table per node. Can be disabled via config.
        self.fuse_flows = bool(self.config.get("fuse_flows", True))
        # projection pushdown: source nodes read only the columns used
        # downstream. Can be disabled via config if ever suspect.
        self.project_pushdown = bool(self.config.get("project_pushdown", True))
        # Incremental flow cache: reuse a node's materialised output across
        # previews/runs when its subtree + projection + the data epoch are all
        # unchanged. Content-addressed (each cached table is named by its
        # fingerprint) and LRU-capped; a data mutation bumps the epoch and
        # drops the cache. Big win when iterating a downstream node (pivot,
        # chart) over heavy stable upstream work.
        self.flow_cache = bool(self.config.get("flow_cache", True))
        self.flow_cache_max = int(self.config.get("flow_cache_max", 32) or 0)
        # Adaptive budgets supply conservative defaults and shrink effective
        # limits under live memory/disk pressure. Explicit settings remain
        # ceilings and can be used with adaptive mode disabled.
        self.adaptive_resources = bool(
            self.config.get("adaptive_resources", True))
        _auto_budget = resourcebudget.recommend(tmputil.instance_dir())
        try:
            _fc_default = os.environ.get(
                "SAMQL_FLOW_CACHE_MB",
                _auto_budget["recommended_flow_cache_mb"])
            _fc_mb = int(self.config.get("flow_cache_mb", _fc_default) or 0)
        except Exception:
            _fc_mb = int(_auto_budget["recommended_flow_cache_mb"])
        self.flow_cache_mb_configured = max(0, _fc_mb)
        self.flow_cache_bytes_max = self.flow_cache_mb_configured * 1024 * 1024
        self.parallel_nodeflows = bool(
            self.config.get("parallel_nodeflows", True))
        try:
            _pw = int(self.config.get(
                "parallel_nodeflow_workers",
                os.environ.get("SAMQL_NODEFLOW_WORKERS",
                               _auto_budget["recommended_parallel_workers"]))
                      or 1)
        except Exception:
            _pw = int(_auto_budget["recommended_parallel_workers"])
        self.parallel_nodeflow_workers = max(1, min(_pw, 16))
        self.persistent_flow_cache = bool(
            self.config.get("persistent_flow_cache", True))
        try:
            _pf_default = os.environ.get(
                "SAMQL_PERSISTENT_FLOW_CACHE_MB",
                _auto_budget["recommended_persistent_cache_mb"])
            _pf_mb = int(self.config.get(
                "persistent_flow_cache_mb", _pf_default) or 0)
        except Exception:
            _pf_mb = int(_auto_budget["recommended_persistent_cache_mb"])
        self.persistent_flow_cache_mb_configured = max(0, _pf_mb)
        try:
            self.persistent_flow_cache_days = max(0, int(self.config.get(
                "persistent_flow_cache_days", 14) or 0))
        except Exception:
            self.persistent_flow_cache_days = 14
        # Flatten JSON on load: when ON, a JSON file is converted then shredded
        # into normalized relational tables (root + child table per nested
        # array). Default OFF -- multi-GB nested JSON is far faster as a single
        # nested Parquet-backed table; turn the Load modal "Flatten into
        # relational tables" toggle on only when you need that shape. SQLite
        # has no nested types, so it always flattens regardless. Applies to
        # every load entry point that does not pass an explicit flatten/shred
        # flag (folder appends, some path loads).
        self.flatten_on_load = bool(self.config.get("flatten_json", False))
        # Load-file size thresholds (Parquet / JSON stream / upload / cache).
        # Persisted overrides win over env vars; applied process-wide so
        # engines.py / filecache / server upload cap all see them.
        try:
            from . import load_thresholds as _LT
            saved_lt = self.config.get("load_thresholds")
            if isinstance(saved_lt, dict) and saved_lt:
                _LT.apply_overrides(saved_lt, replace=True)
        except Exception:
            pass
        # The registry/accounting lives in a dedicated thread-safe cache
        # object; session.py supplies the engine-specific size estimate and
        # table-drop callback. Compatibility aliases keep diagnostics/tests
        # that inspect the registry working while the subsystem is split out.
        self._flow_cache_registry = FlowCache(
            self.flow_cache_max, self.flow_cache_bytes_max,
            self._flow_cache_drop_table)
        _pf_dir = os.environ.get("SAMQL_PERSISTENT_FLOW_CACHE_DIR") or os.path.join(
            os.path.expanduser("~"), ".json_csv_sql_explorer",
            "nodeflow_cache")
        self._persistent_flow_cache_registry = PersistentFlowCache(
            _pf_dir,
            self.persistent_flow_cache_mb_configured * 1024 * 1024,
            self.persistent_flow_cache_days)
        self._resource_budget_last = None
        self._persistent_cache_tainted = False
        self._flow_cache = self._flow_cache_registry.entries
        self._flow_cache_stats = self._flow_cache_registry.stats
        self._flow_cache_lock = self._flow_cache_registry.lock
        self._data_epoch = 0
        # directory-node file cache: path -> (hidden_table, engine, mtime)
        self._dir_files = {}
        # API node: node id -> (hidden table, engine) for its last fetch, and
        # an optional injected fetcher (fetch_json's signature) used instead of
        # the network -- the seam tests and the iterator drive the node through.
        self._api_node_tables = {}
        self._api_fetcher = None
        # optional injected streaming spooler (fetch_to_file's signature):
        # used instead of the network so the spool-to-disk path is testable.
        self._api_spooler = None
        # optional injected sleep (for retry/backoff) so tests don't really wait
        self._api_sleep = None
        # append-from-folder cache: folder -> (hidden_table, engine, signature)
        self._folder_files = {}
        self.db = DBManager(disk_backed=self.low_memory)
        try:
            self.db.execute("SELECT 1")
        except Exception:
            pass
        self.duckdb = None
        self._secrets = None             # lazy SecretStore (DPAPI-backed)
        self._connection_profiles = None  # lazy named profile registry
        self.connections = {}            # name -> SQLServerConnection
        self._hdfs = None                # active WebHDFSClient (or None)
        # Remote "catalog" tables: name + columns only, no data. Querying one
        # in the editor is routed to its connection (passthrough to SQL Server).
        self.catalog_tables = {}         # display name -> info dict
        self._catalog_route = {}         # lowercase bare table name -> conn
        self.current_target = LOCAL_TARGET
        self.temp_files = []
        self._results = {}               # id -> _CachedResult
        self._results_order = []
        # results whose parquet stores are LIVE reuse-view sources right now
        # (rid -> refcount); pinned results are never evicted / freed, and an
        # explicit discard is DEFERRED until the run's views are dropped
        self._reuse_pins = {}
        self._discard_deferred = set()
        self._lock = threading.RLock()
        # DuckDB on disk (bounded memory; cold data spilled to a temp DB
        # file) when low-memory mode or explicitly requested.
        self.duckdb_on_disk = bool(
            self.low_memory or self.config.get("duckdb_on_disk", False))
        # Debounced background housekeeping (checkpoint / reclaim / gc) so
        # SamQL stays lean after drops and queries without slowing them.
        self._cleanup_lock = threading.Lock()
        self._cleanup_timer = None
        self._cleanup_full = False
        self._closed = False
        # Cache table profiles so re-opening a profile tab (or re-profiling
        # an unchanged table) is instant instead of re-scanning. Cleared on
        # any table mutation. Bounded to the most recent few tables.
        self._profile_cache = {}
        self._profile_cache_order = []
        # In-flight cancellable queries: query_id -> engine. A cancel
        # request interrupts that engine so a superseded heavy query stops
        # consuming CPU instead of running to completion.
        self._running = {}
        self._running_lock = threading.Lock()
        # Run ids the user has asked to stop. A loop (iterator / while / folder)
        # checks this between statements so a cancel that lands in the gap
        # between engine calls -- where interrupt() has nothing to hit -- still
        # stops it promptly, not one iteration later.
        self._cancelled_runs = set()
        # Serialises lazy DuckDB construction so two concurrent first-ops can't
        # each build a manager (each opening its own connection + temp file).
        self._duckdb_lock = threading.Lock()
        # The incremental flow cache is shared across request threads. Its
        # dedicated FlowCache owns the lock and always defers engine-table
        # drops until after releasing it, preventing lock-order inversion.
        # Start the stall watchdog once per process: it watches the progress
        # registry and, when an operation stops advancing, dumps thread stacks
        # to a log so a silent hang is diagnosable. Observe-only; never cancels.
        try:
            watchdog.start()
        except Exception:
            pass

    # ---- engines ----------------------------------------------------
    def get_duckdb(self):
        if self.duckdb is None:
            if not HAS_DUCKDB:
                raise RuntimeError(
                    "DuckDB is not installed. Install with: "
                    "pip install duckdb")
            # Double-checked locking: only one thread builds the manager, so a
            # burst of concurrent first-ops can't each open a separate DuckDB
            # connection/temp file (one would win the attribute and the others
            # would leak an open connection).
            with self._duckdb_lock:
                if self.duckdb is None:
                    budget = self._effective_resource_budget()
                    self.duckdb = DuckDBManager(
                        low_memory=self.low_memory,
                        on_disk=self.duckdb_on_disk,
                        memory_limit_mb=(budget.get("engine_memory_mb")
                                         if self.adaptive_resources else None))
                    # .426: always-on -- a persisted opt-out from an
                    # older build must not resurrect on engine rebuild
                    self._concurrent_reads_pref = True
                    pref = getattr(self, "_concurrent_reads_pref", None)
                    if pref is not None:
                        self.duckdb.concurrent_reads = bool(pref)
        return self.duckdb

    def set_flatten_json(self, on):
        """Toggle 'flatten JSON on load' at runtime and persist it.

        When on, a DuckDB JSON load shreds into normalized relational tables
        after convert. When off (the default), the load keeps a single nested
        table. Applies to the next load on every entry point that does not
        pass an explicit flatten/shred flag.
        """
        self.flatten_on_load = bool(on)
        try:
            self.config.set("flatten_json", self.flatten_on_load)
        except Exception:
            pass
        return self.flatten_on_load

    def flatten_json_enabled(self):
        """Whether a JSON load flattens into normalized tables (True) or keeps
        a single nested DuckDB table (False)."""
        return bool(getattr(self, "flatten_on_load", False))

    def set_concurrent_reads(self, on):
        # .426: async reads are ALWAYS ON. The whole product leans on
        # them now (grid pages, catalog peeks, counts); the old opt-out
        # is accepted for API compatibility and clamped to True.
        on = True
        """Toggle DuckDB concurrent reads at runtime. When on, the tables
        panel's catalog reconcile, per-table row counts and schema peeks run on
        a separate cursor instead of queuing behind a running build. Applies to
        the live DuckDB engine immediately and is remembered for one created
        later. SQLite is unaffected (single connection)."""
        self._concurrent_reads_pref = bool(on)
        if self.duckdb is not None:
            self.duckdb.concurrent_reads = bool(on)
        return self._concurrent_reads_pref

    def concurrent_reads_enabled(self):
        """Whether DuckDB read-only paths currently run on a concurrent
        cursor: the live engine's flag if DuckDB exists, else the pending
        runtime preference, else the environment-variable default."""
        if self.duckdb is not None:
            return bool(self.duckdb.concurrent_reads)
        pref = getattr(self, "_concurrent_reads_pref", None)
        if pref is not None:
            return bool(pref)
        from .engines import _env_truthy, CONCURRENT_READS_ENV
        return _env_truthy(CONCURRENT_READS_ENV, True)

    def status(self):
        """Live activity snapshot for the dashboard and diagnostics: in-flight
        operations (with rows processed and seconds since last progress),
        whether each engine's connection lock is currently held, the process
        thread count, the restore flag, and the most recent stall the watchdog
        recorded. Lock-free apart from a microsecond non-blocking probe of each
        engine lock, so it stays responsive even when an engine is busy."""
        def _busy(eng):
            if eng is None:
                return False
            lk = getattr(eng, "write_lock", None)
            if lk is None:
                return False
            try:
                got = lk.acquire(blocking=False)
            except Exception:
                return True
            if got:
                try:
                    lk.release()
                except Exception:
                    pass
                return False
            return True

        try:
            ops = opreg.snapshot()
        except Exception:
            ops = []
        try:
            stall = watchdog.last_stall()
            stall_log = watchdog.log_path()
        except Exception:
            stall, stall_log = None, None
        return {
            "operations": ops,
            "engines": {
                "sqlite": {"active": True, "busy": _busy(self.db)},
                "duckdb": {"active": self.duckdb is not None,
                           "busy": _busy(self.duckdb)},
            },
            "threads": threading.active_count(),
            "restoring": bool(getattr(self, "restoring", False)),
            "concurrent_reads": self.concurrent_reads_enabled(),
            "flatten_json": self.flatten_json_enabled(),
            "last_stall": stall,
            "stall_log": stall_log,
        }

    def reset_engines(self, which="all"):
        """Recover from a wedged engine without restarting the app.

        Interrupts any in-flight work, then swaps in brand-new, empty engines
        (fresh DuckDB connection + temp file; fresh SQLite) and rebuilds the
        previously-loaded tables from the manifest in the background. Temp/flow
        intermediates are disposable and are not rebuilt; tables whose source
        file has since moved are skipped (exactly as a normal restart would).

        Crucially we never wait on the old engine's connection lock -- a truly
        wedged engine would never release it -- so the old manager is simply
        orphaned. A thread stuck inside the old connection keeps running
        detached while all new work uses the fresh engine. Both engines are
        rebuilt together so the manifest replay can't duplicate tables.
        Returns immediately; rebuild progress shows via the 'restoring' flag
        and the activity dashboard."""
        try:
            self.interrupt_loads()
        except Exception:
            pass
        # UNCLOG: interrupt every REGISTERED in-flight query first.
        # Concurrent reads run on their own cursors -- the main-connection
        # interrupt below can't reach them, so a stall there would survive
        # the reset without this sweep.
        try:
            for qid in list(self._running):
                try:
                    self.cancel_query(qid)
                except Exception:
                    pass
        except Exception:
            pass
        reset = []
        old_duck = self.duckdb
        self.duckdb = None
        if old_duck is not None:
            try:
                old_duck.interrupt()
            except Exception:
                pass
            reset.append("duckdb")
        old_db = self.db
        try:
            old_db.interrupt()
        except Exception:
            pass
        self.db = DBManager(disk_backed=self.low_memory)
        reset.append("sqlite")
        try:
            self._invalidate_profiles()
            self._invalidate_counts()
        except Exception:
            pass
        # clear any in-flight op records tied to the now-orphaned engines so
        # the dashboard/watchdog don't keep showing them
        try:
            with self._running_lock:
                self._running.clear()
        except Exception:
            pass
        # the orphaned engines' op records can never progress again --
        # abandon them, then forget the recorded stall so the watchdog
        # re-checks from scratch (it re-flags only if NEW work stalls)
        try:
            for o in opreg.snapshot():
                opreg.end(op_id=o.get("id"))
        except Exception:
            pass
        try:
            watchdog.clear_stall()
        except Exception:
            pass
        self._restore_started = False
        try:
            rebuilding = bool(self.manifest.all())
        except Exception:
            rebuilding = False
        try:
            self.restore_session()
        except Exception:
            pass
        return {"ok": True, "reset": reset, "rebuilding": rebuilding}

    def nuclear_reset(self):
        """.523: the Activity-modal KILL SWITCH. Nothing survives: every
        in-flight query / load / op is interrupted, BOTH engines are
        orphaned (never wait on a wedged lock) and hard-poked on the way
        out, every cached result, session temp file, family record and
        flow intermediate is destroyed, and the restore manifest is
        CLEARED -- so the state afterwards is exactly a fresh launch,
        which starts EMPTY. SQLite is replaced immediately; DuckDB is left
        None so the double-checked first-use block builds a brand-new
        engine exactly as startup does. On-disk history, saved workflows
        and settings are launch state too, so they stay."""
        stats = {"queries": 0, "results": 0, "temps": 0,
                 "tables_forgotten": 0}
        try:
            self.interrupt_loads()
        except Exception:
            pass
        try:
            for qid in list(self._running):
                stats["queries"] += 1
                try:
                    self.cancel_query(qid)
                except Exception:
                    pass
        except Exception:
            pass
        # orphan + hard-interrupt both engines; never touch their locks
        old_duck, self.duckdb = self.duckdb, None
        if old_duck is not None:
            try:
                old_duck._cancel.set()
            except Exception:
                pass
            for f in ("interrupt", "recycle"):
                try:
                    getattr(old_duck, f)()
                except Exception:
                    pass
        old_db, self.db = self.db, DBManager(disk_backed=self.low_memory)
        for f in ("interrupt", "recycle"):
            try:
                getattr(old_db, f)()
            except Exception:
                pass
        # every cached result (and its backing store) dies
        try:
            stats["results"] = len(self._results)
            for cr in list(self._results.values()):
                for f in ("close", "cleanup"):
                    try:
                        getattr(cr.store, f)()
                    except Exception:
                        pass
            self._results.clear()
        except Exception:
            pass
        # session-scoped temp files
        for pth in list(getattr(self, "temp_files", []) or []):
            try:
                os.unlink(pth)
                stats["temps"] += 1
            except Exception:
                pass
        try:
            del self.temp_files[:]
        except Exception:
            pass
        # runtime bookkeeping back to construction values
        try:
            self._table_family.clear()
        except Exception:
            pass
        try:
            self._table_ui_order.clear()
        except Exception:
            pass
        try:
            with self._running_lock:
                self._running.clear()
        except Exception:
            pass
        # fresh launch starts EMPTY: clear the restore manifest (this is
        # what a graceful shutdown does) -- NO background rebuild.
        try:
            stats["tables_forgotten"] = len(self.manifest.all())
        except Exception:
            pass
        try:
            self.manifest.clear()
        except Exception:
            pass
        # dashboard / watchdog: abandon orphaned op records + stall memory
        try:
            for o in opreg.snapshot():
                opreg.end(op_id=o.get("id"))
        except Exception:
            pass
        try:
            watchdog.clear_stall()
        except Exception:
            pass
        try:
            self._invalidate_profiles()
            self._invalidate_counts()
        except Exception:
            pass
        self._restore_started = False
        return {"ok": True, "nuked": stats}

    def optional_features(self):
        import importlib.util as _ilu
        from .secretstore import dpapi_available
        return {
            "duckdb": HAS_DUCKDB,
            "pyarrow": _ilu.find_spec("pyarrow") is not None,
            "sqlglot": _ilu.find_spec("sqlglot") is not None,
            "pandas": _ilu.find_spec("pandas") is not None,
            "msal": _ilu.find_spec("msal") is not None,
            "pyodbc": _ilu.find_spec("pyodbc") is not None,
            "openpyxl": _ilu.find_spec("openpyxl") is not None,
            "secrets": dpapi_available(),
        }

    @property
    def secrets(self):
        """Lazy DPAPI-backed store for saved connection passwords."""
        if self._secrets is None:
            from .secretstore import SecretStore
            self._secrets = SecretStore()
        return self._secrets

    @property
    def connection_profiles(self):
        """Lazy registry of named SQL/API connection profiles (no passwords)."""
        if self._connection_profiles is None:
            from .connection_profiles import ConnectionProfileStore
            self._connection_profiles = ConnectionProfileStore()
        return self._connection_profiles

    # ---- tables tree ------------------------------------------------
    def tables_tree(self):
        out = []
        # Reconcile caches with the live DB so user-created/dropped
        # tables (via raw SQL) are reflected.
        try:
            self.db.sync_catalog()
        except Exception:
            pass
        if self.duckdb is not None:
            try:
                self.duckdb.sync_catalog()
            except Exception:
                pass
        # SQLite
        for name in sorted(self.db.table_columns):
            if name.startswith("__"):
                continue  # internal staging (reconcile __nb_*) — not advertised
            cols = self.db.table_columns.get(name, [])
            # .464: busy-skip -- a running statement must never park the
            # sidebar; a miss shows no types now and fills in next pass.
            types = self.db.types_cached(name)
            out.append({
                "engine": "sqlite",
                "name": name,
                "source": self.db.table_sources.get(name, ""),
                "row_count": self._cached_count(self.db, "sqlite", name),
                "columns": _cols_with_hints(cols, types),
            })
        # DuckDB
        if self.duckdb is not None:
            for name in sorted(self.duckdb.table_columns):
                if name.startswith("__"):
                    continue  # internal staging (reconcile __nb_*)
                cols = self.duckdb.table_columns.get(name, [])
                types = self.duckdb.types_cached(name)
                out.append({
                    "engine": "duckdb",
                    "name": name,
                    "source": self.duckdb.table_sources.get(name, ""),
                    "row_count": self._cached_count(
                        self.duckdb, "duckdb", name),
                    "columns": _cols_with_hints(cols, types),
                    # .501: explicit flatten parentage for the sidebar family
                    # tree (path-only child names carry no parent in the name)
                    "parent": self._table_family.get(name),
                })
        # remote catalog tables: names only -- columns are fetched lazily (via
        # /api/catalog/columns when a table is expanded) so a catalog of
        # thousands of tables stays a light payload. Carry the database /
        # schema / connection so the UI can roll them up under a collapsible
        # group, plus a column count for display and the qualified name.
        for key, info in sorted(self.catalog_tables.items()):
            conn = info.get("conn", "")
            db = info.get("database") or ""
            out.append({
                "engine": "remote",
                "name": key,
                "source": "SQL Server: %s" % conn,
                "row_count": None,
                "remote": True,
                "conn": conn,
                "database": db,
                "schema": info.get("schema", ""),
                "group": db or ("SQL Server: %s" % conn if conn
                                else "SQL Server"),
                "qualified": info.get("qualified", key),
                "col_count": len(info.get("columns", [])),
                "columns": [],
            })
        return self._apply_table_ui_order(out)

    def set_table_ui_order(self, items):
        """Persist Loaded-tables drag order for this session.

        ``items`` is a list of ``{engine, name}`` (local tables only). Remote
        catalog entries are ignored. Tables not listed keep appearing after
        the ranked set (stable relative order from the catalog walk).
        """
        order = []
        seen = set()
        for it in items or []:
            if not isinstance(it, dict):
                continue
            eng = str(it.get("engine") or "").strip()
            name = str(it.get("name") or "").strip()
            if not eng or not name or eng == "remote":
                continue
            key = "%s:%s" % (eng, name)
            if key in seen:
                continue
            seen.add(key)
            order.append(key)
        self._table_ui_order = order
        return {"ok": True, "order": list(order)}

    def _apply_table_ui_order(self, out):
        """Reorder ``tables_tree`` rows by ``_table_ui_order`` when set."""
        rank = getattr(self, "_table_ui_order", None) or []
        if not rank:
            return out
        idx = {k: i for i, k in enumerate(rank)}
        ranked, unranked, remotes = [], [], []
        for t in out:
            if t.get("remote"):
                remotes.append(t)
                continue
            key = "%s:%s" % (t.get("engine", ""), t.get("name", ""))
            if key in idx:
                ranked.append(t)
            else:
                unranked.append(t)
        ranked.sort(key=lambda t: idx["%s:%s" % (t.get("engine", ""),
                                                 t.get("name", ""))])
        return ranked + unranked + remotes

    def _rename_table_ui_order(self, engine, old, new):
        order = getattr(self, "_table_ui_order", None) or []
        if not order:
            return
        old_k = "%s:%s" % (engine, old)
        new_k = "%s:%s" % (engine, new)
        self._table_ui_order = [
            (new_k if k == old_k else k) for k in order if k != new_k
        ]

    def _drop_table_ui_order(self, engine, name):
        order = getattr(self, "_table_ui_order", None) or []
        if not order:
            return
        key = "%s:%s" % (engine, name)
        self._table_ui_order = [k for k in order if k != key]

    def column_field_tree(self, engine, table, column):
        """The nested field tree for one column, parsed from its raw-case
        DuckDB type, so the sidebar can render the schema as an expandable tree
        instead of one giant type string. Returns {type, fields:[...]} where
        each field is {depth, name, type, kind, path, note}. Empty for a flat
        (scalar / SQLite) column.

        Flatten-off loads often store deep nesting as opaque JSON / JSON[]
        (maximum_depth=2), so DESCRIBE alone yields an empty tree. In that
        case sample live cell values and build the tree from them."""
        from .diagnostics import (parse_duckdb_type, flatten_type_tree,
                                  access_recipes)
        mgr = self.duckdb if engine == "duckdb" else self.db
        if mgr is None or not hasattr(mgr, "_column_types_raw"):
            return {"type": "", "fields": []}
        try:
            raw = mgr._column_types_raw(table).get(column, "")
        except Exception:
            raw = ""
        if not raw:
            return {"type": "", "fields": []}
        try:
            node = parse_duckdb_type(raw)
            rows = flatten_type_tree(column, node, max_nodes=1500)
            # per-node access recipes (needs the root row for context)
            access_recipes(column, rows)
        except Exception:
            return {"type": raw, "fields": []}
        # drop the root row (the column itself already shows in the tree)
        if rows and rows[0].get("depth") == 0:
            rows = rows[1:]
        # Call via the class so unit tests that bind this method onto a
        # SimpleNamespace still resolve the static helper.
        if engine == "duckdb" and Session._field_tree_needs_json_sample(
                raw, rows):
            sampled = Session._sample_column_field_tree(
                self, mgr, table, column, raw)
            if sampled is not None:
                return sampled
        return {"type": raw, "fields": rows}

    def table_field_tree(self, engine, table):
        """Unified Field Explorer tree for one loaded table.

        Flatten-off JSON often expands to several top-level columns on a
        *single* catalog table (``id``, ``legs``, ``nest``, …). Field Explorer
        used to list each nested column as its own "source"; this returns one
        tree so the UI can show the table once with every column + nested
        field underneath. Each row carries ``column`` (owning top-level name)
        so preview / shred stay column-scoped while multi-select spans the
        whole table.
        """
        mgr = self.duckdb if engine == "duckdb" else self.db
        if mgr is None or not hasattr(mgr, "_column_types_raw"):
            return {"fields": [], "error": "Engine not available."}
        try:
            raw_cols = mgr._column_types_raw(table) or {}
        except Exception as e:
            return {"fields": [], "error": err_str(e)}
        if not raw_cols:
            return {"fields": [], "error": "No columns."}

        out = []
        for col, typ in raw_cols.items():
            sub = self.column_field_tree(engine, table, col)
            nested = list(sub.get("fields") or [])
            # IDE-style SELECT idents (bare when safe). Always-quote made
            # multi-select emit ``"code" AS …`` instead of ``code AS …``.
            q = self._fe_quote_select_ident(col)
            if nested:
                root_kind = "struct"
                tu = (typ or "").strip().upper()
                if tu.endswith("[]") or tu in ("JSON", "JSON[]"):
                    for n in nested:
                        if int(n.get("depth") or 0) == 1 and n.get("kind") in (
                                "array", "array-scalar"):
                            root_kind = n.get("kind") or "array"
                            break
                out.append({
                    "depth": 1,
                    "name": col,
                    "type": typ or sub.get("type") or "",
                    "kind": root_kind,
                    "path": col,
                    "note": None,
                    "column": col,
                    "access": {"first": q, "sel": q, "unnests": []},
                })
                for n in nested:
                    row = dict(n)
                    row["depth"] = int(n.get("depth") or 0) + 1
                    row["column"] = col
                    out.append(row)
            else:
                out.append({
                    "depth": 1,
                    "name": col,
                    "type": typ or "",
                    "kind": "scalar",
                    "path": col,
                    "note": None,
                    "column": col,
                    "access": {"first": q, "sel": q, "unnests": []},
                })
        return {"ok": True, "fields": out, "table": table}

    # Live-cell field discovery: how many rows to pull when DESCRIBE stops at
    # opaque JSON (flatten-off maximum_depth). Kept modest — union of shapes,
    # not a full scan.
    _FIELD_TREE_SAMPLE_ROWS = 80

    @staticmethod
    def _fe_quote_select_ident(name):
        """Quote a Field Explorer SELECT identifier like the IDE / recipes.

        Safe bare names stay bare (including mixed case). Reserved words and
        non-plain identifiers are double-quoted. Do not use always-on
        ``sqlutil.quote_ident`` here — multi-select splices ``access.sel``
        as-is and over-quoting diverges from IDE insert style.
        """
        from .diagnostics import _needs_quote
        from .shred import _DUCKDB_RESERVED
        s = str(name)
        if _needs_quote(s) or s.lower() in _DUCKDB_RESERVED:
            return '"%s"' % s.replace('"', '""')
        return s

    @staticmethod
    def _field_tree_needs_json_sample(raw_type, fields):
        """True when DESCRIBE has no deep schema (opaque JSON / JSON leaves).

        Flatten-off ``maximum_depth=2`` often yields STRUCT shells with typed
        scalars *and* opaque JSON leaves. Sampling must still run whenever any
        leaf is JSON — otherwise the sidebar stops at those leaves and never
        shows cashflow-style keys nested underneath."""
        tu = (raw_type or "").strip().upper()
        if tu == "JSON" or tu == "JSON[]" or tu.endswith(" JSON") \
                or " JSON)" in tu or tu.endswith("JSON[]") \
                or " JSON," in tu or "JSON " in tu:
            if not fields:
                return True
            # Any opaque JSON leaf means DESCRIBE stopped early — sample live
            # cells so nested keys under those leaves appear in the field tree.
            # (Previously required *all* leaves to be JSON, which skipped mixed
            # STRUCT(id INTEGER, nest JSON) columns — the common flatten-off
            # shape.)
            if any("JSON" in str(f.get("type") or "").upper()
                   for f in fields):
                return True
            # A lone "(element)" under JSON[] is not a useful field tree.
            names = {f.get("name") for f in fields}
            if names <= {"(element)"} and "JSON" in tu:
                return True
        return False

    def _sample_column_field_tree(self, mgr, table, column, raw_type):
        """Sample live values and build a field tree for opaque JSON columns."""
        from .diagnostics import json_values_to_field_tree, access_recipes
        from .sqlutil import quote_ident as _qid
        try:
            sql = ('SELECT %s FROM %s LIMIT %d'
                   % (_qid(column), _qid(table),
                      int(Session._FIELD_TREE_SAMPLE_ROWS)))
            reader = getattr(mgr, "read", None) or getattr(mgr, "execute_read",
                                                          None)
            if reader is not None:
                _c, rows = reader(sql)
            else:
                _c, rows = mgr.execute(sql)
        except Exception:
            return None
        values = [r[0] for r in (rows or []) if r]
        if not values:
            return None
        # Sidebar `path` tooltips: STRUCT shells keep dotted labels; pure
        # JSON / JSON[] use ->>. Field Explorer *queries* always need JSON
        # recipes here — STRUCT-style UNNEST fails on opaque JSON arrays.
        tu = (raw_type or "").upper().strip()
        path_style = ("struct" if tu.startswith("STRUCT") else "json")
        cast_to_json = "STRUCT" in tu
        root_is_list = tu.endswith("[]")
        try:
            ft = json_values_to_field_tree(values, colname=column,
                                          access_style=path_style)
        except Exception:
            return None
        nodes = list(ft.get("nodes") or [])
        root_node = None
        if nodes and nodes[0].get("depth") == 0:
            root_node = nodes[0]
            nodes = nodes[1:]
        if not nodes:
            return None
        try:
            # Rebuild a root row so access_recipes has column context, then
            # drop it again for the sidebar / Field Explorer.
            root_kind = ("array" if root_is_list
                         else "struct" if path_style == "struct"
                         else "scalar")
            # Sampled shape wins for opaque JSON (array/object cell).
            if root_node and root_node.get("kind") in ("array", "array-scalar"):
                root_kind = "array"
            elif root_node and root_node.get("kind") == "struct":
                root_kind = "struct"
            root = {"depth": 0, "name": column, "type": raw_type,
                    "kind": root_kind, "path": column, "note": None,
                    "double_encoded": bool(
                        (root_node or {}).get("double_encoded"))}
            full = [root] + [
                dict(n, depth=int(n.get("depth") or 0)) for n in nodes]
            access_recipes(column, full, access_style="json",
                           cast_to_json=cast_to_json,
                           root_is_list=root_is_list)
            nodes = full[1:]
        except Exception:
            pass
        return {"type": raw_type, "fields": nodes,
                "sampled": ft.get("sampled"), "source": "sampled-values"}

    def preview_column_access(self, engine, table, column, field_idx=None,
                              field_path=None):
        """Execute Field Explorer First SQL (LIMIT 1) and return a working recipe.

        Tries the primary access recipe, then the VARCHAR ``alt``, then a
        mechanical ``["JSON"]``→``["VARCHAR"]`` rewrite. Used so copy-SQL is
        never a NULL/error path for opaque JSON.
        """
        from .diagnostics import (field_access_sql, access_recipe_variants)
        from .sqlutil import quote_ident as _qid
        mgr = self.duckdb if engine == "duckdb" else self.db
        if mgr is None:
            return {"ok": False, "error": "No engine."}
        tree = self.column_field_tree(engine, table, column)
        fields = list(tree.get("fields") or [])
        if not fields:
            return {"ok": False, "error": "No nested fields for this column.",
                    "type": tree.get("type"), "fields": []}
        sel = None
        if field_idx is not None:
            try:
                sel = fields[int(field_idx)]
            except Exception:
                sel = None
        if sel is None and field_path:
            for f in fields:
                if (f.get("path") or "") == field_path \
                        or (f.get("name") or "") == field_path:
                    sel = f
                    break
        if sel is None:
            sel = fields[0]
        access = dict(sel.get("access") or {})
        variants = access_recipe_variants(access)

        def _run(sql):
            reader = getattr(mgr, "read", None) or getattr(
                mgr, "execute_read", None)
            if reader is not None:
                _c, rows = reader(sql)
            else:
                _c, rows = mgr.execute(sql)
            return rows

        last_err = None
        tried = []
        for variant in variants:
            sql = field_access_sql(table, variant, which="first")
            if not sql:
                continue
            tried.append(sql)
            try:
                rows = _run(sql)
                sample = rows[0][0] if rows and rows[0] else None
                if sample is not None:
                    all_sql = field_access_sql(table, variant, which="all")
                    return {
                        "ok": True, "sample": sample, "sql": sql,
                        "all_sql": all_sql, "access": variant,
                        "field": {k: sel.get(k) for k in (
                            "depth", "name", "type", "kind", "path")},
                        "style": "alt" if variant is not access else "primary",
                        "type": tree.get("type"),
                        "fields": fields,
                    }
                last_err = "NULL result"
            except Exception as e:
                last_err = err_str(e)

        # Mechanical fallback: rewrite JSON schema hops to VARCHAR + json().
        for sql in list(tried):
            alt = (sql.replace("'[\"JSON\"]'", "'[\"VARCHAR\"]'")
                   .replace("['\"JSON\"']", "['\"VARCHAR\"']"))
            if alt == sql:
                continue
            # Wrap bare eN select targets when rewriting.
            import re as _re
            alt2 = _re.sub(
                r"SELECT\s+(e\d+)\s+FROM",
                r"SELECT json(\1) FROM", alt, count=1)
            for candidate in (alt2, alt):
                if candidate in tried:
                    continue
                tried.append(candidate)
                try:
                    rows = _run(candidate)
                    sample = rows[0][0] if rows and rows[0] else None
                    if sample is not None:
                        return {
                            "ok": True, "sample": sample, "sql": candidate,
                            "access": access, "style": "rewrite",
                            "field": {k: sel.get(k) for k in (
                                "depth", "name", "type", "kind", "path")},
                            "type": tree.get("type"),
                            "fields": fields,
                        }
                    last_err = "NULL result"
                except Exception as e:
                    last_err = err_str(e)

        return {
            "ok": False,
            "error": last_err or "Could not preview this field.",
            "tried": tried[:6],
            "access": access,
            "field": {k: sel.get(k) for k in (
                "depth", "name", "type", "kind", "path")} if sel else None,
            "type": tree.get("type"),
            "fields": fields,
        }

    @staticmethod
    def _safe_count(engine, name):
        sql = f'SELECT COUNT(*) FROM "{name}"'
        # engine.read() is concurrent on DuckDB when the flag is on (a separate
        # cursor, so the count doesn't queue behind a build) and the locked
        # execute() otherwise / on SQLite. Fall back to execute() on any failure
        # (e.g. a TEMP table a fresh cursor can't see) -- correctness first.
        reader = getattr(engine, "read", None)
        if reader is not None:
            try:
                _c, rows = reader(sql)
                return int(rows[0][0]) if rows else None
            except Exception:
                pass
        try:
            _c, rows = engine.execute(sql)
            return int(rows[0][0]) if rows else None
        except Exception:
            return None

    def _cached_count(self, engine, label, name):
        # (module helper below coerces fakes/mocks safely)
        key = (label, name)
        if key in self._count_cache:
            return self._count_cache[key]
        # .464: a cache MISS during a long-running statement must not
        # park the sidebar behind that statement. If the engine is
        # busy, show no count for now (the tree renders a blank) and
        # let the next refresh fill it in; do NOT cache the miss.
        if _stmt_stepping(engine):
            return None  # a statement is stepping; never queue behind it
        # .464: probe the engine lock without ever punishing a test
        # double -- only a REAL "held by someone else" answer skips;
        # an odd/absent lock computes exactly as before.
        lock = getattr(engine, "write_lock", None)
        state = "nolock"
        if lock is not None:
            try:
                state = ("acquired"
                         if lock.acquire(blocking=False)
                         else "busy")
            except TypeError:
                try:
                    state = ("acquired" if lock.acquire(False)
                             else "busy")
                except Exception:
                    state = "nolock"
            except Exception:
                state = "nolock"
        if state == "busy":
            # .464 pass II: this probe RACED the loader's fire-and-forget
            # indexer -- each CREATE INDEX holds the lock for well under
            # a millisecond, and a tree refresh landing inside one of
            # those blinks showed a BLANK count for a table that loaded
            # fine. Micro-retry between statements (bounded, still
            # non-blocking each try) so an indexing burst can't blank
            # the sidebar, while a genuinely long statement keeps the
            # no-parking guarantee: after ~30ms we still yield None.
            # (.465 tune: ~6ms total. The indexer's holds are sub-ms,
            # so the FIRST retry usually wins; under a genuinely long
            # statement the whole ladder costs six milliseconds, not
            # thirty -- the 464b sustained-busy gauntlet was paying the
            # old ladder thousands of times.)
            deadline = time.monotonic() + 0.006
            while time.monotonic() < deadline:
                time.sleep(0.0015)
                if _stmt_stepping(engine):
                    return None
                try:
                    if lock.acquire(blocking=False):
                        state = "acquired"
                        break
                except Exception:
                    return None
            if state != "acquired":
                return None
        try:
            val = self._safe_count(engine, name)
        finally:
            if state == "acquired":
                try:
                    lock.release()
                except Exception:
                    pass
        self._count_cache[key] = val
        return val

    def _prime_count(self, kind, name, n):
        """The loader KNOWS how many rows it just wrote -- prime the
        count cache so the very next tree refresh never has to probe a
        lock the background indexer may be holding. Call AFTER
        _invalidate_counts (which clears the whole cache)."""
        try:
            if name and n is not None:
                self._count_cache[(kind, name)] = int(n)
        except Exception:
            pass

    def _invalidate_counts(self):
        self._count_cache.clear()
        # Any data mutation also invalidates the incremental flow cache: bump
        # the epoch (so every future fingerprint differs) and drop the cached
        # intermediate tables. Guarded so it's safe if called very early.
        if hasattr(self, "_flow_cache"):
            self._data_epoch += 1
            self._flow_cache_clear()

    def _table_exists(self, eng, name):
        """Cheap check that a (temp) table is still queryable on ``eng``."""
        try:
            eng.execute('SELECT 1 FROM "%s" LIMIT 0' % name)
            return True
        except Exception:
            return False

    def _flow_cache_get(self, fp):
        return self._flow_cache_registry.get(fp)

    @staticmethod
    def _flow_type_width(type_name):
        t = str(type_name or "").upper()
        if any(x in t for x in ("BOOL", "TINYINT")):
            return 1
        if "SMALLINT" in t:
            return 2
        if any(x in t for x in ("BIGINT", "INTEGER", "INT", "DOUBLE",
                                "FLOAT", "REAL", "DATE", "TIME")):
            return 8
        if any(x in t for x in ("DECIMAL", "NUMERIC", "HUGEINT", "UUID")):
            return 16
        # Strings/nested values vary widely. Sixty-four bytes is intentionally
        # conservative without scanning payloads merely to size a cache.
        return 64

    def _flow_cache_estimate_bytes(self, table, engine_target):
        """Cheap, conservative table-size estimate for byte-budgeted LRU."""
        try:
            eng, _kind = self._engine_obj(engine_target)
            _cols, rows_out = eng.execute(
                'SELECT COUNT(*) FROM "%s"' % table)
            rows = max(0, int(rows_out[0][0] if rows_out else 0))
            types = eng.column_types(table) or {}
            row_bytes = sum(self._flow_type_width(t) for t in types.values())
            return max(4096, rows * max(8, row_bytes))
        except Exception:
            return 4096

    def _effective_resource_budget(self):
        budget = resourcebudget.effective_limits(
            self.flow_cache_mb_configured,
            self.persistent_flow_cache_mb_configured,
            self.parallel_nodeflow_workers,
            adaptive=self.adaptive_resources,
            temp_dir=tmputil.instance_dir())
        self._resource_budget_last = budget
        return budget

    def _sync_flow_cache_limits(self):
        budget = self._effective_resource_budget()
        self.flow_cache_bytes_max = int(budget["flow_cache_mb"]) * 1024 * 1024
        self._flow_cache_registry.configure(
            self.flow_cache_max, self.flow_cache_bytes_max)
        persistent_mb = int(budget["persistent_cache_mb"])
        # PersistentFlowCache uses zero as "no byte ceiling". Under adaptive
        # disk pressure, however, an effective zero means "stop consuming this
        # volume", not "become unlimited". A one-byte ceiling prunes existing
        # files and prevents a new positive-sized Parquet from surviving.
        persistent_bytes = persistent_mb * 1024 * 1024
        if (self.adaptive_resources
                and self.persistent_flow_cache_mb_configured > 0
                and persistent_mb <= 0):
            persistent_bytes = 1
        self._persistent_flow_cache_registry.configure(
            persistent_bytes, self.persistent_flow_cache_days)
        duck = getattr(self, "duckdb", None)
        if duck is not None and self.adaptive_resources:
            try:
                duck.apply_resource_memory_mb(budget.get("engine_memory_mb"))
            except Exception:
                pass
        return budget

    def _flow_cache_put(self, fp, table, engine_target):
        approx = self._flow_cache_estimate_bytes(table, engine_target)
        self._sync_flow_cache_limits()
        return self._flow_cache_registry.put(
            fp, table, engine_target, approx)

    def _flow_cache_drop_table(self, engine_target, name):
        try:
            eng, _k = self._engine_obj(engine_target)
            try:
                eng.execute('DROP VIEW IF EXISTS "%s"' % name)
            except Exception:
                pass
            eng.execute('DROP TABLE IF EXISTS "%s"' % name)
        except Exception:
            pass

    def _flow_cache_clear(self):
        registry = getattr(self, "_flow_cache_registry", None)
        if registry is not None:
            registry.clear()

    @staticmethod
    def _sql_path(path):
        return sqlutil.sql_path(path)

    def _stable_path_signature(self, path):
        """A bounded, content-relevant signature for a persistent source.

        Size/mtime alone can collide after an in-place rewrite or timestamp
        preservation. Hashing the first and last 64 KiB keeps lookup bounded
        while making accidental stale reuse materially less likely.
        """
        try:
            path = os.path.realpath(path)
            if os.path.isfile(path):
                st = os.stat(path)
                h = hashlib.sha256()
                size = int(st.st_size)
                # Sample across the whole source rather than only its ends.
                # ctime/inode catch ordinary in-place rewrites even when a
                # producer preserves mtime; the samples make accidental stale
                # reuse highly unlikely without re-reading a multi-GB file.
                offsets = sorted(set([
                    0,
                    max(0, size // 4 - 32 * 1024),
                    max(0, size // 2 - 32 * 1024),
                    max(0, (size * 3) // 4 - 32 * 1024),
                    max(0, size - 64 * 1024),
                ]))
                with open(path, "rb") as fh:
                    for offset in offsets:
                        fh.seek(offset)
                        h.update(str(offset).encode("ascii"))
                        h.update(fh.read(64 * 1024))
                return ["file", path, size, int(st.st_mtime_ns),
                        int(getattr(st, "st_ctime_ns", 0) or 0),
                        int(getattr(st, "st_ino", 0) or 0),
                        h.hexdigest()[:32]]
            if os.path.isdir(path):
                rows = []
                # Bound directory signatures: enough for ordinary load folders
                # without turning cache lookup into a full workload itself.
                for root, dirs, files in os.walk(path):
                    dirs.sort()
                    for name in sorted(files):
                        fp = os.path.join(root, name)
                        try:
                            st = os.stat(fp)
                        except Exception:
                            continue
                        rows.append((os.path.relpath(fp, path),
                                     int(st.st_size), int(st.st_mtime_ns)))
                        if len(rows) >= 10000:
                            return ["dir", path, rows, "truncated"]
                return ["dir", path, rows]
        except Exception:
            return None
        return None

    @staticmethod
    def _persistent_graph_strings(value):
        """Yield user-configured strings that may contain SQL expressions."""
        if isinstance(value, str):
            yield value
        elif isinstance(value, dict):
            for item in value.values():
                yield from Session._persistent_graph_strings(item)
        elif isinstance(value, (list, tuple)):
            for item in value:
                yield from Session._persistent_graph_strings(item)

    def _persistent_graph_is_deterministic(self, graph, eng):
        """Conservatively reject graph expressions unstable across restarts.

        DuckDB publishes stability metadata for built-in functions. Any
        VOLATILE, side-effecting, query-time-dependent, macro, pragma, or table
        function call makes the graph session-only. Bare temporal keywords are
        checked separately because SQL permits them without parentheses.
        False positives merely skip an optimization; false negatives could
        restore stale data, so this intentionally biases toward safety.
        """
        temporal = re.compile(
            r"\b(?:current_date|current_time|current_timestamp|localtime|"
            r"localtimestamp|transaction_timestamp|statement_timestamp|"
            r"clock_timestamp)\b", re.IGNORECASE)
        call_re = re.compile(
            r"(?<![A-Za-z0-9_])([A-Za-z_][A-Za-z0-9_]*(?:\.[A-Za-z_]"
            r"[A-Za-z0-9_]*)?)\s*\(")
        try:
            unstable = set(eng.unstable_function_names())
        except Exception:
            # Metadata uncertainty must not make a persistent result less safe.
            return False
        for node in (graph.get("nodes") or []):
            typ = str(node.get("type") or "").lower()
            cfg = node.get("config") or {}
            if typ == "sample" and str(cfg.get("mode") or "head").lower() == "random":
                return False
            for text in self._persistent_graph_strings(cfg):
                if temporal.search(text):
                    return False
                for match in call_re.finditer(text):
                    name = match.group(1).split(".")[-1].lower()
                    if name in unstable:
                        return False
        return True

    def _persistent_flow_salt(self, graph, engine_target):
        """Return a stable source salt, or ``None`` for a volatile graph.

        Persistence is intentionally conservative. API/SQL/iterator/file-glob
        nodes may depend on time, network or side effects, so their outputs stay
        session-only. Loaded file/folder inputs are accepted only when their
        source can be stat-fingerprinted.
        """
        if engine_target != DUCKDB_TARGET or not self.persistent_flow_cache:
            return None
        registry = self._persistent_flow_cache_registry
        def skip():
            registry.record_skip()
            return None
        if self._persistent_cache_tainted:
            return skip()
        if (self.adaptive_resources
                and int(self._effective_resource_budget().get(
                    "persistent_cache_mb") or 0) <= 0):
            return skip()
        volatile = {"apinode", "sqlserver", "sharepoint", "webscrape",
                    "sql", "iterator", "while", "filebrowser",
                    "directory", "appendfolder", "shred", "write"}
        sources = []
        eng, _kind = self._engine_obj(engine_target)
        if not self._persistent_graph_is_deterministic(graph, eng):
            return skip()
        table_sources = getattr(eng, "table_sources", {}) or {}
        for node in (graph.get("nodes") or []):
            typ = str(node.get("type") or "").lower()
            if typ in volatile:
                return skip()
            cfg = node.get("config") or {}
            tables = []
            if typ == "input" and cfg.get("table"):
                tables.append(cfg.get("table"))
            elif typ == "inputs":
                tables.extend(cfg.get("tables") or [])
            for table in tables:
                source = table_sources.get(table)
                sig = self._stable_path_signature(source) if source else None
                if sig is None:
                    return skip()
                try:
                    types = eng.column_types(table) or {}
                    columns = list((getattr(eng, "table_columns", {}) or {})
                                   .get(table) or types.keys())
                    schema = [[str(col), str(types.get(col) or "")]
                              for col in columns]
                except Exception:
                    return skip()
                sources.append([str(table), sig, schema])
        if not sources and not any(
                str(n.get("type") or "").lower() == "createtable"
                for n in (graph.get("nodes") or [])):
            return skip()
        # Include product semantics in the salt. A deterministic graph can
        # produce different SQL after an upgrade even when its user config is
        # unchanged; versioning prevents reuse of a pre-upgrade relation.
        try:
            from . import __version__ as product_version, BUILD as product_build
        except Exception:
            product_version, product_build = "unknown", "unknown"
        rep = json.dumps({"v": PersistentFlowCache.FORMAT_VERSION,
                          "product": product_version,
                          "build": product_build,
                          "sources": sorted(sources),
                          "engine": "duckdb"},
                         sort_keys=True, separators=(",", ":"), default=str)
        return hashlib.sha256(rep.encode("utf-8")).hexdigest()[:24]

    def _persistent_flow_restore(self, eng, persistent_fp, table_name,
                                 persistent_tables=False):
        registry = self._persistent_flow_cache_registry
        path = registry.acquire(persistent_fp)
        if not path:
            return False
        try:
            kind = "TABLE" if persistent_tables else "TEMP TABLE"
            eng.execute('DROP TABLE IF EXISTS "%s"' % table_name)
            eng.execute('CREATE %s "%s" AS SELECT * FROM read_parquet(\'%s\')'
                        % (kind, table_name, self._sql_path(path)))
            return True
        except Exception:
            # Corrupt/incompatible entries are discarded after the active pin
            # is released. A concurrent trim cannot remove the file mid-read.
            registry.release(path)
            path = None
            registry.discard(persistent_fp)
            return False
        finally:
            if path:
                registry.release(path)

    def _persistent_flow_publish(self, eng, persistent_fp, table_name):
        if not persistent_fp:
            return False
        def writer(path):
            conn = getattr(eng, "conn", None)
            owner = getattr(eng, "manager", eng)
            cancel = getattr(owner, "_cancel", None)
            sql = ("COPY (SELECT * FROM \"%s\") TO '%s' "
                   "(FORMAT PARQUET, COMPRESSION ZSTD)" %
                   (table_name, self._sql_path(path)))
            if conn is not None:
                with _ExecKeepalive(conn, cancel, threading.get_ident()):
                    eng.execute(sql)
            else:
                eng.execute(sql)
        return self._persistent_flow_cache_registry.publish(
            persistent_fp, writer)

    def flow_cache_info(self):
        """Diagnostics for volatile + restart-persistent NodeFlow caches."""
        budget = self._sync_flow_cache_limits()
        return {"enabled": bool(self.flow_cache),
                "configured_mb_max": self.flow_cache_mb_configured,
                "persistent_configured_mb_max":
                    self.persistent_flow_cache_mb_configured,
                "adaptive_resources": bool(self.adaptive_resources),
                "parallel_nodeflows": bool(self.parallel_nodeflows),
                "parallel_workers_configured": self.parallel_nodeflow_workers,
                "parallel_workers_effective": budget["parallel_workers"],
                "resource_budget": budget,
                "persistent_enabled": bool(self.persistent_flow_cache),
                "persistent": self._persistent_flow_cache_registry.info(),
                **self._flow_cache_registry.info()}

    def load_thresholds_info(self):
        """Effective load-file thresholds for the Storage & memory UI."""
        from . import load_thresholds as LT
        return {"ok": True, "thresholds": LT.effective_map(),
                "overrides": LT.overrides_snapshot()}

    def load_preflight(self, path):
        """Advise before a large file load (engine, temp disk, format limits).

        Returns warnings / blockers without starting the load. Used by the
        Load Data UI so multi-GB files get guidance early.
        """
        from . import resourcebudget
        from . import tmputil
        from .engines import HAS_DUCKDB
        warnings = []
        blockers = []
        path = str(path or "").strip()
        info = {
            "ok": True,
            "path": path,
            "exists": False,
            "size": 0,
            "size_mb": 0,
            "ext": "",
            "duckdb": bool(HAS_DUCKDB),
            "disk_free_gb": None,
            "warnings": warnings,
            "blockers": blockers,
        }
        if not path or not os.path.isfile(path):
            blockers.append("File not found.")
            info["ok"] = False
            return info
        info["exists"] = True
        try:
            size = os.path.getsize(path)
        except OSError:
            size = 0
        info["size"] = size
        info["size_mb"] = round(size / (1024 * 1024.0), 1)
        ext = os.path.splitext(path)[1].lower().lstrip(".")
        info["ext"] = ext
        try:
            snap = resourcebudget.snapshot(temp_dir=tmputil.instance_dir())
            info["disk_free_gb"] = snap.get("disk_free_gb")
            free = int(snap.get("disk_free") or 0)
        except Exception:
            free = 0
        # Temp volume should hold ~2× the file for NDJSON rewrite / Parquet.
        if size >= 256 * 1024 * 1024 and free and free < size * 2:
            warnings.append(
                "Temp disk has only %.1f GiB free; this %.1f GiB file may "
                "need ~2× headroom for conversion. Free space or set temp "
                "onto a larger volume."
                % (free / (1024 ** 3), size / (1024 ** 3)))
        if ext in ("xlsx", "xlsm", "xls"):
            warnings.append(
                "Excel is not ideal for multi-GB / multi-million-row data "
                "(~1,048,576 row limit). Prefer CSV or Parquet for large "
                "datasets.")
            if size >= 100 * 1024 * 1024:
                warnings.append(
                    "This workbook is large for openpyxl streaming — "
                    "export to CSV/Parquet first if the load is slow or "
                    "fails.")
        if ext in ("csv", "tsv", "txt", "json", "ndjson", "jsonl") \
                and size >= 512 * 1024 * 1024 and not HAS_DUCKDB:
            blockers.append(
                "DuckDB is required for files this large. Install duckdb "
                "(pip install duckdb) so SamQL can convert to on-disk Parquet.")
            info["ok"] = False
        elif ext in ("csv", "tsv", "txt", "json", "ndjson", "jsonl") \
                and size >= 256 * 1024 * 1024:
            warnings.append(
                "Large file — SamQL will convert to an on-disk Parquet cache "
                "(DuckDB). Prefer Load by path over drag-drop for the "
                "biggest files.")
        if size >= 4 * 1024 * 1024 * 1024:
            warnings.append(
                "File is over 4 GiB. Use Load Data → File (path). Drag-drop "
                "uploads are capped (see Storage & memory → Load thresholds).")
        return info

    def configure_load_thresholds(self, updates=None, reset=False):
        """Apply and persist load-file threshold overrides.

        ``updates`` is a dict of field → number. ``reset=True`` clears all
        overrides (env vars / built-in defaults apply again).
        """
        from . import load_thresholds as LT
        if reset:
            LT.clear_overrides()
            try:
                if "load_thresholds" in self.config.data:
                    del self.config.data["load_thresholds"]
                    self.config.save()
            except Exception:
                pass
        elif updates:
            LT.apply_overrides(updates, replace=False)
            snap = LT.overrides_snapshot()
            self.config.set("load_thresholds", snap)
        return self.load_thresholds_info()

    def configure_flow_cache(self, enabled=None, max_entries=None,
                             max_mb=None, clear=False, reset_stats=False,
                             adaptive_resources=None,
                             parallel_nodeflows=None, parallel_workers=None,
                             persistent_enabled=None, persistent_max_mb=None,
                             persistent_days=None, clear_persistent=False):
        """Apply and persist NodeFlow cache/resource settings."""
        if enabled is not None:
            self.flow_cache = bool(enabled)
            self.config.set("flow_cache", self.flow_cache)
        if max_entries is not None:
            self.flow_cache_max = max(0, min(int(max_entries), 10000))
            self.config.set("flow_cache_max", self.flow_cache_max)
        if max_mb is not None:
            mb = max(0, min(int(max_mb), 1024 * 1024))
            self.flow_cache_mb_configured = mb
            self.config.set("flow_cache_mb", mb)
        if adaptive_resources is not None:
            self.adaptive_resources = bool(adaptive_resources)
            self.config.set("adaptive_resources", self.adaptive_resources)
        if parallel_nodeflows is not None:
            self.parallel_nodeflows = bool(parallel_nodeflows)
            self.config.set("parallel_nodeflows", self.parallel_nodeflows)
        if parallel_workers is not None:
            self.parallel_nodeflow_workers = max(
                1, min(int(parallel_workers), 16))
            self.config.set("parallel_nodeflow_workers",
                            self.parallel_nodeflow_workers)
        if persistent_enabled is not None:
            self.persistent_flow_cache = bool(persistent_enabled)
            self.config.set("persistent_flow_cache",
                            self.persistent_flow_cache)
        if persistent_max_mb is not None:
            self.persistent_flow_cache_mb_configured = max(
                0, min(int(persistent_max_mb), 1024 * 1024))
            self.config.set("persistent_flow_cache_mb",
                            self.persistent_flow_cache_mb_configured)
        if persistent_days is not None:
            self.persistent_flow_cache_days = max(
                0, min(int(persistent_days), 3650))
            self.config.set("persistent_flow_cache_days",
                            self.persistent_flow_cache_days)
        self._sync_flow_cache_limits()
        cleared = 0
        persistent_cleared = 0
        if clear or not self.flow_cache:
            cleared = self._flow_cache_registry.clear(
                reset_stats=bool(reset_stats))
        if clear_persistent or not self.persistent_flow_cache:
            persistent_cleared = self._persistent_flow_cache_registry.clear(
                reset_stats=bool(reset_stats))
        info = self.flow_cache_info()
        info.update({"ok": True, "cleared": cleared,
                     "persistent_cleared": persistent_cleared})
        return info


    # ---- file loading ----------------------------------------------
    def default_destination(self):
        """Preferred engine for newly loaded data.

        DuckDB is used by default whenever it is installed, since it
        handles large files (millions of rows / multi-GB) far better
        than the row-by-row SQLite path. Falls back to SQLite when
        DuckDB is unavailable so the app still works with a bare
        interpreter.
        """
        return "duckdb" if HAS_DUCKDB else "sqlite"

    def _nested_list_columns(self, mgr, table):
        """Top-level array columns of ``table`` (LIST of anything) -- the
        shreddable ones."""
        try:
            raw = mgr._column_types_raw(table)
        except Exception as e:
            # .423: a failed type scan used to masquerade as "no nested
            # array columns" -- return the reason so the note is honest
            return {"__error__": str(e)}
        return [c for c, t in raw.items()
                if str(t).strip().endswith("[]")]

    def _shred_after_load(self, loaded, root_id=None):
        """The Load modal's "flatten" toggle: after a nested DuckDB load,
        shred every array column into relational tables (the .371 engine --
        one vectorized pass per table over the parquet cache). The LOAD's
        success is never at stake: a shred error is reported on the item; a
        cancel stops with what was created so far."""
        mgr = self.duckdb
        if mgr is None:
            return loaded
        for item in loaded or []:
            # .423: the gate used a storage-string allowlist -- and the
            # query-in-place JSON loader stamps a value that was never
            # on it, so the on-box "flatten" checkbox skipped the sophis
            # load SILENTLY (no note, no shred, Tables stayed at 1).
            # The real predicate is simply: a DuckDB table. Whether it
            # has anything to shred is decided by the candidate scan
            # below, which already leaves an honest note when empty.
            if item.get("engine") != "duckdb" or not item.get("name"):
                continue
            table = item.get("name")
            # .474/.494/.495: flatten the WHOLE TABLE (not per-column). Produces
            # the base (top-level scalars + single structs inlined), the
            # <base>_joinkeys hub, one table per list -- AND, for a list whose
            # elements carry further nested arrays, the full deep hierarchy
            # (.494). Every table also gets a single-column surrogate key
            # _sk + _parent_sk (.495) so children join to parents on one column
            # (child._parent_sk = parent._sk); the compound _rid + ordinals stay
            # for source-row grouping and array position.
            r = self.flatten_table(table, base=table, root_id=root_id)
            if r.get("error"):
                item["shred_error"] = r["error"]
                continue
            if r.get("cancelled"):
                item["shred_cancelled"] = True
                continue
            made = [x["name"] for x in (r.get("created") or [])]
            if made:
                item["shredded"] = made
                if r.get("root_id"):
                    item["root_id_stats"] = r["root_id"]
            elif r.get("note"):
                item["shred_note"] = r["note"]
        return loaded

    def load_file(self, path, destination="auto", base_name=None,
                  progress=None, delimiter=None, mode="materialize",
                  sheet=None, header_row=1, op_kind="load", exclude=None,
                  flatten=None, shred=False, root_id=None):
        if destination in (None, "", "auto", "default"):
            destination = self.default_destination()
        _label = base_name or os.path.basename(path or "") or "file"
        _oid = None
        try:
            _oid = opreg.begin(op_kind, target=_label)
        except Exception:
            _oid = None
        try:
            # UI Flatten toggle uses shred=True with flatten=False. Load full
            # STRUCT/LIST types (not depth-capped JSON) so relational shred
            # sees real arrays/structs — and stays the fast Parquet path for
            # large files — instead of mis-shredding JSON scalar columns.
            loaded = loaders.load_file(
                self, path, destination=destination,
                base_name=base_name, progress=progress,
                delimiter=delimiter, mode=mode,
                sheet=sheet, header_row=header_row,
                exclude=exclude, flatten=flatten,
                root_id=root_id,
                full_nested=bool(shred) and flatten is False)
            if shred:
                loaded = self._shred_after_load(loaded, root_id=root_id)
        finally:
            try:
                opreg.end(_oid)
            except Exception:
                pass
        self._invalidate_profiles()
        self._invalidate_counts()   # new/replaced data -> drop count + flow caches
        try:
            self._prime_count(loaded.get("engine"),
                              loaded.get("table"), loaded.get("rows"))
        except Exception:
            pass
        return loaded

    # extensions load_file knows how to read
    _FOLDER_EXTS = {"csv", "tsv", "txt", "json", "ndjson", "jsonl",
                    "parquet", "pq", "xlsx", "xlsm", "xls"}

    def find_loadable_files(self, folder, recursive=False):
        """List supported data files in a folder (sorted). Non-recursive by
        default; recursive walks subfolders."""
        out = []
        if not folder or not os.path.isdir(folder):
            return out
        if recursive:
            for root, _dirs, names in os.walk(folder):
                for nm in sorted(names):
                    ext = os.path.splitext(nm)[1].lower().lstrip(".")
                    if ext in self._FOLDER_EXTS:
                        out.append(os.path.join(root, nm))
        else:
            for nm in sorted(os.listdir(folder)):
                fp = os.path.join(folder, nm)
                ext = os.path.splitext(nm)[1].lower().lstrip(".")
                if ext in self._FOLDER_EXTS and os.path.isfile(fp):
                    out.append(fp)
        return out

    def load_folder(self, folder, destination="auto", recursive=False,
                    on_file=None, delimiter=None, should_cancel=None,
                    op_kind="load"):
        """Load every supported data file in a folder, each as its own table.
        ``on_file(index, count, filename)`` (optional) is called before each
        file so a caller can report progress. ``delimiter`` (CSV only) forces a
        field separator for every delimited file in the batch. Files that fail
        are collected rather than aborting the batch. ``should_cancel`` (an
        optional no-arg predicate) is checked before each file and after a file
        errors out; when it returns true the batch stops cleanly (pair it with
        ``interrupt_loads`` to abort the file already in flight). Returns
        {ok, loaded, rows, files, errors, cancelled} or {error}."""
        if not folder or not os.path.isdir(folder):
            return {"error": "Folder not found."}
        files = self.find_loadable_files(folder, recursive)
        if not files:
            return {"error": "No CSV / JSON / Parquet files found in that "
                    "folder."}
        all_loaded, errors = [], []
        cancelled = False
        for i, fp in enumerate(files):
            if should_cancel is not None and should_cancel():
                cancelled = True
                break
            if on_file is not None:
                try:
                    on_file(i, len(files), os.path.basename(fp))
                except Exception:
                    pass
            try:
                all_loaded.extend(self.load_file(fp, destination=destination,
                                                 delimiter=delimiter,
                                                 op_kind=op_kind))
            except Exception as e:
                if should_cancel is not None and should_cancel():
                    cancelled = True
                    break
                errors.append({"file": os.path.basename(fp),
                               "error": err_str(e)})
        return {"ok": True, "loaded": all_loaded,
                "rows": sum((t.get("rows") or 0) for t in all_loaded),
                "files": len(files), "errors": errors, "cancelled": cancelled}

    # ---- session restore -------------------------------------------
    def record_load(self, kind, path, destination="auto", base_name=None,
                    recursive=False, origin=None):
        """Remember a file/folder load so the next launch can rebuild it.
        ``origin`` (.550) is the ORIGINAL source when known -- a browser
        upload's real filename, or a native/folder load's full path --
        kept alongside the converted/parked ``path`` for Properties."""
        try:
            self.manifest.add(kind, path, destination=destination,
                              base_name=base_name, recursive=recursive,
                              origin=origin)
        except Exception:
            pass

    def restore_session(self):
        """Replay the persisted load manifest in the background, so a restart
        comes back with the same loaded tables. Safe to call once; does
        nothing if the manifest is empty. Sources that have since moved or
        disappeared are skipped without aborting the rest."""
        if self._restore_started:
            return
        self._restore_started = True
        actions = self.manifest.all()
        if not actions:
            return
        self.restoring = True

        def run():
            done = 0
            try:
                for a in actions:
                    # Yield to real user work: if the user has kicked off a
                    # load/query/run, let it go first instead of making it wait
                    # on the engine lock behind the restore. Restore resumes
                    # once the user is idle. Bounded so a long-running user op
                    # can't starve restore forever.
                    self._restore_wait_if_busy()
                    kind = a.get("kind")
                    path = a.get("path")
                    dest = a.get("destination") or "auto"
                    try:
                        if kind == "folder" and path and os.path.isdir(path):
                            res = self.load_folder(
                                path, destination=dest,
                                recursive=bool(a.get("recursive")),
                                op_kind="restore")
                            if isinstance(res, dict):
                                done += len(res.get("loaded") or [])
                        elif kind == "file" and path and os.path.isfile(path):
                            self.load_file(path, destination=dest,
                                           base_name=a.get("base_name"),
                                           op_kind="restore")
                            done += 1
                    except Exception:
                        pass  # a changed/missing source can't abort restore
            finally:
                self._restored_count = done
                self.restoring = False

        threading.Thread(target=run, daemon=True).start()

    def _restore_wait_if_busy(self, poll=0.15, max_wait=120.0):
        """Block while a real user operation is in flight, so session restore
        defers to the user's first action instead of holding the engine lock
        ahead of it. Only non-restore ops count as 'busy'. Bounded by
        ``max_wait`` so a genuinely long user op can't stall restore forever."""
        waited = 0.0
        while waited < max_wait:
            try:
                kinds = opreg.active_kinds()
            except Exception:
                return
            if not (kinds - {"restore"}):
                return
            time.sleep(poll)
            waited += poll

    # ---- NodeFlow (visual data flow) -------------------------------
    def relation_columns(self, sql):
        """Column names of an arbitrary SELECT, via a zero-row probe on the
        engine that owns its tables. Used by the graph compiler."""
        target = self._choose_local_engine(sql)
        eng, _kind = self._engine_obj(target)
        return self._col_names(eng, sql=sql)

    def _flow_has_pivot_upstream(self, graph, node_id, port):
        """True if (node_id, port) is a pivot/python or has one upstream.

        Pivot and Python output columns are only known after they run, so
        such ports need materialising to read their columns; everything else
        can be probed cheaply with a zero-row compile."""
        from . import nodeflow
        nodes = nodeflow._node_map(graph)
        seen = set()
        runtime_types = ("pivot", "python")

        def walk(nid, prt):
            if (nid, prt) in seen:
                return False
            seen.add((nid, prt))
            node = nodes.get(nid)
            if node is None:
                return False
            if node.get("type") in runtime_types:
                return True
            for ip in nodeflow.NODE_PORTS.get(node.get("type"), {}).get(
                    "in", []):
                sn, sp = nodeflow.upstream(graph, nid, ip)
                if sn and walk(sn, sp):
                    return True
            return False

        return walk(node_id, port)

    def nodeflow_columns_batch(self, graph, requests):
        """Columns for several (node, port) pairs in one request, sharing a
        single zero-row probe cache across the whole batch -- so the inspector
        for a multi-input node (join / multijoin) makes one round-trip instead
        of one per port, and shared upstream subtrees are probed once."""
        graph = applyvars.resolve_graph(graph)
        from . import nodeflow
        out = []
        eng_name = ("duckdb"
                    if self._flow_engine_target(graph) == DUCKDB_TARGET
                    else "sqlite")
        _ccache = {}

        def _cols(sql):
            hit = _ccache.get(sql)
            if hit is not None:
                return hit
            r = self.relation_columns(sql)
            _ccache[sql] = r
            return r

        for req in (requests or []):
            nid = (req or {}).get("node")
            prt = (req or {}).get("port") or "out"
            if not nid:
                out.append({"node": nid, "port": prt, "error": "missing node"})
                continue
            try:
                if self._flow_has_pivot_upstream(graph, nid, prt):
                    # pivot columns need a real materialise; use the single path
                    res = self.nodeflow_columns(graph, nid, prt)
                    if res.get("error"):
                        out.append({"node": nid, "port": prt,
                                    "error": res["error"]})
                    else:
                        out.append({"node": nid, "port": prt,
                                    "columns": res.get("columns") or []})
                    continue
                sql = nodeflow.compile_port(graph, nid, prt, _cols,
                                            None, eng_name)
                out.append({"node": nid, "port": prt, "columns": _cols(sql)})
            except nodeflow.NodeflowError as e:
                out.append({"node": nid, "port": prt, "error": str(e)})
            except Exception as e:
                out.append({"node": nid, "port": prt,
                            "error": err_str(e)})
        return {"results": out}

    def nodeflow_columns(self, graph, node_id, port):
        """Columns of (node_id, port) -- used by the inspectors to offer the
        upstream fields. Probes cheaply via a zero-row compile, and only
        materialises (which is far more expensive) when a pivot upstream means
        the columns are shaped at run time."""
        from . import nodeflow
        try:
            # .475: a shred node's ports read tables its pre-pass creates;
            # ensure the family exists before we introspect or compile
            # (cheap and idempotent -- reuses existing tables unless the
            # node asks to refresh).
            try:
                self._shred_flow_prepass(graph, [(node_id, port)])
            except Exception:
                pass  # introspection is best-effort; a real run re-raises
            if self._flow_has_pivot_upstream(graph, node_id, port):
                et = self._flow_engine_target(graph)
                created = []
                tmp = self._materialize_flow(graph, node_id, port, et, created)
                eng, _kind = self._engine_obj(et)
                cols = self._col_names(eng, table=tmp)
                self._drop_flow_temps(et, created)
                return {"columns": cols}
            eng_name = ("duckdb"
                        if self._flow_engine_target(graph) == DUCKDB_TARGET
                        else "sqlite")
            # cache zero-row probes for the duration of this one compile so a
            # shared upstream subtree isn't re-probed once per consumer
            _ccache: dict = {}

            def _cols(sql):
                hit = _ccache.get(sql)
                if hit is not None:
                    return hit
                out = self.relation_columns(sql)
                _ccache[sql] = out
                return out

            sql = nodeflow.compile_port(graph, node_id, port,
                                        _cols, None, eng_name)
            return {"columns": _cols(sql)}
        except nodeflow.NodeflowError as e:
            return {"error": str(e)}
        except Exception as e:
            return {"error": err_str(e)}

    def run_nodeflow_browse(self, graph, node_id, query_id=None):
        """Materialise a browse node's input and profile its columns (reusing
        the profiler), returning a TableProfile for the canvas to show."""
        from . import nodeflow
        from .profiler import profile_table
        created = []
        et = LOCAL_TARGET
        try:
            et = self._flow_engine_target(graph)
            sn, sp = nodeflow.upstream(graph, node_id, "in")
            if sn is None:
                return {"error": "Connect an input to the browse node."}
            eng, _kind = self._engine_obj(et)
            self._register_run(
                query_id, eng, surface="node",
                label=self._flow_node_label(graph, node_id))
            tmp = self._materialize_flow(graph, sn, sp, et, created)
        except Exception as e:
            self._flow_cleanup(query_id, et, created)
            return self._flow_exc(e)
        eng, _kind = self._engine_obj(et)
        try:
            out = profile_table(eng, "flow", source='"%s"' % tmp)
        except Exception as e:
            out = {"error": err_str(e)}
        self._flow_cleanup(query_id, et, created)
        return out

    def validate_nodeflow(self, graph, node_id, checks, query_id=None):
        """Run data-quality checks against a validate node's input (which
        otherwise passes data through). ``checks`` is a list of
        {type, col?, n?}; type is rows_min/rows_max/not_null/unique. Returns
        {ok, total_rows, results:[{type,target,pass,detail}]} or {error}."""
        from . import nodeflow
        created = []
        et = LOCAL_TARGET
        try:
            et = self._flow_engine_target(graph)
            sn, sp = nodeflow.upstream(graph, node_id, "in")
            if sn is None:
                return {"error": "Connect an input to the validate node."}
            eng, _kind = self._engine_obj(et)
            self._register_run(
                query_id, eng, surface="node",
                label=self._flow_node_label(graph, node_id))
            tmp = self._materialize_flow(graph, sn, sp, et, created)
        except Exception as e:
            self._flow_cleanup(query_id, et, created)
            return self._flow_exc(e)

        def scalar(sql):
            _c, rows = eng.execute(sql)
            return rows[0][0] if rows else None

        out = {"results": []}
        try:
            total = int(scalar('SELECT COUNT(*) FROM "%s"' % tmp) or 0)
            out["total_rows"] = total
            cols = set(self._col_names(eng, table=tmp))
            for chk in (checks or []):
                ctype = (chk.get("type") or "").lower()
                col = (chk.get("col") or "").strip()
                try:
                    n = int(chk.get("n")) if chk.get("n") not in (None, "") \
                        else None
                except Exception:
                    n = None
                res = {"type": ctype,
                       "target": col or (str(n) if n is not None else ""),
                       "pass": False, "detail": ""}
                if ctype == "rows_min":
                    res["pass"] = n is not None and total >= n
                    res["detail"] = "%d rows (need >= %s)" % (total, n)
                elif ctype == "rows_max":
                    res["pass"] = n is not None and total <= n
                    res["detail"] = "%d rows (need <= %s)" % (total, n)
                elif ctype == "not_null":
                    if col not in cols:
                        res["detail"] = "no such column"
                    else:
                        nulls = int(scalar(
                            'SELECT COUNT(*) FROM "%s" WHERE %s IS NULL'
                            % (tmp, nodeflow._q(col))) or 0)
                        res["pass"] = nulls == 0
                        res["detail"] = "%d null(s)" % nulls
                elif ctype == "unique":
                    if col not in cols:
                        res["detail"] = "no such column"
                    else:
                        distinct = int(scalar(
                            'SELECT COUNT(DISTINCT %s) FROM "%s"'
                            % (nodeflow._q(col), tmp)) or 0)
                        nonnull = int(scalar(
                            'SELECT COUNT(%s) FROM "%s"'
                            % (nodeflow._q(col), tmp)) or 0)
                        dups = nonnull - distinct
                        res["pass"] = dups == 0
                        res["detail"] = "%d duplicate value(s)" % dups
                else:
                    res["detail"] = "unknown check"
                out["results"].append(res)
            out["ok"] = (all(r["pass"] for r in out["results"])
                         if out["results"] else True)
        except Exception as e:
            out = {"error": err_str(e)}
        self._flow_cleanup(query_id, et, created)
        return out

    def run_nodeflow_chart(self, graph, node_id, spec, query_id=None):
        """Materialise a chart node's input and build a chart from it (reusing
        chart_data), returning ChartData for the canvas to render."""
        from . import nodeflow
        created = []
        et = LOCAL_TARGET
        try:
            et = self._flow_engine_target(graph)
            sn, sp = nodeflow.upstream(graph, node_id, "in")
            if sn is None:
                return {"error": "Connect an input to the chart node."}
            eng0, _k0 = self._engine_obj(et)
            self._register_run(
                query_id, eng0, surface="node",
                label=self._flow_node_label(graph, node_id))
            tmp = self._materialize_flow(graph, sn, sp, et, created)
        except Exception as e:
            self._flow_cleanup(query_id, et, created)
            return self._flow_exc(e)
        eng, kind = self._engine_obj(et)
        cs = dict(spec or {})
        cs.pop("result_id", None)
        cs["table"] = tmp
        cs["engine"] = kind
        if query_id:
            cs["query_id"] = query_id
        out = self.chart_data(cs)
        self._flow_cleanup(query_id, et, created)
        return out

    def run_nodeflow_reconcile(self, graph, node_id, keys, compare=None,
                               balance=None, query_id=None):
        """Materialise a reconcile node's two inputs and run the reconcile
        engine over them, returning the field-level report."""
        from . import nodeflow
        if not keys:
            return {"error": "Pick at least one key field to reconcile on."}
        created = []
        et = LOCAL_TARGET
        try:
            et = self._flow_engine_target(graph)
            ls = nodeflow.upstream(graph, node_id, "left")
            rs = nodeflow.upstream(graph, node_id, "right")
            if ls[0] is None or rs[0] is None:
                return {"error": "Connect both the left and right inputs of "
                        "the reconcile node."}
            eng0, _k0 = self._engine_obj(et)
            self._register_run(
                query_id, eng0, surface="node",
                label=self._flow_node_label(graph, node_id))
            ltmp = self._materialize_flow(graph, ls[0], ls[1], et, created)
            rtmp = self._materialize_flow(graph, rs[0], rs[1], et, created)
        except Exception as e:
            self._flow_cleanup(query_id, et, created)
            return self._flow_exc(e)
        eng, _kind = self._engine_obj(et)
        # the reconcile engine resolves the engine via the table registry, so
        # register the temp tables for the duration of the call (they're
        # hidden from the sidebar by their "__" prefix), then unregister
        eng.table_columns[ltmp] = self._col_names(eng, table=ltmp)
        eng.table_columns[rtmp] = self._col_names(eng, table=rtmp)
        try:
            return self.reconcile({
                "left": ltmp, "right": rtmp,
                "keys": list(keys),
                "compare": list(compare or []),
                "balance": balance,
            })
        finally:
            eng.table_columns.pop(ltmp, None)
            eng.table_columns.pop(rtmp, None)
            self._flow_cleanup(query_id, et, created)

    def _engine_of_table(self, name):
        """'duckdb' / 'sqlite' / None for a loaded table."""
        try:
            if (self.duckdb is not None
                    and name in getattr(self.duckdb, "table_columns", {})):
                return "duckdb"
        except Exception:
            pass
        if name in getattr(self.db, "table_columns", {}):
            return "sqlite"
        return None

    def _flow_engine_target(self, graph):
        """The single engine a flow runs on, derived from its input tables and
        any loaded tables referenced by SQL nodes."""
        from . import nodeflow
        eng = None

        def _note(e):
            nonlocal eng
            if e:
                if eng and eng != e:
                    raise nodeflow.NodeflowError(
                        "All input tables in a flow must live in the same "
                        "engine (DuckDB or SQLite).")
                eng = e

        def _scan(nodes):
            for n in (nodes or []):
                if not isinstance(n, dict):
                    continue
                typ = n.get("type")
                cfg = n.get("config") or {}
                if (typ == "input" or typ == "directory"
                        or typ == "appendfolder" or typ == "apinode"
                        or typ in ("sqlserver", "sharepoint", "webscrape")):
                    t = (cfg.get("table") or cfg.get("err_table")
                         or "").strip()
                    eng_hint = None
                    if typ in ("apinode", "sqlserver", "sharepoint", "webscrape"):
                        # Hidden source tables live in whichever engine the
                        # fetch loaded them into -- trust that record.
                        rec = self._api_node_tables.get(n.get("id")) or {}
                        eng_hint = rec.get("engine")
                    _note(eng_hint
                          or (self._engine_of_table(t) if t else None))
                elif typ == "filebrowser":
                    # reads files directly via DuckDB
                    _note("duckdb")
                elif typ == "sql":
                    # the {{in}} token isn't a table name, so it's ignored here;
                    # any real loaded table the query reads pins the engine.
                    for name in self._table_names_in(cfg.get("sql") or ""):
                        _note(self._engine_of_table(name))
                elif typ in ("group", "iterator"):
                    # a container's body holds the inputs / API nodes that pin
                    # the engine. The container iterator wraps its children in a
                    # synthetic "group" node, so descend into nested children --
                    # otherwise a nested API node is invisible here and the flow
                    # wrongly defaults to the local engine while the data sits
                    # in DuckDB ("no such table __nbapi_...").
                    _scan(cfg.get("children"))

        _scan(graph.get("nodes"))
        return DUCKDB_TARGET if eng == "duckdb" else LOCAL_TARGET

    def _independent_flow_target_groups(self, graph, targets):
        """Partition targets whose upstream node sets do not overlap."""
        edges = graph.get("edges") or []
        back = {}
        for edge in edges:
            src = (edge.get("from") or {}).get("node")
            dst = (edge.get("to") or {}).get("node")
            if src is not None and dst is not None:
                back.setdefault(dst, set()).add(src)

        def ancestors(nid):
            seen, stack = set(), [nid]
            while stack:
                cur = stack.pop()
                if cur in seen:
                    continue
                seen.add(cur)
                stack.extend(back.get(cur, ()))
            return seen

        items = [(target, ancestors(target[0])) for target in targets]
        groups = []
        for target, aset in items:
            hits = [i for i, (_ts, nodes) in enumerate(groups)
                    if nodes.intersection(aset)]
            if not hits:
                groups.append(([target], set(aset)))
                continue
            first = hits[0]
            groups[first][0].append(target)
            groups[first][1].update(aset)
            for i in reversed(hits[1:]):
                groups[first][0].extend(groups[i][0])
                groups[first][1].update(groups[i][1])
                groups.pop(i)
        return [g[0] for g in groups]

    def _parallel_flow_groups(self, graph, targets, engine_target, collect,
                              target_limits):
        """Build independent DuckDB target groups on separate connections."""
        if (engine_target != DUCKDB_TARGET or not self.parallel_nodeflows
                or len(targets) < 2):
            return None
        budget = self._sync_flow_cache_limits()
        workers = max(1, int(budget.get("parallel_workers") or 1))
        if workers <= 1:
            return None
        groups = self._independent_flow_target_groups(graph, targets)
        if len(groups) <= 1:
            return None
        node_by_id = {n.get("id"): n for n in (graph.get("nodes") or [])}
        unsafe = {"apinode", "iterator", "while", "shred", "write",
                  "directory", "appendfolder", "sql", "filebrowser",
                  "createtable"}
        # Parallel workers may only execute pure SQL DAG branches. Nodes with
        # network/file side effects remain on the serial path.
        for group in groups:
            ancestors = set()
            for target in group:
                for g in self._independent_flow_target_groups(graph, [target]):
                    ancestors.update(t[0] for t in g)
            # The helper above returns only targets; walk all upstream here.
            back = {}
            for e in (graph.get("edges") or []):
                src = (e.get("from") or {}).get("node")
                dst = (e.get("to") or {}).get("node")
                if src is not None and dst is not None:
                    back.setdefault(dst, set()).add(src)
            stack = [t[0] for t in group]
            while stack:
                nid = stack.pop()
                if nid in ancestors:
                    continue
                ancestors.add(nid)
                stack.extend(back.get(nid, ()))
            if any(str((node_by_id.get(nid) or {}).get("type") or "").lower()
                   in unsafe for nid in ancestors):
                return None

        from concurrent.futures import ThreadPoolExecutor, as_completed
        manager, _kind = self._engine_obj(engine_target)
        completed_names = []
        result = {}
        lock = threading.Lock()

        def worker(group):
            branch = manager.branch_cursor()
            local_names = []
            try:
                limits = {k: v for k, v in (target_limits or {}).items()
                          if k in group}
                out = self._materialize_flows(
                    graph, group, engine_target, local_names,
                    target_limits=limits, _allow_parallel=False,
                    _engine_override=branch, _persistent_tables=True)
                return out, local_names
            except Exception:
                # A failed worker never reaches the parent's completed-name
                # list. Drop anything it managed to create before propagating
                # the error, otherwise hidden regular tables survive the run.
                self._drop_flow_temps(engine_target, local_names)
                raise
            finally:
                branch.close()

        try:
            with ThreadPoolExecutor(
                    max_workers=min(workers, len(groups)),
                    thread_name_prefix="nodeflow-branch") as pool:
                futures = [pool.submit(worker, group) for group in groups]
                for future in as_completed(futures):
                    out, names = future.result()
                    with lock:
                        result.update(out)
                        completed_names.extend(names)
        except Exception:
            self._drop_flow_temps(engine_target, completed_names)
            raise
        if collect is not None:
            collect.extend(completed_names)
        return result

    def _materialize_flows(self, graph, targets, engine_target,
                           collect=None, target_limits=None,
                           _allow_parallel=True, _engine_override=None,
                           _persistent_tables=False):
        """Build every (node_id, port) in ``targets`` and everything upstream
        of them on ``engine_target`` in ONE pass, returning
        ``{(node_id, port): temp_table_name}``.

        Because the per-node temp cache is shared across all targets, a node
        that feeds more than one target is computed exactly once for the whole
        batch -- this is what makes Run all share heavy upstream subgraphs
        instead of recomputing them per terminal.

        Runs of SQL-only, single-consumer nodes are *fused* into one nested
        query rather than written out as a temp table each, so the engine can
        optimise across the whole run. A node is materialised (a real temp
        table, shared) at every boundary: a pivot (its columns are shaped at
        run time), a node whose output feeds more than one consumer (computed
        once, not re-inlined), and each flow target itself. If a fused query
        hits an execution error, the whole flow is rebuilt node-by-node so the
        failure is attributed to a single node. Names of every temp table
        created are appended to ``collect`` so the caller can drop them."""
        targets = list(targets)
        target_limits = target_limits or {}
        if _allow_parallel and _engine_override is None:
            parallel = self._parallel_flow_groups(
                graph, targets, engine_target, collect, target_limits)
            if parallel is not None:
                return parallel
        # resolve ${name} workflow-variable tokens once, up front: everything
        # below compiles a graph whose tokens are already filled in
        graph = applyvars.resolve_graph(graph)
        # SQL Server / SharePoint / Web scrape: reconnect + import before
        # compile so Dashboard / Run all work with a saved profile + password
        # without a prior interactive Fetch.
        self._ensure_source_nodes_fetched(
            graph, [nid for nid, _port in targets], query_id=None)
        # SHRED nodes create their relational tables before anything composes
        self._shred_flow_prepass(graph, targets)
        from . import nodeflow
        import re as _re
        import secrets as _secrets
        eng, _kind = self._engine_obj(engine_target)
        if _engine_override is not None:
            eng = _engine_override
        # a token unique to this materialisation pass: temp-table names include
        # it so two flows running at once (parallel Run all, or two browser tabs)
        # can't collide on the deterministic per-node names or drop each other's
        # tables. Constant within a pass, so a node feeding several consumers
        # still de-dupes to one shared temp table within the pass.
        _tok = _secrets.token_hex(3)

        _cols_cache: dict = {}

        def cols_of(sql):
            hit = _cols_cache.get(sql)
            if hit is not None:
                return hit
            out = self._col_names(eng, sql=sql)
            _cols_cache[sql] = out
            return out

        nodes = {n["id"]: n for n in (graph.get("nodes") or [])
                 if n.get("id")}
        # column liveness for projection pushdown: which columns each node's
        # output actually feeds downstream (target materialised in full). Source
        # nodes read only those, so a wide table behind a select isn't fully
        # scanned. Off -> empty map -> every source stays SELECT *.
        if getattr(self, "project_pushdown", True):
            try:
                needed_map = nodeflow.needed_columns(graph, list(targets))
            except Exception:
                needed_map = {}
        else:
            needed_map = {}
        # Incremental flow cache: a fingerprint per node output over its
        # subtree + projection + data epoch + engine. A boundary node whose
        # fingerprint is already cached is reused instead of recomputed. Empty
        # (no caching) when the cache is off or fingerprinting fails.
        use_volatile_cache = bool(
            getattr(self, "flow_cache", False) and _engine_override is None)
        fps = {}
        if use_volatile_cache:
            try:
                fps = nodeflow.flow_fingerprints(
                    graph, needed_map, getattr(self, "_data_epoch", 0),
                    "duckdb" if engine_target == DUCKDB_TARGET else "sqlite")
            except Exception:
                fps = {}
        persistent_fps = {}
        persistent_salt = self._persistent_flow_salt(graph, engine_target)
        if persistent_salt:
            try:
                persistent_fps = nodeflow.flow_fingerprints(
                    graph, needed_map, "persist-v1-" + persistent_salt,
                    "duckdb", digest_chars=40)
            except Exception:
                persistent_fps = {}
        cache_checkpoints = bool(fps or persistent_fps)
        checkpoint_nodes = set(_FLOW_CHECKPOINT_TYPES)
        if cache_checkpoints:
            try:
                checkpoint_nodes |= nodeflow.incremental_checkpoint_nodes(
                    graph, targets)
            except Exception:
                # The original blocking-node checkpoint policy remains a safe
                # fallback if an unusual legacy graph defeats the planner.
                pass
        # how many input edges each (node, port) feeds; >1 means it's shared,
        # so it's materialised once instead of inlined into each consumer.
        consumers = {}
        for e in (graph.get("edges") or []):
            f = e.get("from") or {}
            k = (f.get("node"), f.get("port"))
            if k[0] is not None:
                consumers[k] = consumers.get(k, 0) + 1

        def tmp_for(nid, prt):
            return "__nbflow_%s_%s__%s" % (
                _tok,
                _re.sub(r"[^A-Za-z0-9]", "", str(nid)),
                _re.sub(r"[^A-Za-z0-9]", "", str(prt)))

        def run_pass(fuse):
            built = {}

            def node_sql(nid, prt, stack):
                node = nodes.get(nid)
                if node is None:
                    raise nodeflow.NodeflowError("That node no longer exists.")
                typ = node.get("type")
                # a pivot reads its input twice (distinct values + the
                # crosstab), a SQL node's {{in}} token may appear more than
                # once, and a Python node executes outside SQL — their
                # inputs are always materialised, never inlined.
                force = typ in ("pivot", "sql", "python")

                def get_input(in_port):
                    sn, sp = nodeflow.upstream(graph, nid, in_port)
                    if sn is None:
                        return None
                    if force:
                        return '"%s"' % build(sn, sp, stack)
                    return resolve(sn, sp, stack)

                if typ == "pivot":
                    in_rel = get_input("in")
                    if in_rel is None:
                        raise nodeflow.NodeflowError(
                            "Connect an input to the pivot node.")
                    return self._pivot_crosstab_sql(
                        eng, in_rel, node.get("config") or {},
                        "duckdb" if engine_target == DUCKDB_TARGET else "sqlite")
                if typ == "python":
                    in_rel = get_input("in")
                    cfg = node.get("config") or {}
                    side = "__nbpy_%s_%s" % (
                        _tok,
                        _re.sub(r"[^A-Za-z0-9]", "", str(nid)))
                    from . import python_node as _pynode
                    try:
                        _pynode.run_into_table(
                            eng, side,
                            in_rel=in_rel,
                            code=cfg.get("code") or "",
                            timeout_s=cfg.get("timeout_s") or 30,
                        )
                    except _pynode.PythonNodeError as e:
                        raise nodeflow.NodeRunError(
                            str(e), node_id=nid, node_type="python") from e
                    except Exception as e:
                        if _is_interrupt(e):
                            raise
                        raise nodeflow.NodeRunError(
                            nodeflow.explain_node_error(
                                node, err_str(e)),
                            node_id=nid, node_type="python") from e
                    if collect is not None:
                        collect.append(side)
                    return 'SELECT * FROM "%s" AS _py' % side
                live_out = needed_map.get((nid, prt))
                sql = nodeflow.node_output_sql(
                    node, prt, get_input, cols_of,
                    "duckdb" if engine_target == DUCKDB_TARGET else "sqlite",
                    needed=live_out)
                # Keep dependency-only fields available while this node runs,
                # then remove them before an intermediate checkpoint is cached.
                # This makes projection pushdown effective across filter/sort/
                # group boundaries instead of only at the original source.
                if live_out is not None:
                    sql = nodeflow.project_output_sql(sql, live_out, cols_of)
                return sql

            def resolve(nid, prt, stack):
                """A FROM-able expression for (nid, prt): a shared temp-table
                name, or an inline subquery when it's a fusable single
                consumer."""
                key = (nid, prt)
                if key in built:
                    return '"%s"' % built[key]
                node = nodes.get(nid)
                if node is None:
                    raise nodeflow.NodeflowError("That node no longer exists.")
                _ctyp = node.get("type")
                can_fuse = (fuse and _ctyp not in ("pivot", "python")
                            and not (cache_checkpoints
                                     and nid in checkpoint_nodes
                                     and (fps.get(nid)
                                          or persistent_fps.get(nid)))
                            and consumers.get(key, 0) <= 1)
                if not can_fuse:
                    return '"%s"' % build(nid, prt, stack)
                if key in stack:
                    raise nodeflow.NodeflowError(
                        "This flow has a loop -- remove the cycle.")
                return "(%s)" % node_sql(nid, prt, stack + [key])

            def build(nid, prt, stack):
                def _exec_create(nm, q):
                    # per-node heartbeat (.410 journal/node audit): long
                    # multi-node flows were a SILENT op -- no beats, so
                    # the Activity card sat blank and the watchdog could
                    # stall-flag a healthy flow mid-node. Name the node
                    # being built and pulse the op before each CREATE.
                    try:
                        opreg.beat()
                    except Exception:
                        pass
                    try:
                        eng.execute('DROP VIEW IF EXISTS "%s"' % nm)
                    except Exception:
                        pass
                    eng.execute('DROP TABLE IF EXISTS "%s"' % nm)
                    create_kind = "TABLE" if _persistent_tables else "TEMP TABLE"
                    if fuse:
                        # composed query spanning several nodes -- a failure
                        # here can't be pinned to one node, so let the caller
                        # rebuild node-by-node (which lands in the branch below)
                        eng.execute('CREATE %s "%s" AS %s' % (create_kind, nm, q))
                        return
                    # node-by-node pass: this CREATE is exactly one node, so a
                    # failure is that node's -- attribute it with a clear,
                    # node-named explanation (and a cast hint for type errors).
                    try:
                        eng.execute('CREATE %s "%s" AS %s' % (create_kind, nm, q))
                    except Exception as e:
                        if _is_interrupt(e):
                            raise
                        nd = nodes.get(nid) or {}
                        raise nodeflow.NodeRunError(
                            nodeflow.explain_node_error(
                                nd, err_str(e)),
                            node_id=nid, node_type=nd.get("type")) from e

                key = (nid, prt)
                if key in built:
                    return built[key]
                if key in stack:
                    raise nodeflow.NodeflowError(
                        "This flow has a loop -- remove the cycle.")
                fp = fps.get(nid)
                persistent_fp = persistent_fps.get(nid)
                # A row-limited preview target is intentionally incomplete and
                # must never enter either full-result cache. Its upstream
                # checkpoints may still be reused.
                row_limit = target_limits.get(key)
                if row_limit:
                    fp = None
                    persistent_fp = None
                port_key = _re.sub(r"[^A-Za-z0-9]", "", str(prt))
                if fp:
                    # Multi-output nodes have one node fingerprint but distinct
                    # relations per output port.
                    fp = "%s_%s" % (fp, port_key)
                if persistent_fp:
                    persistent_fp = "%s_%s" % (persistent_fp, port_key)
                # cache hit: reuse the materialised table if it still exists
                if fp:
                    cached = self._flow_cache_get(fp)
                    if cached is not None:
                        if self._table_exists(eng, cached):
                            built[key] = cached
                            self._flow_cache_registry.record_hit()
                            return cached
                        # An engine recycle or external DROP can invalidate a
                        # backing table. Remove the dead registry entry so it
                        # stops consuming budget and being probed every run.
                        self._flow_cache_registry.discard(
                            fp, drop=False, stale=True)
                if persistent_fp:
                    restored = "__nfpersist_%s_%s" % (_tok, persistent_fp[:16])
                    if self._persistent_flow_restore(
                            eng, persistent_fp, restored,
                            persistent_tables=_persistent_tables):
                        built[key] = restored
                        kept = False
                        if fp and use_volatile_cache:
                            self._flow_cache_registry.record_miss()
                            kept = self._flow_cache_put(
                                fp, restored, engine_target)
                        if not kept and collect is not None:
                            collect.append(restored)
                        return restored
                sql = node_sql(nid, prt, stack + [key])
                if row_limit:
                    sql = "SELECT * FROM (%s) AS _preview LIMIT %d" % (
                        sql, max(1, int(row_limit)))
                if fp:
                    # content-addressed cache table; lifetime is the cache's
                    # (LRU), so it is NOT added to the per-pass drop list.
                    name = "__nbcache_%s" % fp
                    _exec_create(name, sql)
                    built[key] = name
                    self._flow_cache_registry.record_miss()
                    kept = self._flow_cache_put(fp, name, engine_target)
                    if persistent_fp:
                        self._persistent_flow_publish(
                            eng, persistent_fp, name)
                    if not kept and collect is not None:
                        # A table larger than the whole byte budget remains
                        # valid for this run, but is cleaned up like a normal
                        # per-run temporary instead of being cached then
                        # immediately evicted out from under the caller.
                        collect.append(name)
                    return name
                name = tmp_for(nid, prt)
                _exec_create(name, sql)
                built[key] = name
                if persistent_fp:
                    self._persistent_flow_publish(eng, persistent_fp, name)
                if collect is not None:
                    collect.append(name)
                return name

            out = {}
            for (nid, prt) in targets:
                out[(nid, prt)] = build(nid, prt, [])
            return out

        first = bool(getattr(self, "fuse_flows", True))
        try:
            return run_pass(first)
        except nodeflow.NodeflowError:
            raise  # deterministic; fusion wouldn't change the outcome
        except Exception as e:
            if _is_interrupt(e) or not first:
                raise
            # a composition / SQL error under fusion: drop partial temps and
            # rebuild node-by-node so the error localises to one node.
            if collect:
                self._drop_flow_temps(engine_target, list(collect))
                del collect[:]
            return run_pass(False)

    def _shred_flow_prepass(self, graph, targets):
        """Run every SHRED node feeding a flow target before composing: the
        CREATEs are a side effect, so they happen once, up front, with the
        node id attached to any failure. ``refresh`` re-shreds every run;
        otherwise an existing family root means the work is already done."""
        from . import nodeflow
        edges = (graph or {}).get("edges") or []
        nodes = {n.get("id"): n for n in (graph or {}).get("nodes") or []}
        want = {t[0] for t in targets}
        # walk upstream from the targets so an unwired shred node never runs
        back = {}
        for e in edges:
            back.setdefault((e.get("to") or {}).get("node"), []).append(
                (e.get("from") or {}).get("node"))
        seen = set()
        stack = list(want)
        while stack:
            nid = stack.pop()
            if nid in seen:
                continue
            seen.add(nid)
            stack.extend(back.get(nid, []))
        for nid in seen:
            n = nodes.get(nid)
            if not n or (n.get("type") or "") != "shred":
                continue
            cfg = n.get("config") or {}
            table = (cfg.get("table") or "").strip()
            base = (cfg.get("base") or "").strip() or table
            label = cfg.get("label") or "shred"
            if not table:
                raise nodeflow.NodeRunError(
                    'The "%s" node needs a loaded (nested) table.'
                    % label, node_id=nid, node_type="shred")
            if not cfg.get("refresh"):
                eng = self.duckdb
                existing = {str(t).lower() for t in
                            (getattr(eng, "table_columns", {}) or {})}
                # the base table is the family anchor; its presence means
                # the flatten already ran (base + joinkeys + per-list).
                if base.lower() in existing:
                    continue
            # .475: the shred NODE flattens the WHOLE table via the very same
            # flatten_table the Load-modal toggle uses -- so it inherits every
            # improvement for free: the base + <base>_joinkeys + per-list
            # tables, the DEEP hierarchy when a list's elements carry their own
            # nested arrays (.494), and the single-column surrogate keys _sk +
            # _parent_sk on every table (.495), so a child joins to its parent
            # on one column (child._parent_sk = parent._sk). No node-specific
            # shred logic to keep in sync -- it is one code path.
            r = self.flatten_table(table, base=base)
            if r.get("error"):
                raise nodeflow.NodeRunError(
                    'The "%s" node could not flatten %s: %s'
                    % (label, table, r["error"]),
                    node_id=nid, node_type="shred")
            if r.get("cancelled"):
                raise nodeflow.NodeflowError("cancelled")

    def _materialize_flow(self, graph, node_id, port, engine_target,
                          collect=None, row_limit=None):
        """Single-target convenience wrapper over :meth:`_materialize_flows`.

        ``row_limit`` is used by fast previews: only the terminal relation is
        capped, while reusable upstream cache checkpoints stay complete.
        """
        limits = {(node_id, port): row_limit} if row_limit else None
        out = self._materialize_flows(
            graph, [(node_id, port)], engine_target, collect,
            target_limits=limits)
        return out[(node_id, port)]

    def _flow_cleanup(self, query_id, engine_target, names):
        """End-of-flow-op teardown: stop tracking the run for cancellation and
        drop the temp tables it created. This unregister-then-drop pair recurred
        in every flow handler's success and error paths."""
        self._unregister_run(query_id)
        self._drop_flow_temps(engine_target, names)

    def _drop_flow_temps(self, engine_target, names):
        """Drop the temp tables a flow created. Safe to call once the flow's
        result has been read into a cached result (paging never re-queries
        them), so intermediates don't accumulate across previews/runs."""
        if not names:
            return
        try:
            eng, _kind = self._engine_obj(engine_target)
        except Exception:
            return
        for nm in names:
            try:
                try:
                    eng.execute('DROP VIEW IF EXISTS "%s"' % nm)
                except Exception:
                    pass
                eng.execute('DROP TABLE IF EXISTS "%s"' % nm)
            except Exception:
                pass

    def _pivot_crosstab_sql(self, eng, in_rel, cfg, kind="sqlite"):
        """Crosstab SQL for a pivot node. Spreads the distinct *combinations*
        of one or more column fields into columns, computing one or more
        measures (value+agg) per cell, grouped by zero or more row fields.
        With ``subtotals`` on it adds a Total column and, per row grouping,
        subtotal rows plus a grand-total row (ROLLUP on DuckDB; an equivalent
        UNION ALL on SQLite). ``in_rel`` is a FROM-able relation (a quoted
        temp-table name, which a pivot's input always is, so re-scanning it for
        the SQLite subtotal emulation is cheap).

        Back-compat: a single ``col`` is accepted for ``cols``, and a single
        ``value``/``agg`` for ``values`` -- an old single-field pivot config
        produces byte-for-byte the same columns it did before."""
        from . import nodeflow
        rows = [r for r in (cfg.get("rows") or []) if r]
        cols = [c for c in (cfg.get("cols") or []) if c]
        if not cols and (cfg.get("col") or "").strip():
            cols = [(cfg.get("col") or "").strip()]
        if not cols:
            raise nodeflow.NodeflowError(
                "Pick a column field for the pivot node to spread into columns.")
        measures = []
        for m in (cfg.get("values") or []):
            if isinstance(m, str):
                measures.append((m.strip(), "sum"))
            else:
                measures.append(((m.get("field") or m.get("value") or "").strip(),
                                 (m.get("agg") or "sum").strip().lower()))
        if not measures:
            measures = [((cfg.get("value") or "").strip(),
                         (cfg.get("agg") or "sum").strip().lower())]
        subtotals = bool(cfg.get("subtotals")) and bool(rows)
        outline = bool(cfg.get("outline")) and bool(rows)
        use_rollup = subtotals or outline
        textcast = "VARCHAR" if kind == "duckdb" else "TEXT"
        aggmap = {"sum": "SUM", "avg": "AVG", "min": "MIN", "max": "MAX",
                  "count": "COUNT"}
        multi_meas = len(measures) > 1

        def q(s):
            return '"' + str(s).replace('"', '""') + '"'

        def lit(v):
            if isinstance(v, bool):
                return "1" if v else "0"
            if isinstance(v, (int, float)):
                return repr(v)
            return "'" + str(v).replace("'", "''") + "'"

        def mlabel(f, a):
            return "count" if (a == "count" or not f) else "%s(%s)" % (a, f)

        # distinct combinations of the column field(s)
        notnull = " AND ".join("%s IS NOT NULL" % q(c) for c in cols)
        sel_cols = ", ".join(q(c) for c in cols)
        order = ", ".join(str(i + 1) for i in range(len(cols)))
        CAP = 300
        try:
            _vc, crows = eng.execute(
                "SELECT DISTINCT %s FROM %s AS _p WHERE %s ORDER BY %s LIMIT %d"
                % (sel_cols, in_rel, notnull, order, CAP + 1))
        except Exception as e:
            raise nodeflow.NodeflowError(
                "Couldn't read the pivot column(s): %s" % e)
        combos = [tuple(r) for r in (crows or [])]
        if not combos:
            raise nodeflow.NodeflowError(
                "The pivot column(s) have no values to spread into columns.")
        if len(combos) > CAP:
            raise nodeflow.NodeflowError(
                "That pivot would spread into more than %d column groups. Add a "
                "filter upstream or use fewer column fields." % CAP)
        total_cols = len(combos) * len(measures) + (len(measures) if subtotals else 0)
        if total_cols > 1024:
            raise nodeflow.NodeflowError(
                "That pivot would produce %d columns (%d groups x %d measures). "
                "Reduce the column fields or measures, or filter upstream."
                % (total_cols, len(combos), len(measures)))

        def cell_for(pred, f, a):
            fn = aggmap.get(a, "SUM")
            if a == "count" or not f:
                return "SUM(CASE WHEN %s THEN 1 ELSE 0 END)" % pred
            return "%s(CASE WHEN %s THEN %s END)" % (fn, pred, q(f))

        # the crosstab measure columns: one per (combination x measure)
        cells = []  # (expr, header)
        for combo in combos:
            pred = " AND ".join("%s = %s" % (q(cols[i]), lit(combo[i]))
                                for i in range(len(cols)))
            clabel = " / ".join("" if x is None else str(x) for x in combo)
            for (f, a) in measures:
                hdr = (clabel + " / " + mlabel(f, a)) if multi_meas else clabel
                cells.append((cell_for(pred, f, a), hdr))
        # optional grand-total column(s) (sum across all column groups)
        if subtotals:
            for (f, a) in measures:
                fn = aggmap.get(a, "SUM")
                texpr = ("COUNT(*)" if (a == "count" or not f)
                         else "%s(%s)" % (fn, q(f)))
                thdr = ("Total / " + mlabel(f, a)) if multi_meas else "Total"
                cells.append((texpr, thdr))
        cell_sql = ["%s AS %s" % (e, q(h)) for (e, h) in cells]

        # ---- no rollup needed: a single GROUP BY (or one aggregate row) ----
        if not use_rollup:
            sel = [q(r) for r in rows] + cell_sql
            sql = "SELECT %s FROM %s AS _p" % (", ".join(sel), in_rel)
            if rows:
                sql += " GROUP BY %s" % ", ".join(q(r) for r in rows)
            return sql

        # ---- subtotals / outline: ROLLUP over the row dims --------------
        if kind == "duckdb":
            inner_sel = (
                ["%s AS %s" % (q(r), q("_r%d" % i)) for i, r in enumerate(rows)]
                + ["GROUPING(%s) AS %s" % (q(r), q("_g%d" % i))
                   for i, r in enumerate(rows)]
                + cell_sql)
            inner = "SELECT %s FROM %s AS _p GROUP BY ROLLUP(%s)" % (
                ", ".join(inner_sel), in_rel, ", ".join(q(r) for r in rows))
        else:
            # SQLite has no ROLLUP: UNION ALL one query per rollup level
            # (all dims grouped, then each shorter prefix, then the grand
            # total), tagging the rolled-up dims with a 1 in their _g flag.
            branches = []
            for k in range(len(rows), -1, -1):
                bsel, grp = [], []
                for i, r in enumerate(rows):
                    if i < k:
                        bsel.append("%s AS %s" % (q(r), q("_r%d" % i)))
                        grp.append(q(r))
                    else:
                        bsel.append("NULL AS %s" % q("_r%d" % i))
                for i in range(len(rows)):
                    bsel.append("%d AS %s" % (0 if i < k else 1, q("_g%d" % i)))
                bsel += cell_sql
                b = "SELECT %s FROM %s AS _p" % (", ".join(bsel), in_rel)
                if grp:
                    b += " GROUP BY %s" % ", ".join(grp)
                branches.append(b)
            inner = " UNION ALL ".join(branches)

        # ---- outline (Tableau-style nested sub-rows): one indented label
        # column. The ROLLUP prefix rows become parent header rows shown ABOVE
        # their children; each level is indented by depth. Without subtotals the
        # parent rows are pure headers (measures blank) and the grand total is
        # dropped; with subtotals the parents carry their subtotal values and a
        # 'Grand Total' row is kept. -------------------------------------------
        if outline:
            INDENT = "\u00a0\u00a0\u00a0"  # 3 non-breaking spaces per depth
            label_parts = ["CASE", "WHEN %s = 1 THEN 'Grand Total'" % q("_g0")]
            for k in range(1, len(rows)):
                label_parts.append(
                    "WHEN %s = 1 THEN %s || CAST(%s AS %s)"
                    % (q("_g%d" % k), lit(INDENT * (k - 1)),
                       q("_r%d" % (k - 1)), textcast))
            label_parts.append(
                "ELSE %s || CAST(%s AS %s) END"
                % (lit(INDENT * (len(rows) - 1)),
                   q("_r%d" % (len(rows) - 1)), textcast))
            msel = ["%s AS %s" % (" ".join(label_parts), q("Row Labels"))]
            leaf = "(" + " AND ".join("%s = 0" % q("_g%d" % i)
                                      for i in range(len(rows))) + ")"
            for (_, h) in cells:
                if subtotals:
                    msel.append(q(h))
                else:
                    msel.append("CASE WHEN %s THEN %s ELSE NULL END AS %s"
                                % (leaf, q(h), q(h)))
            where = "" if subtotals else (" WHERE %s = 0" % q("_g0"))
            # parent-before-children pre-order, grand total last
            ot = ["%s ASC" % q("_g0"), "%s ASC" % q("_r0")]
            for i in range(1, len(rows)):
                ot.append("%s DESC" % q("_g%d" % i))
                ot.append("%s ASC" % q("_r%d" % i))
            return "SELECT %s FROM (%s) AS _piv%s ORDER BY %s" % (
                ", ".join(msel), inner, where, ", ".join(ot))

        # ---- subtotals layout: separate row columns with 'Total' markers,
        # each subtotal row placed AFTER its group. ----------------------------
        msel = []
        for i, r in enumerate(rows):
            gi, ri = q("_g%d" % i), q("_r%d" % i)
            if i == 0:
                expr = ("CASE WHEN %s = 1 THEN 'Total' ELSE CAST(%s AS %s) END"
                        % (gi, ri, textcast))
            else:
                gp = q("_g%d" % (i - 1))
                expr = ("CASE WHEN %s = 1 AND %s = 0 THEN 'Total' "
                        "WHEN %s = 1 THEN '' ELSE CAST(%s AS %s) END"
                        % (gi, gp, gi, ri, textcast))
            msel.append("%s AS %s" % (expr, q(r)))
        msel += [q(h) for (_, h) in cells]
        order_terms = []
        for i in range(len(rows)):
            order_terms.append(q("_g%d" % i))
            order_terms.append(q("_r%d" % i))
        return "SELECT %s FROM (%s) AS _piv ORDER BY %s" % (
            ", ".join(msel), inner, ", ".join(order_terms))

    def _flow_brace_misuse(self, graph, node_id):
        """Up-front {{name}} misuse check for a run/preview of (node_id). Scans
        the target node and everything feeding it transitively -- so a stray,
        half-built node elsewhere in the canvas can't block this run -- against
        the workflow's variable values. Returns a clear error string, or None
        when every {{name}} is used correctly. See applyvars.brace_misuse:
        {{name}} inserts a quoted SQL string, so a numeric value wants ${name}
        and an API node's URL wants ${name}."""
        from . import applyvars
        try:
            ctx = applyvars.collect_vars(graph)
            rev = {}
            for e in graph.get("edges") or []:
                fr = (e.get("from") or {}).get("node")
                to = (e.get("to") or {}).get("node")
                if fr and to:
                    rev.setdefault(to, []).append(fr)
            scope = {node_id}
            stack = [node_id]
            while stack:
                n = stack.pop()
                for p in rev.get(n, []):
                    if p not in scope:
                        scope.add(p)
                        stack.append(p)
            sub = [n for n in (graph.get("nodes") or [])
                   if isinstance(n, dict) and n.get("id") in scope]
            return applyvars.brace_misuse({"nodes": sub}, ctx)
        except Exception:
            return None

    def nodeflow_lineage(self, graph):
        """Build the column-level data-flow lineage for a node graph and write
        it to a two-tab .xlsx (Pass-through / Derived). Returns
        ``{ok, path, ...}`` or ``{ok: False, error}``."""
        from . import lineage
        return lineage.export_lineage_xlsx(self, graph)

    def _flow_exc(self, e, graph=None):
        """Map a flow-op exception to its error reply: a NodeflowError
        becomes the friendly flow message, an engine interrupt becomes
        the cancelled marker, and anything else becomes "TypeName:
        message". This is the two-except body that recurred verbatim in
        every flow handler -- centralising it also guarantees a cancel
        is never mis-reported as an error."""
        if isinstance(e, NodeflowError):
            return self._flow_err(e, graph)
        if _is_interrupt(e):
            return {"error": "cancelled", "cancelled": True}
        return {"error": _duckdb_oom_flow_message(
            err_str(e), getattr(self, "duckdb", None))}

    def _flow_err(self, e, graph=None):
        """Error envelope for a flow failure. A NodeRunError also carries the
        id of the node that failed, so the UI can highlight the exact culprit
        instead of the run target. When the graph is supplied, a bare
        'missing accumulator table' is rewritten to the 'run the iterator
        first' hint (that case has a clearer, more specific explanation than
        the generic node-failed wrapper)."""
        msg = str(e)
        if graph is not None:
            msg = _iter_missing_accum_hint(msg, graph)
        # A NodeRunError already carries the raw engine text (e.g. a DuckDB
        # OutOfMemoryException from a node's CREATE ... AS): enrich it with the
        # same failed-alloc/limit + shred/flatten/filter-first guidance the
        # IDE/Journal query path surfaces, so a large-data flow OOM is
        # actionable rather than a bare allocation error.
        msg = _duckdb_oom_flow_message(msg, getattr(self, "duckdb", None))
        out = {"error": msg}
        nid = getattr(e, "node_id", None)
        if nid:
            out["node"] = nid
        return out

    def run_nodeflow(self, graph, node_id, port, query_id=None,
                     preview=False, preview_limit=None):
        """Materialise a NodeFlow graph up to (node_id, port) into temp tables,
        then return a normal query-result envelope so the UI previews it in the
        grid (paging / charts / save-as-table all work). Cancellable via
        ``query_id`` (cancel_query interrupts the running build statement)."""
        from . import nodeflow
        _bad_brace = self._flow_brace_misuse(graph, node_id)
        if _bad_brace:
            return {"error": _bad_brace}
        created = []
        et = LOCAL_TARGET
        try:
            if preview:
                try:
                    preview_limit = max(1, min(int(
                        preview_limit or os.environ.get(
                            "SAMQL_PREVIEW_ROWS", DISPLAY_LIMIT)), 100000))
                except (TypeError, ValueError):
                    preview_limit = DISPLAY_LIMIT
            else:
                preview_limit = None
            et = self._flow_engine_target(graph)
            eng, _kind = self._engine_obj(et)
            self._register_run(
                query_id, eng, surface="node",
                label=self._flow_node_label(graph, node_id))
            tmp = self._materialize_flow(
                graph, node_id, port, et, created,
                row_limit=preview_limit)
        except nodeflow.NodeflowError as e:
            self._flow_cleanup(query_id, et, created)
            return self._flow_err(e, graph)
        except Exception as e:
            self._flow_cleanup(query_id, et, created)
            if _is_interrupt(e):
                return {"error": "cancelled", "cancelled": True}
            return {"error": _duckdb_oom_flow_message(
                _iter_missing_accum_hint(err_str(e), graph),
                getattr(self, "duckdb", None))}
        res = self.run_query('SELECT * FROM "%s"' % tmp, target=et,
                             query_id=query_id)
        if preview_limit and not res.get("error"):
            res.update({"preview": True,
                        "preview_limit": preview_limit,
                        "preview_limited":
                            int(res.get("total_rows") or 0) >= preview_limit})
        self._unregister_run(query_id)
        # the result is fully cached now, so the intermediates can go
        self._drop_flow_temps(et, created)
        return res

    def run_nodeflows(self, graph, requests, query_id=None,
                      preview=False, preview_limit=None):
        """Run several NodeFlow terminals in one materialisation pass.

        Independent DuckDB branches are scheduled on separate child
        connections; targets sharing upstream nodes stay in one group so the
        shared work is still computed once.  Each result is returned in the
        ordinary pageable query envelope, tagged with its node and port.
        """
        from . import nodeflow
        norm = []
        seen = set()
        for item in list(requests or [])[:64]:
            if not isinstance(item, dict):
                continue
            nid = item.get("node") or item.get("node_id")
            port = item.get("port") or "out"
            if not nid or (nid, port) in seen:
                continue
            seen.add((nid, port))
            norm.append((nid, port))
        if not norm:
            return {"error": "No NodeFlow targets were supplied."}
        for nid, _port in norm:
            bad = self._flow_brace_misuse(graph, nid)
            if bad:
                return {"error": bad, "node": nid}
        if preview:
            try:
                preview_limit = max(1, min(int(
                    preview_limit or os.environ.get(
                        "SAMQL_PREVIEW_ROWS", DISPLAY_LIMIT)), 100000))
            except (TypeError, ValueError):
                preview_limit = DISPLAY_LIMIT
        else:
            preview_limit = None
        limits = ({target: preview_limit for target in norm}
                  if preview_limit else None)
        created = []
        et = LOCAL_TARGET
        try:
            et = self._flow_engine_target(graph)
            eng, _kind = self._engine_obj(et)
            self._register_run(
                query_id, eng, surface="node",
                label="NodeFlow batch (%d branches)" % len(norm))
            built = self._materialize_flows(
                graph, norm, et, created, target_limits=limits)
            results = []
            for nid, port in norm:
                if self._run_is_cancelled(query_id):
                    return {"error": "cancelled", "cancelled": True,
                            "results": results}
                res = self.run_query(
                    'SELECT * FROM "%s"' % built[(nid, port)],
                    target=et, query_id=query_id)
                res["node"] = nid
                res["port"] = port
                if preview_limit and not res.get("error"):
                    res.update({
                        "preview": True,
                        "preview_limit": preview_limit,
                        "preview_limited":
                            int(res.get("total_rows") or 0) >= preview_limit,
                    })
                results.append(res)
                if res.get("cancelled"):
                    return {"error": "cancelled", "cancelled": True,
                            "results": results}
            return {"ok": True, "results": results}
        except nodeflow.NodeflowError as e:
            return self._flow_err(e, graph)
        except Exception as e:
            if _is_interrupt(e):
                return {"error": "cancelled", "cancelled": True}
            return {"error": _duckdb_oom_flow_message(
                _iter_missing_accum_hint(err_str(e), graph),
                getattr(self, "duckdb", None))}
        finally:
            self._flow_cleanup(query_id, et, created)

    def _stream_relation_to_file(self, eng, sql, path, fmt, batch=10000):
        """Stream the rows of ``sql`` straight to ``path`` in ``fmt`` (csv /
        tsv / json / xlsx), holding only one batch in memory at a time. Returns
        the number of data rows written. Raises on error (the caller is
        responsible for cleanup); the cursor is always closed."""
        import csv as _csv
        import json as _json
        n = 0
        cursor = None
        # A DuckDB cursor is a separate connection that can't see this
        # connection's TEMP tables, and the export reads a temp table the flow
        # just materialised. Stream on the main connection (same_conn) while
        # holding the engine lock, so the read sees the temp table and nothing
        # else touches the connection mid-stream. SQLite cursors share the
        # connection (temp tables already visible), so they need no special
        # handling.
        same = getattr(eng, "ENGINE_KIND", None) == "duckdb"
        # .472: on DuckDB, csv / tsv / json go out as ONE engine COPY
        # (the relation is a temp table, so it runs on the main
        # connection under the lock) instead of batching rows through
        # Python -- the NodeFlow Export path measured in minutes.
        if same and fmt in ("csv", "tsv", "json", "ndjson"):
            opts = {"csv": "FORMAT CSV, HEADER",
                    "tsv": "FORMAT CSV, HEADER, DELIMITER '\t'",
                    "json": "FORMAT JSON, ARRAY true",
                    "ndjson": "FORMAT JSON"}[fmt]
            with eng.write_lock:
                row = eng.conn.execute(
                    "COPY (%s) TO '%s' (%s)"
                    % (sql, sqlutil.sql_path(path), opts)).fetchone()
            return int(row[0]) if row else 0
        lock = getattr(eng, "write_lock", None) if same else None
        if lock is not None:
            lock.acquire()
        try:
            cols, first, cursor = (
                eng.execute_cursor(sql, batch=batch, same_conn=True) if same
                else eng.execute_cursor(sql, batch=batch))
            cols = list(cols or [])

            def _batches():
                yield first or []
                if cursor is not None:
                    while True:
                        b = cursor.fetchmany(batch)
                        if not b:
                            break
                        opreg.beat()
                        yield b

            if fmt in ("csv", "tsv"):
                delim = "\t" if fmt == "tsv" else ","
                with open(path, "w", newline="", encoding="utf-8-sig") as f:
                    w = _csv.writer(f, delimiter=delim)
                    w.writerow(cols)
                    for b in _batches():
                        for r in b:
                            w.writerow(["" if v is None else v for v in r])
                            n += 1
            elif fmt == "xlsx":
                try:
                    import openpyxl
                except Exception:
                    raise RuntimeError(
                        "Exporting .xlsx needs the 'openpyxl' package "
                        "(pip install openpyxl). CSV / TSV / JSON exports "
                        "work without it.")
                wb = openpyxl.Workbook(write_only=True)
                ws = wb.create_sheet("data")
                ws.append(cols)
                # .522 audit: same guard as the .451 result exporter -- a
                # string starting with "=" must stay TEXT, never become a
                # live Excel formula (spreadsheet injection).
                from openpyxl.cell.cell import WriteOnlyCell
                for b in _batches():
                    for r in b:
                        out = []
                        for v in r:
                            c = WriteOnlyCell(
                                ws, value="" if v is None else v)
                            if getattr(c, "data_type", None) == "f":
                                c.data_type = "s"
                            out.append(c)
                        ws.append(out)
                        n += 1
                wb.save(path)
            elif fmt == "parquet":
                # Parquet needs the whole column to infer a stable type, so
                # this branch collects rather than streams one batch at a time.
                # The engine gives each column a single type, so inference is
                # clean. pyarrow is optional; a clear error is raised if absent.
                try:
                    import pyarrow as pa
                    import pyarrow.parquet as pq
                except Exception:
                    raise RuntimeError(
                        "Parquet export requires pyarrow (pip install pyarrow)")
                data = {c: [] for c in cols}
                for b in _batches():
                    for r in b:
                        for i, c in enumerate(cols):
                            data[c].append(
                                json_safe(r[i]) if i < len(r) else None)
                        n += 1
                pq.write_table(pa.table(data), path)
            elif fmt == "ndjson":
                with open(path, "w", encoding="utf-8") as f:
                    for b in _batches():
                        for r in b:
                            f.write(_json.dumps(
                                {c: v for c, v in zip(cols, r)},
                                default=str))
                            f.write("\n")
                            n += 1
            else:  # json
                with open(path, "w", encoding="utf-8") as f:
                    f.write("[\n")
                    firstrow = True
                    for b in _batches():
                        for r in b:
                            obj = {c: v for c, v in zip(cols, r)}
                            f.write(("" if firstrow else ",\n")
                                    + _json.dumps(obj, default=str))
                            firstrow = False
                            n += 1
                    f.write("\n]\n")
        finally:
            # for the DuckDB same-conn path the "cursor" IS the engine
            # connection -- never close it; just drop the lock.
            if cursor is not None and not same:
                try:
                    cursor.close()
                except Exception:
                    pass
            if lock is not None:
                lock.release()
        return n

    def _resolve_export_dir(self, out_dir):
        """Empty out_dir → user's Downloads (shared ``tmputil.downloads_dir``).
        A non-empty path must already exist as a directory."""
        import os
        from . import tmputil
        out = (out_dir or "").strip()
        if out:
            return out if os.path.isdir(out) else None
        return tmputil.downloads_dir()

    def export_nodeflow(self, graph, node_id, out_dir, fmt="csv",
                        base_name=None, query_id=None):
        """Run an output node's upstream flow and write it to out_dir as CSV or
        JSON. Empty out_dir writes to the user's Downloads folder.
        Returns {ok, path, file, rows} or {error}."""
        from . import nodeflow
        import os
        import csv as _csv
        import json as _json
        import re as _re
        out = self._resolve_export_dir(out_dir)
        if not out:
            return {"error": "Pick an output folder."}
        fmt = (fmt or "csv").lower()
        if fmt not in ("csv", "tsv", "json", "ndjson", "xlsx", "parquet"):
            return {"error": "Output format must be CSV, TSV, JSON, NDJSON, "
                    "XLSX or Parquet."}
        _bad_brace = self._flow_brace_misuse(graph, node_id)
        if _bad_brace:
            return {"error": _bad_brace}
        created = []
        et = LOCAL_TARGET
        try:
            et = self._flow_engine_target(graph)
            eng, _kind = self._engine_obj(et)
            self._register_run(
                query_id, eng, kind="export", surface="node",
                label=self._flow_node_label(graph, node_id))
            tmp = self._materialize_flow(graph, node_id, "out", et, created)
        except Exception as e:
            self._flow_cleanup(query_id, et, created)
            return self._flow_exc(e)
        bn = _re.sub(r"[^A-Za-z0-9_-]+", "_",
                     (base_name or "output")).strip("_") or "output"
        path = os.path.join(out, bn + "." + fmt)
        try:
            n = self._stream_relation_to_file(
                eng, 'SELECT * FROM "%s"' % tmp, path, fmt)
        except Exception as e:
            self._flow_cleanup(query_id, et, created)
            if _is_interrupt(e):
                return {"error": "cancelled", "cancelled": True}
            return {"error": err_str(e)}
        self._flow_cleanup(query_id, et, created)
        return {"ok": True, "path": path, "file": bn + "." + fmt, "rows": n}

    def export_nodeflow_many(self, graph, items, query_id=None):
        """Export several output nodes in ONE shared materialisation pass, so
        any upstream subgraph feeding more than one of them is computed once
        for the whole batch instead of per output. ``items`` is a list of
        ``{node_id, folder, fmt?, base_name?}`` -- each writes to its own
        folder/format. Returns ``{ok, results: [...]}`` where each result is
        ``{node_id, ok, path, file, rows}`` or ``{node_id, error}``; or a
        top-level ``{error}`` if the shared build itself fails (the caller can
        then fall back to exporting each node on its own so the failure
        localises to one node)."""
        from . import nodeflow
        import os
        import re as _re
        items = list(items or [])
        if not items:
            return {"error": "No outputs to export."}
        norm = []
        seen = {}
        for it in items:
            nid = it.get("node_id")
            folder = self._resolve_export_dir(it.get("folder") or "")
            if not folder:
                return {"error": "Pick an output folder for every output node."}
            fmt = (it.get("fmt") or "csv").lower()
            if fmt not in ("csv", "tsv", "json", "ndjson", "xlsx",
                           "parquet"):
                return {"error":
                        "Output format must be CSV, TSV, JSON, NDJSON, "
                        "XLSX or Parquet."}
            _bad_brace = self._flow_brace_misuse(graph, nid)
            if _bad_brace:
                return {"error": _bad_brace}
            bn = _re.sub(r"[^A-Za-z0-9_-]+", "_",
                         (it.get("base_name") or "output")).strip("_") \
                or "output"
            # de-collide files written into the same folder in this batch, so
            # two outputs sharing a name don't silently clobber each other.
            fk = (folder.lower(), (bn + "." + fmt).lower())
            if fk in seen:
                seen[fk] += 1
                bn = "%s_%d" % (bn, seen[fk])
            else:
                seen[fk] = 1
            norm.append({"node_id": nid, "folder": folder, "fmt": fmt,
                         "base": bn})
        created = []
        et = LOCAL_TARGET
        try:
            et = self._flow_engine_target(graph)
            eng, _kind = self._engine_obj(et)
            self._register_run(
                query_id, eng, surface="node",
                label="export (%d)" % len(items or []))
            targets = [(it["node_id"], "out") for it in norm]
            built = self._materialize_flows(graph, targets, et, created)
        except Exception as e:
            self._flow_cleanup(query_id, et, created)
            return self._flow_exc(e)
        # the shared upstream is built exactly once; write each output from its
        # own materialised temp. One file failing doesn't abort the rest.
        results = []
        for it in norm:
            tmp = built.get((it["node_id"], "out"))
            fname = it["base"] + "." + it["fmt"]
            path = os.path.join(it["folder"], fname)
            try:
                n = self._stream_relation_to_file(
                    eng, 'SELECT * FROM "%s"' % tmp, path, it["fmt"])
                results.append({"node_id": it["node_id"], "ok": True,
                                "path": path, "file": fname, "rows": n})
            except Exception as e:
                if _is_interrupt(e):
                    self._flow_cleanup(query_id, et, created)
                    return {"error": "cancelled", "cancelled": True}
                results.append({"node_id": it["node_id"],
                                "error": err_str(e)})
        self._flow_cleanup(query_id, et, created)
        return {"ok": True, "results": results}

    def _write_into_table(self, eng, name, src, mode, replace_keys):
        """Write rows from table ``src`` into ``name``. mode 'overwrite' (or a
        target that doesn't exist yet) drops + recreates; 'append' inserts,
        aligning by column NAME -- a column the source lacks is filled with NULL
        and an extra source column is dropped. With replace_keys the append
        first deletes rows whose key tuple appears in src, so a re-run replaces
        rather than duplicates. Raises NodeflowError on an unknown key column.
        Shared by the write node and the iterator's accumulator."""
        from . import nodeflow

        def _exists(t):
            try:
                eng.execute('SELECT * FROM "%s" LIMIT 0' % t)
                return True
            except Exception:
                return False

        if mode == "overwrite" or not _exists(name):
            eng.execute('DROP TABLE IF EXISTS "%s"' % name)
            eng.execute('CREATE TABLE "%s" AS SELECT * FROM "%s"' % (name, src))
            return
        tcols = self._col_names(eng, table=name)
        sset = set(self._col_names(eng, table=src))
        bad = [k for k in (replace_keys or []) if k not in tcols]
        if bad:
            raise nodeflow.NodeflowError(
                "Replace-key column(s) not found in the table: "
                + ", ".join(bad))
        if replace_keys:
            kp = ", ".join('"%s"' % k for k in replace_keys)
            eng.execute('DELETE FROM "%s" WHERE (%s) IN '
                        '(SELECT DISTINCT %s FROM "%s")' % (name, kp, kp, src))
        collist = ", ".join('"%s"' % c for c in tcols)
        sellist = ", ".join(
            ('"%s"' % c) if c in sset else "NULL" for c in tcols)
        eng.execute('INSERT INTO "%s" (%s) SELECT %s FROM "%s"'
                    % (name, collist, sellist, src))

    def _reduce_accumulator(self, eng, name, keys, aggs):
        """Collapse the accumulator to one row per key, folding each measure
        with a distributive aggregation (sum/min/max/count). Re-run after each
        pass so the table never grows past the number of distinct keys -- the
        point of a chunked map-reduce. Only distributive aggregations are
        offered: averages must be folded as sum/count, and distinct-counts
        cannot be folded across chunks at all."""
        from . import nodeflow
        allowed = {"sum", "min", "max", "count"}
        keys = [str(k).strip() for k in (keys or []) if str(k).strip()]
        if not keys:
            raise nodeflow.NodeflowError(
                "Reduce mode needs at least one key column.")
        measures = []
        for a in (aggs or []):
            col = (a.get("col") or "").strip()
            fn = (a.get("fn") or "sum").strip().lower()
            if not col:
                continue
            if fn not in allowed:
                raise nodeflow.NodeflowError(
                    "Reduce uses only sum/min/max/count (got %r); fold "
                    "averages as sum and count yourself." % fn)
            measures.append((col, fn))
        if not measures:
            raise nodeflow.NodeflowError(
                "Reduce mode needs at least one measure to aggregate.")
        tcols = self._col_names(eng, table=name)
        tset = {str(c).lower() for c in tcols}
        bad = [c for c in keys + [m[0] for m in measures]
               if c.lower() not in tset]
        if bad:
            raise nodeflow.NodeflowError(
                "Reduce column(s) not found in the accumulator: "
                + ", ".join(bad))
        ksel = ", ".join('"%s"' % k for k in keys)
        msel = ", ".join('%s("%s") AS "%s"' % (fn.upper(), col, col)
                         for col, fn in measures)
        tmp = name + "__reduce_tmp"
        eng.execute('DROP TABLE IF EXISTS "%s"' % tmp)
        eng.execute('CREATE TABLE "%s" AS SELECT %s, %s FROM "%s" GROUP BY %s'
                    % (tmp, ksel, msel, name, ksel))
        eng.execute('DROP TABLE "%s"' % name)
        eng.execute('ALTER TABLE "%s" RENAME TO "%s"' % (tmp, name))

    def run_nodeflow_to_table(self, graph, node_id, name, query_id=None):
        """Run a write node's upstream flow and store the result as a loaded
        table (in the flow's own engine) so it shows in the tables list and can
        be queried / joined. Overwrites a same-named table. Cancellable.
        Returns {ok, table, engine, rows, all} or {error}."""
        from . import nodeflow, applyvars
        import re as _re
        _bad_brace = self._flow_brace_misuse(graph, node_id)
        if _bad_brace:
            return {"error": _bad_brace}
        # the target name may itself use ${vars} (e.g. tx_${region})
        try:
            name = applyvars.substitute_text(
                name or "", applyvars.collect_vars(graph))
        except Exception:
            pass
        # the write node's own settings: overwrite (default) vs append, and an
        # optional set of key columns for an idempotent (delete-then-insert)
        # append so re-runs don't duplicate rows
        wnode = None
        for _n in (graph.get("nodes") or []):
            if isinstance(_n, dict) and _n.get("id") == node_id:
                wnode = _n
                break
        wcfg = (wnode or {}).get("config") or {}
        write_mode = (wcfg.get("mode") or "overwrite").strip().lower()
        if write_mode not in ("overwrite", "append"):
            write_mode = "overwrite"
        replace_keys = [str(k).strip()
                        for k in (wcfg.get("replace_keys") or [])
                        if str(k).strip()]
        nm = sanitize_ident(name, "flow_result")
        created = []
        et = LOCAL_TARGET
        try:
            et = self._flow_engine_target(graph)
            eng, _kind = self._engine_obj(et)
            self._register_run(
                query_id, eng, surface="node",
                label=self._flow_node_label(graph, node_id))
            tmp = self._materialize_flow(graph, node_id, "out", et, created)
        except Exception as e:
            self._flow_cleanup(query_id, et, created)
            return self._flow_exc(e)
        try:
            self._write_into_table(eng, nm, tmp, write_mode, replace_keys)
            eng.sync_catalog()
            try:
                eng.table_sources[nm] = "nodeflow"
            except Exception:
                pass
            _c, _r = eng.execute('SELECT COUNT(*) FROM "%s"' % nm)
            n = int(_r[0][0]) if _r else 0
        except nodeflow.NodeflowError as e:
            self._flow_cleanup(query_id, et, created)
            return self._flow_err(e)
        except Exception as e:
            self._flow_cleanup(query_id, et, created)
            if _is_interrupt(e):
                return {"error": "cancelled", "cancelled": True}
            return {"error": "Couldn't write the table: %s: %s"
                    % (type(e).__name__, e)}
        self._flow_cleanup(query_id, et, created)
        self._invalidate_profiles()
        self._invalidate_counts()
        self._prime_count("duckdb" if et == DUCKDB_TARGET else "sqlite",
                          nm, n)
        return {"ok": True, "table": nm, "mode": write_mode,
                "engine": "duckdb" if et == DUCKDB_TARGET else "sqlite",
                "rows": n, "all": self.tables_tree()}

    # ---- iterator ---------------------------------------------------
    def _node_config(self, graph, node_id):
        for n in (graph.get("nodes") or []):
            if isinstance(n, dict) and n.get("id") == node_id:
                return n.get("config") or {}
        return {}

    def _set_node_config(self, graph, node_id, patch):
        for n in (graph.get("nodes") or []):
            if isinstance(n, dict) and n.get("id") == node_id:
                cfg = dict(n.get("config") or {})
                cfg.update(patch or {})
                n["config"] = cfg
                return

    def _api_nodes_upstream(self, graph, start_node):
        """Every apinode that is an ancestor of (or is) ``start_node`` -- the
        API nodes the iterator must refresh before running each pass."""
        edges = graph.get("edges") or []
        nodes = {n.get("id"): n for n in (graph.get("nodes") or [])
                 if isinstance(n, dict)}
        seen, stack, api = set(), [start_node], []
        while stack:
            nid = stack.pop()
            if nid is None or nid in seen:
                continue
            seen.add(nid)
            n = nodes.get(nid)
            if n and n.get("type") == "apinode":
                api.append(nid)
            for e in edges:
                if (e.get("to") or {}).get("node") == nid:
                    frm = (e.get("from") or {}).get("node")
                    if frm:
                        stack.append(frm)
        return api

    def _volatile_source_nodes_upstream(self, graph, start_nodes):
        """Upstream (and self) source nodes that must be fetched before a
        materialize -- SQL Server / SharePoint / Web scrape. API nodes are
        handled separately by iterators; here we cover Dashboard / Run all
        so a saved profile + DPAPI password is enough without a prior Fetch.
        """
        want = {"sqlserver", "sharepoint", "webscrape"}
        edges = graph.get("edges") or []
        nodes = {n.get("id"): n for n in (graph.get("nodes") or [])
                 if isinstance(n, dict)}

        def _collect_nested(n, out_list):
            if not isinstance(n, dict):
                return
            if n.get("type") in want:
                out_list.append(n)
            if n.get("type") in ("group", "iterator"):
                for ch in (n.get("config") or {}).get("children") or []:
                    _collect_nested(ch, out_list)

        seen, stack, out = set(), list(start_nodes or []), []
        while stack:
            nid = stack.pop()
            if nid is None or nid in seen:
                continue
            seen.add(nid)
            n = nodes.get(nid)
            if n:
                _collect_nested(n, out)
            for e in edges:
                if (e.get("to") or {}).get("node") == nid:
                    frm = (e.get("from") or {}).get("node")
                    if frm:
                        stack.append(frm)
        # Stable unique by id (keep first).
        seen_ids, uniq = set(), []
        for n in out:
            nid = n.get("id")
            if nid is None or nid in seen_ids:
                continue
            seen_ids.add(nid)
            uniq.append(n)
        return uniq

    def _ensure_source_nodes_fetched(self, graph, start_nodes, query_id=None):
        """Fetch upstream SQL Server / SharePoint / Web scrape nodes in place
        so ``config.table`` is set before compile. Mutates ``graph`` node
        configs. Raises ``nodeflow.NodeflowError`` on failure."""
        from . import nodeflow
        sources = self._volatile_source_nodes_upstream(graph, start_nodes)
        if not sources:
            return
        # Index every node (including nested children) for config patches.
        by_id = {}

        def _index(nodes):
            for n in nodes or []:
                if not isinstance(n, dict):
                    continue
                nid = n.get("id")
                if nid is not None:
                    by_id[nid] = n
                cfg = n.get("config") or {}
                if n.get("type") in ("group", "iterator"):
                    _index(cfg.get("children"))

        _index(graph.get("nodes"))
        for n in sources:
            nid = n.get("id")
            typ = n.get("type")
            cfg = dict(n.get("config") or {})
            label = (cfg.get("label") or typ or "source")
            fr = self.fetch_source_node(
                typ, nid, cfg, graph=graph, query_id=query_id)
            if fr.get("cancelled"):
                raise nodeflow.NodeflowError("cancelled")
            if fr.get("error") or not fr.get("ok"):
                raise nodeflow.NodeflowError(
                    'Could not fetch "%s" (%s): %s' % (
                        label, typ, fr.get("error") or "unknown error"))
            live = by_id.get(nid) or n
            live_cfg = dict(live.get("config") or {})
            live_cfg["table"] = fr.get("table") or ""
            if fr.get("engine"):
                live_cfg["engine"] = fr.get("engine")
            if fr.get("columns") is not None:
                live_cfg["columns"] = fr.get("columns")
            if fr.get("rows") is not None:
                live_cfg["rows"] = fr.get("rows")
            live["config"] = live_cfg

    def _daterange_values(self, driver):
        import datetime as _dt
        from . import nodeflow
        start = (driver.get("start") or "").strip()
        end = (driver.get("end") or "").strip()
        step = (driver.get("step") or "day").strip().lower()
        if not start or not end:
            return []
        try:
            d0 = _dt.date.fromisoformat(start)
            d1 = _dt.date.fromisoformat(end)
        except ValueError:
            raise nodeflow.NodeflowError(
                "The date range needs ISO dates (YYYY-MM-DD).")
        out, d = [], d0
        for _ in range(100000):     # guard against a runaway range
            if d > d1:
                break
            out.append(d.isoformat())
            if step == "week":
                d = d + _dt.timedelta(days=7)
            elif step == "month":
                import calendar as _cal
                y, m = d.year, d.month + 1
                if m > 12:
                    y, m = y + 1, 1
                d = _dt.date(y, m, min(d.day, _cal.monthrange(y, m)[1]))
            else:
                d = d + _dt.timedelta(days=1)
        return out

    def _distinct_column_values(self, table, column):
        from . import nodeflow
        table = (table or "").strip()
        column = (column or "").strip()
        if not table or not column:
            raise nodeflow.NodeflowError(
                "The column driver needs a table and a column.")
        eng_name = self._engine_of_table(table)
        if not eng_name:
            raise nodeflow.NodeflowError("Table not found: %s" % table)
        e = self.duckdb if eng_name == "duckdb" else self.db
        _c, rows = e.execute(
            'SELECT DISTINCT "%s" FROM "%s" WHERE "%s" IS NOT NULL '
            'ORDER BY 1' % (column, table, column))
        return [r[0] for r in (rows or [])]

    def _row_values(self, driver, max_passes):
        """Rows of a table as per-pass variable bindings -- one dict per row,
        keyed by each selected column's name (or its explicit alias). Powers the
        iterator's 'rows' driver, so a single pass can bind several ${vars} at
        once (one pass per row). Reads at most max_passes+1 rows so the cap is
        detectable without pulling a huge table into memory."""
        from . import nodeflow
        driver = driver or {}
        table = (driver.get("table") or "").strip()
        if not table:
            raise nodeflow.NodeflowError("The rows driver needs a table.")
        pairs = []
        for c in (driver.get("columns") or []):
            if isinstance(c, dict):
                col = (c.get("col") or "").strip()
                alias = (c.get("var") or col).strip()
            else:
                col = str(c).strip()
                alias = col
            if col:
                pairs.append((col, alias))
        if not pairs:
            raise nodeflow.NodeflowError(
                "The rows driver needs at least one column to bind.")
        eng_name = self._engine_of_table(table)
        if not eng_name:
            raise nodeflow.NodeflowError("Table not found: %s" % table)
        e = self.duckdb if eng_name == "duckdb" else self.db
        try:
            lim = max(1, min(int(max_passes or 1000), 100000))
        except (TypeError, ValueError):
            lim = 1000
        collist = ", ".join('"%s"' % col for col, _ in pairs)
        distinct = "DISTINCT " if driver.get("distinct") else ""
        sql = 'SELECT %s%s FROM "%s"' % (distinct, collist, table)
        order_by = (driver.get("order_by") or "").strip()
        if order_by in {col for col, _ in pairs}:
            sql += ' ORDER BY "%s"' % order_by
        sql += " LIMIT %d" % (lim + 1)
        try:
            _c, rows = e.execute(sql)
        except Exception as ex:
            raise nodeflow.NodeflowError(
                "Couldn't read the rows driver from %s (%s)." % (table, ex))
        return [{alias: (r[i] if i < len(r) else None)
                 for i, (_, alias) in enumerate(pairs)}
                for r in (rows or [])]

    def _iterator_values(self, driver, max_passes):
        """The list of values an iterator loops over, from its driver config.
        Returns (values, note); note is set when the list was capped."""
        from . import nodeflow
        driver = driver or {}
        kind = (driver.get("kind") or "list").strip().lower()
        if kind == "list":
            vals = [str(v).strip() for v in (driver.get("values") or [])
                    if str(v).strip() != ""]
        elif kind == "daterange":
            vals = self._daterange_values(driver)
        elif kind == "column":
            vals = self._distinct_column_values(
                driver.get("table"), driver.get("column"))
        elif kind == "rows":
            vals = self._row_values(driver, max_passes)
        else:
            raise nodeflow.NodeflowError("Unknown iterator driver: %r" % kind)
        note = None
        if max_passes and len(vals) > max_passes:
            note = ("Stopped at the %d-pass cap (%d values were available)."
                    % (max_passes, len(vals)))
            vals = vals[:max_passes]
        return vals, note

    def run_iterator(self, graph, node_id, query_id=None):
        """Run an iterator node: loop a driver's values and, for each one,
        re-resolve the graph with the loop variable set to that value, refresh
        any API nodes in the body, run the body, and append the result into a
        single accumulator table. The graph stays acyclic -- the body feeds the
        iterator's input and the loop happens here. Returns
        {ok, passes, attempted, rows, table, engine, errors, note, all} or
        {error}."""
        from . import nodeflow, applyvars
        import re as _re
        cfg = self._node_config(graph, node_id)
        # NEW container model: if the iterator holds inner children (like a
        # group), run them once per row of the wired "variables" input, with
        # that row's columns bound as scalar ${vars}. The classic driver +
        # upstream-body iterator (no children) falls through to the code below,
        # unchanged, so existing iterators keep working.
        if [c for c in (cfg.get("children") or [])
                if isinstance(c, dict) and c.get("type")]:
            return self._run_iterator_container(graph, node_id, cfg, query_id)
        var = (cfg.get("var") or "").strip()
        driver_kind = ((cfg.get("driver") or {}).get("kind")
                       or "list").strip().lower()
        # A table wired to the iterator's top "vars" input drives the loop: one
        # pass per row, the row's columns bound as scalar ${vars}. It overrides
        # the configured driver; the classic driver is the fallback when the
        # top input is not wired.
        vn, vp = nodeflow.upstream(graph, node_id, "vars")
        if vn is None and not var and driver_kind != "rows":
            return {"error": "Wire a values table to the iterator's top input, "
                    "or name the loop variable (e.g. as_of) and pick a driver."}
        bn, bp = nodeflow.upstream(graph, node_id, "in")
        if bn is None:
            return {"error": "Connect the body (the flow to repeat) to the "
                    "iterator's input."}
        accum = sanitize_ident(cfg.get("table"))
        if not accum:
            return {"error": "Give the iterator's output table a name."}
        replace_keys = [str(k).strip() for k in (cfg.get("replace_keys") or [])
                        if str(k).strip()]
        try:
            max_passes = int(cfg.get("max_passes") or 1000)
        except (TypeError, ValueError):
            max_passes = 1000
        max_passes = max(1, min(max_passes, 100000))
        continue_on_error = bool(cfg.get("continue_on_error"))
        reset_first = cfg.get("reset_first")
        reset_first = True if reset_first is None else bool(reset_first)
        accumulate = (cfg.get("accumulate") or "append").strip().lower()
        reduce_keys = [str(k).strip() for k in (cfg.get("reduce_keys") or [])
                       if str(k).strip()]
        reduce_aggs = cfg.get("reduce_aggs") or []

        try:
            if vn is not None:
                values = self._materialize_var_rows(
                    graph, vn, vp, distinct=bool(cfg.get("vars_distinct")),
                    cap=max_passes)
                note = None
                if not values:
                    return {"error": "The values input produced no rows to "
                            "iterate over."}
            else:
                values, note = self._iterator_values(
                    cfg.get("driver"), max_passes)
        except nodeflow.NodeflowError as e:
            return {"error": str(e)}
        except Exception as e:
            return {"error": err_str(e)}
        if not values:
            return {"error": "The driver produced no values to iterate over."}

        api_ids = self._api_nodes_upstream(graph, bn)
        self._register_run(query_id, None)
        errors, passes, last_engine, first = [], 0, None, True
        try:
            for idx, v in enumerate(values, 1):
                if self._run_is_cancelled(query_id):
                    return {"error": "cancelled", "cancelled": True}
                try:
                    opreg.advance(idx, len(values), unit="pass",
                                  op_id=query_id)
                except Exception:
                    pass
                if isinstance(v, dict):
                    # rows driver: bind each selected column as its own ${var};
                    # if a loop variable is also named, bind it to the 1-based
                    # row index (handy for per-pass output names).
                    extra = {str(k): ("" if val is None else str(val))
                             for k, val in v.items()}
                    if var:
                        extra[var] = str(idx)
                else:
                    extra = {var: "" if v is None else str(v)}
                rg = applyvars.resolve_graph(graph, extra=extra)
                created, et = [], None
                try:
                    for aid in api_ids:
                        fr = self.fetch_api_node(
                            aid, self._node_config(rg, aid), graph=rg,
                            should_cancel=(
                                lambda: self._run_is_cancelled(query_id)))
                        if fr.get("error"):
                            raise nodeflow.NodeflowError(
                                "API node fetch failed: " + fr["error"])
                        # propagate BOTH outputs so the body can read the data
                        # (out) or, with continue-on-error, the per-pass error
                        # rows (err) -- enabling an error accumulator.
                        self._set_node_config(rg, aid, {
                            "table": fr.get("table"),
                            "err_table": fr.get("err_table")})
                    et = self._flow_engine_target(rg)
                    eng, _kind = self._engine_obj(et)
                    tmp = self._materialize_flow(rg, bn, bp, et, created)
                    mode = "overwrite" if (first and reset_first) else "append"
                    self._write_into_table(eng, accum, tmp, mode, replace_keys)
                    if accumulate == "reduce":
                        self._reduce_accumulator(
                            eng, accum, reduce_keys, reduce_aggs)
                    eng.sync_catalog()
                    try:
                        eng.table_sources[accum] = "nodeflow"
                    except Exception:
                        pass
                    last_engine = ("duckdb" if et == DUCKDB_TARGET
                                   else "sqlite")
                    passes += 1
                    first = False
                except Exception as e:
                    if _is_interrupt(e):
                        raise
                    msg = (str(e) if isinstance(e, nodeflow.NodeflowError)
                           else err_str(e))
                    errors.append({"value": extra[var], "error": msg})
                    if not continue_on_error:
                        break
                finally:
                    self._drop_flow_temps(et, created)
        except Exception as e:
            self._unregister_run(query_id)
            if _is_interrupt(e):
                return {"error": "cancelled", "cancelled": True}
            return {"error": err_str(e)}
        self._unregister_run(query_id)
        self._invalidate_profiles()
        self._invalidate_counts()
        rows = None
        if last_engine and passes:
            e = self.duckdb if last_engine == "duckdb" else self.db
            try:
                _c, r = e.execute('SELECT COUNT(*) FROM "%s"' % accum)
                rows = int(r[0][0]) if r else None
            except Exception:
                rows = None
        ok = passes > 0 and (continue_on_error or not errors)
        return {"ok": ok, "passes": passes, "attempted": len(values),
                "rows": rows, "table": accum if passes else None,
                "engine": last_engine, "errors": errors, "note": note,
                "all": self.tables_tree()}

    def _materialize_var_rows(self, graph, node, port, distinct=False,
                              cap=1000):
        """Materialise the iterator's wired "variables" input once and read its
        rows as per-pass scalar bindings -- one dict per row keyed by column
        name. Each row drives one pass with its columns bound as ${vars}. Reads
        at most cap+1 rows so the pass cap is detectable without pulling a huge
        table into memory."""
        rg = applyvars.resolve_graph(graph)
        et = self._flow_engine_target(rg)
        eng, _kind = self._engine_obj(et)
        created = []
        try:
            tmp = self._materialize_flow(rg, node, port, et, created)
            try:
                lim = max(1, min(int(cap or 1000), 100000))
            except (TypeError, ValueError):
                lim = 1000
            d = "DISTINCT " if distinct else ""
            cols, rows = eng.execute(
                'SELECT %s* FROM "%s" LIMIT %d' % (d, tmp, lim + 1))
            out = []
            for r in rows:
                out.append({str(cols[i]): (r[i] if i < len(r) else None)
                            for i in range(len(cols))})
            return out
        finally:
            self._drop_flow_temps(et, created)

    def _fetch_api_children(self, rg, children, query_id=None):
        """Fetch every API node nested in a container's children (recursing
        into nested groups/iterators), setting each one's config.table in place
        so the group compile can read it like an input. ${vars} in the URL,
        params and headers were already substituted by resolve_graph, so each
        pass fetches that row's endpoint. Raises NodeflowError on a failure."""
        from . import nodeflow
        for ch in (children or []):
            if not isinstance(ch, dict):
                continue
            t = ch.get("type")
            ccfg = ch.get("config") or {}
            if t == "apinode":
                fr = self.fetch_api_node(
                    ch.get("id"), ccfg, graph=rg,
                    should_cancel=lambda: self._run_is_cancelled(query_id))
                if fr.get("error"):
                    raise nodeflow.NodeflowError(
                        "API node fetch failed: " + fr["error"])
                nc = dict(ccfg)
                nc["table"] = fr.get("table")
                nc["err_table"] = fr.get("err_table")
                ch["config"] = nc
            elif t in ("group", "iterator"):
                self._fetch_api_children(rg, ccfg.get("children"), query_id)

    def _run_iterator_container(self, graph, node_id, cfg, query_id=None):
        """Container iterator: the iterator holds inner children (a mini
        pipeline, like a group) that run once per row of the wired "variables"
        input. Each row's columns are bound as scalar ${vars}, surfaced into
        every nested child config (filter / formula / SQL expressions, API URLs,
        file-browser globs, table names) by resolve_graph. An optional side
        "in" input feeds the inner pipeline's primary input. Each pass's output
        is appended (or reduced) into one accumulator table. Returns the same
        shape as run_iterator."""
        from . import nodeflow, applyvars
        import re as _re
        import copy as _copy

        children = [c for c in (cfg.get("children") or [])
                    if isinstance(c, dict) and c.get("type")]
        if not children:
            return {"error": "Add at least one node inside the iterator "
                    "(drag nodes into it)."}
        bindings = cfg.get("bindings") or {}

        accum = sanitize_ident(cfg.get("table"))
        if not accum:
            return {"error": "Give the iterator's output table a name."}
        replace_keys = [str(k).strip() for k in (cfg.get("replace_keys") or [])
                        if str(k).strip()]
        try:
            max_passes = int(cfg.get("max_passes") or 100000)
        except (TypeError, ValueError):
            max_passes = 100000
        max_passes = max(1, min(max_passes, 100000))
        continue_on_error = bool(cfg.get("continue_on_error"))
        reset_first = cfg.get("reset_first")
        reset_first = True if reset_first is None else bool(reset_first)
        accumulate = (cfg.get("accumulate") or "append").strip().lower()
        reduce_keys = [str(k).strip() for k in (cfg.get("reduce_keys") or [])
                       if str(k).strip()]
        reduce_aggs = cfg.get("reduce_aggs") or []

        # --- the variables: rows of the wired "vars" input (each row -> one
        #     pass, its columns bound as scalar ${vars}); fall back to the
        #     classic driver when the top input is not wired. ---
        vn, vp = nodeflow.upstream(graph, node_id, "vars")
        var_rows = []
        if vn is not None:
            try:
                var_rows = self._materialize_var_rows(
                    graph, vn, vp, distinct=bool(cfg.get("vars_distinct")),
                    cap=max_passes)
            except nodeflow.NodeflowError as e:
                return {"error": str(e)}
            except Exception as e:
                return {"error": err_str(e)}
            if not var_rows:
                return {"error": "The variables input produced no rows to "
                        "iterate over."}
        else:
            var = (cfg.get("var") or "").strip()
            dkind = ((cfg.get("driver") or {}).get("kind") or "list")
            if not var and dkind != "rows":
                return {"error": "Wire a variables table to the iterator's top "
                        "input, or name a loop variable and choose a driver."}
            try:
                values, _n = self._iterator_values(cfg.get("driver"),
                                                   max_passes)
            except nodeflow.NodeflowError as e:
                return {"error": str(e)}
            except Exception as e:
                return {"error": err_str(e)}
            if not values:
                return {"error": "The driver produced no values to iterate "
                        "over."}
            for v in values:
                if isinstance(v, dict):
                    var_rows.append({str(k): v[k] for k in v})
                else:
                    var_rows.append({var: v})

        total = len(var_rows)
        note = None
        if max_passes and total > max_passes:
            note = ("Stopped at the %d-pass cap (%d rows were available)."
                    % (max_passes, total))
            var_rows = var_rows[:max_passes]

        # optional side data input -> the inner pipeline's primary input
        sn, sp = nodeflow.upstream(graph, node_id, "in")

        # template: the main graph + a synthetic group carrying the inner
        # children; wire the side input (if any) to the group's primary input.
        # Resolved fresh each pass so ${vars} land in every nested child config.
        base = _copy.deepcopy(graph)
        inner = {"id": "__iter_body", "type": "group",
                 "config": {"children": children, "bindings": bindings}}
        base.setdefault("nodes", []).append(inner)
        if sn is not None:
            base.setdefault("edges", []).append(
                {"from": {"node": sn, "port": sp},
                 "to": {"node": "__iter_body", "port": "in"}})

        # Before looping: catch a name mismatch and say so plainly. Once the
        # first row's columns are bound, any ${name} still left in the body is
        # an unbound variable -- the engine would otherwise fail every pass with
        # the cryptic 'unrecognized token: "$"'. Name the offending variable and
        # list the fields the values row actually provides, so a typo or a
        # column-name mismatch is obvious instead of looking like a sync bug.
        probe_extra = _rename_row_vars(var_rows[0], cfg.get("var_rename"))
        _probe = applyvars.resolve_graph(base, extra=probe_extra)
        _missing = applyvars.unresolved_tokens(
            self._node_config(_probe, "__iter_body").get("children"))
        if _missing:
            _avail = ", ".join(sorted(probe_extra.keys())) or "(none)"
            return {"error":
                    "The loop references ${%s}, but the values row has no such "
                    "field. Available this pass: %s. Make the name match a "
                    "column of the values table exactly -- and for a text "
                    "value use {{%s}} (auto-quoted) rather than a bare ${%s}."
                    % (_missing[0], _avail, _missing[0], _missing[0])}
        # {{name}} quotes the value as a SQL string. With real per-pass values
        # in hand, reject the two misuses up front (instead of silently quoting
        # a number, or putting SQL quotes in an API URL): a numeric {{name}}
        # should be ${name}, and any {{name}} in an API node should be ${name}.
        # The ctx mirrors what the body actually sees -- variable-node values
        # plus this pass's bindings (bindings win) -- so a numeric variable
        # used inside the loop is caught the same way an iterator one is.
        _brace_ctx = applyvars.collect_vars(base)
        _brace_ctx.update(probe_extra)
        _bad_brace = applyvars.brace_misuse(base, _brace_ctx)
        if _bad_brace:
            return {"error": _bad_brace}

        self._register_run(query_id, None)
        errors, passes, last_engine, first = [], 0, None, True
        try:
            for idx, row in enumerate(var_rows, 1):
                if self._run_is_cancelled(query_id):
                    return {"error": "cancelled", "cancelled": True}
                # Determinate progress: pass idx of len(var_rows). The count is
                # known up front, so the Run bar shows a true percentage.
                try:
                    opreg.advance(idx, len(var_rows), unit="pass",
                                  op_id=query_id)
                except Exception:
                    pass
                extra = _rename_row_vars(row, cfg.get("var_rename"))
                rg = applyvars.resolve_graph(base, extra=extra)
                created, et = [], None
                try:
                    inner_cfg = self._node_config(rg, "__iter_body")
                    self._fetch_api_children(rg, inner_cfg.get("children"),
                                             query_id)
                    et = self._flow_engine_target(rg)
                    eng, _kind = self._engine_obj(et)
                    tmp = self._materialize_flow(
                        rg, "__iter_body", "out", et, created)
                    mode = "overwrite" if (first and reset_first) else "append"
                    self._write_into_table(eng, accum, tmp, mode, replace_keys)
                    if accumulate == "reduce":
                        self._reduce_accumulator(
                            eng, accum, reduce_keys, reduce_aggs)
                    eng.sync_catalog()
                    try:
                        eng.table_sources[accum] = "nodeflow"
                    except Exception:
                        pass
                    last_engine = ("duckdb" if et == DUCKDB_TARGET
                                   else "sqlite")
                    passes += 1
                    first = False
                except Exception as e:
                    if _is_interrupt(e):
                        raise
                    msg = (str(e) if isinstance(e, nodeflow.NodeflowError)
                           else err_str(e))
                    msg = _iter_unquoted_var_hint(msg, extra)
                    errors.append({
                        "value": ", ".join("%s=%s" % (k, extra[k])
                                           for k in extra),
                        "error": msg})
                    if not continue_on_error:
                        break
                finally:
                    self._drop_flow_temps(et, created)
        except Exception as e:
            self._unregister_run(query_id)
            if _is_interrupt(e):
                return {"error": "cancelled", "cancelled": True}
            return {"error": err_str(e)}
        self._unregister_run(query_id)
        self._invalidate_profiles()
        self._invalidate_counts()
        rows = None
        if last_engine and passes:
            e = self.duckdb if last_engine == "duckdb" else self.db
            try:
                _c, r = e.execute('SELECT COUNT(*) FROM "%s"' % accum)
                rows = int(r[0][0]) if r else None
            except Exception:
                rows = None
        ok = passes > 0 and (continue_on_error or not errors)
        return {"ok": ok, "passes": passes, "attempted": total,
                "rows": rows, "table": accum if passes else None,
                "engine": last_engine, "errors": errors, "note": note,
                "all": self.tables_tree()}


    def run_while(self, graph, node_id, query_id=None):
        """Run a while/until controller: repeat the wired body until an
        iteration adds no new rows to the accumulator (a fixpoint), or until the
        iteration cap is hit. Each iteration re-resolves the graph (so the body
        can read the accumulator it is building and compute the next batch) and
        appends into a single accumulator table -- the same append / replace-key
        / reduce machinery the iterator uses. The 1-based iteration number is
        bound to the loop variable when one is named. For a clean fixpoint the
        body should return only new rows (e.g. an antijoin against the
        accumulator) or use replace-keys, so re-seen rows don't keep the loop
        alive; a body that reads the accumulator should seed it first and turn
        'reset first' off. Returns {ok, iterations, converged, rows, table,
        engine, note, all} or {error}."""
        from . import nodeflow, applyvars
        import re as _re
        cfg = self._node_config(graph, node_id)
        var = (cfg.get("var") or "").strip()
        bn, bp = nodeflow.upstream(graph, node_id, "in")
        if bn is None:
            return {"error": "Connect the body (the flow to repeat) to the "
                    "controller's input."}
        accum = sanitize_ident(cfg.get("table"))
        if not accum:
            return {"error": "Give the controller's output table a name."}
        replace_keys = [str(k).strip() for k in (cfg.get("replace_keys") or [])
                        if str(k).strip()]
        try:
            max_iters = int(cfg.get("max_iters") or 100)
        except (TypeError, ValueError):
            max_iters = 100
        max_iters = max(1, min(max_iters, 100000))
        reset_first = cfg.get("reset_first")
        reset_first = True if reset_first is None else bool(reset_first)
        accumulate = (cfg.get("accumulate") or "append").strip().lower()
        reduce_keys = [str(k).strip() for k in (cfg.get("reduce_keys") or [])
                       if str(k).strip()]
        reduce_aggs = cfg.get("reduce_aggs") or []

        api_ids = self._api_nodes_upstream(graph, bn)
        self._register_run(query_id, None)
        iterations, last_engine, first, converged = 0, None, True, False
        try:
            for i in range(1, max_iters + 1):
                if self._run_is_cancelled(query_id):
                    return {"error": "cancelled", "cancelled": True}
                # Progress toward the iteration cap. A while-loop may converge
                # early (then the run just finishes), but reporting i of the cap
                # gives a moving, bounded percentage rather than a blank spinner.
                try:
                    opreg.advance(i, max_iters, unit="iteration",
                                  op_id=query_id)
                except Exception:
                    pass
                # the body may read the very accumulator we are mutating, so its
                # result must be recomputed against the latest rows each round.
                # The flow cache keys on graph text (identical every iteration),
                # so without this it would hand back the first round's output.
                self._flow_cache_clear()
                extra = {var: str(i)} if var else {}
                rg = applyvars.resolve_graph(graph, extra=extra)
                created, et = [], None
                try:
                    for aid in api_ids:
                        fr = self.fetch_api_node(
                            aid, self._node_config(rg, aid), graph=rg,
                            should_cancel=(
                                lambda: self._run_is_cancelled(query_id)))
                        if fr.get("error"):
                            raise nodeflow.NodeflowError(
                                "API node fetch failed: " + fr["error"])
                        self._set_node_config(rg, aid, {
                            "table": fr.get("table"),
                            "err_table": fr.get("err_table")})
                    et = self._flow_engine_target(rg)
                    eng, _kind = self._engine_obj(et)
                    tmp = self._materialize_flow(rg, bn, bp, et, created)
                    mode = "overwrite" if (first and reset_first) else "append"
                    before = 0
                    if not (first and reset_first):
                        try:
                            _c, rr = eng.execute(
                                'SELECT COUNT(*) FROM "%s"' % accum)
                            before = int(rr[0][0]) if rr else 0
                        except Exception:
                            before = 0
                    self._write_into_table(eng, accum, tmp, mode, replace_keys)
                    if accumulate == "reduce":
                        self._reduce_accumulator(
                            eng, accum, reduce_keys, reduce_aggs)
                    eng.sync_catalog()
                    try:
                        eng.table_sources[accum] = "nodeflow"
                    except Exception:
                        pass
                    after = 0
                    try:
                        _c, rr = eng.execute(
                            'SELECT COUNT(*) FROM "%s"' % accum)
                        after = int(rr[0][0]) if rr else 0
                    except Exception:
                        after = 0
                    last_engine = ("duckdb" if et == DUCKDB_TARGET
                                   else "sqlite")
                    iterations += 1
                    first = False
                    if after <= before:    # nothing new this round -> fixpoint
                        converged = True
                        break
                finally:
                    self._drop_flow_temps(et, created)
        except Exception as e:
            self._unregister_run(query_id)
            if _is_interrupt(e):
                return {"error": "cancelled", "cancelled": True}
            return {"error": err_str(e)}
        self._unregister_run(query_id)
        self._invalidate_profiles()
        self._invalidate_counts()
        rows = None
        if last_engine and iterations:
            e = self.duckdb if last_engine == "duckdb" else self.db
            try:
                _c, r = e.execute('SELECT COUNT(*) FROM "%s"' % accum)
                rows = int(r[0][0]) if r else None
            except Exception:
                rows = None
        note = None
        if not converged and iterations >= max_iters:
            note = ("Stopped at the %d-iteration cap (no fixpoint reached)."
                    % max_iters)
        return {"ok": iterations > 0, "iterations": iterations,
                "converged": converged, "rows": rows,
                "table": accum if iterations else None,
                "engine": last_engine, "note": note,
                "all": self.tables_tree()}


    # ---- query execution / routing ---------------------------------
    def _table_names_in(self, sql):
        try:
            return set(re.findall(
                r'(?:\bfrom\b|\bjoin\b)\s+["\[]?([A-Za-z_][\w$]*)',
                sql or "", re.IGNORECASE))
        except Exception:
            return set()

    def _choose_local_engine(self, sql):
        names = self._table_names_in(sql)
        duck_cols = (self.duckdb.table_columns
                     if self.duckdb is not None else {})
        sqlite_cols = self.db.table_columns
        for n in names:
            if n in duck_cols and n not in sqlite_cols:
                return DUCKDB_TARGET
        return LOCAL_TARGET

    def cross_engine_conflict(self, sql):
        names = self._table_names_in(sql)
        duck_cols = (self.duckdb.table_columns
                     if self.duckdb is not None else {})
        sqlite_cols = self.db.table_columns
        s_only = sorted(n for n in names
                        if n in sqlite_cols and n not in duck_cols)
        d_only = sorted(n for n in names
                        if n in duck_cols and n not in sqlite_cols)
        if s_only and d_only:
            return {"sqlite": s_only, "duckdb": d_only}
        return None

    def _catalog_conn_for(self, sql):
        """The connection a query should run on (passthrough) if it references
        a remote catalog table that isn't also a locally-loaded table; else
        None. Ambiguous (two connections) or purely-local queries stay local."""
        if not self._catalog_route:
            return None
        names = {n.lower() for n in self._table_names_in(sql)}
        if not names:
            return None
        local = {n.lower() for n in self.db.table_columns}
        if self.duckdb is not None:
            local |= {n.lower() for n in self.duckdb.table_columns}
        conns = {self._catalog_route[n] for n in names
                 if n in self._catalog_route and n not in local}
        return next(iter(conns)) if len(conns) == 1 else None

    def _query_mixes_catalog_and_local(self, sql):
        """True if a statement references both a remote catalog table and a
        locally-loaded table -- which can't run together in one engine."""
        if not self._catalog_route:
            return False
        names = {n.lower() for n in self._table_names_in(sql)}
        local = {n.lower() for n in self.db.table_columns}
        if self.duckdb is not None:
            local |= {n.lower() for n in self.duckdb.table_columns}
        has_remote = any(n in self._catalog_route and n not in local
                         for n in names)
        has_local = any(n in local for n in names)
        return has_remote and has_local

    def _engine_owning_missing_table(self, error_msg, current_target):
        if current_target not in (LOCAL_TARGET, DUCKDB_TARGET):
            return None
        missing = None
        for pat in _MISSING_TABLE_PATTERNS:
            m = pat.search(error_msg or "")
            if m:
                missing = m.group(1).strip().strip('"').strip("'")
                break
        if not missing:
            return None
        if "." in missing:
            missing = missing.rsplit(".", 1)[-1]
        if current_target == LOCAL_TARGET:
            if (self.duckdb is not None
                    and missing in self.duckdb.table_columns):
                return DUCKDB_TARGET
        else:
            if missing in self.db.table_columns:
                return LOCAL_TARGET
        return None

    def _col_names(self, eng, table=None, sql=None):
        """Column names of a table (by ``table`` name) or of a subquery
        (``sql``) via a 0-row probe. One home for the LIMIT-0 introspection that
        was inlined across many call sites. Returns a list (empty if none)."""
        if table is not None:
            cols, _ = eng.execute('SELECT * FROM "%s" LIMIT 0' % table)
        else:
            cols, _ = eng.execute("SELECT * FROM (%s) AS _z LIMIT 0" % sql)
        return list(cols or [])

    def _engine_obj(self, target):
        if target == LOCAL_TARGET:
            return self.db, "sqlite"
        if target == DUCKDB_TARGET:
            return self.get_duckdb(), "duckdb"
        conn = self.connections.get(target)
        if conn is None:
            raise RuntimeError(f'Connection "{target}" is not active.')
        return conn, "remote"

    # ---- WebHDFS connector -------------------------------------------- #
    # Browse a WebHDFS endpoint, regroup <root>/<date>/<feed>/<file> into a
    # feed-first tree, and combine selected feeds across selected dates into
    # local tables. The transport (samql_core.hdfs.WebHDFSClient) is a faithful
    # port of the original desktop tool's HDFS tab (plain urllib, optional
    # user.name, no Kerberos library). See samql_core/hdfs.py.

    def hdfs_connect(self, url, user=None):
        """Open a WebHDFS endpoint and list its top level (the folders to pick,
        e.g. output / source). A successful LISTSTATUS doubles as a reachability
        check. Replaces any previous HDFS client."""
        from . import hdfs
        try:
            base, path = hdfs.parse_webhdfs_url(url)
        except Exception as e:
            # .479 audit: hdfs has no err_str -- this would AttributeError
            # on any malformed WebHDFS URL. err_str is the errfmt helper,
            # already imported at module scope.
            return {"error": err_str(e)}
        client = hdfs.WebHDFSClient(base, user=user)
        try:
            entries = client.list_dir(path)
        except Exception as e:
            return {"error": hdfs.friendly_hdfs_error(e)}
        self._hdfs = client
        folders = sorted(e["pathSuffix"] for e in entries
                         if e.get("type") == "DIRECTORY" and e.get("pathSuffix"))
        files = sorted(e["pathSuffix"] for e in entries
                       if e.get("type") == "FILE" and e.get("pathSuffix"))
        return {"ok": True, "base": base, "path": path, "folders": folders,
                "files": files}

    def hdfs_browse(self, path):
        """List one directory's sub-folders and files (for drilling in)."""
        from . import hdfs
        if self._hdfs is None:
            return {"error": "Not connected to HDFS."}
        try:
            entries = self._hdfs.list_dir(path or "/")
        except Exception as e:
            return {"error": hdfs.friendly_hdfs_error(e)}
        dirs = sorted(e["pathSuffix"] for e in entries
                      if e.get("type") == "DIRECTORY" and e.get("pathSuffix"))
        files = sorted(e["pathSuffix"] for e in entries
                       if e.get("type") == "FILE" and e.get("pathSuffix"))
        return {"ok": True, "path": path, "dirs": dirs, "files": files}

    def hdfs_load_file(self, remote_path, destination="auto", mode="view",
                      base_name=None, cancel=None, progress=None):
        """Load ONE HDFS file you browsed to (CSV, TSV, JSON, or Parquet) by
        streaming it to a local temp file and handing it to the standard file
        loader, which dispatches on the extension.

        Default mode='view' makes a zero-copy DuckDB view over the downloaded
        file (read_csv_auto / read_json / read_parquet): it's queried in place,
        lazily, and is never copied into a table or held in memory -- the
        optimal path for big files. The streamed temp lives in the per-instance
        dir so the view keeps reading it (and gets swept on shutdown / next
        start, like other session data). mode='materialize' copies it into a
        real table instead, and the temp is removed afterwards. Without DuckDB,
        view mode falls back to a materialised load (CSV/TSV/JSON -> SQLite;
        Parquet needs DuckDB and surfaces a friendly error). Both the stream and
        the load are cancellable. Returns {ok, tables, all} or {error}."""
        from . import hdfs, tmputil
        from .loaders import LoadCancelled
        if self._hdfs is None:
            return {"error": "Not connected to HDFS."}
        rp = (remote_path or "").strip()
        if not rp:
            return {"error": "No file selected."}
        stem = os.path.splitext(os.path.basename(rp.rstrip("/")))[0]
        # Always sanitize -- same rules as path/upload loads (parens, dots,
        # spaces, etc. must not become fragile catalog identifiers).
        from .loaders import base_name_for as _safe_table
        name = _safe_table(base_name or stem or "hdfs_file")
        ext = os.path.splitext(rp)[1].lower() or ".csv"
        tmp = tmputil.new_tempfile("hdfs_", ext)

        def _rm():
            try:
                os.unlink(tmp)
            except Exception:
                pass

        try:
            self._hdfs_fetch_to_file(rp, tmp, cancel=cancel, progress=progress)
        except LoadCancelled:
            _rm()
            raise
        except Exception as e:
            _rm()
            return {"error": hdfs.friendly_hdfs_error(e)}
        try:
            loaded = self.load_file(tmp, destination=destination,
                                    base_name=name, mode=mode)
        except Exception as e:
            _rm()
            return {"error": err_str(e)}
        # A view keeps reading the temp; a materialized load copied the data in,
        # so its temp can go.
        is_view = any(isinstance(t, dict) and t.get("view")
                      for t in (loaded or []))
        if not is_view:
            _rm()
        self._invalidate_profiles()
        self._invalidate_counts()
        return {"ok": True, "tables": loaded, "all": self.tables_tree()}

    def _hdfs_fetch_to_file(self, remote_path, dest, chunk=65536,
                            cancel=None, progress=None, base=0):
        """Stream a WebHDFS OPEN to a local file in chunks (the proven path).

        *cancel* (callable -> bool) is checked between chunks so a long fetch
        stops promptly; *progress* (callable(bytes_done)) reports cumulative
        bytes, *base* being the bytes already streamed by earlier files in the
        same job. Returns the byte count for this file. Raises LoadCancelled if
        cancellation is requested mid-stream."""
        from .loaders import LoadCancelled
        resp = self._hdfs.open_stream(remote_path)
        got = 0
        try:
            with open(dest, "wb") as out:
                while True:
                    if cancel is not None and cancel():
                        raise LoadCancelled()
                    buf = resp.read(chunk)
                    if not buf:
                        break
                    out.write(buf)
                    got += len(buf)
                    if progress is not None:
                        progress(base + got)
        finally:
            try:
                resp.close()
            except Exception:
                pass
        return got

    def _reusable_store_path(self, result_id):
        """The parquet-store path for a result that is SAFE to reuse as data:
        a DuckDB parquet store, not capped (.349 -- a capped store is not the
        full result), file still on disk. None otherwise. Shared by the
        journal chain reuse (TEMP views) and reconcile input staging."""
        cr = self._results.get(result_id)
        store = getattr(cr, "store", None)
        path = getattr(store, "_path", None)
        if (cr is None or path is None
                or getattr(cr, "engine", "") != "duckdb"
                or getattr(cr, "capped", False)
                or not os.path.exists(path)):
            return None
        return path

    def _setup_reuse_views(self, engine, reuse):
        """R1 (journal chain reuse): map {cell_name: result_id} to TEMP VIEWs
        over each result's parquet store, so a chained cell reads its
        upstream's ALREADY-COMPUTED result instead of re-executing its SQL.

        Returns (stale_names, cleanup). A name lands in stale_names -- and no
        views are created -- when any referenced result is missing (evicted),
        not a DuckDB parquet store, or was CAPPED (.349: a capped store is not
        the full result, so reusing it would silently change answers); the
        caller then falls back to full CTE inlining. ``cleanup()`` drops the
        views best-effort."""
        stale, plans = [], []
        # .513: guard against the LIVE catalog, not the mutable cache dict --
        # a concurrent sync_catalog rebuild could momentarily hide a real
        # table and let a reuse view shadow it.
        existing = None
        try:
            cur = engine.conn.execute(
                "SELECT lower(table_name) FROM information_schema.tables "
                "WHERE table_schema='main'")
            existing = {r[0] for r in cur.fetchall()}
        except Exception:
            pass
        if existing is None:
            existing = {str(t).lower()
                        for t in (getattr(engine, "table_columns", {}) or {})}
        for name, rid in (reuse or {}).items():
            path = self._reusable_store_path(rid)
            # a cell name that matches a REAL table/view is never reused as a
            # TEMP view: the view would shadow the user's table for the
            # duration of the query. Bounce it; the client inlines instead.
            if path is None or str(name).lower() in existing:
                stale.append(name)
                continue
            fwd = sqlutil.sql_path(path)
            qn = '"%s"' % str(name).replace('"', '""')
            plans.append((qn, fwd))
        if stale:
            return stale, lambda: None
        made = []
        # pin the source results: an eviction / free / discard while their
        # views are live would yank the file out from under the running query
        pinned = []
        pins = getattr(self, "_reuse_pins", None)
        lock = getattr(self, "_lock", None)
        if pins is not None and lock is not None:
            with lock:
                for _n, rid in (reuse or {}).items():
                    pins[rid] = pins.get(rid, 0) + 1
                    pinned.append(rid)

        def cleanup():
            # .513: a Stop can interrupt these DROPs themselves, orphaning a
            # temp view that then SHADOWS real tables forever. Retry once
            # after a beat; anything that still survives is reaped by the
            # engine's shadow purge on the next query.
            remaining = list(made)
            for attempt in (0, 1):
                failed = []
                for qn in remaining:
                    try:
                        engine.conn.execute("DROP VIEW IF EXISTS %s" % qn)
                    except Exception:
                        failed.append(qn)
                remaining = failed
                if not remaining:
                    break
                time.sleep(0.05)
            if pins is not None and lock is not None:
                flush = []
                with lock:
                    for rid in pinned:
                        n = pins.get(rid, 0) - 1
                        if n <= 0:
                            pins.pop(rid, None)
                            if rid in self._discard_deferred:
                                self._discard_deferred.discard(rid)
                                flush.append(rid)
                        else:
                            pins[rid] = n
                for rid in flush:
                    try:
                        self.discard_result(rid)
                    except Exception:
                        pass
        try:
            for qn, fwd in plans:
                engine.conn.execute(
                    "CREATE OR REPLACE TEMP VIEW %s AS "
                    "SELECT * FROM read_parquet('%s')" % (qn, fwd))
                made.append(qn)
        except Exception:
            cleanup()
            return [n for n, _ in (reuse or {}).items()], lambda: None
        return [], cleanup

    def run_query(self, sql, target="auto", read_only=False, query_id=None,
                  dialect=None, reuse=None, surface=None, label=None,
                  preview_limit=None):
        """Meta wrapper: a run that arrives WITH a surface (ide / journal)
        registers itself so the stat modal lists it live, grouped and
        labelled; flow-internal calls pass no surface and register at the
        flow level instead (no double-begin on the same id)."""
        if surface and query_id:
            self._clear_stale_engine_cancel(except_qid=query_id)
            self._register_run(query_id, None, surface=surface, label=label)
            try:
                return self._run_query_inner(sql, target, read_only,
                                             query_id, dialect, reuse,
                                             preview_limit=preview_limit)
            finally:
                # Keep the cancelled flag through mid-send abort (_REQ_LOCAL).
                self._end_run_keep_cancel(query_id)
        self._clear_stale_engine_cancel(except_qid=query_id)
        return self._run_query_inner(sql, target, read_only, query_id,
                                     dialect, reuse,
                                     preview_limit=preview_limit)

    _TXN_RE = re.compile(r"^\s*(BEGIN|COMMIT|ROLLBACK|END)\b", re.I)

    def run_script(self, sql, target="auto", read_only=False,
                   query_id=None, dialect=None, surface=None,
                   label=None):
        """.458: Run-all with PER-STATEMENT results. Splits with the
        audited scanner and runs each statement through the normal
        single-run path -- same connection (temp state persists), same
        query_id (cancel reaches whichever statement is current) --
        collecting a ledger: index, a trimmed preview, elapsed ms, and
        for row-producing statements their own pageable result_id.
        The RETURN SHAPE stays back-compatible: the top level is the
        LAST row-producing statement's payload (or the last
        statement's), with "statements" added. Scripts that manage
        their own transactions (BEGIN/COMMIT/ROLLBACK) fall back to
        the classic one-batch run -- per-statement autocommit would
        change their semantics -- and the ledger says so."""
        spans = split_sql_statements_spans(sql or "")
        stmts = [sp[2].strip() for sp in spans if sp[2].strip()]
        if len(stmts) <= 1:
            r = self.run_query(sql, target=target, read_only=read_only,
                               query_id=query_id, dialect=dialect,
                               surface=surface, label=label)
            r["statements"] = [{"index": 0,
                                "sql_preview": (stmts[0][:300]
                                                if stmts else ""),
                                "ms": r.get("elapsed_ms"),
                                "result_id": r.get("result_id"),
                                "total_rows": r.get("total_rows"),
                                "error": r.get("error")}]
            return r
        if any(self._TXN_RE.match(st) for st in stmts):
            r = self.run_query(sql, target=target, read_only=read_only,
                               query_id=query_id, dialect=dialect,
                               surface=surface, label=label)
            r["statements"] = [{
                "index": 0, "sql_preview": stmts[0][:300],
                "ms": r.get("elapsed_ms"),
                "result_id": r.get("result_id"),
                "total_rows": r.get("total_rows"),
                "note": ("transaction script: executed as one "
                         "batch"),
                "error": r.get("error")}]
            return r
        if surface and query_id:
            self._register_run(query_id, None, surface=surface,
                               label=label)
        ledger = []
        final = None
        last_rows = None
        try:
            for i, st in enumerate(stmts):
                t0 = time.monotonic()
                r = self.run_query(st, target=target,
                                   read_only=read_only,
                                   query_id=query_id,
                                   dialect=dialect)
                ms = int((time.monotonic() - t0) * 1000)
                ent = {"index": i, "sql_preview": st[:300],
                       "ms": ms}
                if r.get("error"):
                    ent["error"] = r["error"]
                    ledger.append(ent)
                    final = dict(r)
                    final["failed_statement"] = i
                    break
                if r.get("result_id"):
                    ent["result_id"] = r["result_id"]
                    ent["total_rows"] = r.get("total_rows")
                    ent["columns"] = r.get("columns")
                    last_rows = r
                else:
                    ent["total_rows"] = r.get("total_rows", 0)
                ledger.append(ent)
                final = r
        finally:
            if surface and query_id:
                self._unregister_run(query_id)
        base = dict(last_rows if last_rows is not None else
                    (final or {}))
        if final is not None and final.get("error"):
            base = dict(final)
        base["statements"] = ledger
        return base

    def _run_query_inner(self, sql, target="auto", read_only=False,
                         query_id=None, dialect=None, reuse=None,
                         preview_limit=None):
        """Run SQL and cache the result. Returns a dict with the first
        page of rows plus metadata. ``target`` may be 'auto',
        '__local__', '__duckdb__', or a remote connection name."""
        sql = (sql or "").strip()
        if not sql:
            return {"error": "Nothing to run."}
        # The JSON layer may send target as null or "" -> treat as auto.
        if target in (None, "", "auto"):
            target = "auto"
        if read_only and classify_sql_statement(sql) == "write":
            return {"error": "Read-only mode: write statements are blocked."}
        stmt_kind = classify_sql_statement(sql)
        if stmt_kind == "write":
            # A free-form write can mutate a file-backed table without changing
            # its source-file stat signature. Disable restart-persistent reuse
            # for the remainder of this session; the next restart reloads the
            # source and safely re-enables it.
            self._persistent_cache_tainted = True
            # a write may change row counts -> drop the cached counts so the
            # next /api/tables recomputes them
            self._invalidate_counts()

        # A query that references a remote catalog table runs on that SQL
        # Server connection (passthrough); the result is cached like any
        # other so the grid / paging / charts work unchanged.
        if target in ("auto", LOCAL_TARGET):
            if self._query_mixes_catalog_and_local(sql):
                return {"error": "A query against SQL Server catalog tables "
                        "runs on the server, so it can't also reference a "
                        "locally-loaded table in the same statement. Query "
                        "them separately, or load the catalog table's data "
                        "first."}
            cat_conn = self._catalog_conn_for(sql)
            if cat_conn is not None:
                target = cat_conn

        if target in ("auto", LOCAL_TARGET):
            conflict = self.cross_engine_conflict(sql)
            if conflict is not None:
                return {"error": "cross_engine_conflict",
                        "detail": conflict}

        first_target = target
        if target in ("auto", LOCAL_TARGET):
            first_target = self._choose_local_engine(sql)

        # Optional source-dialect transpile: let the user write SQL in another
        # dialect (e.g. Spark SQL) and run it on the resolved local engine.
        # Gated on sqlglot; a construct sqlglot can't translate surfaces as an
        # error rather than silently running mangled SQL. A remote (SQL Server
        # catalog) target is left untouched -- it speaks its own dialect.
        if (dialect and dialect not in ("native", "auto", "")
                and first_target in (DUCKDB_TARGET, LOCAL_TARGET)):
            write = "duckdb" if first_target == DUCKDB_TARGET else "sqlite"
            res = self.transpile_sql(sql, dialect, write)
            if not res.get("ok"):
                return {"error": "Couldn't translate %s SQL to %s: %s"
                        % (dialect, write, res.get("result") or "")}
            sql = (res.get("result") or "").strip() or sql

        original_sql = sql
        preview_n = None
        if preview_limit not in (None, "", 0, "0"):
            try:
                preview_n = max(1, min(int(preview_limit), 100000))
            except (TypeError, ValueError):
                preview_n = DISPLAY_LIMIT
            if stmt_kind != "read" or not self._is_single_read(sql):
                return {"error": "Preview supports one SELECT statement at a time."}
            inner_preview = sql.strip().rstrip(";").strip()
            sql = ("SELECT * FROM (%s) AS _samql_preview LIMIT %d"
                   % (inner_preview, preview_n))

        t0 = time.time()
        # R1 (journal chain reuse): stand up TEMP VIEWs over fresh upstream
        # results for the duration of the exec. DuckDB-only; anything not
        # reusable bounces back as reuse_stale so the caller recomposes with
        # full CTE inlining instead.
        reuse_cleanup = None
        if reuse:
            eng = self.duckdb if first_target == DUCKDB_TARGET else None
            if eng is None:
                return {"reuse_stale": sorted(reuse.keys())}
            with eng.write_lock:
                stale, reuse_cleanup = self._setup_reuse_views(eng, reuse)
            if stale:
                return {"reuse_stale": sorted(stale)}
        try:
            try:
                # a run using reuse views must stay on the locked main
                # connection: TEMP views are invisible to a fresh cursor
                cols, rows, total, engine_kind = self._exec_target(
                    sql, first_target, query_id=query_id,
                    prefer_locked=reuse_cleanup is not None)
            except Exception as e:
                if _is_interrupt(e):
                    self.history.add(original_sql, target=first_target,
                                     error="cancelled",
                                     elapsed_sec=round(time.time() - t0, 3))
                    self._post_interrupt_hygiene(first_target)
                    return {"error": "cancelled", "cancelled": True}
                suggested = self._engine_owning_missing_table(
                    err_str(e), first_target)
                if (suggested is not None and suggested != first_target
                        and reuse_cleanup is None):
                    try:
                        cols, rows, total, engine_kind = self._exec_target(
                            sql, suggested, query_id=query_id,
                            prefer_locked=reuse_cleanup is not None)
                        first_target = suggested
                    except Exception as e2:
                        if _is_interrupt(e2):
                            self._post_interrupt_hygiene(first_target)
                            return {"error": "cancelled", "cancelled": True}
                        return self._query_error(original_sql, first_target, e2, t0)
                else:
                    return self._query_error(original_sql, first_target, e, t0)
        finally:
            if reuse_cleanup is not None:
                try:
                    with self.duckdb.write_lock:
                        reuse_cleanup()
                except Exception:
                    pass

        elapsed = time.time() - t0
        if cols is None:
            # Non-SELECT statement (DDL/DML): the catalog may have changed.
            try:
                engine_obj, _ = self._engine_obj(first_target)
                if hasattr(engine_obj, "sync_catalog"):
                    engine_obj.sync_catalog()
            except Exception:
                pass
            self.history.add(original_sql, target=first_target, row_count=None,
                             elapsed_sec=round(elapsed, 3))
            self._invalidate_profiles()
            self._schedule_cleanup(full=True)
            return {"columns": [], "rows": [], "total_rows": 0,
                    "result_id": None, "elapsed_ms": int(elapsed * 1000),
                    "engine": engine_kind, "statement": "ok",
                    "truncated": False}

        rid = self._cache_result(cols, rows, total, original_sql, first_target,
                                 engine_kind)
        capped = bool(getattr(rows, "capped", False))
        cap_n = getattr(rows, "cap", None)
        self.history.add(original_sql, target=first_target, row_count=total,
                         elapsed_sec=round(elapsed, 3))
        self._schedule_cleanup(full=False)
        # Return only the first chunk (the grid lazy-loads more in LAZY_CHUNK
        # steps as you scroll). Shipping the full DISPLAY_LIMIT up front meant a
        # `SELECT *` over a flatten-off nested table serialized thousands of
        # multi-KB struct cells in one response -- enough to hang/reload the
        # tab. One chunk keeps the first response bounded; nothing is lost.
        page = self.page(rid, 0, INITIAL_PAGE_ROWS)
        page.update({
            "result_id": rid,
            "elapsed_ms": int(elapsed * 1000),
            "engine": engine_kind,
            "target": first_target,
            # lets the client skip a needless /api/tables refresh after a
            # pure read (a SELECT cannot have changed the catalog)
            "stmt_kind": stmt_kind,
            "truncated": total > DISPLAY_LIMIT,
            "result_capped": capped,
            "result_cap": cap_n if capped else None,
        })
        if preview_n is not None:
            page.update({"preview": True,
                         "preview_limit": preview_n,
                         "preview_limited": total >= preview_n})
        return page

    @staticmethod
    def _is_single_read(sql):
        """True only for a single read (SELECT-like) statement, so it is
        safe to wrap in COPY (...) TO parquet."""
        try:
            parts = [p for p in split_statements(sql) if p and p.strip()]
        except Exception:
            return False
        if len(parts) != 1:
            return False
        try:
            return classify_sql_statement(parts[0]) == "read"
        except Exception:
            return False

    def _result_row_cap(self):
        """A hard ceiling on how many rows a single query result may
        materialize, so a runaway explode (e.g. UNNEST cross-product on a nested
        column) can't produce billions of rows and pin the engine forever. The
        materialization stops at the ceiling and the result is flagged capped.
        Override with SAMQL_MAX_RESULT_ROWS (0 / negative disables the cap)."""
        v = os.environ.get("SAMQL_MAX_RESULT_ROWS")
        if v is None:
            return 10_000_000
        try:
            n = int(v)
            return n if n > 0 else None
        except (TypeError, ValueError):
            return 10_000_000

    def _result_byte_cap(self):
        """Hard ceiling on materialized result Parquet size (bytes).

        Complements the row cap so a wide nested explode can't fill the disk
        before hitting 10M rows. Override with ``SAMQL_MAX_RESULT_GB``
        (default 8; ``0`` disables)."""
        v = os.environ.get("SAMQL_MAX_RESULT_GB")
        if v is None:
            return 8 * 1024 * 1024 * 1024
        try:
            gb = float(v)
            return None if gb <= 0 else int(gb * 1024 * 1024 * 1024)
        except (TypeError, ValueError):
            return 8 * 1024 * 1024 * 1024

    class _ReadCursorHandle:
        """Precise cancel target for one DuckDB read cursor.

        ``interrupt()`` also trips a per-run event consumed by the native
        execute keepalive. DuckDB COPY work can observe an interrupt only at a
        later checkpoint, so the keepalive reissues the interrupt until this
        exact cursor exits. The engine-global cancel event is deliberately not
        used: cancelling one concurrent read must not poison or stop siblings.
        """
        __slots__ = ("_rc", "cancel_event")

        def __init__(self, rc):
            self._rc = rc
            self.cancel_event = threading.Event()

        def interrupt(self):
            self.cancel_event.set()
            try:
                self._rc.interrupt()
            except Exception:
                pass

    _BARE_SELECT = re.compile(
        r'^\s*SELECT\s+\*\s+FROM\s+'
        r'("(?:[^"]|"")+"|[A-Za-z_]\w*)\s*;?\s*$',
        re.IGNORECASE)

    # Simple SELECT list FROM one identifier, optional LIMIT.
    # Used to widen zero-copy to column projections over a parquet source.
    _SIMPLE_SELECT = re.compile(
        r'^\s*SELECT\s+(?P<cols>.+?)\s+FROM\s+'
        r'(?P<table>"(?:[^"]|"")+"|[A-Za-z_]\w*)'
        r'(?:\s+LIMIT\s+(?P<limit>\d+))?'
        r'\s*;?\s*$',
        re.IGNORECASE | re.DOTALL)

    @staticmethod
    def _parse_simple_select_cols(cols_sql):
        """Return ``'*'``, a list of bare column names, or ``None`` if the
        select list is an expression / alias / function (not zero-copy safe)."""
        s = (cols_sql or "").strip()
        if not s:
            return None
        if s == "*":
            return "*"
        out = []
        for part in s.split(","):
            p = part.strip()
            if not p or "(" in p or ")" in p:
                return None
            if re.search(r"(?i)\bas\b", p):
                return None
            if p.startswith('"') and p.endswith('"') and len(p) >= 2:
                out.append(p[1:-1].replace('""', '"'))
                continue
            if re.match(r"^[A-Za-z_]\w*$", p):
                out.append(p)
                continue
            return None
        return out or None

    def _parquet_table_source(self, engine, ident):
        """Resolve ``table_sources[ident]`` to a .parquet path, or None."""
        if ident.startswith('"'):
            ident = ident[1:-1].replace('""', '"')
        sources = getattr(engine, "table_sources", {}) or {}
        src = sources.get(ident)
        if src is None:
            low = ident.lower()
            for k, v in sources.items():
                if str(k).lower() == low:
                    src = v
                    break
        if not src or not str(src).lower().endswith(".parquet"):
            return None
        return str(src)

    def _zero_copy_source(self, engine, sql):
        """Serve a parquet-backed table as the result store with no COPY.

        Matches:
          * bare ``SELECT * FROM t`` (historical path)
          * simple projections ``SELECT a, b FROM t`` (no WHERE/JOIN/…)

        The on-box 2026-07-02 failure: rematerializing 1.7 GB of nested structs
        into a second parquet starved every other request. Projection still
        borrows the source file; ParquetResultStore pushes column lists at
        read time.
        """
        m = self._SIMPLE_SELECT.match(sql or "")
        if not m:
            return None
        want = self._parse_simple_select_cols(m.group("cols"))
        if want is None:
            return None
        src = self._parquet_table_source(engine, m.group("table"))
        if not src:
            return None
        try:
            cur = engine.conn.cursor()
            try:
                fwd = sqlutil.sql_path(src)
                cur.execute(
                    "SELECT count(*) FROM read_parquet('%s')" % fwd)
                total = int(cur.fetchone()[0])
                cur.execute(
                    "DESCRIBE SELECT * FROM read_parquet('%s')" % fwd)
                all_cols = [r[0] for r in cur.fetchall()]
            finally:
                try:
                    cur.close()
                except Exception:
                    pass
        except Exception:
            return None
        if want == "*":
            cols = all_cols
        else:
            # Case-insensitive match against described columns; preserve
            # the SELECT-list order the user asked for.
            by_low = {str(c).lower(): c for c in all_cols}
            cols = []
            for name in want:
                hit = by_low.get(name.lower())
                if hit is None:
                    return None  # unknown column → normal path / error there
                cols.append(hit)
        lim_raw = m.groupdict().get("limit")
        lim = None
        if lim_raw is not None:
            try:
                lim = max(0, int(lim_raw))
            except Exception:
                lim = None
        from .rows import ParquetResultStore
        store = ParquetResultStore(engine, src, cols, owns_path=False)
        if lim is not None and lim < total:
            store._n = lim
            store.capped = True
            store.cap = lim
            total = lim
        else:
            store._n = total
            store.capped = False
        return cols, store, total, "duckdb"

    def _exec_duckdb_parquet(self, engine, sql, query_id=None,
                             prefer_locked=False):
        """Materialize a DuckDB result to a temporary Parquet file and
        return a ParquetResultStore over it. The query runs exactly once
        (the COPY). Raises on any problem so the caller can fall back.
        A bare SELECT * over a parquet-backed view short-circuits to the
        source file itself (see _zero_copy_source).

        A safety LIMIT (``_result_row_cap``) bounds the COPY so a runaway
        explode terminates at the ceiling instead of writing forever (which
        would hold the connection lock and wedge the whole app); the result is
        flagged ``capped`` so the UI can say so.

        CONCURRENCY: when the engine's concurrent-reads flag is on (and the
        caller didn't ask for the locked path), the COPY runs on its OWN
        cursor -- DuckDB's documented one-cursor-per-thread MVCC model -- so a
        read never serializes behind a build/load holding the main connection.
        Rails: a run using R1 reuse views stays on the main connection (TEMP
        views are invisible to a fresh cursor); an INTERRUPT re-raises as a
        cancel (never silently retried); any other cursor failure (e.g. a
        TEMP table the cursor can't see) falls back to the locked path with a
        fresh temp file; and the cursor registers itself under ``query_id``
        so cancel_query() interrupts the right statement."""
        fast = self._zero_copy_source(engine, sql)
        if fast is not None:
            return fast
        inner = sql.strip().rstrip(";").strip()
        if not inner:
            raise RuntimeError("empty query")
        cap = self._result_row_cap()
        # cap + 1 so we can tell "exactly at the ceiling" from "over it".
        limit_clause = " LIMIT %d" % (cap + 1) if cap else ""

        def _attempt(exec_conn, lock, handle):
            path = tmputil.new_tempfile("qr_", ".parquet")
            try:
                os.unlink(path)
            except Exception:
                pass
            fwd = sqlutil.sql_path(path)
            prev_run = None
            already_cancelled = False
            if handle is not None and query_id:
                with self._running_lock:
                    prev_run = self._running.get(query_id)
                    self._running[query_id] = handle
                    already_cancelled = query_id in getattr(self, "_cancelled_runs", set())
                # Stop can arrive in the small gap after _exec_target
                # registers the run but before this exact cursor exists.
                # Deliver that queued cancellation immediately on bind.
                if already_cancelled:
                    self._interrupt_entry(handle)
            try:
                ctx = lock if lock is not None else _NullCtx()
                with ctx:
                    # No row_number() window here: a global window forces a
                    # full single-threaded sequencing of the result before a
                    # byte is written and defeats parallel COPY. Parquet
                    # preserves write order within a file, so the store
                    # synthesizes the stable "__rn" at READ time from
                    # file_row_number instead (rows.py). The COPY can run for
                    # minutes on a huge result, so it heartbeats (no false
                    # "stalled" card) and takes cancel nudges via the
                    # keepalive.
                    cancel_signal = getattr(
                        handle, "cancel_event",
                        getattr(engine, "_cancel", None))
                    keepalive_interval = (0.25 if hasattr(
                        handle, "cancel_event") else 3.0)
                    with _ExecKeepalive(exec_conn, cancel_signal,
                                        threading.get_ident(),
                                        interval=keepalive_interval):
                        # ORDER BY in the inner SQL must survive parallel COPY:
                        # temporarily force insertion order so file_row_number
                        # matches the query's sort when paging later.
                        has_order = bool(re.search(
                            r"(?is)\border\s+by\b", inner))
                        if has_order:
                            try:
                                exec_conn.execute(
                                    "SET preserve_insertion_order = true")
                            except Exception:
                                pass
                        bcap_fn = getattr(self, "_result_byte_cap", None)
                        bcap = bcap_fn() if callable(bcap_fn) else None

                        def _copy_cancelled():
                            return bool(query_id) and self._run_is_cancelled(
                                query_id)

                        try:
                            from .engines import exec_copy_parquet, _is_interrupt
                            try:
                                exec_copy_parquet(
                                    exec_conn,
                                    "SELECT * FROM (%s)%s"
                                    % (inner, limit_clause),
                                    fwd,
                                    max_bytes=bcap,
                                    should_cancel=_copy_cancelled,
                                    interrupt_fn=getattr(
                                        exec_conn, "interrupt", None))
                            except RuntimeError:
                                raise
                            except Exception as e:
                                # Only fall back when the engine rejected
                                # jumbo options / unexpected SQL — not on
                                # cancel or byte-ceiling aborts.
                                if _copy_cancelled() or _is_interrupt(e):
                                    raise
                                exec_conn.execute(
                                    "COPY (SELECT * FROM (%s)%s) "
                                    "TO '%s' (FORMAT PARQUET)"
                                    % (inner, limit_clause, fwd))
                        finally:
                            if has_order:
                                try:
                                    exec_conn.execute(
                                        "SET preserve_insertion_order = false")
                                except Exception:
                                    pass
                        # Post-COPY byte check (race with monitor).
                        if bcap:
                            try:
                                sz = os.path.getsize(path)
                            except OSError:
                                sz = 0
                            if sz > bcap:
                                try:
                                    os.unlink(path)
                                except Exception:
                                    pass
                                raise RuntimeError(
                                    "Query result Parquet is %.1f GiB — over the "
                                    "%.1f GiB ceiling (SAMQL_MAX_RESULT_GB). "
                                    "Narrow the SELECT, add LIMIT, or raise the "
                                    "ceiling."
                                    % (sz / (1024 ** 3), bcap / (1024 ** 3)))
                        cur = exec_conn.execute(
                            f"SELECT * FROM read_parquet('{fwd}') LIMIT 0")
                        allcols = ([d[0] for d in cur.description]
                                   if cur.description else [])
                        if not allcols:
                            raise RuntimeError("parquet store: no columns")
                        cnt = exec_conn.execute(
                            f"SELECT count(*) FROM read_parquet('{fwd}')"
                        ).fetchone()
                        total = int(cnt[0]) if cnt else 0
                return path, allcols, total
            except BaseException:
                try:
                    os.unlink(path)
                except Exception:
                    pass
                raise
            finally:
                if handle is not None and query_id:
                    with self._running_lock:
                        if self._running.get(query_id) is handle:
                            if prev_run is not None:
                                self._running[query_id] = prev_run
                            else:
                                self._running.pop(query_id, None)

        use_cursor = (bool(getattr(engine, "concurrent_reads", False))
                      and not prefer_locked)
        path = allcols = total = None
        if use_cursor:
            # bounded cursors: no free slot (a Run storm) -> degrade to the
            # locked path instead of opening yet another connection.
            # .426 (lock-aware): degrading is only cheap when the main
            # connection is IDLE -- if its write lock is held (a build or
            # long CTAS in flight), the locked path queues behind exactly
            # the work async reads exist to dodge, so wait briefly for a
            # slot first and only then surrender.
            slots = getattr(engine, "_read_slots", None)
            if not _acquire_read_slot(slots, engine):
                use_cursor = False
        if use_cursor:
            rc = None
            try:
                rc = engine.conn.cursor()
                path, allcols, total = _attempt(
                    rc, None, self._ReadCursorHandle(rc))
            except BaseException as e:
                if _is_interrupt(e):
                    raise  # a cancel is a cancel -- never silently rerun
                if query_id and self._run_is_cancelled(query_id):
                    raise InterruptedError("cancelled") from e
                path = None  # anything else: retry on the locked main conn
            finally:
                if rc is not None:
                    try:
                        rc.close()
                    except Exception:
                        pass
                if slots is not None:
                    try:
                        slots.release()
                    except Exception:
                        pass
        if path is None:
            # Bind the main engine as the exact target only when we actually
            # take the serialized path. This preserves scoped cancellation
            # during the earlier cursor-acquisition race.
            path, allcols, total = _attempt(engine.conn, engine.write_lock,
                                            engine)
        capped = bool(cap) and total > cap
        if capped:
            total = cap  # expose the ceiling; the +1 sentinel row stays unread
        cols = allcols
        store = ParquetResultStore(engine, path, cols)
        store.capped = capped
        store.cap = cap
        return cols, store, total, "duckdb"

    def _exec_target(self, sql, target, query_id=None, prefer_locked=False):
        engine, kind = self._engine_obj(target)
        registration = engine
        if (query_id and kind == "duckdb"
                and getattr(engine, "concurrent_reads", False)
                and not prefer_locked):
            # A concurrent DuckDB read has no cursor yet at this point. Keep a
            # thread-aware DEFERRED registration instead of a bare engine:
            # falling back to engine.interrupt() here would cancel siblings.
            # _exec_duckdb_parquet binds the exact cursor moments later and
            # immediately delivers any cancellation that arrived in the gap.
            registration = {
                "engine": engine,
                "tid": threading.get_ident(),
                "deferred": True,
            }
        if query_id:
            with self._running_lock:
                self._running[query_id] = registration
        try:
            return self._exec_target_inner(sql, target, engine, kind,
                                           query_id=query_id,
                                           prefer_locked=prefer_locked)
        finally:
            if query_id:
                with self._running_lock:
                    self._running.pop(query_id, None)

    def reinterrupt_if_cancelled(self, query_id):
        """Watchdog escalation hook: when a STALLED op was already
        cancel-requested, fire the engine interrupt again (a giant native
        fetch can miss the first one). Returns True when re-fired."""
        if query_id not in self._cancelled_runs:
            return False
        with self._running_lock:
            handle = self._running.get(query_id)
        if handle is None:
            return False
        self._interrupt_entry(handle)
        return True

    @staticmethod
    def _interrupt_entry(entry):
        """Interrupt the MOST PRECISE thing a _running entry points at:
        a per-run cursor handle interrupts exactly that statement; a
        {engine, tid} registration interrupts that thread's own-cursor
        native op when one exists, else the engine; a bare engine (or
        anything with .interrupt) is the coarse fallback."""
        if entry is None:
            return
        tid = None
        target = entry
        deferred = False
        if isinstance(entry, dict):
            tid = entry.get("tid")
            deferred = bool(entry.get("deferred"))
            target = entry.get("handle") or entry.get("engine")
        if target is None:
            return
        try:
            ops = getattr(target, "_native_ops", None)
            if tid is not None and ops:
                h = None
                lk = getattr(target, "_native_ops_lock", None)
                if lk is not None:
                    with lk:
                        h = ops.get(tid)
                else:
                    h = ops.get(tid)
                if h is not None:
                    h.interrupt()
                    return
            # A deferred concurrent read has not created its precise cursor
            # yet. Do not use the engine-wide hammer: cancel_query has already
            # recorded the id, and the cursor bind will consume it.
            if deferred:
                return
            target.interrupt()
        except Exception:
            pass

    def cancel_query(self, query_id, scope="run"):
        """Stop in-flight work for ONE run (.411): interrupt exactly the
        run's own statement -- its read cursor, its thread's native-op
        cursor, or its engine as the last resort -- and flag the id so
        cooperative loops unwind in the gaps between statements. With
        five things now running concurrently (.402/.404), the old
        "interrupt every engine" behavior meant one Stop killed them
        all; that global sweep is still available as scope="all" (the
        Stop-everything affordances use it)."""
        with self._running_lock:
            entry = self._running.get(query_id)
            if query_id:
                self._cancelled_runs.add(query_id)
        self._interrupt_entry(entry)
        if scope == "all":
            try:
                self.interrupt_loads()
            except Exception:
                pass
        return {"ok": True, "cancelled": True}

    def flag_run_cancelled(self, query_id):
        """Flag a run id as cancelled so its loop's cooperative check
        (``_run_is_cancelled``, tested between passes) unwinds it -- WITHOUT
        cancel_query's global engine hammer. This is what a *scoped* per-card
        cancel of a run-job (an iterator / API run on the activity tray) calls,
        so stopping one card never aborts another task's engine work; the
        card's own owner-scoped interrupt handles its in-flight statement."""
        if query_id:
            with self._running_lock:
                self._cancelled_runs.add(query_id)

    def _run_is_cancelled(self, query_id):
        """True if Stop was pressed for this run id -- checked by long loops
        between statements so they stop promptly even between engine calls."""
        if not query_id:
            return False
        with self._running_lock:
            return query_id in self._cancelled_runs

    def _clear_stale_engine_cancel(self, eng=None, except_qid=None):
        """Clear sticky engine cancel Events left by a prior Stop / stall cancel.

        ``DuckDBManager.interrupt`` sets ``_cancel`` and never clears it.
        ``_BeatDaemon`` then re-interrupts every subsequent statement while the
        flag stays set -- so a cancelled load made the next right-click Profile
        (and other dashboard reads) auto-cancel. Clear only when no other run
        is still in flight (a live cancel may still need BeatDaemon nudges).
        """
        with self._running_lock:
            others = [k for k in self._running if k != except_qid]
            if others:
                return
        engines = []
        if eng is not None:
            engines.append(eng)
        else:
            engines.extend([self.db, getattr(self, "duckdb", None)])
        for e in engines:
            if e is None:
                continue
            try:
                ev = getattr(e, "_cancel", None)
                if ev is not None and ev.is_set():
                    ev.clear()
            except Exception:
                pass

    def interrupt_loads(self):
        """Best-effort interrupt of any engine work currently in flight, used
        to cancel a running file load. A load runs a single big statement on
        DuckDB (CREATE TABLE AS ... read_csv) or a stream of batched inserts on
        SQLite -- ``interrupt()`` aborts whichever is running. Safe to call when
        nothing is in flight (it's a no-op)."""
        try:
            self.db.interrupt()
        except Exception:
            pass
        if self.duckdb is not None:
            try:
                self.duckdb.interrupt()
            except Exception:
                pass

    def snapshot_table_names(self):
        """The set of all table names across both engines right now -- paired
        with ``drop_tables_created_since`` to undo a cancelled load's partial
        table(s)."""
        names = set(getattr(self.db, "table_columns", {}) or {})
        if self.duckdb is not None:
            names |= set(getattr(self.duckdb, "table_columns", {}) or {})
        return names

    def drop_tables_created_since(self, before):
        """Drop any tables that have appeared since the ``before`` snapshot.
        Used to clean up after a cancelled load -- the streaming SQLite insert
        already rolls back its own partial table, but a DuckDB load (or a
        multi-table JSON load that got part way) may have committed one, so this
        is the catch-all. Returns the names dropped."""
        try:
            after = self.snapshot_table_names()
        except Exception:
            return []
        dropped = []
        for name in (after - set(before or ())):
            in_duck = (self.duckdb is not None
                       and name in getattr(self.duckdb, "table_columns", {}))
            try:
                self.drop_table("duckdb" if in_duck else "sqlite", name)
                dropped.append(name)
            except Exception:
                pass
        return dropped

    def _qualify_catalog_sql(self, sql, conn_name):
        """Rewrite bare references to this connection's catalog tables to their
        fully-qualified [database].[schema].[table] names. A user types the
        bare table name (e.g. ``FROM AggCreditCard``) because that's what the
        sidebar shows, but SQL Server can't resolve it unless it lives in the
        connection's default schema -- so we expand it to the real object name
        before passthrough. Already-qualified names and aliases are left as-is.
        """
        qual = {}
        for key, info in self.catalog_tables.items():
            if info.get("conn") != conn_name:
                continue
            q = info.get("qualified")
            if not q:
                continue
            qual[key.lower()] = q
            tbl = (info.get("table") or "").lower()
            if tbl:
                qual.setdefault(tbl, q)
        if not qual:
            return sql
        pat = re.compile(
            r'(\b(?:from|join)\b\s+)([\[\"]?)([A-Za-z_][\w$]*)([\]\"]?)',
            re.IGNORECASE)

        def repl(m):
            q = qual.get(m.group(3).lower())
            return (m.group(1) + q) if q else m.group(0)

        return pat.sub(repl, sql)

    def _exec_target_inner(self, sql, target, engine, kind,
                           query_id=None, prefer_locked=False):
        # A catalog (SQL Server) passthrough: expand bare table names to their
        # fully-qualified object names so the server can resolve them.
        if kind == "remote":
            sql = self._qualify_catalog_sql(sql, target)
        # Fast, memory-light path for DuckDB reads: spill the result to a
        # temporary Parquet file and page from it columnar-side, instead of
        # converting millions of rows into Python tuples and a row store.
        # ORDER BY stays on this path too: COPY with preserve_insertion_order
        # writes rows in query order so file_row_number paging matches.
        if (kind == "duckdb" and getattr(self, "use_parquet_results", True)
                and self._is_single_read(sql)):
            try:
                return self._exec_duckdb_parquet(
                    engine, sql, query_id=query_id,
                    prefer_locked=prefer_locked)
            except Exception as e:
                # An interrupted COPY is the completed cancellation path.
                # Falling through used to RUN THE SAME LONG QUERY AGAIN via
                # execute_cursor(), leaving Stop apparently hung and sometimes
                # moving the rerun onto the shared main connection.
                if _is_interrupt(e):
                    raise
                if self._run_is_cancelled(query_id):
                    raise InterruptedError("cancelled") from e
                # An out-of-memory failure is a property of the query, not the
                # cursor: re-running it through the generic in-memory drain
                # below would execute the same explode a second time on a
                # HEAVIER (row-materialising) path and hold the connection
                # longer. Surface it now so _query_error attaches the OOM
                # guidance immediately. Other cursor problems (a TEMP table a
                # fresh cursor can't see, a jumbo-option rejection) still fall
                # through to the drain fallback unchanged.
                if _looks_like_oom(err_str(e)):
                    raise
                pass  # non-cancel cursor problem: generic drain fallback
        cols, first, cursor = engine.execute_cursor(
            sql, batch=self._RESULT_BATCH)
        if cols is None:
            return None, None, 0, kind
        if cursor is None:
            return cols, first, len(first), kind
        # Drain the remainder. Keep small/medium results as a fast
        # in-memory list and only spill genuinely large ones into an
        # off-heap store, so typical queries never pay for a temp
        # database while massive ones stay memory-bounded.
        spill_at = 10000 if self.low_memory else 50000
        fetch = self._FETCH_BATCH
        cap = self._result_row_cap()
        out = list(first)
        store = None
        capped = False
        try:
            while True:
                batch = cursor.fetchmany(fetch)
                if not batch:
                    break
                batch = [tuple(r) for r in batch]
                if store is not None:
                    store.extend(batch)
                else:
                    out.extend(batch)
                    if len(out) >= spill_at:
                        store = DiskBackedRows(block=fetch)
                        store.extend(out)
                        out = []
                # Safety ceiling: stop draining a runaway result rather than
                # pulling rows forever (which also holds the connection). The
                # cursor is closed in finally, ending the underlying scan.
                held = len(store) if store is not None else len(out)
                if cap and held >= cap:
                    capped = True
                    break
        finally:
            try:
                cursor.close()
            except Exception:
                pass
        if store is not None:
            store.capped = capped
            store.cap = cap
            return cols, store, len(store), kind
        return cols, out, len(out), kind

    def _query_error(self, sql, target, exc, t0):
        msg = err_str(exc)
        if target == DUCKDB_TARGET:
            msg = _duckdb_oom_query_message(
                msg, getattr(self, "duckdb", None))
        self.history.add(sql, target=target, error=msg,
                         elapsed_sec=round(time.time() - t0, 3))
        return {"error": msg}

    # Cap the rows held resident by in-memory (list) result stores, on top
    # of the count cap, so many medium results can't pile up in RAM.
    def _enforce_memory_budget(self):
        row_budget = 50000 if self.low_memory else 300000
        # also cap *bytes* so a few very wide results can't blow up memory
        # while still being under the row cap
        byte_budget = (128 if self.low_memory else 768) * 1024 * 1024
        # drop ids already released elsewhere (e.g. by an export) so they
        # can't dangle in the order list and KeyError below
        if any(r not in self._results for r in self._results_order):
            self._results_order = [r for r in self._results_order
                                   if r in self._results]
        while len(self._results_order) > 1:
            rows = 0
            nbytes = 0
            for r in self._results_order:
                st = self._results[r].store
                if isinstance(st, list):
                    rows += len(st)
                    if st:
                        nbytes += _estimate_row_bytes(st[0]) * len(st)
            if rows <= row_budget and nbytes <= byte_budget:
                break
            victim = None
            for rid in self._results_order[:-1]:   # never evict the newest
                if (isinstance(self._results[rid].store, list)
                        and not self._reuse_pins.get(rid)):
                    victim = rid
                    break
            if victim is None:
                break
            cr = self._results.pop(victim, None)
            self._results_order.remove(victim)
            if cr is not None:
                cr.close()
        # Parquet result stores live on DISK: the count cap alone lets a few
        # multi-GB explodes pile up tens of GB of qr_ files. Bound their total
        # size (SAMQL_RESULTS_GB, default 20, clamped 1..512), evicting the
        # oldest first -- never the newest, never a pinned reuse source.
        try:
            gb = float(os.environ.get("SAMQL_RESULTS_GB") or 20)
        except Exception:
            gb = 20.0
        if gb <= 0:
            return  # 0 disables the disk budget (same convention as the
                    # result row cap); the count cap still applies
        disk_budget = int(min(gb, 512.0) * (1024 ** 3))

        def _psize(rid):
            st = self._results[rid].store
            if not getattr(st, "_owns_path", True):
                return 0   # borrowed source file: not our disk to budget
            path = getattr(st, "_path", None)
            if not path:
                return 0
            try:
                return os.path.getsize(path)
            except Exception:
                return 0
        while len(self._results_order) > 1:
            total = sum(_psize(r) for r in self._results_order)
            if total <= disk_budget:
                break
            victim = None
            for rid in self._results_order[:-1]:
                if _psize(rid) > 0 and not self._reuse_pins.get(rid):
                    victim = rid
                    break
            if victim is None:
                break
            cr = self._results.pop(victim, None)
            self._results_order.remove(victim)
            if cr is not None:
                cr.close()

    # ---- background housekeeping ------------------------------------
    def _schedule_cleanup(self, full=False):
        if getattr(self, "_closed", False):
            return
        with self._cleanup_lock:
            if full:
                self._cleanup_full = True
            if self._cleanup_timer is not None:
                return
            t = threading.Timer(1.5, self._run_cleanup)
            t.daemon = True
            self._cleanup_timer = t
            t.start()

    def _run_cleanup(self):
        with self._cleanup_lock:
            full = self._cleanup_full
            self._cleanup_full = False
            self._cleanup_timer = None
        try:
            self._reclaim(full=full)
        except Exception:
            pass

    def _reclaim(self, full=False):
        """Keep the engines lean: flush/reclaim the DuckDB on-disk file,
        refresh SQLite planner stats (and reclaim freed pages on a
        disk-backed DB after a drop), then collect garbage and return
        memory to the OS. Runs in the background, debounced."""
        import gc
        if self.duckdb is not None:
            # Don't checkpoint while a query or flow run is using the DuckDB
            # connection. The checkpoint needs the connection lock, so running
            # it under load makes the in-flight (or the very next) run wait on
            # that lock -- a 2-row createtable could sit on "Running ..." behind
            # a slow on-disk checkpoint. Defer to the next idle moment instead;
            # DuckDB auto-checkpoints the WAL regardless, so nothing is lost.
            with self._running_lock:
                busy = bool(self._running)
            if busy:
                self._schedule_cleanup(full=full)
            else:
                try:
                    self.duckdb.checkpoint()
                except Exception:
                    pass
        try:
            with self.db.write_lock:
                self.db.conn.execute("PRAGMA optimize")
                self.db.conn.commit()
        except Exception:
            pass
        if full and self.low_memory:
            try:
                with self.db.write_lock:
                    self.db.conn.execute("VACUUM")
                    self.db.conn.commit()
            except Exception:
                pass
        gc.collect()
        self._malloc_trim()

    def _evict_results_referencing(self, name):
        """Drop cached results whose SQL mentions ``name`` (best effort),
        so a dropped/changed table doesn't leave stale results around."""
        if not name:
            return
        low = str(name).lower()
        with self._lock:
            for rid in list(self._results_order):
                cr = self._results.get(rid)
                if cr is None:
                    continue
                if low in (cr.sql or "").lower():
                    self._results.pop(rid, None)
                    if rid in self._results_order:
                        self._results_order.remove(rid)
                    try:
                        cr.close()
                    except Exception:
                        pass

    def _cache_result(self, cols, store, total, sql, target, engine_kind):
        rid = uuid.uuid4().hex[:12]
        cr = _CachedResult(rid, cols, store, total, sql, target,
                           engine_kind)
        with self._lock:
            self._results[rid] = cr
            self._results_order.append(rid)
            # .448 [PLAN PASS 5] fix: the cap trim evicted PINNED
            # results -- exactly the yank-the-store-under-a-running-
            # chain failure _reuse_pins exists to prevent. Trim the
            # oldest UNPINNED result instead; if everything in excess
            # is pinned, run over-cap until the pins release.
            while len(self._results_order) > self.MAX_CACHED_RESULTS:
                victim_rid = None
                for cand in self._results_order:
                    if not self._reuse_pins.get(cand):
                        victim_rid = cand
                        break
                if victim_rid is None:
                    break  # all pinned; hold over-cap
                self._results_order.remove(victim_rid)
                victim = self._results.pop(victim_rid, None)
                if victim is not None:
                    victim.close()
            self._enforce_memory_budget()
        return rid

    def page(self, result_id, offset=0, limit=DISPLAY_LIMIT,
             sort_col=None, descending=False, filters=None, columns=None,
             query_id=None):
        cr = self._results.get(result_id)
        if cr is None:
            return {"error": "result expired"}
        # .443 audit fix: a garbage offset/limit ("abc") raised
        # ValueError straight through the route as a 500. Coerce
        # safely; nonsense gets a clean error like a bad result id
        # does.
        try:
            offset = max(0, int(offset))
            limit = max(1, min(int(limit), DISPLAY_LIMIT))
        except (TypeError, ValueError):
            return {"error": "offset and limit must be integers"}
        # Register a cancel target for long sort/filter materialize + page
        # fetches so Stop interrupts DuckDB, not only the HTTP send.
        own_run = bool(query_id)
        handle = None
        eng = getattr(getattr(cr, "store", None), "_engine", None)
        if own_run and eng is not None:
            class _PageHandle:
                __slots__ = ("_eng",)

                def __init__(self, e):
                    self._eng = e

                def interrupt(self):
                    try:
                        self._eng.interrupt()
                    except Exception:
                        pass

            handle = _PageHandle(eng)
            self._register_run(query_id, handle, kind="page",
                               target="result page")
        try:
            if own_run and self._run_is_cancelled(query_id):
                return {"cancelled": True}
            if filters:
                view, total = cr.filtered_view(filters, sort_col, descending)
            else:
                view = cr.view(sort_col, descending)
                total = cr.total
            if own_run and self._run_is_cancelled(query_id):
                return {"cancelled": True}
            chunk = view[offset:offset + limit]
            # Optional column projection: return only the requested columns
            # (in the requested order). Keeps the wire payload and the client's
            # row objects small for very wide results. Sorting/filtering still
            # reference the full column set, so this never changes which rows
            # come back, only their width.
            proj = None
            out_cols = cr.cols
            if columns:
                idx = [cr.cols.index(c) for c in columns if c in cr.cols]
                if idx:
                    proj = idx
                    out_cols = [cr.cols[i] for i in idx]
            # .514: CAP FIRST, convert after. json_safe deep-converts every raw
            # cell -- on a flatten-off nested column that is a multi-MB dict per
            # cell, so a 1000-row page burned MINUTES of CPU copying values the
            # cap then threw away (the on-box 93s "Query 1" stall: the op sat in
            # page serialization while cancel and health starved). Truncating the
            # raw values first makes the convert step O(displayed bytes).
            if proj is not None:
                raw = [[r[i] for i in proj] for r in chunk]
            else:
                raw = [list(r) for r in chunk]
            raw, cell_capped = _cap_page_rows(raw)
            rows = [[json_safe(v) for v in r] for r in raw]
            return {
                "columns": out_cols,
                "rows": rows,
                "offset": offset,
                "total_rows": total,
                "filtered": bool(filters),
                "sql": cr.sql,
                "engine": cr.engine,
                "cell_capped": cell_capped,
                "result_capped": bool(cr.capped) and not filters,
                "result_cap": cr.cap if (cr.capped and not filters) else None,
            }
        except Exception as e:
            from .engines import _is_interrupt
            if own_run and (self._run_is_cancelled(query_id)
                            or _is_interrupt(e)):
                return {"cancelled": True}
            raise
        finally:
            if own_run:
                self._end_run_keep_cancel(query_id)

    def shred_preflight(self, engine, table):
        """Every gate between a table and 'Create relational tables',
        with reasons -- so a decline is EXPLAINED, never silent (on-box
        2026-07-02: an empty picker with no why). Pure; creates nothing."""
        from .diagnostics import parse_duckdb_type
        from .shred import plan_shred
        out = {"table": table, "gates": [], "columns": [],
               "candidates": [], "verdict": ""}

        def gate(name, ok, detail):
            out["gates"].append({"name": name, "ok": bool(ok),
                                 "detail": detail})
            return ok
        mgr = self.duckdb if engine == "duckdb" else None
        if not gate("engine", mgr is not None,
                    "Shredding runs on the DuckDB engine."
                    if mgr is None else "DuckDB engine active."):
            out["verdict"] = "Load the file into DuckDB first."
            return out
        src = (getattr(mgr, "table_sources", {}) or {}).get(table)
        parquet_ok = bool(src) and str(src).lower().endswith(".parquet")
        gate("parquet source", parquet_ok,
             ("Backed by " + str(src)) if parquet_ok else
             ("No Parquet backing" + (" (source: %s)" % src if src else "")
              + ". Shred reads the cached Parquet directly. A table loaded "
              "with 'Copy into a table' has no backing -- reload via Load "
              "Data \u2192 File as query-in-place, or turn on 'Flatten "
              "into relational tables' at load time."))
        try:
            raw_types = mgr._column_types_raw(table) if mgr else {}
        except Exception as e:
            raw_types = {}
            gate("column types", False,
                 "Couldn't read column types: " + err_str(e))
        for name, raw in (raw_types or {}).items():
            row = {"name": name, "type": raw or "", "arrays": 0,
                   "reason": ""}
            if not raw:
                row["reason"] = ("type unreadable (DESCRIBE blank and "
                                 "typeof() fallback empty)")
            else:
                try:
                    if raw.count("(") != raw.count(")"):
                        row["reason"] = ("type string looks TRUNCATED "
                                         "(unbalanced parens) -- deeper "
                                         "arrays may be invisible")
                    node = parse_duckdb_type(raw)
                    plan = plan_shred(name, node, base=table)
                    row["arrays"] = len(plan.get("tables") or [])
                    row["tables"] = [t["name"] for t
                                     in (plan.get("tables") or [])]
                    if row["arrays"]:
                        out["candidates"].append(name)
                    elif row["reason"]:
                        pass   # truncation note already explains it
                    else:
                        up = raw.upper()
                        if up.startswith("MAP("):
                            row["reason"] = ("MAP: key/value pairs, not an "
                                             "array of records -- query "
                                             "with map_extract instead")
                        elif "[]" not in raw and "LIST" not in up:
                            row["reason"] = "no nested array levels"
                        else:
                            row["reason"] = ("array of scalars only -- "
                                             "UNNEST directly")
                except Exception as e:
                    row["reason"] = ((row["reason"] + "; ")
                                     if row["reason"] else "") + \
                        "type didn't parse: " + err_str(e)
            out["columns"].append(row)
        gate("nested array columns", bool(out["candidates"]),
             (", ".join(out["candidates"]) + " can shred")
             if out["candidates"] else
             "No column has an array-of-records level.")
        if not parquet_ok:
            out["verdict"] = ("Blocked by the Parquet-source gate -- see "
                              "its detail for the fix.")
        elif not out["candidates"]:
            out["verdict"] = ("No shreddable columns; see per-column "
                              "reasons.")
        else:
            planned = sum(c.get("arrays", 0) for c in out["columns"]
                          if c["name"] in out["candidates"])
            out["verdict"] = ("Ready: %s -> %d table%s planned "
                              "(incl. the joinkeys hub)."
                              % (", ".join(out["candidates"]), planned,
                                 "" if planned == 1 else "s"))
        return out

    def list_functions(self):
        """Every SQL function the two engines expose RIGHT NOW -- live
        from PRAGMA function_list (SQLite, including the SamQL-registered
        regexp helpers) and duckdb_functions() (DuckDB), cached for the
        session since the sets are static. Feeds the Documentation
        modal's SQL-functions tab (.433)."""
        cached = getattr(self, "_fn_docs_cache", None)
        if cached is not None:
            return cached
        sq = []
        try:
            seen = {}
            for row in self.db.conn.execute(
                    "PRAGMA function_list").fetchall():
                name = str(row[0])
                narg = row[2] if len(row) > 2 else None
                prev = seen.get(name)
                if prev is None:
                    seen[name] = {"name": name, "args": narg}
                elif prev.get("args") != narg:
                    prev["args"] = -1  # overloaded -> variadic-ish
            sq = sorted(seen.values(), key=lambda d: d["name"])
        except Exception:
            sq = []
        dk = []
        note = None
        # .473: the DuckDB manager is created lazily on first use --
        # opening the docs before touching DuckDB left self.duckdb None,
        # so the function list came back empty AND GOT CACHED, so it
        # stayed empty for the whole session (the reported "no DuckDB
        # functions"). Spin the engine up here (it's cheap and the user
        # is explicitly asking to see its functions).
        mgr = self.duckdb
        if mgr is None and HAS_DUCKDB:
            try:
                mgr = self.get_duckdb()
            except Exception:
                mgr = None
        if mgr is None:
            note = ("DuckDB engine not available in this session."
                    if not HAS_DUCKDB
                    else "DuckDB engine could not be started.")
        else:
            try:
                with mgr.native_op_cursor() as (xc, lk):
                    with lk:
                        xc.execute(
                            "SELECT DISTINCT function_name, function_type "
                            "FROM duckdb_functions() ORDER BY 1")
                        rows = xc.fetchall()
                dk = [{"name": str(n), "type": str(t)} for n, t in rows]
            except Exception as e:
                note = "DuckDB function list failed: %s" % err_str(e)
        out = {"sqlite": sq, "duckdb": dk,
               "counts": {"sqlite": len(sq), "duckdb": len(dk)}}
        if note:
            out["note"] = note
        # .473: only cache a COMPLETE answer -- never let an empty or
        # errored DuckDB list stick for the rest of the session.
        if dk and sq:
            self._fn_docs_cache = out
        return out

    def _parquet_rowcount(self, mgr, src_parquet):
        """Top-level row count of the parquet backing (metadata-fast) --
        the .430 OOM-batching fallback slices file_row_number by it."""
        try:
            fwd = src_parquet.replace("\\", "/").replace("'", "''")
            with mgr.native_op_cursor() as (xc, lk):
                with lk:
                    xc.execute("SELECT count(*) FROM read_parquet('%s')"
                               % fwd)
                    row = xc.fetchone()
            return int(row[0]) if row else 0
        except Exception:
            return 0

    def shred_plan(self, engine, table, column, base=None):
        """The relational-shred plan for a nested column: one table per array
        level with deterministic keys. Pure -- nothing is created.

        Opaque JSON / JSON[] (flatten-off) columns have no DESCRIBE LIST type;
        sample live cells and plan from the inferred shape, marking the plan
        for JSON-access CTAS over Parquet."""
        from .diagnostics import (parse_duckdb_type, json_samples_to_type_node,
                                  column_needs_json_shred)
        from .shred import plan_shred
        from .sqlutil import quote_ident as _qid
        mgr = self.duckdb if engine == "duckdb" else None
        if mgr is None:
            return {"error": "Shredding runs on the DuckDB engine."}
        src = (getattr(mgr, "table_sources", {}) or {}).get(table)
        needs_cache = not (src and str(src).lower().endswith(".parquet"))
        # .471: a JSON (or API-spool, or CTAS) table has no parquet
        # backing -- flatten-on-load hit this refusal the moment a
        # nested JSON was loaded with flatten toggled. Any live object
        # can shred now: shred_run first materializes a one-shot
        # parquet cache FROM THE LOADED OBJECT ITSELF (so the loader's
        # exact column typing rides along and file_row_number gives
        # the stable _rid), then runs the unchanged pipeline on it.
        # Parquet-backed tables keep the zero-copy fast path.
        try:
            raw = mgr._column_types_raw(table).get(column, "")
            node = parse_duckdb_type(raw)
        except Exception as e:
            return {"error": "Couldn't read the column's type: " + err_str(e)}
        json_access = False
        root_is_json_list = False
        # Opaque / shallow DESCRIBE: infer LIST/STRUCT from sampled cells.
        if column_needs_json_shred(raw, node) and engine == "duckdb":
            try:
                sql = ('SELECT %s FROM %s LIMIT %d'
                       % (_qid(column), _qid(table),
                          int(Session._FIELD_TREE_SAMPLE_ROWS)))
                reader = getattr(mgr, "read", None) or getattr(
                    mgr, "execute_read", None)
                if reader is not None:
                    _c, rows = reader(sql)
                else:
                    _c, rows = mgr.execute(sql)
                values = [r[0] for r in (rows or []) if r]
                syn = json_samples_to_type_node(values) if values else None
                if syn and syn.get("t") == "list":
                    node = syn
                    json_access = True
                    tu = (raw or "").strip().upper()
                    root_is_json_list = (tu == "JSON[]"
                                         or tu.replace(" ", "") == "JSON[]")
            except Exception:
                pass
        plan = plan_shred(column, node,
                          base=(base or table or "rec"))
        plan["source"] = src
        plan["needs_cache"] = needs_cache
        plan["json_access"] = json_access
        plan["root_is_json_list"] = root_is_json_list
        plan["raw_type"] = raw
        return plan


    @staticmethod
    def _repair_json_prefix(raw):
        """.521: a drag-dropped file is sniffed from a PREFIX upload, and a
        truncated JSON array / object stream is invalid as-is. Keep only
        COMPLETE records: ndjson cuts at the last newline; an array (or a
        concatenated object stream) is balance-scanned (string/escape aware)
        and cut at the last object that closed at the top level, then
        re-wrapped as a valid array. Pure; unit-tested."""
        try:
            txt = raw.decode("utf-8", "ignore")
        except Exception:
            return raw
        i = 0
        n = len(txt)
        while i < n and txt[i] in " \t\r\n\ufeff":
            i += 1
        if i >= n:
            return b"[]"
        head = txt[i]
        if head not in "[{":
            # ndjson / jsonl: whole lines only
            cut = txt.rfind("\n")
            return (txt[:cut] if cut > 0 else txt).encode("utf-8")
        body = txt[i + 1:] if head == "[" else txt[i:]
        depth = 0
        in_str = False
        esc = False
        last_end = -1
        rec_start = -1
        recs = []
        for j, ch in enumerate(body):
            if esc:
                esc = False
                continue
            if ch == "\\":
                esc = in_str
                continue
            if ch == '"':
                in_str = not in_str
                continue
            if in_str:
                continue
            if ch in "[{":
                if depth == 0 and ch == "{":
                    rec_start = j
                depth += 1
            elif ch in "]}":
                depth -= 1
                if depth == 0 and ch == "}":
                    last_end = j
                    if rec_start >= 0:
                        recs.append(body[rec_start:j + 1])
                elif depth < 0:
                    break
        if last_end < 0 and head == "{" and not recs:
            # .526 on-box: a WRAPPER OBJECT around one giant array
            # ({"json": [ ... ]}) never closes inside the prefix, so no
            # record ever ended at depth 0. Find the first '[' (string-
            # aware), keep the wrapper text up to and including it, repair
            # the ARRAY BODY after it, and close the wrapper braces -- the
            # sniff then sees the true promoted shape.
            in_str = esc = False
            depth_o = 0
            bi = -1
            for j, ch in enumerate(body):
                if esc:
                    esc = False
                    continue
                if ch == "\\":
                    esc = in_str
                    continue
                if ch == '"':
                    in_str = not in_str
                    continue
                if in_str:
                    continue
                if ch == "{":
                    depth_o += 1
                elif ch == "}":
                    depth_o -= 1
                elif ch == "[":
                    bi = j
                    break
            if bi >= 0:
                arr = body[bi + 1:]
                in_str = esc = False
                depth = 0
                le2 = -1
                for j, ch in enumerate(arr):
                    if esc:
                        esc = False
                        continue
                    if ch == "\\":
                        esc = in_str
                        continue
                    if ch == '"':
                        in_str = not in_str
                        continue
                    if in_str:
                        continue
                    if ch in "[{":
                        depth += 1
                    elif ch in "]}":
                        depth -= 1
                        if depth == 0 and ch == "}":
                            le2 = j
                        elif depth < 0:
                            break
                if le2 >= 0:
                    return (body[:bi + 1] + arr[:le2 + 1] + "]"
                            + "}" * max(depth_o, 1)).encode("utf-8")
            return b"[]"
        if last_end < 0:
            return b"[]"
        if head == "[":
            # commas are already in place inside an array slice
            return ("[" + body[:last_end + 1] + "]").encode("utf-8")
        # object STREAM / ndjson: the separators were newlines (or nothing)
        # -- rebuild as a strict comma-joined array of the complete records
        return ("[" + ",".join(recs) + "]").encode("utf-8")

    def sniff_root_sample(self, raw, name="sample.json"):
        """.521: sniff from an uploaded prefix (drag-drop). Repairs the
        truncation, writes a throwaway temp file, and delegates to
        sniff_root_candidates."""
        fixed = self._repair_json_prefix(raw or b"")
        ext = ".jsonl" if str(name).lower().endswith((".jsonl", ".ndjson")) \
            else ".json"
        tmp = tmputil.new_tempfile("sniff_", ext)
        try:
            with open(tmp, "wb") as fh:
                fh.write(fixed)
            return self.sniff_root_candidates(tmp)
        finally:
            try:
                os.remove(tmp)
            except OSError:
                pass

    def sniff_root_candidates(self, path):
        """.521: pre-load schema sniff for the Load modal's optional
        unique-identifier dropdown. DESCRIBEs a small read_json_auto sample,
        derives the root_id candidates (1:1 paths, first-element list paths,
        and MAP columns), then fills each map candidate's LIVE key names
        from the sample. Read-only; never loads anything."""
        # .525 on-box: self.duckdb is LAZY (None until first use) -- a
        # sniff on a fresh launch said "DuckDB is not available". Use the
        # lazy getter, which builds the engine exactly like any first op.
        try:
            mgr = self.get_duckdb()
        except Exception:
            mgr = None
        if mgr is None:
            return {"error": "DuckDB is not available."}
        try:
            from .diagnostics import parse_duckdb_type as _P
            from . import shred as _shred
            fp = sqlutil.sql_path(path)
            # .525 on-box: sql_path returns the ESCAPED PATH WITHOUT
            # quotes (callers wrap) -- bare C:\... parsed as C then ':'
            # ("syntax error at or near ':'"). Quote it like every other
            # call site.
            src = "read_json_auto('%s', sample_size = 2048)" % fp
            _, rows = mgr.execute("DESCRIBE SELECT * FROM %s" % src)
            columns = [(r[0], _P(str(r[1]))) for r in rows]
            promote_col = None
            if len(columns) == 1 and (columns[0][1] or {}).get("t") == "list" \
                    and ((columns[0][1].get("of") or {}).get("t")
                         == "struct"):
                promote_col = columns[0][0]
            cands = _shred.root_id_candidates(columns,
                                              promote_col=promote_col)
            _qq = _shred._qq

            def _frm_and_alias(in_list):
                if promote_col is not None:
                    frm = ("(SELECT unnest(t0.%s) AS e0 FROM %s AS t0) AS s0"
                           % (_qq(promote_col), src))
                    alias = "e0"
                else:
                    frm = "%s AS t0" % src
                    alias = "t0"
                if in_list:
                    lp = ".".join(_qq(x) for x in in_list)
                    frm = ("(SELECT unnest(%s.%s) AS eL FROM %s) AS s1"
                           % (alias, lp, frm))
                    alias = "eL"
                return frm, alias

            for c in cands:
                if not c.get("map"):
                    continue
                keys = []
                try:
                    frm, alias = _frm_and_alias(c.get("in_list"))
                    mexpr = "%s.%s" % (alias,
                                       ".".join(_qq(x) for x in c["steps"]))
                    q = ("SELECT DISTINCT k FROM (SELECT unnest(map_keys("
                         "%s)) AS k FROM %s WHERE %s IS NOT NULL) LIMIT 40"
                         % (mexpr, frm, mexpr))
                    _, krows = mgr.execute(q)
                    keys = sorted({str(r[0]) for r in krows
                                   if r and r[0] is not None})
                except Exception:
                    keys = []
                c["keys"] = keys
            return {"ok": True, "promote": promote_col is not None,
                    "candidates": cands}
        except Exception as e:
            return {"error": "sniff failed: %s" % str(e)[:300]}

    def _flatten_table_columns(self, mgr, table):
        """Parse + sample-enrich columns for flatten / root_id (same as flatten_table)."""
        from .diagnostics import (parse_duckdb_type as _P,
                                  json_samples_to_type_node,
                                  column_needs_json_shred)
        from .sqlutil import quote_ident as _qid
        try:
            raw = mgr._column_types_raw(table)
        except Exception as e:
            return None, None, {"error": "Couldn't read the table's columns: "
                                + err_str(e)}
        columns = []
        json_access_cols = set()
        for name, typ in raw.items():
            node = _P(typ)
            if column_needs_json_shred(typ, node):
                try:
                    sql = ('SELECT %s FROM %s LIMIT %d'
                           % (_qid(name), _qid(table),
                              int(Session._FIELD_TREE_SAMPLE_ROWS)))
                    reader = getattr(mgr, "read", None) or getattr(
                        mgr, "execute_read", None)
                    if reader is not None:
                        _c, rows = reader(sql)
                    else:
                        _c, rows = mgr.execute(sql)
                    values = [r[0] for r in (rows or []) if r]
                    syn = json_samples_to_type_node(values) if values else None
                    if syn and syn.get("t") in ("list", "struct"):
                        node = syn
                        json_access_cols.add(name)
                except Exception:
                    pass
            columns.append((name, node))
        return columns, json_access_cols, None

    @staticmethod
    def _flatten_path_sql_exprs(path_parts):
        """SQL expressions that may address a Field Explorer nest path.

        Tries STRUCT/LIST dotted access first, then JSON ``->`` / ``::JSON``
        extracts for opaque JSON columns (flatten-off depth caps).
        """
        from .diagnostics import _json_path_seg
        from .sqlutil import quote_ident as _qid
        if not path_parts:
            return []
        out = [".".join(_qid(p) for p in path_parts)]
        if len(path_parts) >= 2:
            col = _qid(path_parts[0])
            jptr = "$" + "".join(
                _json_path_seg(p) for p in path_parts[1:])
            out.append("%s -> '%s'" % (col, jptr))
            out.append("%s::JSON -> '%s'" % (col, jptr))
        return out

    def _sample_enrich_flatten_path(self, mgr, table, path_parts):
        """Sample a Field Explorer nest path (often opaque JSON under a
        depth-capped STRUCT) into a real list/struct type node.

        Returns (synth_name, node, json_access, sql_expr) or None when the
        path cannot be enriched into a flattenable nest.
        """
        from .diagnostics import json_samples_to_type_node
        from .sqlutil import quote_ident as _qid
        if not path_parts:
            return None
        reader = getattr(mgr, "read", None) or getattr(
            mgr, "execute_read", None)
        stem = "_".join(
            re.sub(r"[^A-Za-z0-9_]+", "_", p).strip("_") or "x"
            for p in path_parts)
        synth = (stem + "_nest") if stem else "fx_nest"
        for acc in self._flatten_path_sql_exprs(path_parts):
            try:
                sql = ('SELECT %s FROM %s LIMIT %d'
                       % (acc, _qid(table),
                          int(Session._FIELD_TREE_SAMPLE_ROWS)))
                if reader is not None:
                    _c, rows = reader(sql)
                else:
                    _c, rows = mgr.execute(sql)
                values = [r[0] for r in (rows or [])
                          if r and r[0] is not None]
                if not values:
                    continue
                syn = json_samples_to_type_node(values)
                if not syn or syn.get("t") not in ("list", "struct"):
                    continue
                # Depth-capped loads store nested arrays as JSON text/cells.
                use_json = True
                try:
                    _, tr = mgr.execute(
                        "SELECT typeof(%s) FROM %s LIMIT 1"
                        % (acc, _qid(table)))
                    tu = (str(tr[0][0]) if tr and tr[0] else "").upper()
                    use_json = ("JSON" in tu) or tu in ("VARCHAR", "CHAR")
                except Exception:
                    use_json = True
                return synth, syn, use_json, acc
            except Exception:
                continue
        return None

    def _flatten_promote_col(self, columns):
        """Single wrapping LIST(STRUCT) column to promote, else None."""
        _list_specs = []
        for name, node in columns:
            if (node or {}).get("t") == "list":
                _list_specs.append({"list_path": [name], "node": node})
        promote = (len(_list_specs) == 1
                   and ((_list_specs[0]["node"].get("of") or {}).get("t")
                        == "struct"))
        if not promote:
            return None
        lp = _list_specs[0]["list_path"]
        return lp[0] if len(lp) == 1 else None

    def table_root_id_options(self, engine, table):
        """Unique-identifier candidates for Field Explorer flatten picker.

        Same schema walk as load-time sniff, against the loaded table (with
        opaque-JSON sample enrichment). Map candidates include live keys.
        """
        from . import shred as _shred
        from .sqlutil import quote_ident as _qid
        mgr = self.duckdb if engine == "duckdb" else self.db
        if mgr is None:
            return {"error": "Engine not available."}
        columns, _ja, err = self._flatten_table_columns(mgr, table)
        if err:
            return err
        promote_col = self._flatten_promote_col(columns)
        cands = _shred.root_id_candidates(columns, promote_col=promote_col)
        # Fill map keys from a small sample (same idea as sniff_root_candidates).
        for c in cands:
            if not c.get("map"):
                continue
            keys = []
            try:
                if promote_col is not None:
                    frm = ("(SELECT unnest(%s) AS e0 FROM %s) AS s0"
                           % (_qid(promote_col), _qid(table)))
                    alias = "e0"
                else:
                    frm = "%s AS t0" % _qid(table)
                    alias = "t0"
                in_list = c.get("in_list")
                if in_list:
                    lp = ".".join(_qid(x) for x in in_list)
                    frm = ("(SELECT unnest(%s.%s) AS eL FROM %s) AS s1"
                           % (alias, lp, frm))
                    alias = "eL"
                mexpr = "%s.%s" % (alias,
                                   ".".join(_qid(x) for x in c["steps"]))
                q = ("SELECT DISTINCT k FROM (SELECT unnest(map_keys("
                     "%s)) AS k FROM %s WHERE %s IS NOT NULL) LIMIT 40"
                     % (mexpr, frm, mexpr))
                _, krows = mgr.execute(q)
                keys = sorted({str(r[0]) for r in krows
                               if r and r[0] is not None})
            except Exception:
                keys = []
            c["keys"] = keys
        return {"ok": True, "candidates": cands,
                "promote": promote_col is not None,
                "promote_col": promote_col}

    def table_root_id_stats(self, engine, table, root_id):
        """Uniqueness probe for a chosen root_id on a loaded table.

        Returns counts so the Field Explorer picker can warn on duplicates
        before Confirm Flatten.
        """
        from . import shred as _shred
        from .sqlutil import quote_ident as _qid
        mgr = self.duckdb if engine == "duckdb" else self.db
        if mgr is None:
            return {"error": "Engine not available."}
        columns, _ja, err = self._flatten_table_columns(mgr, table)
        if err:
            return err
        promote_col = self._flatten_promote_col(columns)
        cand = _shred.validate_root_choice(
            columns, promote_col, root_id)
        if cand is None:
            return {"error": "That field was not found in this table's schema."}
        mk = cand.get("map_key")
        if promote_col is not None:
            alias = "e0"
            expr = _shred.render_root_expr(cand, alias, map_key=mk)
            frm = ("(SELECT unnest(%s) AS e0 FROM %s) AS s0"
                   % (_qid(promote_col), _qid(table)))
        else:
            alias = "t0"
            expr = _shred.render_root_expr(cand, alias, map_key=mk)
            frm = "%s AS t0" % _qid(table)
        try:
            _, rows = mgr.execute(
                "SELECT COUNT(*), COUNT(%s), COUNT(DISTINCT %s) FROM %s"
                % (expr, expr, frm))
        except Exception as e:
            return {"error": "Couldn't check uniqueness: " + err_str(e)}
        if not rows:
            return {"error": "Couldn't check uniqueness."}
        records, nonnull, distinct = (int(rows[0][0]), int(rows[0][1]),
                                      int(rows[0][2]))
        duplicated = max(0, nonnull - distinct)
        nulls = max(0, records - nonnull)
        unique = (records > 0 and nulls == 0 and duplicated == 0
                  and distinct == records)
        return {
            "ok": True,
            "label": cand.get("label"),
            "records": records,
            "nonnull": nonnull,
            "distinct": distinct,
            "duplicated": duplicated,
            "nulls": nulls,
            "unique": unique,
            "choice": {
                "steps": cand.get("steps"),
                "in_list": cand.get("in_list"),
                "map": bool(cand.get("map")),
                "map_key": cand.get("map_key"),
                "label": cand.get("label"),
            },
        }

    # Field Explorer opaque-JSON paths look like:
    #   json ->> '$._embedded.items'   or   _embedded -> '$.items'
    # Naively splitting on "." mangles these into garbage STRUCT parts and
    # walk-up then enriches the whole column (sibling HAL _links → href/rel).
    _FE_JSON_EXTRACT_PATH_RE = re.compile(
        r'^((?:"[^"]+")|(?:[A-Za-z_][A-Za-z0-9_]*))'
        r'(?:\s*::\s*JSON)?'
        r'\s*(?:->>|->)\s*'
        r"'(\$[^\']*)'"
        r'\s*$',
        re.IGNORECASE,
    )

    @staticmethod
    def _json_ptr_keys(jptr):
        """Split a DuckDB JSON pointer (``$._embedded.items``) into keys."""
        s = (jptr or "").strip()
        if not s or s == "$":
            return []
        if s.startswith("$"):
            s = s[1:]
        keys = []
        buf = []
        in_q = False
        i = 0
        while i < len(s):
            ch = s[i]
            if ch == '"':
                in_q = not in_q
                buf.append(ch)
                i += 1
                continue
            if ch == "." and not in_q:
                if buf:
                    keys.append("".join(buf))
                    buf = []
                i += 1
                continue
            if ch == "[" and not in_q:
                # Keep ``name[0]`` together; start a new key at bare ``[``.
                if buf:
                    keys.append("".join(buf))
                    buf = []
                j = s.find("]", i)
                if j < 0:
                    buf.append(s[i:])
                    break
                keys.append(s[i:j + 1])
                i = j + 1
                continue
            buf.append(ch)
            i += 1
        if buf:
            keys.append("".join(buf))
        out = []
        for k in keys:
            if not k:
                continue
            if len(k) >= 2 and k[0] == '"' and k[-1] == '"':
                out.append(k[1:-1].replace('""', '"'))
            else:
                # ``items[0]`` → key ``items`` (index is display-only here)
                m = re.match(r'^([A-Za-z_][A-Za-z0-9_]*)(?:\[\d+\])?$', k)
                out.append(m.group(1) if m else k)
        return out

    @staticmethod
    def _flatten_path_parts(path, column=None):
        """Normalize a Field Explorer field path into column-rooted parts.

        Accepts STRUCT dotted paths (``_embedded.items``) and opaque JSON
        extract paths (``json ->> '$._embedded.items'``).
        """
        if path is None or path is False:
            return None
        if isinstance(path, (list, tuple)):
            parts = [str(p) for p in path if p not in (None, "")]
        else:
            s = str(path).strip()
            m = Session._FE_JSON_EXTRACT_PATH_RE.match(s)
            if m:
                col = m.group(1)
                if len(col) >= 2 and col[0] == '"' and col[-1] == '"':
                    col = col[1:-1].replace('""', '"')
                parts = [col] + Session._json_ptr_keys(m.group(2))
            else:
                parts = [p for p in s.replace("/", ".").split(".")
                         if p]
        # Field-tree (element) wrappers are display-only; never STRUCT keys.
        parts = [p for p in parts if p and p != "(element)"]
        if not parts:
            return None
        if column and parts[0] != column:
            parts = [column] + parts
        return parts

    @staticmethod
    def _flatten_keep_source_base(path_parts, column, table):
        """Family base name for Field Explorer keep_source flatten.

        The hub is named ``<stem>_flattened`` by the planner; this returns
        the stem only (no ``_flat`` / ``_joinkeys`` suffix).
        """
        if path_parts:
            stem = "_".join(
                re.sub(r"[^A-Za-z0-9_]+", "_", p).strip("_") or "x"
                for p in path_parts)
        elif column:
            stem = (re.sub(r"[^A-Za-z0-9_]+", "_", column).strip("_")
                    or "col")
        else:
            stem = (re.sub(r"[^A-Za-z0-9_]+", "_", table or "tbl").strip("_")
                    or "tbl")
        return stem

    @staticmethod
    def _filter_flatten_plan_for_path(plan, path_parts):
        """Keep hub + only list/map/jsonside tables for ``path_parts``."""
        if not path_parts:
            return plan
        kept = []
        for t in plan.get("tables") or []:
            kind = t.get("kind")
            if kind in ("base", "joinkeys"):
                kept.append(t)
                continue
            if kind == "list":
                lp = list(t.get("list_path") or [])
                if (len(lp) >= len(path_parts)
                        and lp[:len(path_parts)] == path_parts):
                    kept.append(t)
                continue
            if kind in ("map", "jsonside"):
                col = t.get("col") or t.get("column")
                if col and path_parts == [col]:
                    kept.append(t)
        return dict(plan, tables=kept)

    def flatten_table(self, table, base=None, query_id=None, root_id=None,
                      keep_source=False, column=None, path=None):
        """.474: the Load-modal "flatten" toggle, done at TABLE level.
        Turns one nested table into the relational set the model wants:
        a base table of every top-level scalar PLUS single-struct
        columns flattened inline; a <base>_flattened records hub of _rid
        + inlined scalars; and one <list>_flattened table per top-level
        list, exploded with the element's fields inlined.

        Reads through the loaded table's Parquet cache so file_row_number
        gives a stable _rid (materializing a one-shot cache first if the
        table isn't already parquet-backed -- e.g. a real CTAS table).
        Returns {ok, created:[{name,rows}], base} or {error}/{cancelled}.

        Field Explorer passes ``keep_source=True`` plus optional ``column`` /
        ``path`` so the original nested table stays loaded and only the
        selected nest produces a relational family (+ Master_Keys and
        Join_Keys when a unique identifier is chosen). Load-time flatten
        keeps the default (replace source under the file name).
        """
        from . import shred as _shred
        from .engines import _is_interrupt as _is_intr
        mgr = self.duckdb
        if mgr is None:
            return {"error": "Flatten runs on the DuckDB engine."}
        # Nested flatten/shred CTAS needs a generous engine ceiling. Adaptive
        # budget sync can otherwise leave a crushed limit after a large load.
        try:
            from .engines import ensure_heavy_op_engine_memory
            ensure_heavy_op_engine_memory(mgr)
        except Exception:
            pass
        path_parts = self._flatten_path_parts(path, column)
        if column is not None:
            column = str(column)
            if not column:
                column = None
        # Field Explorer: family base is path/column-derived so it never
        # steals the source table name; load-time still uses base=table.
        # May be recomputed after path walk-up finds a shorter enrichable nest.
        if keep_source and (base is None or base == table):
            base = self._flatten_keep_source_base(path_parts, column, table)
        base = base or table or "rec"
        columns, json_access_cols, err = self._flatten_table_columns(mgr, table)
        if err:
            return err
        # root_id is validated against the FULL table schema (any level);
        # planning may be narrowed to the selected column/path.
        full_columns = columns
        # True when a list element's type contains an array at any struct
        # depth (needs compound-key deep shred, not one-level shallow).
        def _elem_has_nested_array(list_node):
            stack = [(list_node or {}).get("of") or {}]
            while stack:
                n = stack.pop()
                t = (n or {}).get("t")
                if t == "list":
                    return True
                if t == "struct":
                    for _fn, _fnode in (n.get("fields") or []):
                        stack.append(_fnode)
            return False

        # Field Explorer path scope: depth-capped flatten-off JSON often
        # stores the nest as opaque JSON under a STRUCT. Project that nest
        # into the flatten cache as a synthetic top-level column and plan
        # only that nest — source table untouched.
        _path_proj = None  # (sql_path, synth_name)
        # Extra cache projections: deep list fields peeled out of a
        # path-projected STRUCT so plan_shred owns their full subtree
        # (e.g. terms.legs → cashflows / adjustments / history).
        _path_peels = []  # [(sql_expr, peel_name), ...]
        if path_parts:
            # Walk up prefixes when a leaf path is not STRUCT-addressable
            # (FE (element) children often have path=null → "phones" becomes
            # counterparty.phones). Prefer the longest enrichable nest so
            # contacts stays a child table, not residual JSON on joinkeys.
            # Opaque FE paths (``json ->> '$.…'``) are normalized above; do
            # not settle for the bare column when a deeper nest was requested
            # — that pulls in sibling HAL ``_links`` (href/rel) tables.
            _enriched = None
            _enrich_parts = None
            _path_sql = None
            _orig_path_parts = list(path_parts)
            _min_end = 1
            if len(_orig_path_parts) > 1:
                # Keep at least the first nest key under the column when the
                # caller asked for a nested path (items, not whole json).
                _min_end = 2
            for _end in range(len(path_parts), _min_end - 1, -1):
                _try = path_parts[:_end]
                _got = self._sample_enrich_flatten_path(mgr, table, _try)
                if _got is not None:
                    _enriched = _got
                    _enrich_parts = _try
                    break
            # Last resort: bare column (phones → counterparty walk-up).
            if (_enriched is None and len(_orig_path_parts) > 1
                    and _min_end > 1):
                _try = path_parts[:1]
                _got = self._sample_enrich_flatten_path(mgr, table, _try)
                if _got is not None:
                    _enriched = _got
                    _enrich_parts = _try
            if _enriched is not None and _enrich_parts is not None:
                if (keep_source and _enrich_parts != path_parts):
                    base = self._flatten_keep_source_base(
                        _enrich_parts, column, table)
                path_parts = _enrich_parts
                _synth, _snode, _sjson, _path_sql = _enriched
                from .sqlutil import quote_ident as _qid_path
                if not _path_sql:
                    _path_sql = ".".join(_qid_path(p) for p in path_parts)
                _path_proj = (_path_sql, _synth)
                columns = [(_synth, _snode)]
                json_access_cols = {_synth} if _sjson else set()
                # Struct nests: shallow plan_flatten only emits ONE
                # compound-key child level under a list-in-struct, so
                # deeper arrays-of-objects are dropped (notes only). Peel
                # each nested-array-bearing list field into a top-level
                # cache column and route it through deep shred — same
                # posture as flattening that list path directly.
                if (_snode or {}).get("t") == "struct":
                    _kept_fields = []
                    _peeled_cols = []
                    for _fname, _fnode in (_snode.get("fields") or []):
                        if ((_fnode or {}).get("t") == "list"
                                and _elem_has_nested_array(_fnode)):
                            _peel = (
                                _synth + "_"
                                + (re.sub(r"[^A-Za-z0-9_]+", "_",
                                          str(_fname)).strip("_")
                                   or "list"))
                            # Peel off the projected nest expression.
                            if "->" in _path_sql or "::" in _path_sql:
                                from .diagnostics import _json_path_seg
                                _peel_sql = (
                                    "json_extract(%s, '$%s')"
                                    % (_path_sql,
                                       _json_path_seg(_fname)))
                            else:
                                _peel_sql = "%s.%s" % (
                                    _path_sql, _qid_path(_fname))
                            _path_peels.append((_peel_sql, _peel))
                            _peeled_cols.append((_peel, _fnode))
                            if _sjson:
                                json_access_cols.add(_peel)
                        else:
                            _kept_fields.append((_fname, _fnode))
                    if _peeled_cols:
                        _snode = dict(_snode, fields=_kept_fields)
                        columns = [(_synth, _snode)] + _peeled_cols
                # Already scoped via projection; do not path-filter the plan.
                path_parts = None
            elif column is not None:
                plan_columns = [(n, nd) for n, nd in columns if n == column]
                if not plan_columns:
                    return {"error": "column %r was not found on table %r"
                                     % (column, table)}
                columns = plan_columns
                json_access_cols = {c for c in json_access_cols
                                    if c == column}
        elif column is not None:
            plan_columns = [(n, nd) for n, nd in columns if n == column]
            if not plan_columns:
                return {"error": "column %r was not found on table %r"
                                 % (column, table)}
            columns = plan_columns
            json_access_cols = {c for c in json_access_cols if c == column}
        # .501: uniquify planned names against the live catalog (minus this
        # base, which re-flatten replaces) so path-only child names ("legs",
        # "json") from two different files can't clobber each other.
        _reserved = set(getattr(mgr, "table_columns", {}) or {})
        # a RE-flatten replaces its OWN family: free those names so the same
        # tables are rebuilt in place instead of suffixing "_2" copies.
        _reserved -= self._family_members(base)
        # .518: the family anchors at the HUB now -- a re-flatten must free
        # the hub-rooted members (and the hub name itself) so a reload lands
        # on the SAME names instead of suffixing "_2" copies.
        _reserved -= self._family_members(base + "_flattened")
        _reserved.discard(base + "_flattened")
        # Also free a legacy *_joinkeys family from older builds so a
        # re-flatten replaces instead of suffixing _2 copies.
        _reserved -= self._family_members(base + "_joinkeys")
        _reserved.discard(base + "_joinkeys")
        # Keep-source: the original nested table must remain reserved so the
        # family never lands on that catalog name.
        if keep_source and table:
            _reserved.add(table)
        plan_full = _shred.plan_flatten_table(columns, base=base,
                                              reserved=_reserved)

        # .494 (on-box "18 tables planned but only 2 came out"): a top-level
        # LIST whose ELEMENTS carry further nested arrays -- an array of
        # records where each record has its own arrays (json[].SwapHistory[],
        # json[].cashFlow.dateAdjustments..., businessCenters VARCHAR[]) --
        # needs a DEEP breakout with compound keys. The shallow flatten unnests
        # ONE level, inlines the element scalars, and leaves the inner arrays
        # behind (the giant residual json/map columns in the picker). Route
        # those columns through the deep shred (plan_shred + build_table_sql):
        # the SAME planner the preflight counts, over the memory-lean .429
        # unnest pipeline, so the full hierarchy actually gets built. Simple
        # top-level lists and the base table keep the light shallow flatten.
        col_nodes = dict(columns)
        deep_cols = [name for name, node in columns
                     if (node or {}).get("t") == "list"
                     and _elem_has_nested_array(node)]
        # Path scope may drop the top-level shallow list from the plan; keep
        # deep naming / eligibility from the unfiltered plan.
        shallow_list_name = {}
        shallow_list_ord = {}
        for _s in plan_full["tables"]:
            if _s.get("kind") == "list":
                _p = _s.get("list_path") or []
                if len(_p) == 1:
                    shallow_list_name[_p[0]] = _s["name"]
                    shallow_list_ord[_p[0]] = _s.get("ord_col")
        plan = (self._filter_flatten_plan_for_path(plan_full, path_parts)
                if path_parts else plan_full)
        if not any(t["kind"] in ("list", "map", "jsonside") or
                   (t["kind"] == "base" and t["inline_structs"])
                   for t in plan["tables"]) and not deep_cols:
            return {"ok": True, "created": [], "base": None,
                    "note": "no nested columns to flatten"}
        deep_set = set(deep_cols)

        # .507 PROMOTION: when the base table would carry NOTHING but keys
        # (no top-level scalars, no inline struct leaves, no map/json side
        # tables) and there is exactly ONE top-level list, that keys-only base
        # is noise (on-box: the file-named table was just _sk/_rid). Skip it
        # and give the list's ROOT the file's name -- the file-named table IS
        # the records table, carrying the record's top-level fields; children
        # keep element-path names.
        _b0 = plan["tables"][0]
        # Only TOP-LEVEL lists count for promotion / deep routing — nested
        # list-inside-list-element children (compound keys) are not
        # independent top-level nests.
        _list_specs = [t for t in plan["tables"]
                       if t.get("kind") == "list"
                       and len(t.get("list_path") or []) == 1
                       and not t.get("parent_ords")]
        _other_kids = [t for t in plan["tables"]
                       if t.get("kind") in ("map", "jsonside")]
        promote = (not _b0.get("scalars") and not _b0.get("inline_structs")
                   and not _other_kids and len(_list_specs) == 1)
        _promo_col = (_list_specs[0]["list_path"][0]
                      if promote and len(_list_specs[0]["list_path"]) == 1
                      else None)
        if _promo_col is None:
            promote = False

        # .525 on-box: a cancel Event left SET by an earlier Stop is
        # sticky on the manager, and _BeatScope interrupts every statement
        # while it is set -- so a fresh flatten died on arrival with
        # "flatten cancelled (tables so far kept)". A fresh flatten is a
        # NEW operation: clear the stale flag exactly like the loaders'
        # fresh-load clears. A REAL cancel of THIS run arrives via
        # query_id / a new Stop, which sets it again.
        try:
            mgr._cancel.clear()
        except Exception:
            pass

        # .521: optional record-level UNIQUE IDENTIFIER (root_id). The
        # client's choice is VALIDATED against this file's own schema (a
        # steps/in_list/map match against our derived candidates -- no
        # client SQL), then rendered once per flavor: e0 = the promoted
        # record element, t0 = the source row.
        # Field Explorer may scope planning to one nest, but the UID picker
        # still offers any level of the SOURCE table -- validate against
        # full_columns with the same promote gate a whole-table flatten uses.
        _rid_cand = None
        _rex_e0 = _rex_t0 = None
        if root_id is not None:
            if column is None:
                _rid_cols = columns
                _rid_promo = _promo_col if promote else None
            else:
                _rid_cols = full_columns
                _fplan = _shred.plan_flatten_table(
                    full_columns, base="__uid__", reserved=set())
                _fb0 = _fplan["tables"][0]
                _flists = [t for t in _fplan["tables"]
                           if t.get("kind") == "list"
                           and len(t.get("list_path") or []) == 1
                           and not t.get("parent_ords")]
                _fother = [t for t in _fplan["tables"]
                           if t.get("kind") in ("map", "jsonside")]
                _fok = (not _fb0.get("scalars")
                        and not _fb0.get("inline_structs")
                        and not _fother and len(_flists) == 1)
                _rid_promo = (_flists[0]["list_path"][0]
                              if _fok and len(_flists[0]["list_path"]) == 1
                              else None)
            _rid_cand = _shred.validate_root_choice(
                _rid_cols, _rid_promo, root_id)
            if _rid_cand is None:
                return {"error": "root_id: that field was not found in "
                                 "this file's schema"}
            _mk = _rid_cand.get("map_key")
            _rex_e0 = _shred.render_root_expr(_rid_cand, "e0", map_key=_mk)
            _rex_t0 = _shred.render_root_expr(_rid_cand, "t0", map_key=_mk)

        def _shallow_spec_replaced(spec):
            """The shallow list child / joinkeys for a DEEP top-level column is
            dropped here -- that column's full shred hierarchy replaces it.

            Compound-key children planned under a deep list (list_path starting
            with the deep column) are also skipped — deep shred owns them.
            """
            if spec.get("kind") == "list":
                p = spec.get("list_path") or []
                return bool(p) and p[0] in deep_set
            if spec.get("kind") == "joinkeys":
                ps = spec.get("list_paths") or []
                return (len(ps) == 1 and len(ps[0]) == 1
                        and ps[0][0] in deep_set)
            return False

        # a stable _rid needs file_row_number, which only read_parquet
        # provides -- materialize a one-shot cache unless the table is
        # already backed by parquet. Path-scoped Field Explorer flatten
        # always projects the selected nest into a synthetic column, so it
        # needs its own cache even when the source is already parquet.
        src = (getattr(mgr, "table_sources", {}) or {}).get(table)
        parquet = src if (src and str(src).lower().endswith(".parquet")
                          and _path_proj is None) else None
        cache_made = None
        if parquet is None:
            try:
                cdir = os.path.join(tempfile.gettempdir(),
                                    "samql_flatten_cache")
                os.makedirs(cdir, exist_ok=True)
                parquet = os.path.join(
                    cdir, "%s_%s.parquet"
                    % (re.sub(r"[^A-Za-z0-9_]+", "_", table)[:40],
                       uuid.uuid4().hex[:8]))
                # keepalive-guarded: this streams the whole table to
                # disk, so it must heartbeat (no false stall) and be
                # cancellable mid-write.
                if _path_proj is not None:
                    _psql, _synth = _path_proj
                    _peel_sel = "".join(
                        ", t.%s AS %s" % (psql, self._rq(pname))
                        for psql, pname in _path_peels)
                    _copy = (
                        "COPY (SELECT t.*, t.%s AS %s%s FROM %s AS t) "
                        "TO '%s' (FORMAT PARQUET)"
                        % (_psql, self._rq(_synth), _peel_sel,
                           self._rq(table),
                           sqlutil.sql_path(parquet)))
                else:
                    _copy = (
                        "COPY (SELECT * FROM %s) TO '%s' (FORMAT PARQUET)"
                        % (self._rq(table), sqlutil.sql_path(parquet)))
                with mgr.native_op_cursor() as (xc, lk):
                    with lk:
                        with _ExecKeepalive(xc, mgr._cancel,
                                            threading.get_ident()):
                            xc.execute(_copy)
                cache_made = parquet
            except Exception as e:
                return {"error": "Couldn't build the flatten cache: "
                        + err_str(e)}

        fwd = sqlutil.sql_path(parquet)
        # read_parquet(...) drops straight into FROM <src> AS t0 with NO
        # wrapping parens. .486: `FROM (read_parquet(...)) AS t0` makes DuckDB
        # treat the parens as a subquery and expect a SELECT inside them -- a
        # bare table function there is a parser error ("syntax error at or near
        # ')'"), which is exactly what the flatten hit on Input_Derivatives.
        # Every other read_parquet in the codebase, including the shred NODE
        # builder, uses this bare form; the flatten was the lone outlier.
        read_src = ("read_parquet('%s', file_row_number = true)" % fwd)
        row_expr = "t0.file_row_number + 1"
        created = []
        self._register_run(query_id, mgr, kind="flatten", target=base)
        # .494: cap threads while the deep (fat-struct) pipelines run -- peak
        # memory multiplies by thread count -- exactly as shred_run does;
        # restored on every exit path. Engaged ONLY when a deep column is
        # present, so the simple flatten keeps full parallelism.
        _unthrottle = None
        if deep_cols:
            from .engines import shred_thread_throttle
            _unthrottle = shred_thread_throttle(
                getattr(mgr, "_conn", None),
                getattr(mgr, "write_lock", None))
        try:
            def _materialize(sql, tname, kind, parent=None,
                             shred_spec=None, shred_col=None,
                             shred_parquet=None, json_access=False):
                """Drop any squatting view, run the CTAS, count rows, record the
                table. Returns None on success, or an early-return dict (cancel
                / error) the caller must propagate.

                For deep shred CTAS, OutOfMemory falls back to row-range
                batches (same .430 path as shred_run) after raising the engine
                memory ceiling once more.
                """
                # pre-drop any view squatting on the target name (a flatten-off
                # load leaves a VIEW; CREATE OR REPLACE TABLE can't replace it).
                try:
                    mgr.execute('DROP VIEW IF EXISTS %s' % self._rq(tname))
                except Exception:
                    pass
                try:
                    mgr.execute(sql)
                except Exception as e:
                    if _is_intr(e) or self._run_is_cancelled(query_id):
                        return {"cancelled": True, "created": created}
                    is_oom = ("OutOfMemory" in err_str(e)
                              or "Out of Memory" in err_str(e))
                    if is_oom:
                        # Raise toward total-RAM budget (not OS available) and
                        # checkpoint so a stuck ~3–4 GiB pool can grow before
                        # the row-range / single retry below.
                        try:
                            from .engines import ensure_heavy_op_engine_memory
                            ensure_heavy_op_engine_memory(mgr)
                        except Exception:
                            pass
                        try:
                            mgr.execute("CHECKPOINT")
                        except Exception:
                            pass
                    recovered = False
                    if (is_oom and shred_spec is not None
                            and shred_col and shred_parquet):
                        # .430-style row-range CTAS for deep flatten children.
                        total = self._parquet_rowcount(mgr, shred_parquet)
                        if total:
                            size = max(1, (total + 7) // 8)
                            lo, first = 1, True
                            schemas = (['"JSON"', '"VARCHAR"']
                                       if json_access else ['"JSON"'])
                            batch_failed = False
                            while lo <= total:
                                if self._run_is_cancelled(query_id):
                                    return {"cancelled": True,
                                            "created": created}
                                hi = min(total, lo + size - 1)
                                ok_batch = False
                                last_err = None
                                for sch in schemas:
                                    bsql = _shred.build_table_sql(
                                        shred_spec, shred_col, shred_parquet,
                                        row_range=(lo, hi),
                                        insert=not first,
                                        json_access=json_access,
                                        json_arr_schema=sch)
                                    try:
                                        mgr.execute(bsql)
                                        ok_batch = True
                                        break
                                    except Exception as e2:
                                        if _is_intr(e2) or self._run_is_cancelled(
                                                query_id):
                                            return {"cancelled": True,
                                                    "created": created}
                                        last_err = e2
                                        if ("OutOfMemory" in err_str(e2)
                                                or "Out of Memory" in err_str(e2)):
                                            break
                                if not ok_batch:
                                    if last_err is not None and (
                                            "OutOfMemory" in err_str(last_err)
                                            or "Out of Memory" in err_str(
                                                last_err)):
                                        if size <= 50:
                                            e = last_err
                                            batch_failed = True
                                            break
                                        size = max(50, size // 2)
                                        continue
                                    e = last_err or e
                                    batch_failed = True
                                    break
                                first = False
                                lo = hi + 1
                            recovered = not batch_failed and lo > total
                    if not recovered:
                        if is_oom and not shred_spec:
                            # Shallow flatten SQL has no row-range builder —
                            # retry once after the memory raise above.
                            try:
                                mgr.execute(sql)
                                recovered = True
                            except Exception as e3:
                                e = e3
                        if not recovered:
                            _sql1 = " ".join(sql.split())
                            if len(_sql1) > 4000:
                                _sql1 = _sql1[:4000] + " ...[+%d chars]" % (
                                    len(_sql1) - 4000)
                            return {"error": "Flatten failed on %s: %s -- SQL: %s"
                                    % (tname, err_str(e), _sql1),
                                    "created": created}
                try:
                    _c, _r = mgr.execute('SELECT COUNT(*) FROM %s'
                                         % self._rq(tname))
                    n = int(_r[0][0]) if _r else 0
                except Exception:
                    n = 0
                created.append({"name": tname, "rows": n, "kind": kind})
                if parent:
                    # .501: the family map feeds the sidebar tree + family
                    # joins (path-only child names carry no parent in the name)
                    self._table_family[tname] = parent
                return None

            _hub_name = next(
                (t["name"] for t in plan["tables"]
                 if t.get("kind") == "joinkeys"),
                base + "_flattened")
            # 1) shallow flatten: base + SIMPLE top-level lists (+ hub).
            #    The list child / hub of any DEEP column is skipped here --
            #    replaced by that column's shred hierarchy in step 2.
            for spec in plan["tables"]:
                if self._run_is_cancelled(query_id):
                    return {"cancelled": True, "created": created}
                if _shallow_spec_replaced(spec):
                    continue
                if promote and spec["kind"] in ("base", "joinkeys"):
                    continue      # .507: the keys-only base is not created
                # .518: the BASE DATA TABLE is never created anymore. Every
                # on-box "the file name still answers with nested json" ghost
                # traced to SOME relation living under the base name; the
                # contract now is that NOTHING does. The joinkeys hub -- the
                # same keys + the record scalars -- is the family's anchor.
                if spec["kind"] == "base":
                    continue
                if promote and spec["kind"] == "list" \
                        and _promo_col not in deep_set:
                    # a promoted SHALLOW list keeps its element-path name;
                    # the file name stays UNBOUND (nothing answers to it)
                    pass
                if _rex_t0:
                    # Path-projected nests promote the nest element, but the
                    # unique id still lives on the source row (t0).
                    spec = dict(spec, root_expr=_rex_t0)
                _ja = False
                if spec.get("kind") == "list":
                    _lp = spec.get("list_path") or []
                    if _lp and _lp[0] in json_access_cols:
                        _ja = True
                elif spec.get("kind") == "jsonside":
                    if (spec.get("column") or "") in json_access_cols:
                        # deepened away from jsonside — shouldn't happen
                        pass
                sql = _shred.build_flatten_sql(
                    spec, read_src, row_expr, json_access=_ja)
                _par = None if spec["kind"] == "joinkeys" else _hub_name
                early = _materialize(
                    sql, spec["name"], spec["kind"], parent=_par)
                if early is not None:
                    return early
                # Heartbeat between child tables so a long Field-Explorer
                # flatten doesn't look stalled on the activity tray.
                try:
                    opreg.beat(rows=len(created), op_id=query_id)
                except Exception:
                    pass

            # 2) deep hierarchy for each nested-array-bearing top-level list.
            #    plan_shred names its ROOT after `base`, so pass the shallow
            #    child name ("<list>_flattened") -> root/hub "<…>_flattened",
            #    children "<path>_flattened". Its root ORDINAL would otherwise
            #    be "<base>_json_ord"; hand it the shallow list's ord
            #    ("json_ord") as root_ord so it matches a one-level flatten.
            #    Each CTAS reads the parquet directly; build_table_sql escapes
            #    the path itself, so hand it the RAW cache path, exactly like
            #    shred_run.
            for name in deep_cols:
                if promote and name == _promo_col:
                    # .507: the deep root IS the file-named records table
                    deep_base = base
                    _kw = {"children_use_base": False}
                else:
                    deep_base = (shallow_list_name.get(name)
                                 or (base + "_" + name))
                    # Shallow list tables are already *_flattened; plan_shred
                    # would otherwise emit <that>_flattened again as hub.
                    if deep_base.lower().endswith("_flattened"):
                        deep_base = deep_base[: -len("_flattened")]
                    _kw = {}
                sub = _shred.plan_shred(
                    name, col_nodes[name], base=deep_base,
                    root_ord=shallow_list_ord.get(name),
                    reserved=_reserved, **_kw)
                _promoted_here = bool(promote and name == _promo_col)
                _sub_hub = next(
                    (t["name"] for t in sub["tables"]
                     if t.get("join_helper")),
                    _shred._with_flattened(deep_base))
                _root_name = next(
                    (t["name"] for t in sub["tables"]
                     if not t.get("join_helper")),
                    deep_base)
                _specs = sub["tables"]
                # .518 / naming: join_helper hub anchors the family. Skip the
                # redundant shred root (and any skip_materialize marks). When
                # root and hub share a name after claiming <base>_flattened,
                # select by join_helper — not by name equality.
                if any(t.get("join_helper") for t in _specs):
                    _specs = (
                        [t for t in _specs if t.get("join_helper")]
                        + [t for t in _specs
                           if not t.get("join_helper")
                           and not t.get("skip_materialize")
                           and t["name"] != _root_name]
                    )
                # Field Explorer path scope: keep hub + ancestors + the
                # selected nest and its descendants (not sibling arrays).
                if path_parts and len(path_parts) > 1:
                    _want = ".".join(path_parts)
                    _kept_deep = []
                    for _t in _specs:
                        if (_t["name"] == _sub_hub
                                or _t.get("join_helper")):
                            _kept_deep.append(_t)
                            continue
                        _p = (str(_t.get("path") or "").split(" (", 1)[0])
                        if (_p == _want or _p.startswith(_want + ".")
                                or _want.startswith(_p + ".")
                                or _p == name):
                            _kept_deep.append(_t)
                    _specs = _kept_deep
                _use_json = name in json_access_cols
                for spec in _specs:
                    if self._run_is_cancelled(query_id):
                        return {"cancelled": True, "created": created}
                    if _rex_e0 or _rex_t0:
                        # Whole-table promote: id is on the record element
                        # (e0). Field Explorer path projection promotes the
                        # selected nest only — id stays on the source row.
                        _use_e0 = (_promoted_here and _path_proj is None
                                   and _rex_e0)
                        spec = dict(spec, root_expr=(
                            _rex_e0 if _use_e0 else _rex_t0))
                    sql = _shred.build_table_sql(
                        spec, name, parquet,
                        json_access=_use_json,
                        json_arr_schema='"JSON"',
                        root_is_json_list=False)
                    _par = spec.get("parent")
                    if spec.get("join_helper") or spec["name"] == _sub_hub:
                        # Promote: hub anchors the family. Else: nest under
                        # the shallow records hub when one was created.
                        _par = None if _promoted_here else _hub_name
                    elif (_par == _root_name
                          or (isinstance(_par, str)
                              and any(t.get("skip_materialize")
                                      and t["name"] == _par
                                      for t in sub["tables"]))):
                        _par = _sub_hub
                    elif _par is None:
                        _par = _hub_name if not _promoted_here else _sub_hub
                    early = _materialize(
                        sql, spec["name"], "shred", parent=_par,
                        shred_spec=spec, shred_col=name,
                        shred_parquet=parquet, json_access=_use_json)
                    if early is not None:
                        return early
        finally:
            if _unthrottle is not None:
                try:
                    _unthrottle()
                except Exception:
                    pass
            self._unregister_run(query_id)
            if cache_made:
                try:
                    os.remove(cache_made)
                except Exception:
                    pass
        # .521: the Master_Keys table -- the DISTINCT, NULL-free list of
        # root_ids, one row each, parented into the family (so a re-flatten
        # replaces it like any other member). If the chosen field is not
        # actually unique per record, the stats say so out loud instead of
        # the dedupe hiding a data problem.
        #
        # Join_Keys is the sibling per-record map: one row per hub record
        # with _sk + _rid + root_id (the chosen unique identifier, stamped
        # as root_id on every family table). Same naming / family parenting
        # as Master_Keys; not a rename of Master_Keys.
        _rid_stats = None
        if _rid_cand is not None and created:
            _hub_tbl = created[0]["name"]
            _taken = {n.lower() for n in
                      (getattr(mgr, "table_columns", {}) or {})}
            _taken -= {c["name"].lower() for c in created}
            # Re-flatten: prior Master_Keys / Join_Keys sit under the hub in
            # _table_family — free those names so CREATE OR REPLACE lands on
            # the same titles instead of suffixing _2 copies.
            _taken -= {n.lower() for n in self._family_members(_hub_tbl)}
            # Load-time replace frees the source name; Field Explorer keep
            # must not let Master_Keys / Join_Keys steal the still-loaded
            # source table.
            if not keep_source:
                _taken.discard(table.lower())
            _mk_name = "Master_Keys"
            _n = 2
            while _mk_name.lower() in _taken:
                _mk_name = "Master_Keys_%d" % _n
                _n += 1
            _mk_sql = ('CREATE OR REPLACE TABLE %s AS '
                       'SELECT DISTINCT "root_id" FROM %s '
                       'WHERE "root_id" IS NOT NULL'
                       % (self._rq(_mk_name), self._rq(_hub_tbl)))
            early = _materialize(_mk_sql, _mk_name, "masterkeys",
                                 parent=_hub_tbl)
            if early is not None:
                return early
            _taken.add(_mk_name.lower())
            _jk_name = "Join_Keys"
            _n = 2
            while _jk_name.lower() in _taken:
                _jk_name = "Join_Keys_%d" % _n
                _n += 1
            _jk_sql = ('CREATE OR REPLACE TABLE %s AS '
                       'SELECT "_sk", "_rid", "root_id" FROM %s'
                       % (self._rq(_jk_name), self._rq(_hub_tbl)))
            early = _materialize(_jk_sql, _jk_name, "joinkeys_uid",
                                 parent=_hub_tbl)
            if early is not None:
                return early
            try:
                _, _rws = mgr.execute(
                    'SELECT COUNT(*), COUNT("root_id"), '
                    'COUNT(DISTINCT "root_id") FROM %s'
                    % self._rq(_hub_tbl))
                if _rws:
                    _a, _b, _c = (int(_rws[0][0]), int(_rws[0][1]),
                                  int(_rws[0][2]))
                    _rid_stats = {"column": "root_id",
                                  "label": _rid_cand.get("label"),
                                  "master_table": _mk_name,
                                  "join_table": _jk_name,
                                  "records": _a, "nonnull": _b,
                                  "distinct": _c,
                                  "duplicated": _b - _c,
                                  "nulls": _a - _b}
            except Exception:
                pass

        # .518: load-time flatten drops the nested SOURCE so nothing answers
        # to the bare file name (hub anchors the family). Field Explorer
        # flatten keeps the source table and only adds the nest family.
        if created:
            if not keep_source:
                try:
                    self._drop_named(mgr, [table])
                except Exception:
                    pass
            _purge = getattr(mgr, "purge_shadow_temp_views", None)
            if _purge is not None:
                try:
                    _purge_names = [c["name"] for c in created]
                    if not keep_source:
                        _purge_names.append(table)
                    _purge(_purge_names)
                except Exception:
                    pass
        try:
            mgr.sync_catalog()
        except Exception:
            pass
        self._invalidate_counts()
        self._invalidate_profiles()
        for c in created:
            self._prime_count("duckdb", c["name"], c["rows"])
        # Listing order: Master_Keys, Join_Keys, then flattened data tables.
        # SQL still builds the hub first (keys SELECT FROM hub); only the
        # returned / catalog-facing order is rearranged.
        def _flatten_list_rank(c):
            k = c.get("kind")
            n = c.get("name") or ""
            if k == "masterkeys" or n == "Master_Keys" \
                    or n.startswith("Master_Keys_"):
                return 0
            if k == "joinkeys_uid" or n == "Join_Keys" \
                    or n.startswith("Join_Keys_"):
                return 1
            return 2
        _data = [c for c in created if _flatten_list_rank(c) == 2]
        _keys = [c for c in created if _flatten_list_rank(c) < 2]
        _keys.sort(key=_flatten_list_rank)
        created = _keys + _data
        base_tbl = next((c["name"] for c in _data), None)
        out = {"ok": True, "created": created, "base": base_tbl,
               "source": table, "kept_source": bool(keep_source)}
        if _rid_stats is not None:
            out["root_id"] = _rid_stats
        return out

    def shred_run(self, engine, table, column, base=None, tables=None,
                  query_id=None):
        """Create the selected relational tables from the plan. Each table is
        ONE vectorized CREATE..AS over the column's parquet backing (the
        historical slowness came from row-by-row normalization; this reads
        only each table's subtree, columnar). Cancellable between tables and
        mid-statement (interrupt); progress on the activity dashboard."""
        from .shred import build_table_sql
        from .engines import _is_interrupt as _is_intr

        def _is_oom(e):
            t = err_str(e)
            return ("OutOfMemory" in t or "Out of Memory" in t)

        plan = self.shred_plan(engine, table, column, base=base)
        if plan.get("error"):
            return plan
        try:
            from .engines import ensure_heavy_op_engine_memory
            ensure_heavy_op_engine_memory(self.duckdb)
        except Exception:
            pass
        if plan.get("needs_cache"):
            # .471: materialize the parquet cache from the loaded
            # object (view or table) so every downstream statement --
            # including the .427 pre-drop rule, which requires that
            # shred never reads the view it may replace -- stays
            # exactly as it was.
            mgr = self.duckdb
            try:
                cdir = os.path.join(tempfile.gettempdir(),
                                    "samql_shred_cache")
                os.makedirs(cdir, exist_ok=True)
                cache = os.path.join(
                    cdir, "%s_%s.parquet"
                    % (re.sub(r"[^A-Za-z0-9_]+", "_", table)[:40],
                       uuid.uuid4().hex[:8]))
                mgr.execute(
                    "COPY (SELECT * FROM %s) TO '%s' (FORMAT PARQUET)"
                    % (self._rq(table), sqlutil.sql_path(cache)))
                plan["source"] = cache
                plan["cache_materialized"] = True
            except Exception as e:
                return {"error": "Couldn't build the shred cache for "
                        "this table: " + err_str(e)}
        want = set(tables or [t["name"] for t in plan["tables"]])
        todo = [t for t in plan["tables"] if t["name"] in want]
        if not todo:
            return {"error": "Nothing selected to create."}
        mgr = self.duckdb
        created = []
        self._register_run(query_id, mgr, kind="shred", target=table)
        # .431: cap threads for the fat-struct pipelines (peak memory
        # multiplies by thread count); restored on every exit path.
        from .engines import shred_thread_throttle
        _unthrottle = shred_thread_throttle(
            getattr(mgr, "_conn", None),
            getattr(mgr, "write_lock", None))
        try:
            for i, spec in enumerate(todo):
                if self._run_is_cancelled(query_id):
                    return {"cancelled": True, "created": created}
                sql = build_table_sql(
                    spec, column, plan["source"],
                    json_access=bool(plan.get("json_access")),
                    json_arr_schema='"JSON"',
                    root_is_json_list=bool(plan.get("root_is_json_list")))
                # .427 (on-box CatalogException "... is not a table"):
                # a query-in-place load is a VIEW, and the root spec
                # deliberately takes the source name (root replaces
                # source -- the established shred semantic). But CREATE
                # OR REPLACE TABLE cannot replace a VIEW, so pre-drop a
                # view squatting on any target name. Safe: every shred
                # CTAS reads read_parquet() directly, never the view.
                if (spec["name"] == table
                        or spec["name"] in getattr(mgr, "view_backing",
                                                   {})):
                    _dropper = getattr(self, "_drop_named", None)
                    if _dropper is not None:  # bare test doubles lack it
                        _dropper(mgr, [spec["name"]])
                try:
                    opreg.advance(i, len(todo), unit="table",
                                  op_id=query_id)
                except Exception:
                    pass
                def _exec_one(_sql):
                    with mgr.native_op_cursor() as (xc, lk):
                        with lk:
                            with _ExecKeepalive(xc, mgr._cancel,
                                                threading.get_ident()):
                                xc.execute(_sql)

                def _build_sql(schema='"JSON"', row_range=None, insert=False):
                    return build_table_sql(
                        spec, column, plan["source"],
                        row_range=row_range, insert=insert,
                        json_access=bool(plan.get("json_access")),
                        json_arr_schema=schema,
                        root_is_json_list=bool(
                            plan.get("root_is_json_list")))

                batched = 0
                try:
                    _exec_one(sql)
                except Exception as e:
                    if _is_intr(e):
                        raise
                    recovered = False
                    # Ordered fallback: JSON from_json schema → VARCHAR +
                    # json(e) for double-encoded arrays.
                    if plan.get("json_access") and not _is_oom(e):
                        try:
                            _exec_one(_build_sql('"VARCHAR"'))
                            recovered = True
                        except Exception as e_alt:
                            if _is_intr(e_alt):
                                raise
                            e = e_alt
                    if not recovered and _is_oom(e):
                        # .430: OOM fallback -- rebuild THIS table in row
                        # ranges (bounded working set), CREATE for the first
                        # range then INSERT for the rest, halving the batch
                        # on a repeat OOM down to a floor.
                        total = self._parquet_rowcount(mgr, plan["source"])
                        if not total:
                            raise
                        size = max(1, (total + 7) // 8)
                        lo, first = 1, True
                        schemas = (['"JSON"', '"VARCHAR"']
                                   if plan.get("json_access")
                                   else ['"JSON"'])
                        while lo <= total:
                            if self._run_is_cancelled(query_id):
                                return {"cancelled": True,
                                        "created": created}
                            hi = min(total, lo + size - 1)
                            ok_batch = False
                            last_batch_err = None
                            for sch in schemas:
                                bsql = _build_sql(
                                    sch, row_range=(lo, hi),
                                    insert=not first)
                                try:
                                    _exec_one(bsql)
                                    ok_batch = True
                                    break
                                except Exception as e2:
                                    if _is_intr(e2):
                                        raise
                                    last_batch_err = e2
                                    if _is_oom(e2):
                                        break
                            if not ok_batch:
                                if last_batch_err is not None \
                                        and _is_oom(last_batch_err):
                                    if size <= 50:
                                        raise last_batch_err
                                    size = max(50, size // 2)
                                    continue
                                raise last_batch_err
                            first = False
                            batched += 1
                            lo = hi + 1
                            try:
                                opreg.beat(rows=hi, op_id=query_id)
                            except Exception:
                                pass
                        recovered = True
                    if not recovered:
                        raise
                created.append({"name": spec["name"],
                                "keys": spec["keys"],
                                "batched": batched or None})
            try:
                mgr.sync_catalog()
            except Exception:
                pass
            # .513: the family must be reachable UNQUALIFIED. If a temp
            # view (an orphaned reuse view) shadows any name this run just
            # created, purge the collisions now.
            _purge = getattr(mgr, "purge_shadow_temp_views", None)
            if _purge is not None:
                try:
                    _purge([t["name"] for t in created])
                except Exception:
                    pass
            self._invalidate_counts()
            self._invalidate_profiles()
            return {"ok": True, "created": created}
        except Exception as e:
            if _is_interrupt(e) or self._run_is_cancelled(query_id):
                return {"cancelled": True, "created": created}
            return {"error": err_str(e), "created": created}
        finally:
            _unthrottle()
            self._unregister_run(query_id)

    def cell_value(self, result_id, row, column, sort_col=None,
                   descending=False, filters=None):
        """The FULL value of one cell, addressed under the same view (sort +
        filters) the grid is showing -- the counterpart to _cap_page_rows'
        display truncation. Bounded by SAMQL_CELL_FETCH_MAX (default 5 MB) so
        a pathological single value can't blow up the browser either; the
        response says when even the fetch is clipped."""
        cr = self._results.get(result_id)
        if cr is None:
            return {"error": "result expired"}
        if column not in cr.cols:
            return {"error": "unknown column"}
        ci = cr.cols.index(column)
        try:
            if filters:
                view, total = cr.filtered_view(filters, sort_col, descending)
            else:
                view = cr.view(sort_col, descending)
                total = cr.total
            row = int(row)
            if row < 0 or row >= max(total, 0):
                return {"error": "row out of range"}
            v = json_safe(view[row][ci])
        except Exception as e:
            return {"error": err_str(e)}
        clipped = False
        if isinstance(v, str):
            try:
                cap = int(os.environ.get("SAMQL_CELL_FETCH_MAX") or 5_000_000)
            except Exception:
                cap = 5_000_000
            if cap > 0 and len(v) > cap:
                v, clipped = v[:cap], True
        return {"value": v, "column": column, "row": row,
                "length": len(v) if isinstance(v, str) else None,
                "clipped": clipped}

    def discard_result(self, result_id):
        with self._lock:
            if self._reuse_pins.get(result_id):
                # a live reuse view reads this store -- defer the discard
                # until the run's cleanup unpins it (the grid keeps paging)
                self._discard_deferred.add(result_id)
                return {"ok": True, "deferred": True}
            cr = self._results.pop(result_id, None)
            if result_id in self._results_order:
                self._results_order.remove(result_id)
        if cr is not None:
            cr.close()
            # closing a result frees memory/temp files; reclaim in the
            # background so the app stays lean.
            self._schedule_cleanup(full=True)

    # ---- statement-at-cursor (for the editor's F5 behaviour) --------
    @staticmethod
    def statement_at_cursor(text, pos):
        spans = split_sql_statements_spans(text)
        found = find_statement_at(spans, pos)
        if not found:
            return {"start": 0, "end": len(text or ""),
                    "sql": (text or "").strip()}
        s, e, st = found
        return {"start": s, "end": e, "sql": st.strip()}

    # ---- profiling --------------------------------------------------
    def profile(self, table, engine="sqlite", query_id=None):
        eng = (self.get_duckdb() if engine == "duckdb"
               else self.db)
        if engine != "duckdb" and table not in self.db.table_columns:
            if (self.duckdb is not None
                    and table in self.duckdb.table_columns):
                eng = self.duckdb
        key = (engine, table)
        # Honour an explicit cancel for this query id before serving cache —
        # otherwise a pre-flagged Stop looks like a successful cached profile.
        if self._run_is_cancelled(query_id):
            return {"cancelled": True}
        cached = self._profile_cache.get(key)
        if cached is not None:
            return cached
        # Register so Stop / window-close (cancel_query) interrupts the heavy
        # aggregate pass; the cancel callable also lets profile_table bail
        # between passes even if an interrupt lands in a swallowed gap.
        # Clear a sticky engine cancel from an earlier stalled load Stop so
        # BeatDaemon does not auto-interrupt this fresh profile.
        self._clear_stale_engine_cancel(eng, except_qid=query_id)
        self._register_run(query_id, eng, kind="profile", target=table)
        try:
            if self._run_is_cancelled(query_id):
                return {"cancelled": True}
            result = profile_table(
                eng, table,
                cancel=lambda: self._run_is_cancelled(query_id))
        except Exception as e:
            if _is_interrupt(e) or self._run_is_cancelled(query_id):
                return {"cancelled": True}
            raise
        finally:
            self._end_run_keep_cancel(query_id)
        self._profile_cache[key] = result
        self._profile_cache_order.append(key)
        while len(self._profile_cache_order) > 16:
            old = self._profile_cache_order.pop(0)
            self._profile_cache.pop(old, None)
        return result

    def profile_field(self, spec):
        """Profile a single column (field) of a result or table. Returns the
        same shape as profile_table but with a single entry in ``columns``,
        plus ``field``. ``spec`` is {result_id|table, engine, column, query_id}.

        Cancel binds the same interrupt target as chart/pivot (result-store
        connection/engine), so Stop reaches the aggregate SQL.
        """
        qid = spec.get("query_id") or ("pf-%s" % uuid.uuid4().hex[:10])
        target = self._agg_interrupt_target(spec)
        self._clear_stale_engine_cancel(target, except_qid=qid)
        self._register_run(
            qid, target, kind="profile", surface="profile",
            label=str(spec.get("column") or "field"))
        try:
            if self._run_is_cancelled(qid):
                return {"cancelled": True, "error": "cancelled",
                        "table": "", "total_rows": 0, "columns": []}
            return self._profile_field_inner(spec, qid)
        except Exception as e:
            if _is_interrupt(e) or self._run_is_cancelled(qid):
                return {"cancelled": True, "error": "cancelled",
                        "table": "", "total_rows": 0, "columns": []}
            raise
        finally:
            self._end_run_keep_cancel(qid)

    def _profile_field_inner(self, spec, qid=None):
        col = spec.get("column")
        if not col:
            return {"error": "No column specified."}
        ctx, cols = self._agg_source(spec)
        if cols is None:
            return {"error": "That result is no longer available — re-run the "
                              "query, then profile the field."}
        if col not in cols:
            return {"error": f"Unknown column: {col}"}
        ci = list(cols).index(col)
        if qid and self._run_is_cancelled(qid):
            return {"cancelled": True, "error": "cancelled",
                    "table": "", "total_rows": 0, "columns": []}

        if ctx is not None:
            run, src, colref, castnum = ctx
            qc = colref(ci)
            srows = run(f"SELECT {qc} FROM {src} "
                        f"WHERE {qc} IS NOT NULL LIMIT 200")
            samples = [r[0] for r in (srows or [])]
            t = _infer_type(samples)
            numeric = t == "NUMERIC"
            parts = ["COUNT(*)", f"COUNT({qc})", f"COUNT(DISTINCT {qc})",
                     f"MIN({qc})", f"MAX({qc})"]
            if numeric:
                parts.append(f"AVG({castnum(qc)})")
                parts.append(f"AVG({castnum(qc)}*{castnum(qc)})")
            agg = (run("SELECT " + ", ".join(parts) + f" FROM {src}")
                   or [[0]])[0]
            total = agg[0] or 0
            nonnull = agg[1] or 0
            distinct = agg[2] or 0
            mn, mx = agg[3], agg[4]
            mean = std = None
            if numeric and len(agg) >= 7 and agg[5] is not None:
                mean = agg[5]
                var = (agg[6] or 0) - mean * mean
                std = var ** 0.5 if var and var > 0 else 0.0
            trows = run(f"SELECT {qc} AS v, COUNT(*) AS c FROM {src} "
                        f"GROUP BY {qc} ORDER BY c DESC LIMIT 10") or []
            top = [(r[0], r[1]) for r in trows]
        else:
            _c, rows = self._source_rows(
                result_id=spec.get("result_id"),
                table=spec.get("table"),
                engine=spec.get("engine", "sqlite"),
                limit=DISPLAY_LIMIT)
            rows = list(rows)
            vals = [r[ci] if ci < len(r) else None for r in rows]
            nonnull_vals = [v for v in vals if v is not None]
            total = len(vals)
            nonnull = len(nonnull_vals)
            distinct = len(set(nonnull_vals))
            t = _infer_type(nonnull_vals[:200])
            numeric = t == "NUMERIC"
            nums = [n for n in (_to_number(v) for v in nonnull_vals)
                    if n is not None]
            try:
                mn = min(nonnull_vals) if nonnull_vals else None
                mx = max(nonnull_vals) if nonnull_vals else None
            except TypeError:
                sv = sorted(str(v) for v in nonnull_vals)
                mn = sv[0] if sv else None
                mx = sv[-1] if sv else None
            mean = std = None
            if numeric and nums:
                mean = sum(nums) / len(nums)
                var = sum((x - mean) ** 2 for x in nums) / len(nums)
                std = var ** 0.5 if var > 0 else 0.0
            counts = {}
            for v in nonnull_vals:
                counts[v] = counts.get(v, 0) + 1
            top = sorted(counts.items(),
                         key=lambda kv: (-kv[1], str(kv[0])))[:10]

        nulls = (total - nonnull) if total else 0
        stats = {
            "column": col,
            "type": t,
            "raw_type": t,
            "date_fmt": None,
            "nulls": nulls,
            "null_pct": round(100 * nulls / total, 1) if total else 0.0,
            "distinct": distinct,
            "distinct_pct": (round(100 * distinct / total, 1)
                             if total else 0.0),
            "min": json_safe(mn) if t in ("NUMERIC", "TEXT") else None,
            "max": json_safe(mx) if t in ("NUMERIC", "TEXT") else None,
            "mean": mean,
            "std": std,
            "outliers": None,
            "top_values": [
                {"value": json_safe(v), "count": c,
                 "pct": round(100 * c / total, 1) if total else 0.0}
                for (v, c) in top
            ],
        }
        if qid and self._run_is_cancelled(qid):
            return {"cancelled": True, "error": "cancelled",
                    "table": "", "total_rows": 0, "columns": []}
        return {"table": col, "total_rows": total,
                "columns": [stats], "field": col}

    def _invalidate_profiles(self):
        self._profile_cache.clear()
        self._profile_cache_order.clear()

    # ---- table operations -------------------------------------------
    def rename_table(self, engine, old, new):
        eng = self.duckdb if engine == "duckdb" else self.db
        new = re.sub(r"[^\w]", "_", new).strip("_") or "table"
        if new in eng.table_columns:
            return {"error": f'A table named "{new}" already exists.'}
        # .441 audit fix: OLD was spliced raw -- a table whose name
        # carried a double-quote broke the ALTER (and the splice was
        # injection-shaped). Quote it properly; NEW stays sanitized to
        # word characters on purpose (catalog-name hygiene).
        from .sqlutil import quote_ident as _qid
        eng.conn.execute(
            f'ALTER TABLE {_qid(old)} RENAME TO "{new}"')
        try:
            eng.conn.commit()
        except Exception:
            pass
        eng.table_columns[new] = eng.table_columns.pop(old, [])
        eng.table_sources[new] = eng.table_sources.pop(old, "")
        origins = getattr(eng, "table_origins", None)
        if isinstance(origins, dict) and old in origins:
            origins[new] = origins.pop(old)
        self._rename_table_ui_order(engine, old, new)
        self._invalidate_profiles()
        self._invalidate_counts()
        return {"ok": True, "name": new}

    def drop_table(self, engine, name):
        eng = self.duckdb if engine == "duckdb" else self.db
        src = eng.table_sources.get(name)   # capture before drop clears it
        eng.drop_table(name)
        self._cleanup_temp_source(src)      # delete a SamQL temp now, not at restart
        self._evict_results_referencing(name)
        try:
            self.manifest.remove_for_table(name)  # so a drop survives restart
        except Exception:
            pass
        self._drop_table_ui_order(engine, name)
        self._invalidate_profiles()
        self._invalidate_counts()
        self._schedule_cleanup(full=True)
        return {"ok": True}

    def _cleanup_temp_source(self, src):
        """Delete a dropped table/view's backing file *iff* SamQL had streamed
        it into this instance's own temp dir (e.g. an HDFS file loaded as a
        query-in-place view). Without this the temp would linger until the
        next-start sweep. A view over the user's own file (anywhere outside the
        instance temp dir) and any non-file source label ("nodeflow", legacy "nodebook", "") are
        left untouched. Best-effort: never raises."""
        if not src or not isinstance(src, str):
            return
        try:
            from . import tmputil
            inst = os.path.abspath(tmputil.instance_dir())
            path = os.path.abspath(src)
            if os.path.isfile(path) and os.path.commonpath([inst, path]) == inst:
                os.unlink(path)
        except Exception:
            pass

    def optimize_to_parquet(self, name, cancel=None, progress=None):
        """Convert a file-backed DuckDB view (e.g. an HDFS CSV/JSON loaded as a
        zero-copy view) into a columnar Parquet file and re-point the view at
        it, so repeat queries get column + row-group pushdown instead of
        re-parsing the text on every scan. DuckDB-only; the source must be a
        convertible file (csv/tsv/json), not already Parquet.

        The Parquet lands in this instance's temp dir, so dropping the table
        later cleans it up (see _cleanup_temp_source), and the old CSV/JSON
        temp is removed here. Cancellable: a Stop aborts the COPY via the
        load-cancel rail; the partial Parquet is deleted and the *original*
        view is left intact (no re-point happens until the COPY finishes).
        Returns {ok, name, rows, all} or {error}.
        """
        from .loaders import LoadCancelled
        if self.duckdb is None:
            return {"error": "Converting to Parquet needs the DuckDB engine."}
        mgr = self.duckdb
        src = mgr.table_sources.get(name)
        src_is_file = bool(src) and isinstance(src, str) and os.path.isfile(src)
        if src_is_file:
            ext = os.path.splitext(src)[1].lower()
            if ext in (".parquet", ".pq"):
                return {"error": "This table is already stored as Parquet."}
            if ext not in (".csv", ".tsv", ".txt", ".json",
                           ".ndjson", ".jsonl", ""):
                return {"error": "Only CSV/TSV/JSON-backed tables can be "
                                 "converted to Parquet."}
        # else: a materialized (in-engine) table with no live backing file --
        # e.g. a normal CSV loaded in "materialize" mode. Still convertible by
        # copying its rows out to Parquet, which frees the in-memory copy and
        # gives repeat queries column + row-group pushdown.
        if cancel and cancel():
            raise LoadCancelled()
        pq = tmputil.new_tempfile("hdfs_", ".parquet")
        try:
            mgr.view_to_parquet(name, pq, cancel=cancel)
        except BaseException:
            # cancel (pre-check or interrupted COPY) or a genuine failure --
            # drop the partial Parquet and let the caller classify it. The
            # original view is untouched (we have not re-pointed yet).
            try:
                os.unlink(pq)
            except OSError:
                pass
            raise
        old_src = src
        origin = (getattr(mgr, "table_origins", {}) or {}).get(name) or (
            old_src if src_is_file else None)
        try:
            mgr.drop_table(name)
            final = mgr.create_view_from_file(name, pq, "parquet")
            remember = getattr(mgr, "_remember_origin", None)
            if remember and origin:
                remember(final, origin)
        except Exception as e:
            try:
                os.unlink(pq)
            except OSError:
                pass
            return {"error": err_str(e)}
        if src_is_file:
            self._cleanup_temp_source(old_src)  # delete the old CSV/JSON temp
        self._invalidate_profiles()
        self._invalidate_counts()
        return {"ok": True, "name": final, "all": self.tables_tree()}

    def materialize(self, name, sql, target="auto", query_id=None,
                    from_result=None):
        """Create (replacing any existing) a named table from a SELECT so a
        notebook cell's result can be handed to the reconcile engine, which
        compares *named tables* only -- it cannot take a subquery. ``name``
        must be a plain identifier and is expected to be caller-namespaced
        (the notebook uses a ``__nb_`` prefix, which the sidebar hides). The
        reconcile engine requires both inputs to live in one engine, so the
        caller materialises co-located inputs by passing a fixed ``target``
        rather than 'auto'. Returns {name, columns, engine} or {error}.
        """
        name = (name or "").strip()
        sel = (sql or "").strip().rstrip(";").strip()
        if not re.match(r"^[A-Za-z_][A-Za-z0-9_]*$", name):
            return {"error": "Invalid materialised table name."}
        if not sel:
            return {"error": "Nothing to materialise."}
        if classify_sql_statement(sel) == "write":
            return {"error": "Only a SELECT can be materialised."}
        # R1 parity for reconcile: when the source cell has a FRESH parquet
        # result, materialise straight from it instead of re-executing the
        # cell's whole compiled chain. Only rewrites when the store is safely
        # reusable AND the staging engine is DuckDB (read_parquet); otherwise
        # the provided SQL runs exactly as before.
        if from_result and target == DUCKDB_TARGET:
            p = self._reusable_store_path(from_result)
            if p:
                fwd = p.replace("\\", "/").replace("'", "''")
                sel = f"SELECT * FROM read_parquet('{fwd}')"
        if target in (None, "", "auto", LOCAL_TARGET):
            eng = self.db
        elif target == DUCKDB_TARGET:
            if self.duckdb is None:
                return {"error": "DuckDB engine is not available."}
            eng = self.duckdb
        else:
            return {"error": "Unknown target: %r" % (target,)}
        self._register_run(query_id, eng, kind="materialize", target=name)
        try:
            if self._run_is_cancelled(query_id):
                return {"cancelled": True}
            with eng.write_lock:
                eng.conn.execute('DROP TABLE IF EXISTS "%s"' % name)
                eng.conn.execute('CREATE TABLE "%s" AS %s' % (name, sel))
                try:
                    eng.conn.commit()
                except Exception:
                    pass
            # register the new table in the in-memory catalog so the reconcile
            # engine (and /api/tables) can see it
            eng.sync_catalog()
            cols = list(eng.table_columns.get(name, []))
            engine_kind = ("duckdb"
                           if (self.duckdb is not None and eng is self.duckdb)
                           else "sqlite")
            return {"name": name, "columns": cols, "engine": engine_kind}
        except Exception as e:
            if _is_interrupt(e) or self._run_is_cancelled(query_id):
                return {"cancelled": True}
            return {"error": err_str(e)}
        finally:
            self._unregister_run(query_id)

    def materialize_result(self, result_id, table_name, target="auto",
                           query_id=None):
        """Save an already-computed query result (by id) as a new local table.
        Works for any result -- SQLite, DuckDB, or a passthrough SQL Server
        result -- because it streams from the cached result store rather than
        re-running the SQL (so a remote result is captured without hitting the
        server again, and a query that joined catalog + local data still
        saves). The name is auto-uniquified. Returns
        {ok, table, rows, engine, columns} or {error}."""
        cr = self._results.get(result_id)
        if cr is None:
            return {"error": "That result is no longer available -- re-run the "
                    "query, then save it."}
        name = (table_name or "").strip()
        if not re.match(r"^[A-Za-z_][A-Za-z0-9_]*$", name):
            return {"error": "Enter a valid table name: letters, numbers and "
                    "underscores, and not starting with a number."}
        cols = list(cr.cols)
        if target in (None, "", "auto"):
            target = self.default_destination()
        if target in ("duckdb", DUCKDB_TARGET):
            if HAS_DUCKDB and self.duckdb is not None:
                eng, kind = self.get_duckdb(), "duckdb"
            else:
                eng, kind = self.db, "sqlite"  # graceful fallback
        elif target in ("sqlite", LOCAL_TARGET):
            eng, kind = self.db, "sqlite"
        else:
            return {"error": "Unknown target engine: %r" % (target,)}
        self._register_run(query_id, eng, kind="materialize", target=name)
        try:
            if self._run_is_cancelled(query_id):
                return {"cancelled": True}
            # .472: a parquet-backed DuckDB result saves with ONE
            # engine statement (CREATE TABLE AS read_parquet) instead
            # of shipping every row through Python into executemany --
            # that Python round-trip was the reported minutes. SQLite
            # keeps its streaming path (measured FASTER than an
            # ATTACH..INSERT SELECT of the spool). Any surprise falls
            # back to streaming.
            tname = None
            pq = getattr(cr.store, "parquet_path", None)
            pq = pq() if callable(pq) else None
            try:
                if (pq and kind == "duckdb"
                        and hasattr(eng, "add_table_from_parquet")):
                    tname, n = eng.add_table_from_parquet(
                        name, cols, pq, source="result")
            except Exception as _e:
                if _is_interrupt(_e) or self._run_is_cancelled(query_id):
                    return {"cancelled": True}
                tname = None
            if tname is None:
                tname, n = eng.add_table_streaming(
                    name, cols, cr.store, source="result")
        except Exception as e:
            if _is_interrupt(e) or self._run_is_cancelled(query_id):
                return {"cancelled": True}
            return {"error": err_str(e)}
        finally:
            self._unregister_run(query_id)
        self._invalidate_counts()
        self._invalidate_profiles()
        self._prime_count(kind, tname, n)
        return {"ok": True, "table": tname, "rows": n, "engine": kind,
                "columns": list(eng.table_columns.get(tname, cols))}

    def change_column_type(self, engine, table, col, new_type, query_id=None):
        eng = self.duckdb if engine == "duckdb" else self.db
        self._register_run(query_id, eng, kind="alter", target=table)
        try:
            if self._run_is_cancelled(query_id):
                return {"cancelled": True}
            ok = eng.change_column_type(table, col, new_type)
        except Exception as e:
            if _is_interrupt(e) or self._run_is_cancelled(query_id):
                return {"cancelled": True}
            raise
        finally:
            self._unregister_run(query_id)
        self._invalidate_profiles()
        self._invalidate_counts()
        return {"ok": bool(ok)}

    def clear_all(self, clear_manifest=False):
        self.db.drop_all()
        if self.duckdb is not None:
            self.duckdb.drop_all()
        with self._lock:
            for cr in self._results.values():
                cr.close()
            self._results.clear()
            self._results_order.clear()
        try:
            self._table_family.clear()
        except Exception:
            pass
        try:
            self._table_ui_order.clear()
        except Exception:
            pass
        if clear_manifest:
            # only when the user explicitly clears everything -- NOT on the
            # internal clear during shutdown, or restore would have nothing
            # left to rebuild from on the next launch
            try:
                self.manifest.clear()
            except Exception:
                pass
        self._invalidate_profiles()
        self._invalidate_counts()
        self._schedule_cleanup(full=True)
        return {"ok": True}

    # ---- memory management -----------------------------------------
    @staticmethod
    def _rss_bytes():
        """Current resident memory of this process, cross-platform."""
        # Linux: resident pages from /proc/self/statm
        try:
            with open("/proc/self/statm") as f:
                resident = int(f.read().split()[1])
            return resident * os.sysconf("SC_PAGE_SIZE")
        except Exception:
            pass
        # Windows: WorkingSetSize via psapi / kernel32
        try:
            import ctypes
            from ctypes import wintypes

            class _PMC(ctypes.Structure):
                _fields_ = [("cb", wintypes.DWORD),
                            ("PageFaultCount", wintypes.DWORD),
                            ("PeakWorkingSetSize", ctypes.c_size_t),
                            ("WorkingSetSize", ctypes.c_size_t),
                            ("QuotaPeakPagedPoolUsage", ctypes.c_size_t),
                            ("QuotaPagedPoolUsage", ctypes.c_size_t),
                            ("QuotaPeakNonPagedPoolUsage", ctypes.c_size_t),
                            ("QuotaNonPagedPoolUsage", ctypes.c_size_t),
                            ("PagefileUsage", ctypes.c_size_t),
                            ("PeakPagefileUsage", ctypes.c_size_t)]
            c = _PMC()
            c.cb = ctypes.sizeof(c)
            k = ctypes.windll.kernel32
            h = k.GetCurrentProcess()
            ok = False
            try:
                ok = ctypes.windll.psapi.GetProcessMemoryInfo(
                    h, ctypes.byref(c), c.cb)
            except Exception:
                ok = False
            if not ok:
                try:
                    ok = k.K32GetProcessMemoryInfo(
                        h, ctypes.byref(c), c.cb)
                except Exception:
                    ok = False
            if ok:
                return int(c.WorkingSetSize)
        except Exception:
            pass
        # Fallback: peak RSS via resource (KB on Linux, bytes on macOS)
        try:
            import resource
            import sys as _sys
            r = resource.getrusage(resource.RUSAGE_SELF).ru_maxrss
            return r if _sys.platform == "darwin" else r * 1024
        except Exception:
            return 0

    @staticmethod
    def _system_total_bytes():
        return total_physical_ram_bytes()

    def _duckdb_bytes(self):
        if self.duckdb is None:
            return None
        # Must not block on the engine's write lock -- the memory poll runs on a
        # timer, so a blocking probe during a load would pile up and exhaust the
        # browser's connection pool (starving /api/status -> "checking...").
        try:
            probe = getattr(self.duckdb, "memory_bytes", None)
            if probe is not None:
                return probe()
            _c, rows = self.duckdb.execute(
                "SELECT sum(memory_usage_bytes) FROM duckdb_memory()")
            if rows and rows[0] and rows[0][0] is not None:
                return int(rows[0][0])
        except Exception:
            pass
        return None

    def memory_usage(self):
        rss = self._rss_bytes()
        total = self._system_total_bytes()
        duck = self._duckdb_bytes()
        return {
            "rss_mb": round(rss / 1048576, 1) if rss else None,
            "total_mb": round(total / 1048576, 1) if total else None,
            "percent": (round(rss / total * 100, 1)
                        if rss and total else None),
            "duckdb_mb": (round(duck / 1048576, 1)
                          if duck is not None else None),
            "cached_results": len(self._results),
        }

    @staticmethod
    def _malloc_trim():
        try:
            import ctypes
            import ctypes.util
            name = ctypes.util.find_library("c")
            if name:
                libc = ctypes.CDLL(name)
                if hasattr(libc, "malloc_trim"):
                    libc.malloc_trim(0)
        except Exception:
            pass

    def free_memory(self):
        """Release memory held by cached query results (keeping only the most
        recent so the visible grid keeps working), then ask each loaded-table
        engine to hand back the memory it can: SQLite runs a memory shrink, and
        DuckDB checkpoints its WAL. Finally run garbage collection and return
        freed pages to the OS where possible. Loaded tables themselves are kept
        intact -- dropping a table (close) is what actually deletes its data."""
        import gc
        before = self._rss_bytes()
        with self._lock:
            keep = set(self._results_order[-1:])
            for rid in list(self._results_order):
                if rid in keep or self._reuse_pins.get(rid):
                    continue
                victim = self._results.pop(rid, None)
                if victim is not None:
                    try:
                        victim.close()
                    except Exception:
                        pass
            self._results_order = [r for r in self._results_order
                                   if r in self._results]
            # Drop any cached in-memory sorted copies of the kept results.
            for cr in self._results.values():
                try:
                    cr._sorted_cache.clear()
                except Exception:
                    pass
        # Ask the engines holding loaded tables to release what they can.
        try:
            # .438 audit fix: every other sqlite maintenance op takes the
            # write lock; this one raced writers.
            with self.db.write_lock:
                self.db.conn.execute("PRAGMA shrink_memory")
        except Exception:
            pass
        if self.duckdb is not None:
            try:
                self.duckdb.execute("CHECKPOINT")
            except Exception:
                pass
        gc.collect()
        self._malloc_trim()
        after = self._rss_bytes()
        info = self.memory_usage()
        info["freed_mb"] = (round(max(0, before - after) / 1048576, 1)
                            if before and after else 0.0)
        info["kept_results"] = len(self._results)
        return info

    # ---- JSON flatten ----------------------------------------------
    def _duck_scalar(self, mgr, sql):
        """Read a single scalar value from DuckDB, tolerating either the
        concurrent-cursor read path or the locked main connection. Returns the
        value, or None on any failure."""
        try:
            _c, rows = mgr.execute_read(sql)
            if rows:
                return rows[0][0]
        except Exception:
            pass
        try:
            with mgr.native_op_cursor() as (xc, lk):
                with lk:
                    row = xc.execute(sql).fetchone()
                    return row[0] if row else None
        except Exception:
            return None

    def _pick_records_array(self, mgr, table, struct_arrays):
        """For a wrapping-object table (a handful of rows carrying one big
        struct array, e.g. ``{"trades":[ ... ]}``), choose which struct-array
        column actually holds the records: prefer a collection-like name, then
        the array with the most total elements. Returns a column name or
        None."""
        from . import flatten as _F
        if not struct_arrays:
            return None
        if len(struct_arrays) == 1:
            return struct_arrays[0]
        pat = re.compile(
            r"(trade|deal|position|instrument|record|item|result|row|entr|"
            r"data|element|leg|cashflow|order|txn|transaction)s?$", re.I)
        named = [c for c in struct_arrays if pat.search(str(c))]
        cands = named or struct_arrays
        best, best_n = None, -1
        for c in cands:
            n = self._duck_scalar(
                mgr, "SELECT coalesce(sum(len(%s)),0) FROM %s"
                % (_F._qi(c), _F._qi(table)))
            try:
                n = int(n or 0)
            except Exception:
                n = 0
            if n > best_n:
                best, best_n = c, n
        return best or cands[0]

    def _duckdb_unnest_flatten(self, table, query_id=None, on_progress=None,
                               cancel=None, replace=False, depth=0,
                               max_depth=3, _registered=False):
        """Flatten a nested DuckDB table into relational tables entirely in
        SQL (UNNEST), with no Python JSON parse and no dump-to-file round trip.
        This is the fast path for large JSON already loaded nested into DuckDB:
        the same C++ engine that read the file does the explode, so it stays
        bounded in memory instead of buffering the file in Python.

        Each array column becomes a child table (struct arrays are expanded
        into columns; scalar arrays get a single ``value`` column). A stable row
        key is always synthesized (``_rowkey``) and carried identically on the
        root and every child, so one ``WHERE``/join links them -- no fragile
        name-based key guessing. A struct-only table (no arrays) flattens to a
        single keyed table with structs left nested (queryable via dot access).
        A wrapping-object shape (few rows, one big struct array) is detected and
        the array promoted to the root.

        Returns a result dict ``{ok, tables, all, method, key, synthesized}``
        on success, ``{"cancelled": True}`` if interrupted, or ``None`` to tell
        the caller to fall back to the JSON-dump + Python flatten path (nothing
        nested to flatten, or the SQL errored). The ``method``/``key`` fields
        drive the UI notification so the user sees which path ran and the key
        column to filter on."""
        from . import flatten as _F
        mgr = self.duckdb
        if mgr is None:
            return None
        try:
            coltypes = mgr.column_types(table)
        except Exception:
            return None
        # Engage on any nested content -- arrays to explode OR a top-level
        # struct column. A struct-only table still flattens in-engine here (a
        # keyed copy, structs kept nested) rather than dropping to the slow
        # dump + Python path. A truly flat table has nothing to flatten -> None.
        has_list = any(_F._type_is_list(t) for t in coltypes.values())
        has_struct = any(_F._type_is_struct(t) for t in coltypes.values())
        if not coltypes or (not has_list and not has_struct):
            return None

        def _tick():
            if (cancel and cancel()) or self._run_is_cancelled(query_id):
                raise loaders.LoadCancelled()
            opreg.beat(rows=0, op_id=query_id)
            if on_progress is not None:
                on_progress(0, 0)

        def _run(sql):
            with mgr.native_op_cursor() as (xc, lk):
                with lk:
                    with _ExecKeepalive(xc, mgr._cancel,
                                        threading.get_ident()):
                        xc.execute(sql)

        if not _registered:
            self._register_run(query_id, mgr, kind="flatten",
                               target=table)
        created = []      # permanent tables (root + children) -- drop on error
        temp_tables = []  # staging (keyed copy) -- always dropped
        try:
            if self._run_is_cancelled(query_id):
                return {"cancelled": True}
            existing = set(mgr.table_columns.keys())
            if replace:
                # .501: a RE-flatten of the same source replaces its OWN
                # family -- free those names (and drop the stale children up
                # front, CREATE TABLE won't overwrite) so a reload lands on
                # the same table names instead of accumulating "_1" copies.
                _own = self._family_members(table)
                if _own:
                    existing -= _own
                    self._drop_named(mgr, sorted(_own))
                    for _n in _own:
                        self._table_family.pop(_n, None)
            # In replace mode (flatten-on-load) the flattened root takes over
            # the nested source's name: build the root under a temporary name
            # (the source still exists as we read from it), then drop the source
            # and rename the built tables into place. Otherwise (right-click
            # Flatten) keep the nested source and give the root a uniquified
            # name alongside it.
            if replace:
                final_root = table
                base = mgr.reserve_table_name(table + "__flat")
            else:
                final_root = None
                base = mgr.reserve_table_name(table)
            struct_arrays = [c for c, t in coltypes.items()
                             if _F._type_is_struct_list(t)]
            rowcount = self._duck_scalar(
                mgr, "SELECT count(*) FROM %s" % _F._qi(table))

            promoted_from = None
            if rowcount is not None and rowcount <= 8 and struct_arrays:
                # Wrapping-object shape: the array elements are the records.
                # Promote that array into a keyed intermediate (row key baked
                # in as we unnest, so it is stable), then flatten from there.
                records_col = self._pick_records_array(
                    mgr, table, struct_arrays)
                if records_col is None:
                    return None
                _tick()
                inter = mgr.reserve_table_name(table + "__promoted")
                _run("CREATE TABLE %s AS SELECT row_number() OVER () AS %s, "
                     "_e.* FROM (SELECT UNNEST(%s) AS _e FROM %s "
                     "WHERE %s IS NOT NULL) _s"
                     % (_F._qi(inter), _F._qi("_rowkey"), _F._qi(records_col),
                        _F._qi(table), _F._qi(records_col)))
                temp_tables.append(inter)
                promoted_from = records_col
                itypes = mgr.column_types(inter)
                # Flatten the keyed intermediate: _rowkey is a real column here,
                # so no extra staging copy; children link by that same _rowkey.
                plan = _F.duckdb_unnest_plan(
                    inter, itypes, base, key="_rowkey", detect=False,
                    fk_name="_rowkey", existing_names=existing | {inter},
                    child_prefix=("" if depth == 0 else None))
                statements = plan["statements"]
                synth, key_used = True, "_rowkey"
                root_name, children = plan["root"], plan["children"]
                hub_name = plan.get("hub")
            else:
                # Normal shape (row = record). Always synthesize the row key
                # (deterministic; no name-based guessing) and carry it on the
                # root and every child so one WHERE/join works everywhere.
                # each recursion level gets its OWN synthesized key
                # (_rowkey, _rowkey2, _rowkey3 ...) so a child's rows are
                # addressable independently of the parent fk it carries
                synth_name = ("_rowkey" if depth == 0
                              else "_rowkey%d" % (depth + 1))
                plan = _F.duckdb_unnest_plan(table, coltypes, base, key=None,
                                             detect=False,
                                             synth_key=synth_name,
                                             existing_names=existing,
                                             child_prefix=("" if depth == 0
                                                           else None))
                statements = plan["statements"]
                synth = bool(plan["synthesized"])
                key_used = plan["key"]   # the actual column, e.g. "_rowkey"
                root_name, children = plan["root"], plan["children"]
                hub_name = plan.get("hub")

            for st in statements:
                _tick()
                _run(st["sql"])
                if st["kind"] == "staging":
                    temp_tables.append(st["target"])
                else:
                    created.append(st["target"])

            for t in temp_tables:
                try:
                    _run("DROP TABLE IF EXISTS %s" % _F._qi(t))
                except Exception:
                    pass

            # Replace mode: drop the nested source and rename the freshly-built
            # root + children into the source's name space. The built tables are
            # real (CTAS-materialized) tables, so RENAME is safe; the source may
            # be a Parquet-cache VIEW, so its drop tries both view and table.
            # Each rename is independent, and the reported names track whatever
            # actually landed, so a partial failure never desyncs the catalog.
            if replace and final_root is not None:
                self._drop_named(mgr, [table])
                try:
                    _run("ALTER TABLE %s RENAME TO %s"
                         % (_F._qi(base), _F._qi(final_root)))
                    root_name = final_root
                    root_prefix = final_root
                except Exception:
                    root_prefix = base   # root kept its build name
                new_children = {}
                for child, srccol in list(children.items()):
                    # .501: only prefix-swap a child that actually carries the
                    # BUILD prefix. Path-only children ("json") are created
                    # under their final names, and blindly slicing len(base)
                    # off a shorter name produced a rename onto the ROOT.
                    if child.startswith(base + "_"):
                        target = root_prefix + child[len(base):]
                    else:
                        target = child
                    if target != child:
                        try:
                            _run("ALTER TABLE %s RENAME TO %s"
                                 % (_F._qi(child), _F._qi(target)))
                        except Exception:
                            target = child   # keep build name on failure
                    new_children[target] = srccol
                children = new_children
                if hub_name:
                    target = root_prefix + hub_name[len(base):]
                    if target != hub_name:
                        try:
                            _run("ALTER TABLE %s RENAME TO %s"
                                 % (_F._qi(hub_name), _F._qi(target)))
                            hub_name = target
                        except Exception:
                            pass

            try:
                mgr.sync_catalog()
            except Exception:
                pass

            # .423 DEEP RECURSION: a child that still carries LIST or
            # STRUCT columns (the sophis shape: list-of-records whose
            # element fields are records and arrays themselves) is
            # flattened again IN PLACE, one level per pass, each level
            # keyed by its own _rowkeyN chaining rows to the parent.
            extra_tables = []
            if depth < max_depth:
                for child in list(children.keys()):
                    try:
                        ct2 = mgr.column_types(child)
                    except Exception:
                        continue
                    if not any(_F._type_is_list(t) or _F._type_is_struct(t)
                               for t in ct2.values()):
                        continue
                    if len(mgr.table_columns) > 200:
                        break   # hard table budget: pathological docs
                    _tick()
                    sub = self._duckdb_unnest_flatten(
                        child, query_id=query_id,
                        on_progress=on_progress, cancel=cancel,
                        replace=True, depth=depth + 1,
                        max_depth=max_depth, _registered=True)
                    if sub and sub.get("cancelled"):
                        return {"cancelled": True}
                    if sub and sub.get("ok"):
                        extra_tables.extend(
                            t for t in sub.get("tables", [])
                            if t.get("name") != child)

            method = "duckdb-unnest"
            if promoted_from is not None:
                method = "duckdb-unnest (promoted '%s')" % promoted_from
            if depth == 0 and extra_tables:
                method += " deep"

            # .501: record the family (child -> parent) for the sidebar tree
            # and family joins; name-prefix guessing can't parent a path-only
            # child. The recursion records its own levels the same way.
            for _c in children:
                self._table_family[_c] = root_name
            if hub_name:
                self._table_family[hub_name] = root_name
            loaded = []
            seen_nm = set()
            for nm in ([root_name] + list(children.keys())
                       + ([hub_name] if hub_name else [])):
                if nm in seen_nm:
                    continue
                seen_nm.add(nm)
                rows = self._duck_scalar(
                    mgr, "SELECT count(*) FROM %s" % _F._qi(nm))
                cols = list(mgr.column_types(nm).keys())
                loaded.append({"name": nm, "rows": int(rows or 0),
                               "columns": cols, "engine": "duckdb"})
            for t in extra_tables:
                if t.get("name") not in seen_nm:
                    seen_nm.add(t.get("name"))
                    loaded.append(t)
            self._invalidate_profiles()
            self._invalidate_counts()
            return {"ok": True, "tables": loaded, "all": self.tables_tree(),
                    "method": method, "key": key_used, "synthesized": synth}
        except loaders.LoadCancelled:
            self._drop_named(mgr, temp_tables + created)
            return {"cancelled": True}
        except Exception as e:
            if _is_interrupt(e) or self._run_is_cancelled(query_id):
                self._drop_named(mgr, temp_tables + created)
                return {"cancelled": True}
            # Any other failure: clean partials and fall back to Python.
            self._drop_named(mgr, temp_tables + created)
            return None
        finally:
            if not _registered:
                self._unregister_run(query_id)

    def _family_members(self, root):
        """Every table whose parent chain (self._table_family) reaches
        ``root`` -- the set a re-flatten of ``root`` may replace."""
        fam = self._table_family
        out = set()
        for n in list(fam):
            cur, seen = n, set()
            while cur in fam and cur not in seen:
                seen.add(cur)
                cur = fam[cur]
            if cur == root:
                out.add(n)
        return out

    def _drop_named(self, mgr, names):
        """Best-effort drop of a list of DuckDB relations (partial-flatten
        cleanup, or dropping a nested source in replace mode). Tries both the
        table and the view form since a large-file source is a Parquet-cache
        view."""
        from . import flatten as _F
        for t in reversed(list(names or [])):
            # .513: a TEMP view under this name would survive the unqualified
            # drops below AND shadow whatever replaces the name -- kill it
            # first, fully qualified so a real main-schema view is untouched.
            try:
                with mgr.write_lock:
                    mgr.conn.execute(
                        'DROP VIEW IF EXISTS "temp"."main"."%s"'
                        % str(t).replace('"', '""'))
            except Exception:
                pass
            for kw in ("TABLE", "VIEW"):
                try:
                    with mgr.write_lock:
                        mgr.conn.execute(
                            "DROP %s IF EXISTS %s" % (kw, _F._qi(t)))
                except Exception:
                    pass
            # the engine must forget the dropped name too, or view-backing
            # bookkeeping can resurrect it later
            for reg in ("view_backing", "table_sources", "table_columns",
                        "table_origins"):
                try:
                    getattr(mgr, reg, {}).pop(t, None)
                except Exception:
                    pass
        try:
            mgr.sync_catalog()
        except Exception:
            pass

    def flatten_json(self, table, engine="duckdb", query_id=None,
                     on_progress=None, cancel=None, replace=False):
        """Flatten a JSON table into relational tables in the SAME engine the
        source table lives in: a DuckDB JSON table flattens into DuckDB, a
        SQLite one into SQLite.

        .501: ``replace=True`` (flatten-on-load) makes the flattened ROOT take
        over the nested source's name -- the loader always MEANT this ("the
        in-engine flatten replacing the nested table in place") but never
        passed it, so the nested table survived under the file's name, the
        flat root landed beside it as "<file>_2", and a SELECT * on the file
        name returned the raw json column while the sidebar showed the flat
        family (the on-box mismatch). Right-click Flatten keeps the default
        (source preserved, root beside it).

        DuckDB FIRST tries the vectorized in-engine UNNEST flatten (no
        JSON re-parse, no dump round trip, types come from the engine --
        the on-box 2026-07-02 '' -> INT64 conversion failure cannot occur
        there). Only when it reports nothing-to-do does the JSON path run.

        Normally this re-reads the original JSON file backing ``table``.
        When there is no file on disk -- e.g. a table loaded straight from
        the API connector into DuckDB, whose spooled response was a temp
        file we deleted after loading -- we instead dump the DuckDB table
        to a temporary JSON file (nested STRUCT/LIST columns are written
        back out as nested JSON) and flatten that. So flatten now works for
        any DuckDB JSON table, not just file-backed ones. Returns
        {ok, tables, all} or {error}."""
        mgr = self.duckdb if engine == "duckdb" else self.db
        if mgr is None:
            return {"error": "Engine not available."}
        if engine == "duckdb":
            # FAST PATH (.401): the in-engine flatten is now the DEEP
            # engine (full recursive tree) and it works from the
            # ALREADY-TYPED table -- no dump round trip, no JSON re-parse,
            # so a conversion-hostile value (the on-box '' -> INT64)
            # cannot occur. It registers/unregisters its own run and
            # returns None when there is nothing nested to flatten.
            try:
                fast = self._duckdb_unnest_flatten(
                    table, query_id=query_id, on_progress=on_progress,
                    cancel=cancel, replace=replace)
            except Exception as e:
                if _is_interrupt(e) or self._run_is_cancelled(query_id):
                    return {"cancelled": True}
                fast = None
            if fast is not None:
                return fast
        path = mgr.table_sources.get(table)
        tmp_dump = None
        if not _source_is_json_file(path):
            # Not a JSON file we can read directly -- a CSV/Parquet/text source,
            # a Parquet cache from a large load, or an API-loaded table with no
            # file. Dump the engine table to JSON (nested STRUCT/LIST columns are
            # written back out as nested JSON) and flatten that. This works for
            # any DuckDB table; SQLite has no dumper, so it still needs a JSON
            # source file on disk.
            if engine != "duckdb":
                if path:
                    return {"error": f"'{table}' is not a JSON-backed table."}
                return {"error": f"No source file on disk for '{table}'."}
            try:
                path = tmp_dump = self._duckdb_table_to_json(table)
            except Exception as e:
                return {"error": f"Couldn't read '{table}' from DuckDB to "
                        f"flatten: {type(e).__name__}: {e}"}
        # Register the load-target engine so Stop / window-close interrupts the
        # flatten load on the engine it's actually writing to; cancel_query's
        # safety net also interrupts the other engine, covering the (rare)
        # DuckDB dump pre-phase above.
        self._register_run(query_id, mgr, kind="flatten", target=table)
        before = self.snapshot_table_names()

        def _prog(done=0, total=0):
            # Make the flatten cooperatively cancellable: load_json and the
            # flattener call this between record batches and between insert
            # batches, so raising here aborts both the parse and write phases
            # promptly rather than letting the engine finish flattening.
            if (cancel and cancel()) or self._run_is_cancelled(query_id):
                raise loaders.LoadCancelled()
            # Heartbeat the progress registry: without this a long flatten
            # makes no progress signal, so the stall watchdog flags it as hung
            # after its threshold and its row count sits at 0 on the dashboard.
            # Stamping here keeps it off the stall list and shows live rows
            # (and drives the tray card's row count when run as a background job).
            opreg.beat(rows=done, op_id=query_id)
            if total:
                opreg.advance(done=done, total=total, op_id=query_id)
            if on_progress is not None:
                on_progress(done, total)

        try:
            if self._run_is_cancelled(query_id):
                return {"cancelled": True}
            loaded = loaders.load_json(mgr, path, base_name=table,
                                       progress=_prog, engine=engine)
        except Exception as e:
            if _is_interrupt(e) or self._run_is_cancelled(query_id):
                self.drop_tables_created_since(before)
                return {"cancelled": True}
            return {"error": err_str(e)}
        finally:
            self._unregister_run(query_id)
            if tmp_dump:
                try:
                    os.unlink(tmp_dump)
                except Exception:
                    pass
        self._invalidate_profiles()
        self._invalidate_counts()
        return {"ok": True, "tables": loaded, "all": self.tables_tree(),
                "method": "python", "key": None}

    def _duckdb_table_to_json(self, table):
        """Dump a DuckDB table to a temporary JSON file (nested types kept)
        so a table with no file on disk can still be flattened. The caller
        removes the file. Raises on failure."""
        duck = self.duckdb
        if duck is None:
            raise RuntimeError("DuckDB not available")
        path = tmputil.new_tempfile("flatten_", ".json")
        try:
            os.unlink(path)   # DuckDB's COPY writes the file itself
        except Exception:
            pass
        fwd = sqlutil.sql_path(path)
        qi = '"' + str(table).replace('"', '""') + '"'
        with duck.native_op_cursor() as (xc, lk):
            with lk:
                with _ExecKeepalive(xc, duck._cancel,
                                    threading.get_ident()):
                    xc.execute(
                        f"COPY (SELECT * FROM {qi}) TO '{fwd}' "
                        f"(FORMAT JSON)")
        if not os.path.isfile(path):
            raise RuntimeError("DuckDB produced no JSON output for the table")
        return path

    # ---- SQL helpers ------------------------------------------------
    def format_sql(self, sql, dialect=None):
        ok, out = sqlglot_transform(sql, read=dialect, write=dialect,
                                    pretty=True)
        return {"ok": ok, "result": out}

    def transpile_sql(self, sql, read, write):
        ok, out = sqlglot_transform(sql, read=read, write=write,
                                    pretty=True)
        return {"ok": ok, "result": out}

    # ---- export -----------------------------------------------------
    def table_properties(self, engine, table):
        """.538: the right-click Properties payload for a loaded table --
        format + source file (from the restore manifest), row/column
        counts, and HOW it is stored in SamQL (engine, table vs
        query-in-place view, the backing files). Every field
        best-effort; missing pieces are stated, never invented."""
        e = (engine or "").lower()
        if e in ("sqlite", "local", LOCAL_TARGET):
            eng, ename = self.db, "sqlite"
        elif e in ("duckdb", DUCKDB_TARGET):
            eng, ename = getattr(self, "duckdb", None), "duckdb"
        else:
            eng, ename = None, e
        if eng is None:
            return {"error": "engine '%s' is not available" % engine}
        out = {"engine": ename, "name": table}
        # rows / columns
        try:
            out["rows"] = self._cached_count(eng, ename, table)
        except Exception:
            out["rows"] = None
        try:
            out["columns"] = len(eng.column_types(table) or {})
        except Exception:
            out["columns"] = None
        # object kind + backing files
        obj = "table"
        view_sql = None
        try:
            if ename == "duckdb":
                r = eng.execute_read(
                    "select sql from duckdb_views() where view_name = ?",
                    [table]).fetchone()
                if r:
                    obj = "view"
                    view_sql = r[0] or ""
            else:
                r = eng.execute_read(
                    "select type, sql from sqlite_master where name = ?",
                    [table]).fetchone()
                if r and (r[0] or "").lower() == "view":
                    obj = "view"
                    view_sql = r[1] or ""
        except Exception:
            pass
        stored = {"engine": ename, "object": obj}
        try:
            stored["database"] = getattr(eng, "db_path", None) or "in-memory"
        except Exception:
            pass
        if view_sql:
            import re as _re
            m = _re.search(r"read_(?:parquet|json_auto|csv_auto)\('"
                           r"([^']+)'", view_sql)
            if m:
                pq = m.group(1)
                stored["backing_file"] = pq
                try:
                    stored["backing_bytes"] = os.path.getsize(pq)
                except OSError:
                    stored["backing_bytes"] = None
        stored["note"] = (
            "query-in-place: a view reading its converted cache directly"
            if obj == "view" else "materialized rows inside the engine")
        out["stored"] = stored
        # source, from the restore manifest (newest match wins). Older
        # entries carry no base_name; the file STEM is the fallback key
        # (loads auto-name from it), and a folder entry is the last
        # resort for its per-file children.
        import re as _re2

        def _norm(x):
            return _re2.sub(r"[^a-z0-9]+", "_", (x or "").lower()).strip("_")

        tnorm = _norm(table)
        src = None
        folder_src = None
        for e in reversed(self.manifest.all()):
            cand = e.get("base_name") or os.path.splitext(
                os.path.basename(e.get("path") or ""))[0]
            cn = _norm(cand)
            if cn and (cn == tnorm or tnorm.startswith(cn + "_")):
                src = e
                break
            if e.get("kind") == "folder" and folder_src is None:
                folder_src = e
        if src is None and folder_src is not None:
            src = dict(folder_src)
            src["_folder_only"] = True
        if src:
            path = src.get("path") or ""
            ext = os.path.splitext(path)[1].lstrip(".").lower()
            fmt = ext or src.get("kind")
            info = {"format": fmt,
                    "path": path,
                    "file": os.path.basename(path),
                    "folder": os.path.dirname(path),
                    "loaded_at": src.get("ts")}
            # .550: the ORIGINAL source location/name, distinct from the
            # converted/parked ``path`` above. A native or folder load
            # carries a real filesystem path; a browser upload carries
            # only the original filename (browsers withhold the full
            # path for security), so we surface whichever we have and
            # say which it is.
            origin = src.get("origin")
            if origin:
                if os.path.isabs(origin) or (":" in origin[:3]) \
                        or origin.startswith("\\\\"):
                    info["original_path"] = origin
                    info["original_folder"] = os.path.dirname(origin)
                    info["original_file"] = os.path.basename(origin)
                else:
                    info["original_file"] = origin
                    info["original_note"] = (
                        "uploaded via the browser, which provides only "
                        "the file name, not its full path")
            try:
                st = os.stat(path)
                info["bytes"] = st.st_size
                info["modified"] = _dt.datetime.fromtimestamp(
                    st.st_mtime).isoformat(timespec="seconds")
            except OSError:
                info["missing"] = True
            if src.get("_folder_only"):
                info["note"] = ("loaded from this folder; the exact "
                                "per-file source was not recorded")
                info["file"] = None
                out["format"] = None
            out["format"] = fmt
            out["source"] = info
        else:
            out["format"] = None
            out["source"] = {
                "note": "No source file recorded -- this table was "
                        "created in-app (a query result, a flatten, or "
                        "a node output)."}
        return out

    def export(self, result_id, fmt, out_path=None,
               sort_col=None, descending=False, query_id=None):
        cr = self._results.get(result_id)
        if cr is None:
            return {"error": "result expired"}
        fmt = (fmt or "csv").lower()
        if out_path is None:
            out_path = tmputil.new_tempfile("export_", f".{fmt}")
        # Bind the result store's engine/conn (parquet DuckDB / spill SQLite),
        # not session.duckdb -- otherwise Stop never reaches the COPY.
        eng = None
        store = getattr(cr, "store", None)
        if store is not None:
            eng = (getattr(store, "_conn", None)
                   or getattr(store, "_engine", None))
        if eng is None:
            try:
                eng = self.get_duckdb()
            except Exception:
                eng = getattr(self, "db", None)
        # .533: exports register like any run, so the Activity tray shows
        # a cancellable "export" card; Stop flags the id (cooperative
        # checks in the row loops) AND interrupts the engine COPY.
        self._register_run(query_id, eng,
                           kind="export", target="%s export" % fmt)
        try:
            return self._export_inner(cr, fmt, out_path, sort_col,
                                      descending, query_id)
        finally:
            self._end_run_keep_cancel(query_id)

    def _export_inner(self, cr, fmt, out_path,
                      sort_col, descending, query_id):
        # .472: a parquet-backed result exports with ONE engine COPY
        # (csv / parquet / json / ndjson) -- ordered by the stable row
        # number, or by the requested sort -- instead of iterating
        # every row through Python (the reported minutes). xlsx and
        # non-parquet stores keep the streaming writers below.
        def _xc():
            return bool(query_id) and self._run_is_cancelled(query_id)

        if _xc():
            return {"cancelled": True}
        pq = getattr(cr.store, "parquet_path", None)
        pq = pq() if callable(pq) else None
        if pq and fmt in ("csv", "tsv", "parquet", "json", "ndjson"):
            try:
                self._export_from_parquet(
                    cr, pq, fmt, out_path, sort_col, descending)
                return {"ok": True, "path": out_path, "rows": cr.total}
            except Exception as e:
                from .engines import _is_interrupt as _isi
                if _xc() or _isi(e):
                    try:
                        os.unlink(out_path)
                    except OSError:
                        pass
                    return {"cancelled": True}
                pass  # any surprise: the streaming writers below
        view = cr.view(sort_col, descending)
        if _xc():
            return {"cancelled": True}
        view = _CancelIter(view, _xc)
        try:
            if fmt in ("csv", "tsv"):
                self._export_csv(out_path, cr.cols, view,
                                 delim="\t" if fmt == "tsv" else ",")
            elif fmt in ("json", "ndjson"):
                self._export_json(out_path, cr.cols, view,
                                  ndjson=(fmt == "ndjson"))
            elif fmt in ("xlsx", "excel"):
                self._export_xlsx(out_path, cr.cols, view)
            elif fmt == "parquet":
                self._export_parquet(out_path, cr.cols, view)
            else:
                return {"error": f"Unsupported format: {fmt}"}
        except _ExportCancelled:
            try:
                os.unlink(out_path)
            except OSError:
                pass
            return {"cancelled": True}
        return {"ok": True, "path": out_path, "rows": cr.total}

    @staticmethod
    def _export_from_parquet(cr, pq_path, fmt, out_path,
                             sort_col=None, descending=False):
        """One COPY statement straight off the result's parquet backing.
        Runs on the store's own engine via a read cursor (no write-lock
        contention; interruptible via the .471 cursor registration)."""
        eng = cr.store._engine
        qcols = ", ".join('"%s"' % str(c).replace('"', '""')
                          for c in cr.cols)
        order = "file_row_number"
        if sort_col and sort_col in cr.cols:
            order = ('"%s" %s, file_row_number'
                     % (str(sort_col).replace('"', '""'),
                        "DESC" if descending else "ASC"))
        sel = ("SELECT %s FROM read_parquet('%s', file_row_number = true) "
               "ORDER BY %s"
               % (qcols, sqlutil.sql_path(pq_path), order))
        opts = {"csv": "FORMAT CSV, HEADER",
                "tsv": "FORMAT CSV, HEADER, DELIMITER '\t'",
                "parquet": "FORMAT PARQUET",
                "ndjson": "FORMAT JSON",
                "json": "FORMAT JSON, ARRAY true"}[fmt]
        eng.execute_read("COPY (%s) TO '%s' (%s)"
                         % (sel, sqlutil.sql_path(out_path), opts))

    @staticmethod
    def _export_csv(path, cols, view, bom=False, delim=","):
        enc = "utf-8-sig" if bom else "utf-8"
        with open(path, "w", encoding=enc, newline="") as f:
            w = _csv.writer(f, delimiter=delim)
            w.writerow(cols)
            for r in view:
                w.writerow(["" if v is None else json_safe(v) for v in r])

    @staticmethod
    def _export_json(path, cols, view, ndjson=False):
        with open(path, "w", encoding="utf-8") as f:
            if ndjson:
                for r in view:
                    f.write(json.dumps(
                        {c: json_safe(v) for c, v in zip(cols, r)},
                        default=str))
                    f.write("\n")
            else:
                f.write("[\n")
                first = True
                for r in view:
                    if not first:
                        f.write(",\n")
                    first = False
                    f.write(json.dumps(
                        {c: json_safe(v) for c, v in zip(cols, r)},
                        default=str))
                f.write("\n]\n")

    @staticmethod
    def _export_xlsx(path, cols, view):
        try:
            import openpyxl
        except Exception:
            raise RuntimeError(
                "xlsx export requires openpyxl (pip install openpyxl)")
        wb = openpyxl.Workbook(write_only=True)
        ws = wb.create_sheet("results")
        ws.append(list(cols))
        # .451 [PLAN PASS 8] fix: openpyxl promotes any string starting
        # with "=" to a LIVE FORMULA cell -- grid DATA became executable
        # Excel content on export (classic spreadsheet injection), and
        # the literal text was lost. Force every string cell to stay a
        # string; numbers and None pass through untouched.
        from openpyxl.cell.cell import WriteOnlyCell
        for r in view:
            out = []
            for v in r:
                v = json_safe(v)
                c = WriteOnlyCell(ws, value=v)
                if getattr(c, "data_type", None) == "f":
                    c.data_type = "s"
                out.append(c)
            ws.append(out)
        wb.save(path)

    @staticmethod
    def _export_parquet(path, cols, view):
        """Stream rows to Parquet in batches — never accumulate the full
        result as Python column lists (OOM on millions of rows)."""
        try:
            import pyarrow as pa
            import pyarrow.parquet as pq
        except Exception:
            raise RuntimeError(
                "parquet export requires pyarrow (pip install pyarrow)")
        writer = None
        batch_cols = {c: [] for c in cols}
        n_batch = 0
        chunk = 5000

        def _flush():
            nonlocal writer, n_batch, batch_cols
            if n_batch == 0:
                return
            table = pa.table({c: batch_cols[c] for c in cols})
            if writer is None:
                writer = pq.ParquetWriter(path, table.schema)
            writer.write_table(table)
            batch_cols = {c: [] for c in cols}
            n_batch = 0

        for r in view:
            for i, c in enumerate(cols):
                batch_cols[c].append(
                    json_safe(r[i]) if i < len(r) else None)
            n_batch += 1
            if n_batch >= chunk:
                _flush()
        _flush()
        if writer is not None:
            writer.close()
        else:
            # empty result: still write a valid schema-only file
            pq.write_table(
                pa.table({c: pa.array([], type=pa.string()) for c in cols}),
                path)

    # ---- persistence: history / saved -------------------------------
    def history_all(self):
        return self.history.all()

    def history_clear(self):
        self.history.clear()
        return {"ok": True}

    def saved_all(self):
        return self.saved.all()

    def saved_upsert(self, name, sql, tags=None):
        return self.saved.upsert(name, sql, tags)

    def saved_delete(self, name):
        return {"ok": self.saved.delete(name)}

    # ---- saved workflows (IDE script / Journal doc / Node graph) ----
    def _migrate_saved_queries_to_workflows(self):
        """One-time: bring pre-existing IDE saved queries (the legacy
        saved.json store) into the unified Saved Workflows store as kind="ide",
        so they keep showing in the sidebar after the panel switched to the
        workflow store. Runs once (guarded by a config flag), and never
        re-imports an entry the user has since deleted."""
        if self.config.get("ide_saved_migrated"):
            return
        try:
            for s in (self.saved.all() or []):
                name = (s.get("name") or "").strip()
                if not name:
                    continue
                if self.workflows.get(name, "ide") is None:
                    self.workflows.upsert(
                        name,
                        {"sql": s.get("sql") or "", "tags": s.get("tags") or []},
                        "ide",
                    )
        except Exception:
            pass
        self.config.set("ide_saved_migrated", True)

    def workflows_all(self):
        return {"workflows": self.workflows.all()}

    def workflow_get(self, name, kind="node"):
        e = self.workflows.get(name, kind)
        if e is None:
            return {"error": "No saved workflow named %r." % (name or "")}
        return e

    def workflow_save(self, name, graph, kind="node"):
        e = self.workflows.upsert(name, graph or {}, kind)
        if e is None:
            return {"error": "Give the workflow a name to save it."}
        return {"ok": True, "name": e.get("name"), "kind": e.get("kind")}

    def workflow_delete(self, name, kind="node"):
        return {"ok": self.workflows.delete(name, kind)}

    def _post_interrupt_hygiene(self, target):
        """After a cancel lands, leave the engine connection CLEAN: an
        interrupt can strand an open transaction, and the next statement on
        that connection then fails with 'transaction aborted' instead of
        running (part of the cancel-then-Failed-to-fetch sequence seen
        on-box 2026-07-02). Best-effort ROLLBACK; 'no transaction active'
        is the healthy outcome and is swallowed."""
        try:
            eng, _k = self._engine_obj(target)
            conn = getattr(eng, "conn", None)
            lock = getattr(eng, "write_lock", None)
            if conn is None:
                return
            if lock is not None and not lock.acquire(timeout=0.5):
                return   # teardown still owns it; the next run copes
            try:
                conn.execute("ROLLBACK")
            except Exception:
                pass
            finally:
                if lock is not None:
                    lock.release()
        except Exception:
            pass

    def _flow_node_label(self, graph, node_id):
        """Human label for a flow run: the target node's label, else its type,
        else its id -- shown in the stat modal's running-queries list."""
        try:
            for n in (graph or {}).get("nodes") or []:
                if n.get("id") == node_id:
                    cfg = n.get("config") or {}
                    return (cfg.get("label") or n.get("type")
                            or node_id or "flow")
        except Exception:
            pass
        return node_id or "flow"

    def _register_run(self, query_id, engine, kind="query", target=None,
                      surface=None, label=None):
        """Register a flow's engine so cancel_query(query_id) can interrupt the
        currently-executing build statement on it, and register the operation
        in the progress registry so it shows on the activity dashboard and the
        stall watchdog can see it."""
        if query_id and engine is not None:
            with self._running_lock:
                self._running[query_id] = {
                    "engine": engine,
                    "tid": threading.get_ident(),
                }
        if query_id:
            ekind = getattr(engine, "ENGINE_KIND", None) or (
                "sqlite" if engine is not None else None)
            try:
                opreg.begin(kind, target=target, engine=ekind,
                               op_id=query_id,
                               surface=surface, label=label)
            except Exception:
                pass

    def _unregister_run(self, query_id):
        if query_id:
            with self._running_lock:
                self._running.pop(query_id, None)
                self._cancelled_runs.discard(query_id)
            try:
                opreg.end(query_id)
            except Exception:
                pass

    def _end_run_keep_cancel(self, query_id):
        """Drop the in-flight binding but keep the cancelled flag.

        Pivot/chart/reconcile need this so a Stop that lands after the engine
        unwinds still reports as cancelled (not a generic failure).
        """
        if not query_id:
            return
        with self._running_lock:
            self._running.pop(query_id, None)
        try:
            opreg.end(query_id)
        except Exception:
            pass

    def _run_op(self, query_id, kind, target, work):
        """Run a synchronous engine read (reconcile / pivot / chart / field
        profile) as a cancellable, dashboard-visible operation. Registers it
        under a run id -- generated when the caller didn't pass one -- so it
        shows in /api/status with an inline cancel and cancel_query interrupts
        its in-flight engine statement. Returns the cancel sentinel instead of
        raising on interrupt -- and also if a cancel landed during a step the
        callee swallowed -- so a user cancel is never reported as a failure.
        NOTE: this stops *engine* work; a Python-side write loop (result
        export) is not interrupted by it and is handled separately."""
        qid = query_id or ("op-%s-%s" % (kind, uuid.uuid4().hex[:8]))
        # Pivot / chart / reconcile: same sticky-cancel trap as profile.
        self._clear_stale_engine_cancel(except_qid=qid)
        self._register_run(qid, None, kind=kind, target=target)
        try:
            if self._run_is_cancelled(qid):
                return {"error": "cancelled", "cancelled": True}
            result = work()
            if self._run_is_cancelled(qid):
                return {"error": "cancelled", "cancelled": True}
            return result
        except Exception as e:
            if _is_interrupt(e) or self._run_is_cancelled(qid):
                return {"error": "cancelled", "cancelled": True}
            raise
        finally:
            self._end_run_keep_cancel(qid)

    # ---- charting ---------------------------------------------------
    def _engine_kind_for_table(self, table):
        """Where a loaded table actually lives: 'duckdb' or 'sqlite', or None
        if it isn't in either local engine. SQLite is preferred if (unusually)
        the same name exists in both. Lets chart/pivot/profile resolve a
        table's engine by where it really is, so an absent or stale engine
        hint from the UI can't route a DuckDB table to SQLite (or vice versa).
        """
        if table and table in self.db.table_columns:
            return "sqlite"
        if (table and self.duckdb is not None
                and table in self.duckdb.table_columns):
            return "duckdb"
        return None

    def _source_rows(self, result_id=None, table=None, engine="sqlite",
                     limit=200000):
        """Return (cols, iterable_of_rows) for a cached result or a table.
        Used by chart/pivot aggregation. Capped at ``limit`` rows."""
        if result_id is not None:
            cr = self._results.get(result_id)
            if cr is None:
                return None, None
            return list(cr.cols), cr.store
        if table is not None:
            engine = self._engine_kind_for_table(table) or engine
            eng = self.get_duckdb() if engine == "duckdb" else self.db
            cols, first, cursor = eng.execute_cursor(
                f'SELECT * FROM "{table}"', batch=10000)
            if cols is None:
                return [], []
            rows = list(first)
            if cursor is not None:
                while len(rows) < limit:
                    batch = cursor.fetchmany(10000)
                    if not batch:
                        break
                    rows.extend(batch)
                try:
                    cursor.close()
                except Exception:
                    pass
            return cols, rows[:limit]
        return None, None

    def _agg_source(self, spec):
        """Return (ctx, cols) for chart/pivot aggregation. ``ctx`` is a
        SQL-aggregation context (run, src, colref, castnum) when the source
        can aggregate in the engine -- a spilled result, a Parquet result,
        or a table -- else None so the caller uses the small in-memory
        Python path. ``cols`` is the column-name list (or None if gone)."""
        rid = spec.get("result_id")
        if rid is not None:
            cr = self._results.get(rid)
            if cr is None:
                return None, None
            store = cr.store
            ctx = (store._agg_ctx()
                   if hasattr(store, "_agg_ctx") and len(store) > 0
                   else None)
            return ctx, list(cr.cols)
        table = spec.get("table")
        if table:
            kind = (self._engine_kind_for_table(table)
                    or spec.get("engine") or "sqlite")
            eng = self.get_duckdb() if kind == "duckdb" else self.db
            try:
                cols, _ = eng.execute(f'SELECT * FROM "{table}" LIMIT 0')
            except Exception:
                return None, None
            if cols is None:
                return None, []
            names = list(cols)
            qt = '"' + str(table).replace('"', '""') + '"'

            def run(sql):
                _c, rows = eng.execute(sql)
                return rows or []
            colref = (lambda i: '"'
                      + str(names[int(i)]).replace('"', '""') + '"')
            castnum = ((lambda e: f"TRY_CAST({e} AS DOUBLE)")
                       if kind == "duckdb"
                       else (lambda e: f"CAST({e} AS REAL)"))
            return (run, qt, colref, castnum), names
        return None, None

    def chart_data(self, spec):
        """Aggregate a cached result (or table) into chart-ready series.

        spec = {result_id|table, engine, chart_type, x, y, agg, bins,
                limit, query_id}. Aggregation is pushed into the engine
        (SQLite for a spilled result, DuckDB for a Parquet result, or the
        table's own engine) so a multi-million-row source never crosses into
        Python; only small in-memory results take the Python path. Returns
        {chart_type, labels, series:[{name, points|values}]} or {error}.
        Cancelable via query_id like pivot -- Stop interrupts the aggregate.
        """
        qid = spec.get("query_id") or ("chart-%s" % uuid.uuid4().hex[:10])
        eng0 = None
        try:
            eng0 = (self.get_duckdb()
                    if (spec.get("engine") or "sqlite") == "duckdb"
                    else self.db)
        except Exception:
            eng0 = None
        self._register_run(
            qid, eng0, kind="chart", surface="chart",
            label=str(spec.get("table") or spec.get("result_id") or "result"))
        try:
            return self._chart_data_inner(spec, qid)
        finally:
            self._end_run_keep_cancel(qid)

    def _chart_data_inner(self, spec, qid=None):
        chart_type = (spec.get("chart_type") or "bar").lower()
        x = spec.get("x")
        y = spec.get("y")
        s = spec.get("series")  # optional split dimension -> multiple series
        agg = (spec.get("agg") or "sum").lower()
        limit = int(spec.get("limit") or 100000)
        ctx, cols = self._agg_source(spec)
        target = self._agg_interrupt_target(spec)
        if qid and target is not None:
            self._register_run(
                qid, target, kind="chart", surface="chart",
                label=str(spec.get("table")
                          or spec.get("result_id") or "result"))
        if qid and self._run_is_cancelled(qid):
            return {"error": "interrupted", "cancelled": True}
        if cols is None:
            return {"error": "result expired"}
        if x not in cols:
            return {"error": f"Unknown column: {x}"}
        xi = cols.index(x)
        try:
            if chart_type == "candlestick":
                return self._chart_candlestick(spec, cols, xi)
            if chart_type in ("multix", "multi_x"):
                return self._chart_multix(spec, cols)
            if chart_type in ("multiy", "multi_y"):
                return self._chart_multiy(spec, cols)
            yi = cols.index(y) if (y in cols) else None
            si = cols.index(s) if (s and s in cols and s != x) else None
            if qid and self._run_is_cancelled(qid):
                return {"error": "interrupted", "cancelled": True}
            if ctx is not None:
                out = self._chart_sql(ctx, chart_type, x, y, xi, yi, agg, spec, si)
            else:
                _c, rows = self._source_rows(
                    result_id=spec.get("result_id"), table=spec.get("table"),
                    engine=spec.get("engine", "sqlite"), limit=limit)
                if qid and self._run_is_cancelled(qid):
                    return {"error": "interrupted", "cancelled": True}
                out = self._chart_python(rows, chart_type, x, y, xi, yi, agg, spec, si)
            if qid and self._run_is_cancelled(qid):
                return {"error": "interrupted", "cancelled": True}
            return out
        except Exception as e:
            if _is_interrupt(e) or (qid and self._run_is_cancelled(qid)):
                return {"error": "interrupted", "cancelled": True}
            return {"error": str(e)}

    def _finish_series(self, pairs, chart_type, x, y, yi, agg):
        """Order grouped (key, value) pairs deterministically and shape the
        chart series. >200 categories are capped to the top 200 by value
        (readability); otherwise they are ordered by the x key, numerically
        when the keys are numeric (correct for line charts) else as text."""
        if len(pairs) > 200:
            pairs = sorted(pairs, key=lambda p: p[1], reverse=True)[:200]
        else:
            def _sk(p):
                k = p[0]
                if k is None:
                    return (0, 0.0, "")
                try:
                    return (1, float(k), "")
                except (TypeError, ValueError):
                    return (2, 0.0, str(k))
            pairs = sorted(pairs, key=_sk)
        labels = ["" if k is None else str(k) for k, _ in pairs]
        values = [v for _, v in pairs]
        name = f"{agg}({y})" if yi is not None else "count"
        return {"chart_type": chart_type, "x": x, "labels": labels,
                "series": [{"name": name, "values": values}]}

    def _finish_multi_series(self, grouped, chart_type, x, y, yi, agg):
        """Pivot (x, series, value) rows into shared x labels + one values[]
        per series -- this is what backs grouped / stacked bars and multi-line
        charts. x labels and series are both ordered deterministically; missing
        (x, series) cells are 0. Capped to the top 200 x labels (by total) and
        30 series (by total) so a wide split can't produce an unreadable chart
        or a runaway payload."""
        def _sk(k):
            if k is None:
                return (0, 0.0, "")
            try:
                return (1, float(k), "")
            except (TypeError, ValueError):
                return (2, 0.0, str(k))
        cell = {}
        xtot = {}
        stot = {}
        for row in grouped:
            xv, sv = row[0], row[1]
            val = float(row[-1]) if row[-1] is not None else 0.0
            cell[(xv, sv)] = cell.get((xv, sv), 0.0) + val
            xtot[xv] = xtot.get(xv, 0.0) + val
            stot[sv] = stot.get(sv, 0.0) + val
        xkeys = list(xtot.keys())
        if len(xkeys) > 200:
            xkeys = sorted(xkeys, key=lambda k: xtot[k], reverse=True)[:200]
        xkeys = sorted(xkeys, key=_sk)
        skeys = sorted(stot.keys(), key=lambda k: stot[k], reverse=True)[:30]
        skeys = sorted(skeys, key=_sk)
        labels = ["" if k is None else str(k) for k in xkeys]
        series = [{"name": "" if sv is None else str(sv),
                   "values": [cell.get((xk, sv), 0.0) for xk in xkeys]}
                  for sv in skeys]
        return {"chart_type": chart_type, "x": x, "labels": labels,
                "series": series}

    def _chart_multix(self, spec, cols):
        # Two independent (x, y) series, each aggregated and ordered on its own
        # x-axis (one on the bottom, one on the top). Aggregated in Python over
        # up to `limit` rows -- multi-x charts are a couple of trend lines.
        def idx(key):
            c = spec.get(key)
            return cols.index(c) if (c and c in cols) else None
        x1i, y1i = idx("x"), idx("y")
        x2i, y2i = idx("x2"), idx("y2")
        if x1i is None or x2i is None:
            return {"error": "Multiple X axes needs two X columns (x and x2)."}
        _c, rows = self._source_rows(
            result_id=spec.get("result_id"), table=spec.get("table"),
            engine=spec.get("engine", "sqlite"),
            limit=int(spec.get("limit") or 100000))
        agg = (spec.get("agg") or "sum").lower()

        def num(v):
            try:
                return float(v)
            except (TypeError, ValueError):
                return None

        def apply(vs):
            if agg == "count":
                return float(len(vs))
            if not vs:
                return 0.0
            if agg == "avg":
                return sum(vs) / len(vs)
            if agg == "min":
                return min(vs)
            if agg == "max":
                return max(vs)
            return sum(vs)

        def _sk(k):
            if k is None:
                return (0, 0.0, "")
            try:
                return (1, float(k), "")
            except (TypeError, ValueError):
                return (2, 0.0, str(k))

        def series_for(xi, yi):
            groups = {}
            for r in rows:
                k = r[xi]
                lst = groups.get(k)
                if lst is None:
                    lst = []
                    groups[k] = lst
                if yi is not None:
                    v = num(r[yi])
                    if v is not None:
                        lst.append(v)
                else:
                    lst.append(1.0)
            keys = sorted(groups.keys(), key=_sk)
            labels = ["" if k is None else str(k) for k in keys]
            values = [apply(groups[k]) for k in keys]
            return labels, values

        labels1, vals1 = series_for(x1i, y1i)
        labels2, vals2 = series_for(x2i, y2i)
        nm1 = (("%s(%s)" % (agg, spec.get("y"))) if y1i is not None
               else spec.get("x"))
        nm2 = (("%s(%s)" % (agg, spec.get("y2"))) if y2i is not None
               else spec.get("x2"))
        return {"chart_type": "multix", "x": spec.get("x"),
                "labels": labels1, "labels2": labels2,
                "series": [
                    {"name": nm1, "values": vals1, "xAxisIndex": 0},
                    {"name": nm2, "values": vals2, "xAxisIndex": 1}]}

    def _chart_multiy(self, spec, cols):
        # One shared category x-axis with two metrics (y, y2), each aggregated
        # over the same x keys and rendered on its own y-axis (left + right).
        def idx(key):
            c = spec.get(key)
            return cols.index(c) if (c and c in cols) else None
        xi, y1i, y2i = idx("x"), idx("y"), idx("y2")
        if xi is None or y1i is None or y2i is None:
            return {"error": "Multiple Y axes needs an X column and two Y "
                    "columns (y and y2)."}
        _c, rows = self._source_rows(
            result_id=spec.get("result_id"), table=spec.get("table"),
            engine=spec.get("engine", "sqlite"),
            limit=int(spec.get("limit") or 100000))
        agg = (spec.get("agg") or "sum").lower()

        def num(v):
            try:
                return float(v)
            except (TypeError, ValueError):
                return None

        def apply(vs):
            if agg == "count":
                return float(len(vs))
            if not vs:
                return 0.0
            if agg == "avg":
                return sum(vs) / len(vs)
            if agg == "min":
                return min(vs)
            if agg == "max":
                return max(vs)
            return sum(vs)

        def _sk(k):
            if k is None:
                return (0, 0.0, "")
            try:
                return (1, float(k), "")
            except (TypeError, ValueError):
                return (2, 0.0, str(k))

        g1, g2 = {}, {}
        for r in rows:
            k = r[xi]
            g1.setdefault(k, [])
            g2.setdefault(k, [])
            v1 = num(r[y1i])
            if v1 is not None:
                g1[k].append(v1)
            v2 = num(r[y2i])
            if v2 is not None:
                g2[k].append(v2)
        keys = sorted(g1.keys(), key=_sk)
        if len(keys) > 1000:
            keys = keys[:1000]
        labels = ["" if k is None else str(k) for k in keys]
        nm1 = "%s(%s)" % (agg, spec.get("y"))
        nm2 = "%s(%s)" % (agg, spec.get("y2"))
        return {"chart_type": "multiy", "x": spec.get("x"), "labels": labels,
                "series": [
                    {"name": nm1, "values": [apply(g1[k]) for k in keys],
                     "yAxisIndex": 0},
                    {"name": nm2, "values": [apply(g2[k]) for k in keys],
                     "yAxisIndex": 1}]}

    def _chart_candlestick(self, spec, cols, xi):
        # OHLC chart: one row per x (date), four price columns. We don't
        # aggregate -- the source is assumed to already be per-period prices.
        def idx(key):
            c = spec.get(key)
            return cols.index(c) if (c and c in cols) else None
        oi, hi_, li, ci = idx("open"), idx("high"), idx("low"), idx("close")
        if None in (oi, hi_, li, ci):
            return {"error": "Candlestick needs open, high, low and close "
                             "columns."}
        _c, rows = self._source_rows(
            result_id=spec.get("result_id"), table=spec.get("table"),
            engine=spec.get("engine", "sqlite"),
            limit=int(spec.get("limit") or 100000))

        def num(v):
            try:
                return float(v)
            except (TypeError, ValueError):
                return None
        labels, ohlc = [], []
        for r in rows:
            o, h, lo, c = num(r[oi]), num(r[hi_]), num(r[li]), num(r[ci])
            if None in (o, h, lo, c):
                continue
            labels.append("" if r[xi] is None else str(r[xi]))
            ohlc.append([o, c, lo, h])  # ECharts order: open, close, low, high
        if not ohlc:
            return {"error": "No numeric OHLC rows to plot."}
        return {"chart_type": "candlestick", "x": spec.get("x"),
                "labels": labels,
                "series": [{"name": spec.get("x") or "price", "ohlc": ohlc}]}

    def _chart_sql(self, ctx, chart_type, x, y, xi, yi, agg, spec, si=None):
        if chart_type == "scatter":
            if yi is None:
                return {"error": "Scatter needs both an x and a y column."}
            pts = [{"x": a, "y": b} for a, b in agg_xy(ctx, xi, yi, 5000)]
            return {"chart_type": "scatter", "x": x, "y": y,
                    "series": [{"name": f"{y} vs {x}", "points": pts}]}
        if chart_type == "histogram":
            lo, hi = agg_minmax(ctx, xi)
            if lo is None:
                return {"error": "No numeric values to bin."}
            nbins = max(1, int(spec.get("bins") or 20))
            if hi <= lo:
                hi = lo + 1.0
            counts = agg_histogram(ctx, xi, lo, hi, nbins)
            width = (hi - lo) / nbins
            labels = [f"{lo + i*width:.4g}" for i in range(nbins)]
            return {"chart_type": "histogram", "x": x, "labels": labels,
                    "series": [{"name": x, "values": counts}]}
        # split by a second dimension -> one series per distinct split value
        if si is not None and chart_type not in ("pie", "donut"):
            grouped = agg_group_multi(ctx, [xi, si], yi, agg, cap=100000)
            return self._finish_multi_series(grouped, chart_type, x, y, yi, agg)
        # bar / line / pie -> group by x, aggregate y in SQL
        grouped = agg_group_multi(ctx, [xi], yi, agg, cap=100000)
        pairs = [(row[0], float(row[-1]) if row[-1] is not None else 0.0)
                 for row in grouped]
        return self._finish_series(pairs, chart_type, x, y, yi, agg)

    def _chart_python(self, rows, chart_type, x, y, xi, yi, agg, spec, si=None):
        def _num(v):
            try:
                return float(v)
            except (TypeError, ValueError):
                return None
        if chart_type == "scatter":
            pts = []
            for r in rows:
                xv = _num(r[xi])
                yv = _num(r[yi]) if yi is not None else None
                if xv is not None and yv is not None:
                    pts.append({"x": xv, "y": yv})
                if len(pts) >= 5000:
                    break
            return {"chart_type": "scatter", "x": x, "y": y,
                    "series": [{"name": f"{y} vs {x}", "points": pts}]}
        if chart_type == "histogram":
            vals = [v for v in (_num(r[xi]) for r in rows) if v is not None]
            if not vals:
                return {"error": "No numeric values to bin."}
            nbins = max(1, int(spec.get("bins") or 20))
            lo, hi = min(vals), max(vals)
            if hi <= lo:
                hi = lo + 1.0
            width = (hi - lo) / nbins
            counts = [0] * nbins
            for v in vals:
                idx = min(nbins - 1, int((v - lo) / width))
                counts[idx] += 1
            labels = [f"{lo + i*width:.4g}" for i in range(nbins)]
            return {"chart_type": "histogram", "x": x, "labels": labels,
                    "series": [{"name": x, "values": counts}]}
        groups = {}
        for r in rows:
            key = r[xi]
            if key not in groups:
                groups[key] = []
            if yi is not None:
                nv = _num(r[yi])
                if nv is not None:
                    groups[key].append(nv)
            else:
                groups[key].append(1.0)

        def _apply(values):
            if agg == "count":
                return float(len(values))
            if not values:
                return 0.0
            if agg == "sum":
                return sum(values)
            if agg == "avg":
                return sum(values) / len(values)
            if agg == "min":
                return min(values)
            if agg == "max":
                return max(values)
            return sum(values)
        # split by a second dimension -> one series per distinct split value
        if si is not None and chart_type not in ("pie", "donut"):
            mg = {}
            for r in rows:
                k = (r[xi], r[si])
                lst = mg.get(k)
                if lst is None:
                    lst = []
                    mg[k] = lst
                if yi is not None:
                    nv = _num(r[yi])
                    if nv is not None:
                        lst.append(nv)
                else:
                    lst.append(1.0)
            grouped = [(k[0], k[1], _apply(v)) for k, v in mg.items()]
            return self._finish_multi_series(grouped, chart_type, x, y, yi, agg)
        pairs = [(k, _apply(v)) for k, v in groups.items()]
        return self._finish_series(pairs, chart_type, x, y, yi, agg)

    # ---- pivot ------------------------------------------------------
    def pivot(self, spec):
        # .471: a pivot is a REGISTERED, cancellable run like a query
        # or a reconcile -- it shows on the activity dashboard, and
        # cancel_query(query_id) interrupts the aggregate statement
        # mid-flight (the grid assembly also checks the flag).
        qid = spec.get("query_id") or ("pivot-%s" % uuid.uuid4().hex[:10])
        eng0 = None
        try:
            eng0 = (self.get_duckdb()
                    if (spec.get("engine") or "sqlite") == "duckdb"
                    else self.db)
        except Exception:
            eng0 = None
        self._register_run(qid, eng0, kind="pivot", surface="pivot",
                           label=str(spec.get("table")
                                     or spec.get("result_id") or "result"))
        try:
            return self._pivot_inner(spec, qid)
        finally:
            # Keep the cancelled flag so a late cancel is still reported as
            # cancel (mid-send abort via _REQ_LOCAL).
            self._end_run_keep_cancel(qid)

    def _agg_interrupt_target(self, spec):
        """Connection/engine that actually runs chart/pivot aggregates for cancel.

        Current-result aggregates use DiskBackedRows._conn (private SQLite) or
        ParquetResultStore._engine (DuckDB), not session.db -- cancel_query must
        interrupt that target or Stop looks broken.
        """
        rid = spec.get("result_id")
        if rid is not None:
            cr = self._results.get(rid)
            store = getattr(cr, "store", None) if cr is not None else None
            if store is None:
                return None
            conn = getattr(store, "_conn", None)
            if conn is not None:
                return conn
            eng = getattr(store, "_engine", None)
            if eng is not None:
                return eng
            return None
        table = spec.get("table")
        if table:
            kind = (self._engine_kind_for_table(table)
                    or spec.get("engine") or "sqlite")
            try:
                return (self.get_duckdb() if kind == "duckdb" else self.db)
            except Exception:
                return None
        return None

    def _pivot_interrupt_target(self, spec):
        # Back-compat alias used by older tests / call sites.
        return self._agg_interrupt_target(spec)

    def _pivot_inner(self, spec, qid=None):
        """Build a pivot table over a cached result or a loaded table.

        spec = {result_id | table, engine,
                rows:   [field, ...],          # row dims (stack = sub-pivot)
                cols:   [field, ...],          # column dims
                values: [{field, agg}, ...],   # summarize tile (multiple)
                filters:[{field, op, value, value2, values}, ...],
                limit}
        Back-compat: a single ``value`` + ``agg`` is accepted in place of
        ``values``. Grouping + filtering are pushed into the engine when the
        source supports SQL, so large sources stay off the Python heap; only
        the compact grid is assembled here. Returns {columns, rows,
        row_count, truncated?, note?} or {error}.
        """
        row_dims = [c for c in (spec.get("rows") or []) if c]
        col_dims = [c for c in (spec.get("cols") or []) if c]
        values = spec.get("values")
        if not values:
            if spec.get("value"):
                values = [{"field": spec.get("value"),
                           "agg": (spec.get("agg") or "sum")}]
            else:
                values = [{"field": None, "agg": "count"}]
        norm = []
        for v in values:
            if isinstance(v, str):
                norm.append({"field": v, "agg": "sum"})
            else:
                norm.append({"field": v.get("field"),
                             "agg": (v.get("agg") or "sum").lower()})
        values = norm
        filters = [f for f in (spec.get("filters") or [])
                   if f and f.get("field")]

        ctx, cols = self._agg_source(spec)
        # Re-bind cancel to the connection that will execute the GROUP BY.
        target = self._agg_interrupt_target(spec)
        if qid and target is not None:
            self._register_run(
                qid, target, kind="pivot", surface="pivot",
                label=str(spec.get("table")
                          or spec.get("result_id") or "result"))
        if qid and self._run_is_cancelled(qid):
            return {"error": "interrupted", "cancelled": True}
        if cols is None:
            return {"error": "result expired"}
        names = set(cols)
        for c in row_dims + col_dims:
            if c not in names:
                return {"error": f"Unknown column: {c}"}
        for v in values:
            if v["field"] and v["field"] not in names:
                return {"error": f"Unknown value column: {v['field']}"}
        for f in filters:
            if f["field"] not in names:
                return {"error": f"Unknown filter column: {f['field']}"}
        if not row_dims and not col_dims:
            return {"error": "Pivot needs at least one row or column field."}
        cap = int(spec.get("limit") or 200000)
        try:
            if ctx is not None:
                return self._pivot_sql2(ctx, cols, row_dims, col_dims,
                                        values, filters, cap, qid=qid)
            _c, rows = self._source_rows(
                result_id=spec.get("result_id"), table=spec.get("table"),
                engine=spec.get("engine", "sqlite"), limit=cap)
            if qid and self._run_is_cancelled(qid):
                return {"error": "interrupted", "cancelled": True}
            return self._pivot_python2(cols, rows, row_dims, col_dims,
                                       values, filters, qid=qid)
        except Exception as e:
            if _is_interrupt(e) or (qid and self._run_is_cancelled(qid)):
                return {"error": "interrupted", "cancelled": True}
            return {"error": str(e)}

    @staticmethod
    def _vlabel(v):
        f, a = v.get("field"), (v.get("agg") or "sum").lower()
        return "count" if not f else f"{a}({f})"

    @staticmethod
    def _to_num(v):
        try:
            return float(v)
        except (TypeError, ValueError):
            return None

    @staticmethod
    def _pivot_where(filters, cidx, colref, castnum):
        """Build a WHERE clause from filter specs (''=no filter)."""
        def lit(x):
            return "'" + str(x).replace("'", "''") + "'"

        def numlit(x):
            try:
                return repr(float(x))
            except (TypeError, ValueError):
                return lit(x)
        out = []
        for f in filters:
            c = colref(cidx[f["field"]])
            op = (f.get("op") or "equals").lower()
            val, val2 = f.get("value"), f.get("value2")
            vals = f.get("values") or []
            if op in ("isnull", "is_null", "blank"):
                out.append(f"{c} IS NULL")
            elif op in ("notnull", "not_null", "not_blank"):
                out.append(f"{c} IS NOT NULL")
            elif op in ("equals", "eq", "is"):
                out.append(f"{c} = {lit(val)}")
            elif op in ("not_equals", "ne", "is_not"):
                out.append(f"({c} IS NULL OR {c} <> {lit(val)})")
            elif op == "contains":
                out.append(f"{c} LIKE {lit('%' + str(val) + '%')}")
            elif op in ("not_contains", "does_not_contain"):
                out.append(f"({c} IS NULL OR {c} NOT LIKE "
                           f"{lit('%' + str(val) + '%')})")
            elif op in ("starts_with", "startswith"):
                out.append(f"{c} LIKE {lit(str(val) + '%')}")
            elif op in ("ends_with", "endswith"):
                out.append(f"{c} LIKE {lit('%' + str(val))}")
            elif op in ("gt", "greater_than", ">"):
                out.append(f"{castnum(c)} > {numlit(val)}")
            elif op in ("lt", "less_than", "<"):
                out.append(f"{castnum(c)} < {numlit(val)}")
            elif op in ("gte", ">=", "at_least"):
                out.append(f"{castnum(c)} >= {numlit(val)}")
            elif op in ("lte", "<=", "at_most"):
                out.append(f"{castnum(c)} <= {numlit(val)}")
            elif op == "between":
                out.append(f"{castnum(c)} BETWEEN {numlit(val)} "
                           f"AND {numlit(val2)}")
            elif op in ("in", "one_of"):
                if vals:
                    out.append(f"{c} IN ({', '.join(lit(x) for x in vals)})")
            elif op in ("not_in", "none_of"):
                if vals:
                    out.append(f"({c} IS NULL OR {c} NOT IN "
                               f"({', '.join(lit(x) for x in vals)}))")
        return " AND ".join(f"({x})" for x in out)

    @staticmethod
    def _pivot_row_passes(r, cidx, f):
        raw = r[cidx[f["field"]]]
        op = (f.get("op") or "equals").lower()
        val, val2 = f.get("value"), f.get("value2")
        vals = f.get("values") or []
        s = "" if raw is None else str(raw)

        def num(x):
            try:
                return float(x)
            except (TypeError, ValueError):
                return None
        if op in ("isnull", "is_null", "blank"):
            return raw is None or s == ""
        if op in ("notnull", "not_null", "not_blank"):
            return not (raw is None or s == "")
        if op in ("equals", "eq", "is"):
            return s == ("" if val is None else str(val))
        if op in ("not_equals", "ne", "is_not"):
            return s != ("" if val is None else str(val))
        if op == "contains":
            return str(val) in s
        if op in ("not_contains", "does_not_contain"):
            return str(val) not in s
        if op in ("starts_with", "startswith"):
            return s.startswith(str(val))
        if op in ("ends_with", "endswith"):
            return s.endswith(str(val))
        if op in ("in", "one_of"):
            return s in {str(x) for x in vals}
        if op in ("not_in", "none_of"):
            return s not in {str(x) for x in vals}
        nv, tv, tv2 = num(raw), num(val), num(val2)
        if nv is None:
            return False
        if op in ("gt", "greater_than", ">"):
            return tv is not None and nv > tv
        if op in ("lt", "less_than", "<"):
            return tv is not None and nv < tv
        if op in ("gte", ">=", "at_least"):
            return tv is not None and nv >= tv
        if op in ("lte", "<=", "at_most"):
            return tv is not None and nv <= tv
        if op == "between":
            return (tv is not None and tv2 is not None and tv <= nv <= tv2)
        return True

    def _pivot_sql2(self, ctx, cols, row_dims, col_dims, values, filters, cap,
                    qid=None):
        if qid and self._run_is_cancelled(qid):
            return {"error": "interrupted", "cancelled": True}
        run, src, colref, castnum = ctx
        cidx = {c: i for i, c in enumerate(cols)}
        dim_names = row_dims + col_dims
        dimsel = ", ".join(colref(cidx[d]) for d in dim_names)

        def agg_expr(v):
            agg, fld = (v.get("agg") or "sum").lower(), v.get("field")
            if agg == "count":
                return "count(*)" if not fld else f"count({colref(cidx[fld])})"
            if agg in ("count_distinct", "distinct"):
                return ("count(*)" if not fld
                        else f"count(DISTINCT {colref(cidx[fld])})")
            fn = {"sum": "sum", "avg": "avg", "mean": "avg", "min": "min",
                  "max": "max"}.get(agg, "sum")
            return f"{fn}({castnum(colref(cidx[fld]))})"

        aggsel = ", ".join(agg_expr(v) for v in values)
        where = self._pivot_where(filters, cidx, colref, castnum)
        sql = f"SELECT {dimsel}, {aggsel} FROM {src}"
        if where:
            sql += f" WHERE {where}"
        sql += f" GROUP BY {dimsel}"
        if cap:
            sql += f" LIMIT {int(cap)}"
        grouped = run(sql)
        if qid and self._run_is_cancelled(qid):
            return {"error": "interrupted", "cancelled": True}
        return self._pivot_assemble(grouped, row_dims, col_dims, values)

    def _pivot_python2(self, cols, rows, row_dims, col_dims, values, filters,
                       qid=None):
        cidx = {c: i for i, c in enumerate(cols)}
        ri = [cidx[d] for d in row_dims]
        ci = [cidx[d] for d in col_dims]
        nval = len(values)
        acc, order = {}, []
        for i, r in enumerate(rows):
            if qid and i % 2048 == 0 and self._run_is_cancelled(qid):
                return {"error": "interrupted", "cancelled": True}
            if filters and not all(self._pivot_row_passes(r, cidx, f)
                                   for f in filters):
                continue
            rk = tuple("" if r[i] is None else str(r[i]) for i in ri)
            ck = tuple("" if r[i] is None else str(r[i]) for i in ci)
            key = (rk, ck)
            if key not in acc:
                acc[key] = [[] for _ in range(nval)]
                order.append(key)
            for vj, v in enumerate(values):
                fld = v.get("field")
                if fld is None:
                    acc[key][vj].append(1.0)
                else:
                    nv = self._to_num(r[cidx[fld]])
                    if nv is not None:
                        acc[key][vj].append(nv)

        def reduce(vals, agg):
            agg = (agg or "sum").lower()
            if agg == "count":
                return len(vals)
            if agg in ("count_distinct", "distinct"):
                return len(set(vals))
            if not vals:
                return None
            if agg == "sum":
                return sum(vals)
            if agg in ("avg", "mean"):
                return round(sum(vals) / len(vals), 6)
            if agg == "min":
                return min(vals)
            if agg == "max":
                return max(vals)
            return sum(vals)
        grouped = []
        for (rk, ck) in order:
            grouped.append(list(rk) + list(ck) + [
                reduce(acc[(rk, ck)][vj], values[vj].get("agg"))
                for vj in range(nval)])
        return self._pivot_assemble(grouped, row_dims, col_dims, values)

    def _pivot_assemble(self, grouped, row_dims, col_dims, values):
        """Shape grouped rows [dim.., val..] into a pivot grid. Data columns
        are (col-key x value) when column dims exist, else one per value."""
        nrow, ncol, nval = len(row_dims), len(col_dims), len(values)
        MAXCOLS = 400
        cells, row_keys, row_seen, col_keys, col_seen = {}, [], set(), [], set()
        for g in grouped:
            rk = tuple("" if x is None else str(x) for x in g[:nrow])
            ck = tuple("" if x is None else str(x)
                       for x in g[nrow:nrow + ncol])
            vv = list(g[nrow + ncol:nrow + ncol + nval])
            if rk not in row_seen:
                row_seen.add(rk)
                row_keys.append(rk)
            if ck not in col_seen:
                col_seen.add(ck)
                col_keys.append(ck)
            cells[(rk, ck)] = vv
        col_keys.sort()
        row_keys.sort()
        truncated = len(col_keys) > MAXCOLS
        if truncated:
            col_keys = col_keys[:MAXCOLS]
        vlabels = [self._vlabel(v) for v in values]
        header = list(row_dims)
        col_spec = []
        if col_dims:
            for ck in col_keys:
                ckname = " / ".join("" if x is None else str(x) for x in ck)
                if nval == 1:
                    header.append(ckname or "value")
                    col_spec.append((ck, 0))
                else:
                    for vj, vl in enumerate(vlabels):
                        header.append((ckname + " / " + vl) if ckname else vl)
                        col_spec.append((ck, vj))
        else:
            for vj, vl in enumerate(vlabels):
                header.append(vl)
                col_spec.append(((), vj))
        out_rows = []
        for rk in row_keys:
            line = ["" if x is None else x for x in rk]
            for (ck, vj) in col_spec:
                cell = cells.get((rk, ck))
                line.append(cell[vj] if cell is not None else None)
            out_rows.append(line)
        res = {"columns": header, "rows": out_rows, "row_count": len(out_rows)}
        if truncated:
            res["truncated"] = True
            res["note"] = f"Showing the first {MAXCOLS} column groups."
        return res
    # ---- reconcile --------------------------------------------------
    # ---- reconcile --------------------------------------------------
    @staticmethod
    def _rq(name):
        return '"' + str(name).replace('"', '""') + '"'

    def _recon_engine(self, left, right):
        """Pick the engine holding both tables (DuckDB preferred). Raises
        ValueError with a user-facing message if they aren't both present
        in a single engine."""
        duck = self.duckdb
        if (duck is not None and left in duck.table_columns
                and right in duck.table_columns):
            return DUCKDB_TARGET, "duckdb"
        if (left in self.db.table_columns
                and right in self.db.table_columns):
            return LOCAL_TARGET, "sqlite"

        def _present(t):
            return (t in self.db.table_columns) or (
                duck is not None and t in duck.table_columns)
        missing = [t for t in (left, right) if not _present(t)]
        if missing:
            raise ValueError("Table(s) not found: "
                             + ", ".join('"%s"' % t for t in missing))
        lk = self._engine_kind_for_table(left)
        rk = self._engine_kind_for_table(right)
        raise ValueError(
            'Both tables must be in the same engine to compare them: '
            '"%s" is in %s, "%s" is in %s. Re-load or import one so they '
            'share an engine.' % (left, lk, right, rk))

    def _recon_bucket_sql(self, left, right, keys, bucket,
                          field=None, balance=None, cma=None, cmb=None):
        """SQL selecting the underlying rows of one reconcile bucket. For
        a_only/b_only it is the full unmatched row; for matching /
        non_matching it is the key plus each side's value of ``field``
        (and of ``balance`` when given), so a drill-in shows the break.

        ``cma`` / ``cmb`` map a *canonical* field name to the real column in
        table A / table B respectively (used when a field mapping lines up
        differently-named columns). When a name is absent the canonical name
        is used as-is, so the unmapped case is unchanged. Because the per-side
        real columns are referenced inline (rather than through a view), a
        drill-in resolves the same way the report did -- no view-visibility
        surprises across cursors."""
        q = self._rq
        cma = cma or {}
        cmb = cmb or {}

        def acol(name):
            return q(cma.get(name, name))

        def bcol(name):
            return q(cmb.get(name, name))

        join_on = " AND ".join(f"L.{acol(k)} = R.{bcol(k)}" for k in keys)

        # Anti-joins use LEFT JOIN ... IS NULL rather than a correlated NOT
        # EXISTS: on SQLite the correlated form can abort with "abort due to
        # ROLLBACK" while its rows are streamed into the on-disk spill store.
        # .452 [PLAN PASS 9] fix: rows whose OWN key is NULL fail the
        # anti-join on BOTH sides -- the same row surfaced in a_only
        # AND b_only while the totals (EXCEPT semantics) excluded it,
        # so "1 left-only" drilled to 2 rows. NULL-keyed rows are
        # unjoinable, not "only": the drills now exclude them exactly
        # like the totals do, and the summary reports them under
        # null_keys so nothing vanishes silently.
        a_notnull = " AND ".join(
            f"L.{acol(k)} IS NOT NULL" for k in keys)
        b_notnull = " AND ".join(
            f"R.{bcol(k)} IS NOT NULL" for k in keys)
        if bucket == "a_only":
            return (f"SELECT L.* FROM {q(left)} L LEFT JOIN {q(right)} R "
                    f"ON {join_on} WHERE R.{bcol(keys[0])} IS NULL "
                    f"AND {a_notnull}")
        if bucket == "b_only":
            return (f"SELECT R.* FROM {q(right)} R LEFT JOIN {q(left)} L "
                    f"ON {join_on} WHERE L.{acol(keys[0])} IS NULL "
                    f"AND {b_notnull}")
        if bucket not in ("matching", "non_matching"):
            raise ValueError("Unknown bucket: %r" % (bucket,))
        if not field:
            raise ValueError("A field is required for matching / "
                             "non_matching drill-in.")
        eqp = (f"COALESCE(CAST(L.{acol(field)} AS VARCHAR),'') = "
               f"COALESCE(CAST(R.{bcol(field)} AS VARCHAR),'')")
        pred = eqp if bucket == "matching" else "NOT (" + eqp + ")"
        key_sel = ", ".join(f"L.{acol(k)} AS {q(k)}" for k in keys)
        sel = (key_sel + f", L.{acol(field)} AS {q('A_' + field)}, "
               f"R.{bcol(field)} AS {q('B_' + field)}")
        if balance and balance != field:
            sel += (f", L.{acol(balance)} AS {q('A_' + balance)}, "
                    f"R.{bcol(balance)} AS {q('B_' + balance)}")
        return (f"SELECT {sel} FROM {q(left)} L JOIN {q(right)} R "
                f"ON {join_on} WHERE {pred}")

    def reconcile(self, spec):
        qid = spec.get("query_id") or ("recon-%s" % uuid.uuid4().hex[:10])
        # .469: the inner body reports milestone progress against this
        # id -- make sure it rides along even when the caller sent none.
        spec = dict(spec)
        spec["query_id"] = qid
        eng0 = None
        try:
            # .471: register the MANAGER, not the engine's name --
            # _interrupt_entry calls .interrupt() on what's registered,
            # and a string swallowed it silently: Stop never reached a
            # running reconcile.
            _t0, _k0 = self._recon_engine(spec.get("left"),
                                          spec.get("right"))
            eng0, _ = self._engine_obj(_t0)
        except Exception:
            eng0 = None
        self._register_run(qid, eng0, kind="reconcile",
                           surface="reconcile",
                           label="%s vs %s" % (spec.get("left"),
                                               spec.get("right")))
        try:
            return self._reconcile_inner(spec)
        finally:
            self._end_run_keep_cancel(qid)

    def _reconcile_inner(self, spec):
        """Compare two tables on key columns and return a *field-level*
        reconciliation report plus headline totals -- without materializing
        any rows. For each compared field it reports how many key-matched
        rows agree vs differ on that field, and (when a balance field is
        given) the summed balance of each. Headline totals give keys only
        in A, only in B, fully matching rows, rows with any difference, and
        the overall record count. Everything is computed as SQL aggregates
        on whichever engine holds both tables (DuckDB preferred), so it
        stays fast and memory-light even on very large inputs. The
        underlying rows of any cell are fetched on demand via
        ``reconcile_drilldown`` / ``reconcile_profile``.
        """
        # .469: four honest milestones for the reconcile modal's
        # progress bar (validate, row counts, bucket totals, the
        # per-field aggregate). advance() rides the same registry the
        # activity pill reads; the modal polls /api/status for it.
        _rq_id = spec.get("query_id")

        def _step(n):
            # .471: each milestone is also a cancellation gate, so a
            # Stop that lands BETWEEN statements still unwinds the run.
            if _rq_id and self._run_is_cancelled(_rq_id):
                raise RuntimeError("interrupted")
            try:
                opreg.advance(done=n, total=4, op_id=_rq_id)
            except Exception:
                pass
        left = spec.get("left")
        right = spec.get("right")
        keys = list(spec.get("keys") or [])
        compare = list(spec.get("compare") or [])
        balance = spec.get("balance") or None
        colmap_a = dict(spec.get("colmap_a") or {})
        colmap_b = dict(spec.get("colmap_b") or {})
        if not left or not right or not keys:
            return {"error": "left, right and at least one key required."}
        try:
            target, engine_kind = self._recon_engine(left, right)
        except Exception as e:
            return {"error": str(e)}

        q = self._rq
        eng, _ = self._engine_obj(target)

        def acol(name):
            return q(colmap_a.get(name, name))

        def bcol(name):
            return q(colmap_b.get(name, name))

        def flabel(name):
            # Show the canonical (mapped) name plus the underlying A/B
            # originals only when a mapping actually renamed them; an
            # unchanged name is shown bare.
            oa = colmap_a.get(name, name)
            ob = colmap_b.get(name, name)
            if oa == name and ob == name:
                return name
            if oa == ob:
                return f"{name} ({oa})"
            return f"{name} (A: {oa} / B: {ob})"

        # EXCEPT compares by column position, so A's real key columns on the
        # left and B's on the right line up under the canonical order.
        a_keylist = ", ".join(acol(k) for k in keys)
        b_keylist = ", ".join(bcol(k) for k in keys)
        join_on = " AND ".join(f"L.{acol(k)} = R.{bcol(k)}" for k in keys)

        # Reconcile is key-based, so collapse each side to one row per key
        # before joining. Without this, duplicate keys make ``L JOIN R`` fan
        # out (M*N rows per key), inflating the matched count, every per-field
        # tally, and the balance sum.
        def _dedup(table, side_keys):
            cols = ", ".join(side_keys)
            return (f"(SELECT * FROM (SELECT *, ROW_NUMBER() OVER "
                    f"(PARTITION BY {cols} ORDER BY {cols}) AS __samql_rn "
                    f"FROM {q(table)}) _dd WHERE __samql_rn = 1)")

        ded_l = _dedup(left, [acol(k) for k in keys])
        ded_r = _dedup(right, [bcol(k) for k in keys])
        join_clause = f"{ded_l} L JOIN {ded_r} R ON {join_on}"

        # .413 efficiency: summary counts are pure reads; when BOTH
        # inputs are real catalog tables they ride execute_read (own
        # cursor -- no write-lock contention with everything else).
        # Journal-staged inputs are TEMP views (connection-local), so
        # those keep the locked path.
        _both_real = (left in getattr(eng, "table_columns", {})
                      and right in getattr(eng, "table_columns", {}))
        _rd = (eng.execute_read
               if _both_real and hasattr(eng, "execute_read")
               else eng.execute)
        _step(1)

        def _count(sql):
            try:
                _c, rows = _rd(sql)
            except Exception as e:
                raise RuntimeError(str(e))
            if rows and rows[0] and rows[0][0] is not None:
                return int(rows[0][0])
            return 0

        def _eqp(f):
            return (f"COALESCE(CAST(L.{acol(f)} AS VARCHAR),'') = "
                    f"COALESCE(CAST(R.{bcol(f)} AS VARCHAR),'')")

        # balance cast: TRY_CAST on DuckDB (non-numeric -> NULL),
        # CAST AS REAL on SQLite; summed from the left table. Hoisted
        # above the counts (.414) so both count paths can use it.
        bal_expr = None
        if balance:
            _braw = f"L.{acol(balance)}"
            if engine_kind == "duckdb":
                _clean = (f"REPLACE(REPLACE(REPLACE(CAST({_braw} AS "
                          f"VARCHAR), ',', ''), '$', ''), ' ', '')")
                bal_expr = f"TRY_CAST({_clean} AS DOUBLE)"
            else:
                _clean = (f"REPLACE(REPLACE(REPLACE(CAST({_braw} AS "
                          f"TEXT), ',', ''), '$', ''), ' ', '')")
                bal_expr = f"CAST({_clean} AS REAL)"

        def _legacy_counts():
            """The pre-.414 path: five separate statements (two EXCEPT
            scans, the matched count, the per-field aggregate, the
            row-level mismatch count). Kept verbatim as the fallback."""
            a_only = _count(
                f"SELECT COUNT(*) FROM (SELECT {a_keylist} FROM {q(left)} "
                f"EXCEPT SELECT {b_keylist} FROM {q(right)})")
            b_only = _count(
                f"SELECT COUNT(*) FROM (SELECT {b_keylist} FROM {q(right)} "
                f"EXCEPT SELECT {a_keylist} FROM {q(left)})")
            matched = _count(
                f"SELECT COUNT(*) FROM {join_clause}")
            _step(2)
            flat = []
            row_nonmatch = 0
            if compare:
                parts = []
                for f in compare:
                    eqp = _eqp(f)
                    parts.append(f"SUM(CASE WHEN {eqp} THEN 1 ELSE 0 END)")
                    parts.append(f"SUM(CASE WHEN {eqp} THEN 0 ELSE 1 END)")
                    if bal_expr:
                        parts.append(
                            f"SUM(CASE WHEN {eqp} THEN {bal_expr} ELSE 0 END)")
                        parts.append(
                            f"SUM(CASE WHEN {eqp} THEN 0 ELSE {bal_expr} END)")
                _c, rows = _rd(
                    "SELECT " + ", ".join(parts)
                    + f" FROM {join_clause}")
                _step(3)
                flat = list(rows[0]) if rows and rows[0] else []
                diff_any = " OR ".join("NOT (" + _eqp(f) + ")"
                                       for f in compare)
                row_nonmatch = _count(
                    f"SELECT COUNT(*) FROM {join_clause} "
                    f"WHERE {diff_any}")
            return a_only, b_only, matched, flat, row_nonmatch

        def _one_pass_counts():
            """.414: the whole summary as ONE null-safe FULL OUTER JOIN
            aggregate over the deduped sides -- a_only, b_only, matched,
            every per-field tally (with balances), and the row-level
            mismatch, in a single scan of each input. Semantics parity
            with the legacy path: EXCEPT treats NULLs as equal, so
            membership joins null-safe (IS NOT DISTINCT FROM / IS);
            "matched" keeps the plain-equality gate, so NULL-containing
            key tuples land in NO bucket -- exactly as before."""
            ns = ("IS NOT DISTINCT FROM" if engine_kind == "duckdb"
                  else "IS")
            join_ns = " AND ".join(
                f"L.{acol(k)} {ns} R.{bcol(k)}" for k in keys)
            lsrc = f"(SELECT _l.*, 1 AS __samql_la FROM {ded_l} _l)"
            rsrc = f"(SELECT _r.*, 1 AS __samql_rb FROM {ded_r} _r)"
            mpred = ("(L.__samql_la IS NOT NULL AND "
                     "R.__samql_rb IS NOT NULL AND " + join_on + ")")
            parts = [
                "SUM(CASE WHEN R.__samql_rb IS NULL THEN 1 ELSE 0 END)",
                "SUM(CASE WHEN L.__samql_la IS NULL THEN 1 ELSE 0 END)",
                f"SUM(CASE WHEN {mpred} THEN 1 ELSE 0 END)",
            ]
            for f in compare:
                eqp = _eqp(f)
                parts.append(
                    f"SUM(CASE WHEN {mpred} AND {eqp} THEN 1 ELSE 0 END)")
                parts.append(f"SUM(CASE WHEN {mpred} AND NOT ({eqp}) "
                             f"THEN 1 ELSE 0 END)")
                if bal_expr:
                    parts.append(f"SUM(CASE WHEN {mpred} AND {eqp} "
                                 f"THEN {bal_expr} ELSE 0 END)")
                    parts.append(f"SUM(CASE WHEN {mpred} AND NOT ({eqp}) "
                                 f"THEN {bal_expr} ELSE 0 END)")
            if compare:
                diff_any = " OR ".join("NOT (" + _eqp(f) + ")"
                                       for f in compare)
                parts.append(f"SUM(CASE WHEN {mpred} AND ({diff_any}) "
                             f"THEN 1 ELSE 0 END)")
            _step(2)
            _c, rows = _rd(
                "SELECT " + ", ".join(parts)
                + f" FROM {lsrc} L FULL OUTER JOIN {rsrc} R ON {join_ns}")
            _step(3)
            row0 = list(rows[0]) if rows and rows[0] else []
            if not row0:
                row0 = [0, 0, 0]
            a_only = int(row0[0] or 0)
            b_only = int(row0[1] or 0)
            matched = int(row0[2] or 0)
            width = len(compare) * (4 if bal_expr else 2)
            flat = list(row0[3:3 + width])
            row_nonmatch = int(row0[-1] or 0) if compare else 0
            return a_only, b_only, matched, flat, row_nonmatch

        try:
            got = None
            if RECON_SINGLE_PASS and (
                    engine_kind == "duckdb"
                    or (RECON_SINGLE_PASS_SQLITE
                        and sqlite3.sqlite_version_info >= (3, 39))):
                try:
                    got = _one_pass_counts()
                except Exception:
                    got = None   # any surprise -> the legacy statements
            if got is None:
                got = _legacy_counts()
            a_only, b_only, matched, flat, row_nonmatch = got
            _step(4)

            fields = []
            if compare:
                step = 4 if bal_expr else 2
                for i, f in enumerate(compare):
                    base = i * step
                    m = int(flat[base] or 0) if base < len(flat) else 0
                    nm = (int(flat[base + 1] or 0)
                          if base + 1 < len(flat) else 0)
                    if bal_expr:
                        mb = (float(flat[base + 2])
                              if flat[base + 2] is not None else 0.0)
                        nmb = (float(flat[base + 3])
                               if flat[base + 3] is not None else 0.0)
                    else:
                        mb = nmb = None
                    fields.append({
                        "field": f, "label": flabel(f),
                        "a_only": a_only, "b_only": b_only,
                        "non_matching": nm, "matching": m,
                        "sum_matching_balance": mb,
                        "sum_non_matching_balance": nmb,
                    })
        except Exception as e:
            return {"error": str(e)}

        row_match = matched - row_nonmatch
        # .452: NULL-keyed rows are excluded from every bucket (they
        # cannot join); count them per side so they stay VISIBLE.
        _anyn_a = " OR ".join(f"{acol(k)} IS NULL" for k in keys)
        _anyn_b = " OR ".join(f"{bcol(k)} IS NULL" for k in keys)
        null_keys = {
            "a": _count(f"SELECT COUNT(*) FROM {q(left)} "
                        f"WHERE {_anyn_a}"),
            "b": _count(f"SELECT COUNT(*) FROM {q(right)} "
                        f"WHERE {_anyn_b}"),
        }
        out = {
            "engine": engine_kind,
            "keys": keys,
            "balance_field": balance,
            "totals": {
                "a_only": a_only, "b_only": b_only,
                "null_keys": null_keys,
                "matching": row_match, "non_matching": row_nonmatch,
                "total": a_only + b_only + matched,
            },
            "fields": fields,
        }
        self._schedule_cleanup(full=False)
        return out

    def reconcile_drilldown(self, spec):
        qid = (spec.get("query_id")
               or ("recon-dd-%s" % uuid.uuid4().hex[:10]))
        eng0 = None
        try:
            # .471: register the MANAGER (see reconcile above) -- a
            # name string swallowed interrupts silently.
            _t0, _k0 = self._recon_engine(spec.get("left"),
                                          spec.get("right"))
            eng0, _ = self._engine_obj(_t0)
        except Exception:
            eng0 = None
        self._register_run(qid, eng0, kind="reconcile",
                           surface="reconcile",
                           label="drill: %s" % (spec.get("bucket"),))
        try:
            return self._reconcile_drilldown_inner(spec)
        finally:
            self._end_run_keep_cancel(qid)

    def _reconcile_drilldown_inner(self, spec):
        """Materialize the underlying rows of one reconcile bucket to the
        on-disk result store and return {result_id, columns, count}. The
        rows are paged like any other result; empty buckets free at once."""
        left = spec.get("left")
        right = spec.get("right")
        keys = list(spec.get("keys") or [])
        bucket = spec.get("bucket")
        field = spec.get("field")
        balance = spec.get("balance") or None
        colmap_a = dict(spec.get("colmap_a") or {})
        colmap_b = dict(spec.get("colmap_b") or {})
        if (not left or not right or not keys
                or bucket not in ("a_only", "b_only", "matching",
                                  "non_matching")):
            return {"error": "left, right, keys and a valid bucket "
                    "are required."}
        try:
            target, _ = self._recon_engine(left, right)
            sql = self._recon_bucket_sql(left, right, keys, bucket,
                                         field, balance,
                                         cma=colmap_a, cmb=colmap_b)
            cols, store, total, kind = self._exec_target(sql, target)
        except Exception as e:
            return {"error": str(e)}
        if cols is None:
            return {"result_id": None, "columns": [], "count": 0}
        if int(total) <= 0:
            closer = getattr(store, "close", None)
            if callable(closer):
                try:
                    closer()
                except Exception:
                    pass
            return {"result_id": None, "columns": list(cols), "count": 0}
        rid = self._cache_result(cols, store, int(total), sql, target, kind)
        self._schedule_cleanup(full=False)
        return {"result_id": rid, "columns": list(cols), "count": int(total)}

    def reconcile_failures_export(self, spec, query_id=None):
        """.540: ONE CSV of every failed value. For EVERY compared field,
        each mismatching key with its left/right values -- exactly the
        union of the per-field non_matching drill-downs: same inline
        colmap columns, same null-safe COALESCE/VARCHAR predicate, and
        NULL-keyed rows excluded by the key join (.452). Registers a
        cancellable "export" card under query_id; Stop interrupts the
        union statement AND the CSV write."""
        left = spec.get("left")
        right = spec.get("right")
        keys = spec.get("keys") or []
        # .543: the reconcile spec's real key is ``compare`` (what the
        # modal and every report tab carry); ``fields`` stays as an
        # alias. The .540 cut read only the alias, so a genuine report's
        # export answered "fields are required" while showing 71k
        # mismatches on screen.
        fields = spec.get("compare") or spec.get("fields") or []
        colmap_a = spec.get("colmap_a") or {}
        colmap_b = spec.get("colmap_b") or {}
        if not isinstance(keys, (list, tuple)) \
                or not all(isinstance(k, str) for k in keys):
            return {"error": "keys must be a list of column names."}
        if not isinstance(fields, (list, tuple)) \
                or not all(isinstance(f, str) for f in fields):
            return {"error": "compare must be a list of column names."}
        if not isinstance(colmap_a, dict) or not isinstance(colmap_b, dict):
            return {"error": "colmap_a / colmap_b must be objects."}
        keys = list(keys)
        fields = [f for f in fields if f not in keys]
        if not fields and left and right and keys \
                and not colmap_a and not colmap_b:
            # unmapped + unspecified: derive the shared columns exactly
            # like the modal's default selection does
            try:
                target0, _ = self._recon_engine(left, right)
                eng0, _ = self._engine_obj(target0)
                ca = list((getattr(eng0, "table_columns", {}) or {})
                          .get(left) or [])
                cb = set((getattr(eng0, "table_columns", {}) or {})
                         .get(right) or [])
                fields = [c for c in ca if c in cb and c not in keys]
            except Exception:
                fields = []
        if not left or not right or not keys or not fields:
            return {"error": "left, right, keys and compared fields "
                    "are required."}
        q = self._rq
        cma, cmb = dict(colmap_a), dict(colmap_b)

        def acol(n):
            return q(cma.get(n, n))

        def bcol(n):
            return q(cmb.get(n, n))

        join_on = " AND ".join(f"L.{acol(k)} = R.{bcol(k)}" for k in keys)
        key_sel = ", ".join(f"L.{acol(k)} AS {q(k)}" for k in keys)
        parts = []
        for f in fields:
            eqp = (f"COALESCE(CAST(L.{acol(f)} AS VARCHAR),'') = "
                   f"COALESCE(CAST(R.{bcol(f)} AS VARCHAR),'')")
            lit = f.replace("'", "''")
            parts.append(
                f"SELECT {key_sel}, '{lit}' AS {q('field')}, "
                f"CAST(L.{acol(f)} AS VARCHAR) AS {q('left_value')}, "
                f"CAST(R.{bcol(f)} AS VARCHAR) AS {q('right_value')} "
                f"FROM {q(left)} L JOIN {q(right)} R ON {join_on} "
                f"WHERE NOT ({eqp})")
        order = ", ".join(q(k) for k in keys) + ", " + q("field")
        sql = " UNION ALL ".join(parts) + " ORDER BY " + order
        try:
            target, _ = self._recon_engine(left, right)
            eng, _ = self._engine_obj(target)
        except Exception as e:
            return {"error": str(e)}
        out_path = tmputil.new_tempfile("recon_failures_", ".csv")
        self._register_run(query_id, eng, kind="export",
                           surface="reconcile",
                           target="recon failures csv")
        rid = None
        try:
            if query_id and self._run_is_cancelled(query_id):
                return {"cancelled": True}
            try:
                cols, store, total, kind = self._exec_target(sql, target)
            except Exception as e:
                from .engines import _is_interrupt as _isi
                if _isi(e) or (query_id
                               and self._run_is_cancelled(query_id)):
                    return {"cancelled": True}
                return {"error": str(e)}
            cols = list(cols or keys + ["field", "left_value",
                                        "right_value"])
            if int(total or 0) <= 0:
                closer = getattr(store, "close", None)
                if callable(closer):
                    try:
                        closer()
                    except Exception:
                        pass
                import csv as _csv
                with open(out_path, "w", newline="",
                          encoding="utf-8") as fh:
                    _csv.writer(fh).writerow(cols)
                return {"ok": True, "path": out_path, "rows": 0,
                        "fields": len(fields)}
            rid = self._cache_result(cols, store, int(total), sql,
                                     target, kind)
            cr = self._results.get(rid)
            res = self._export_inner(cr, "csv", out_path, None, False,
                                     query_id)
            if res.get("cancelled"):
                return {"cancelled": True}
            if not res.get("ok"):
                return {"error": res.get("error", "Export failed.")}
            return {"ok": True, "path": out_path, "rows": int(total),
                    "fields": len(fields)}
        finally:
            if rid:
                try:
                    self.discard_result(rid)
                except Exception:
                    pass
            self._unregister_run(query_id)

    def reconcile_profile(self, spec):
        """Profile the underlying rows of one reconcile bucket. The rows are
        profiled as a subquery (never materialized as a table)."""
        qid = (spec.get("query_id")
               or ("recon-pf-%s" % uuid.uuid4().hex[:10]))
        eng0 = None
        try:
            _t0, _k0 = self._recon_engine(spec.get("left"),
                                          spec.get("right"))
            eng0, _ = self._engine_obj(_t0)
        except Exception:
            eng0 = None
        self._register_run(qid, eng0, kind="profile",
                           surface="reconcile",
                           label="profile: %s" % (spec.get("bucket"),))
        try:
            if self._run_is_cancelled(qid):
                return {"cancelled": True, "error": "cancelled",
                        "table": "", "total_rows": 0, "columns": []}
            left = spec.get("left")
            right = spec.get("right")
            keys = list(spec.get("keys") or [])
            bucket = spec.get("bucket")
            field = spec.get("field")
            balance = spec.get("balance") or None
            colmap_a = dict(spec.get("colmap_a") or {})
            colmap_b = dict(spec.get("colmap_b") or {})
            if (not left or not right or not keys
                    or bucket not in ("a_only", "b_only", "matching",
                                      "non_matching")):
                return {"table": "", "total_rows": 0, "columns": [],
                        "error": "left, right, keys and a valid bucket "
                        "are required."}
            target, engine_kind = self._recon_engine(left, right)
            sql = self._recon_bucket_sql(left, right, keys, bucket,
                                         field, balance,
                                         cma=colmap_a, cmb=colmap_b)
            eng, _ = self._engine_obj(target)
            label = {
                "a_only": f"A only · {left}",
                "b_only": f"B only · {right}",
                "matching": f"Matching · {field}",
                "non_matching": f"Not matching · {field}",
            }.get(bucket, bucket)
            return profile_table(
                eng, label, source=f"({sql}) AS _r",
                cancel=lambda: self._run_is_cancelled(qid))
        except Exception as e:
            if _is_interrupt(e) or self._run_is_cancelled(qid):
                return {"cancelled": True, "error": "cancelled",
                        "table": "", "total_rows": 0, "columns": []}
            return {"table": "", "total_rows": 0, "columns": [],
                    "error": str(e)}
        finally:
            self._end_run_keep_cancel(qid)
    # ---- REST API loader --------------------------------------------
    def load_api(self, url, base_name="api_data", auth_user=None,
                 auth_pass=None, json_path=None, params=None,
                 destination="auto", query_id=None):
        from .apiload import load_api as _load_api
        # Make the IDE API fetch interruptible by Stop / window-close: the
        # network work isn't an engine statement, so the fetch loops check this
        # flag (between pages, between download chunks) and abort.
        should = ((lambda: self._run_is_cancelled(query_id))
                  if query_id else None)
        if query_id:
            self._register_run(query_id, None, kind="fetch", target=base_name)
        try:
            if query_id and self._run_is_cancelled(query_id):
                return {"cancelled": True}
            res = _load_api(self, url, base_name=base_name,
                            auth_user=auth_user, auth_pass=auth_pass,
                            json_path=json_path, params=params,
                            destination=destination, should_cancel=should)
        finally:
            if query_id:
                self._end_run_keep_cancel(query_id)
        if isinstance(res, dict) and res.get("cancelled"):
            return res
        if isinstance(res, dict) and not res.get("error"):
            self._invalidate_profiles()
            self._invalidate_counts()   # new data -> drop count + flow caches
        return res

    def preview_api(self, url, auth_user=None, auth_pass=None,
                    json_path=None, params=None):
        from .apiload import preview_api as _preview_api
        return _preview_api(url, auth_user=auth_user, auth_pass=auth_pass,
                            json_path=json_path, params=params)

    def import_from_connection(self, name, query, base_name="import",
                               destination="duckdb", query_id=None):
        """Run a query against a registered remote connection (e.g. SQL
        Server) and load its result into a local table. The connection
        enforces its own read-only / GO-aware safeguards (a read-only
        connection blocks non-SELECT batches), and the result streams to
        disk, so this stays memory-bounded on large pulls.

        Engine selection is DuckDB-first with SQLite as the fallback: when
        DuckDB is available we stream into it (better for analytical use);
        if DuckDB isn't installed, or its typed ingest rejects a value, we
        fall back to SQLite (dynamically typed, always available). The
        remote row store is re-iterable, so the fallback re-drains cleanly.
        Returns a LoadResult-shaped dict (so the UI counts it) or {error}.
        """
        conn = self.connections.get(name)
        if conn is None:
            return {"error": 'Connection "%s" is not active.' % name}
        q = (query or "").strip()
        if not q:
            return {"error": "Enter a query to import."}
        # Register the remote connection so Stop / window-close (cancel_query)
        # calls conn.interrupt() -> the active pyodbc cursor's .cancel(); the
        # cancel flag also stops the local-load phase via cancel_query's
        # interrupt-every-engine safety net. A query already executing on the
        # server may take a moment to actually stop, depending on the server.
        self._register_run(query_id, conn, kind="mssql", target=name)
        try:
            if self._run_is_cancelled(query_id):
                return {"cancelled": True}
            cols, rows = conn.execute(q)
            if cols is None:
                return {"error": "That statement returned no result set to "
                        "import."}
            bn = (base_name or "").strip() or "import"
            src = "mssql:%s" % name

            tname, n, engine, cols_now = None, 0, None, list(cols)
            if destination in ("duckdb", "auto") and HAS_DUCKDB:
                try:
                    duck = self.get_duckdb()
                    # Re-bind interrupt to the local engine for ingest -- Stop
                    # during add_table_streaming must reach DuckDB, not only
                    # the remote ODBC handle that already finished execute.
                    self._register_run(query_id, duck, kind="mssql",
                                       target=name)
                    if self._run_is_cancelled(query_id):
                        return {"cancelled": True}
                    tname, n = duck.add_table_streaming(
                        bn, cols, rows, source=src)
                    cols_now = duck.table_columns.get(tname, list(cols))
                    engine = "duckdb"
                except Exception:
                    # DuckDB ingest cleaned up after itself; re-drain to SQLite.
                    tname, n, engine = None, 0, None
            if engine is None:
                self._register_run(query_id, self.db, kind="mssql",
                                   target=name)
                if self._run_is_cancelled(query_id):
                    return {"cancelled": True}
                tname, n = self.db.add_table_streaming(
                    bn, cols, rows, source=src)
                cols_now = self.db.table_columns.get(tname, list(cols))
                engine = "sqlite"
            return {
                "ok": True,
                "table": tname,
                "engine": engine,
                "loaded": [{"file": bn,
                            "tables": [{"name": tname, "rows": n,
                                        "columns": cols_now, "engine": engine}]}],
                "tables": [],
            }
        except Exception as e:
            if _is_interrupt(e) or self._run_is_cancelled(query_id):
                return {"cancelled": True}
            return {"error": str(e)}
        finally:
            self._end_run_keep_cancel(query_id)

    def write_image(self, out_dir, base_name, fmt, data_url):
        """Write a chart/dashboard image (produced client-side by ECharts'
        getDataURL) to disk. ``data_url`` is a data: URI -- base64 for
        png/jpeg, URL-encoded for svg. Empty out_dir writes to Downloads.
        Returns {ok, file} or {error}."""
        import os
        import re
        import base64
        import urllib.parse
        out_dir = self._resolve_export_dir(out_dir)
        if not out_dir:
            return {"error": "Pick an output folder that exists."}
        fmt = (fmt or "png").lower().lstrip(".")
        if fmt not in ("png", "jpeg", "jpg", "svg"):
            return {"error": "Unsupported image format: %s" % fmt}
        if not data_url or not data_url.startswith("data:"):
            return {"error": "No image data was supplied."}
        try:
            header, payload = data_url.split(",", 1)
        except ValueError:
            return {"error": "The image data was malformed."}
        try:
            if ";base64" in header:
                blob = base64.b64decode(payload)
            else:
                blob = urllib.parse.unquote(payload).encode("utf-8")
            if not blob:
                # .453 [PLAN PASS 10]: b64decode silently discards
                # non-alphabet characters, so pure garbage decoded to
                # EMPTY bytes and a 0-byte "image" landed with ok:true.
                return {"error": "Couldn't decode the image: "
                                 "empty payload"}
        except Exception as e:
            return {"error": "Couldn't decode the image: %s: %s"
                    % (type(e).__name__, e)}
        safe = re.sub(r"[^A-Za-z0-9_.-]+", "_", (base_name or "chart").strip()) \
            or "chart"
        ext = "jpg" if fmt in ("jpeg", "jpg") else fmt
        path = os.path.join(out_dir, "%s.%s" % (safe, ext))
        try:
            with open(path, "wb") as fh:
                fh.write(blob)
        except Exception as e:
            return {"error": "Couldn't write the image: %s: %s"
                    % (type(e).__name__, e)}
        return {"ok": True, "file": path, "bytes": len(blob)}

    def load_directory_file(self, path, query_id=None):
        """Read a single file the user picked from a folder (the directory
        node). The file is loaded once into a hidden table (kept out of the
        tables list by its "__" prefix) and reused on later runs unless it
        changes on disk. The directory node then references that table like an
        input. Returns {ok, table, engine, columns, rows} or {error}."""
        import os
        import hashlib
        from . import loaders
        if not path or not os.path.isfile(path):
            return {"error": "That file no longer exists."}
        try:
            mtime = os.path.getmtime(path)
        except OSError as e:
            return {"error": err_str(e)}

        def _count(eng, name):
            try:
                _c, r = eng.execute('SELECT COUNT(*) FROM "%s"' % name)
                return int(r[0][0]) if r else None
            except Exception:
                return None

        cached = self._dir_files.get(path)
        if cached:
            name, eng_name, mt = cached
            eng = self.duckdb if eng_name == "duckdb" else self.db
            if mt == mtime and eng is not None and name in eng.table_columns:
                return {"ok": True, "table": name, "engine": eng_name,
                        "columns": eng.table_columns.get(name, []),
                        "rows": _count(eng, name)}
            # stale (file changed) -- drop the old hidden table and reload
            try:
                eng.drop_table(name)
            except Exception:
                pass
            self._dir_files.pop(path, None)

        _oid = None
        try:
            _oid = opreg.begin("load", target=os.path.basename(path) or path)
        except Exception:
            _oid = None
        # Make the read cancelable (see load_folder_files): register the engine
        # the read will use so cancel_query(query_id) can interrupt it.
        if query_id:
            try:
                eng_c = (self.get_duckdb()
                         if self.default_destination() == "duckdb"
                         else self.db)
            except Exception:
                eng_c = self.db
            if eng_c is not None:
                with self._running_lock:
                    self._running[query_id] = eng_c
        base = "__nbfile_" + hashlib.md5(path.encode("utf-8")).hexdigest()[:10]
        try:
            descs = loaders.load_file(self, path, destination="auto",
                                      base_name=base)
        except Exception as e:
            if _is_interrupt(e):
                return {"error": "cancelled", "cancelled": True}
            return {"error": "Couldn't read that file: %s: %s"
                    % (type(e).__name__, e)}
        finally:
            if query_id:
                with self._running_lock:
                    self._running.pop(query_id, None)
            try:
                opreg.end(_oid)
            except Exception:
                pass
        d = (descs or [None])[0]
        if not d or not d.get("name"):
            return {"error": "No readable data in that file."}
        self._dir_files[path] = (d["name"], d.get("engine", "sqlite"), mtime)
        self._invalidate_counts()
        return {"ok": True, "table": d["name"],
                "engine": d.get("engine", "sqlite"),
                "columns": d.get("columns", []), "rows": d.get("rows")}

    def fetch_api_node(self, node_id, config, graph=None, query_id=None,
                       should_cancel=None):
        """Fetch an API node's endpoint and load the response into a hidden
        table (kept out of the tables list by its "__" prefix) that the node
        then references like an input. ${variables} in the URL, query params
        and headers are resolved first (so an iterator can vary them), and the
        basic-auth password comes from the secret store by key -- it is never
        kept in the saved workflow. Re-fetching replaces the node's table.
        Returns {ok, table, engine, columns, rows, url} or {error}."""
        import hashlib
        from . import applyvars
        cfg = dict(config or {})
        # {{name}} inserts a quoted SQL string literal, which is meaningless in
        # a URL / query string -- the raw value is wanted, so ${name} is the
        # right form here. Flag it up front rather than firing a request with a
        # stray '...' in the URL. (No-op when an iterator already resolved cfg.)
        _bad_brace = applyvars.brace_misuse(
            {"nodes": [{"type": "apinode", "config": cfg}]}, {})
        if _bad_brace:
            return {"error": _bad_brace}
        url = (cfg.get("url") or "").strip()
        if not url:
            return {"error": "Give the API node a URL "
                    "(it can include ${variables})."}
        # resolve ${vars} from the graph's variable nodes (if a graph is given)
        ctx = {}
        try:
            ctx = applyvars.collect_vars(graph or {})
        except Exception:
            ctx = {}

        def sub(v):
            try:
                return applyvars.substitute_text(v, ctx) if isinstance(v, str) \
                    else v
            except Exception:
                return v

        url = sub(url)
        params = [(sub(str(p.get("key", ""))), sub(str(p.get("value", ""))))
                  for p in (cfg.get("params") or [])
                  if str(p.get("key", "")).strip()]
        headers = {sub(str(h.get("key", ""))): sub(str(h.get("value", "")))
                   for h in (cfg.get("headers") or [])
                   if str(h.get("key", "")).strip()}
        auth_user = sub(cfg.get("auth_user") or "") or None
        pwd = None
        sk = (cfg.get("secret_key") or "").strip()
        if sk:
            try:
                pwd = self.secrets.get(sk)
            except Exception:
                pwd = None
        if pwd is None and cfg.get("auth_pass"):
            pwd = cfg.get("auth_pass")   # inline fallback (discouraged)

        base = "__nbapi_" + hashlib.md5(
            str(node_id).encode("utf-8")).hexdigest()[:10]
        # drop the node's previous table before reloading
        prev = self._api_node_tables.get(node_id)
        if prev:
            for nm, en in ((prev.get("table"), prev.get("engine")),
                           (prev.get("err"), prev.get("err_engine"))):
                if not nm:
                    continue
                eng = self.duckdb if en == "duckdb" else self.db
                try:
                    if eng is not None:
                        eng.drop_table(nm)
                except Exception:
                    pass
            self._api_node_tables.pop(node_id, None)

        from .apiload import load_api as _load_api
        try:
            max_pages = int(cfg.get("max_pages") or 50)
        except (TypeError, ValueError):
            max_pages = 50
        continue_on_error = bool(cfg.get("continue_on_error"))
        # Make the fetch interruptible by Stop. The network work isn't an engine
        # statement, so engine interrupt can't touch it; instead the fetch loops
        # check this flag (between pages, and between download chunks) and abort.
        # When a caller (an iterator / while pass) already owns the run
        # registration, it passes should_cancel directly -- we then use it for
        # the cancel check but must NOT re-register the id, since this method's
        # unregister would otherwise clear the caller's progress / cancel
        # registration the moment a pass's fetch returns.
        own_run = should_cancel is None and bool(query_id)
        if should_cancel is None and query_id:
            should_cancel = lambda: self._run_is_cancelled(query_id)
        if own_run:
            self._register_run(query_id, None, kind="fetch", target=base)
        try:
            res = _load_api(self, url, base_name=base, auth_user=auth_user,
                            auth_pass=pwd, params=params or None,
                            destination="auto", json_path=cfg.get("json_path"),
                            headers=headers or None, fetch=self._api_fetcher,
                            spool=self._api_spooler,
                            pagination=cfg.get("pagination"),
                            max_pages=max_pages,
                            retry=cfg.get("retry"), sleep=self._api_sleep,
                            should_cancel=should_cancel)
        finally:
            if own_run:
                self._unregister_run(query_id)
        if isinstance(res, dict) and res.get("cancelled"):
            return {"error": "cancelled", "cancelled": True}

        errname = "__nbapierr_" + hashlib.md5(
            str(node_id).encode("utf-8")).hexdigest()[:10]

        def _make_err_table(eng_name, row):
            """Create the node's fixed-schema error table (0 rows when row is
            None) in ``eng_name``; returns (name, engine)."""
            eng = self.duckdb if eng_name == "duckdb" else self.db
            if eng is None:
                eng, eng_name = self.db, "sqlite"
            try:
                eng.drop_table(errname)
            except Exception:
                pass
            rows = [row] if row is not None else []
            try:
                nm, _n = eng.add_table_streaming(
                    errname, ["ok", "status", "error", "url"], iter(rows),
                    source="api error")
            except Exception:
                return None, eng_name
            return nm, eng_name

        def _default_engine():
            try:
                return self.default_destination()
            except Exception:
                return "sqlite"

        def _soft_fail(msg, status, src_url):
            # continue-on-error: surface the failure on the err port and leave
            # the data (out) port empty, rather than aborting the fetch.
            en, eng_used = _make_err_table(
                _default_engine(), (0, status, msg, src_url))
            self._api_node_tables[node_id] = {
                "table": None, "engine": None,
                "err": en, "err_engine": eng_used}
            self._invalidate_counts()
            return {"ok": True, "fetched": False, "error_captured": msg,
                    "status": status, "table": None, "columns": [], "rows": 0,
                    "err_table": en, "err_rows": 1, "url": src_url}

        if not isinstance(res, dict) or res.get("error"):
            msg = (res or {}).get("error", "API fetch failed.")
            status = (res or {}).get("status")
            src_url = (res or {}).get("url", url)
            if not continue_on_error:
                return {"error": msg, "status": status, "url": src_url}
            return _soft_fail(msg, status, src_url)

        tables = res.get("tables") or []
        root = tables[0] if tables else None
        if not root or not root.get("name"):
            if not continue_on_error:
                return {"error": "The response loaded no table.",
                        "url": res.get("url", url)}
            return _soft_fail("Response loaded no table.", res.get("status"),
                              res.get("url", url))

        eng_name = root.get("engine", "sqlite")
        # a 0-row error table in the same engine keeps the err port's schema
        # stable whether or not the latest fetch failed
        en, eng_used = _make_err_table(eng_name, None)
        self._api_node_tables[node_id] = {
            "table": root["name"], "engine": eng_name,
            "err": en, "err_engine": eng_used}
        self._invalidate_profiles()
        self._invalidate_counts()
        return {"ok": True, "table": root["name"], "engine": eng_name,
                "columns": root.get("columns", []), "rows": root.get("rows"),
                "err_table": en, "err_rows": 0,
                "pages": res.get("pages"), "url": res.get("url", url)}

    def _drop_hidden_source_table(self, node_id):
        prev = self._api_node_tables.get(node_id)
        if not prev:
            return
        for nm, en in ((prev.get("table"), prev.get("engine")),
                       (prev.get("err"), prev.get("err_engine"))):
            if not nm:
                continue
            eng = self.duckdb if en == "duckdb" else self.db
            try:
                if eng is not None:
                    eng.drop_table(nm)
            except Exception:
                pass
        self._api_node_tables.pop(node_id, None)

    def _load_hidden_columns_rows(self, node_id, prefix, columns, rows,
                                  source_label, query_id=None):
        """Materialize columns/rows into a hidden ``__…`` table for source nodes."""
        import hashlib
        from .engines import HAS_DUCKDB
        self._drop_hidden_source_table(node_id)
        base = prefix + hashlib.md5(str(node_id).encode("utf-8")).hexdigest()[:10]
        cols = [str(c) for c in (columns or [])]
        if not cols:
            return {"error": "No columns to load."}
        data = list(rows or [])
        tname, n, engine = None, 0, None
        own_run = bool(query_id)
        if own_run:
            self._register_run(query_id, None, kind="fetch", target=base)
        try:
            if HAS_DUCKDB:
                try:
                    duck = self.get_duckdb()
                    if own_run:
                        self._register_run(query_id, duck, kind="fetch",
                                           target=base)
                    tname, n = duck.add_table_streaming(
                        base, cols, iter(data), source=source_label)
                    engine = "duckdb"
                except Exception:
                    tname, n, engine = None, 0, None
            if engine is None:
                if own_run:
                    self._register_run(query_id, self.db, kind="fetch",
                                       target=base)
                tname, n = self.db.add_table_streaming(
                    base, cols, iter(data), source=source_label)
                engine = "sqlite"
        finally:
            if own_run:
                self._unregister_run(query_id)
        self._api_node_tables[node_id] = {"table": tname, "engine": engine}
        self._invalidate_profiles()
        self._invalidate_counts()
        eng = self.duckdb if engine == "duckdb" else self.db
        col_now = list((getattr(eng, "table_columns", {}) or {}).get(tname) or cols)
        return {
            "ok": True,
            "table": tname,
            "engine": engine,
            "columns": col_now,
            "rows": n,
        }

    def fetch_sqlserver_node(self, node_id, config, query_id=None):
        """Connect via a saved mssql profile, inline node fields, or an active
        connection name and import the node's query into a hidden table.

        Password resolution matches Load Data: one-shot ``pwd`` in the fetch
        body, else DPAPI secret under ``secret_key`` / ``profile_key``.
        """
        cfg = dict(config or {})
        query = (cfg.get("query") or "").strip()
        if not query:
            return {"error": "Enter a SQL query for the SQL Server node."}
        conn_name = (cfg.get("connection") or "").strip()
        profile_key = (cfg.get("profile_key") or "").strip()
        secret_key = (cfg.get("secret_key") or "").strip() or profile_key
        server = (cfg.get("server") or "").strip()
        if not conn_name and profile_key:
            conn_name = profile_key.split(":", 1)[-1] if ":" in profile_key \
                else profile_key
        if not conn_name and server:
            conn_name = server
        if not conn_name:
            return {"error":
                    "Set a server, saved mssql profile, or active connection "
                    "name on the SQL Server node."}

        if conn_name not in self.connections:
            from samql_core.mssql import (SQLServerConnection,
                                          build_mssql_conn_str,
                                          split_domain_user, HAS_PYODBC)
            if not HAS_PYODBC:
                return {"error": "pyodbc is not installed on the server."}

            fields = {}
            pk = profile_key or ("mssql:" + conn_name)
            prof = None
            try:
                prof = self.connection_profiles.get(pk)
            except Exception:
                prof = None
            if prof:
                fields = dict(prof.get("fields") or {})
                if not secret_key:
                    secret_key = pk
            # Inline node config fills / overrides blanks (Load-modal parity).
            for k, ck in (
                ("driver", "driver"),
                ("server", "server"),
                ("port", "port"),
                ("auth", "auth"),
                ("user", "user"),
                ("encrypt", "encrypt"),
                ("trust", "trust"),
                ("multi_subnet", "multi_subnet"),
                ("login_timeout", "login_timeout"),
                ("stmt_timeout", "stmt_timeout"),
                ("read_only", "read_only"),
                ("extra", "extra"),
            ):
                v = cfg.get(ck)
                if v is None or v == "":
                    continue
                fields[k] = v
            if not (fields.get("server") or "").strip() and not prof:
                return {"error":
                        'Connection "%s" is not active and no saved profile '
                        "or server fields were found. Save an mssql profile "
                        "(with password if needed) or connect in Load Data, "
                        "then Fetch again." % conn_name}

            auth = fields.get("auth", "windows") or "windows"
            pwd = (cfg.get("pwd") or "") or ""
            if not pwd and secret_key:
                try:
                    pwd = self.secrets.get(secret_key) or ""
                except Exception:
                    pwd = ""
            if not pwd and pk and pk != secret_key:
                try:
                    pwd = self.secrets.get(pk) or ""
                except Exception:
                    pwd = ""
            if auth in ("sql", "windows_alt") and not pwd:
                return {"error":
                        "No password available for %s. Enter it once and tick "
                        "Save password on the SQL Server node (or Load Data "
                        "profile), then retry." % (secret_key or pk or conn_name)}

            alt_creds = None
            if auth == "windows_alt":
                domain, user = split_domain_user(fields.get("user", ""))
                alt_creds = (domain, user, pwd)
            try:
                conn_str = build_mssql_conn_str(
                    fields.get("driver"), fields.get("server"),
                    fields.get("port", ""), auth, fields.get("user", ""),
                    pwd, bool(fields.get("encrypt", True)),
                    bool(fields.get("trust", True)),
                    bool(fields.get("multi_subnet", False)),
                    fields.get("extra", ""))
                conn = SQLServerConnection(
                    conn_name, conn_str, alt_creds=alt_creds,
                    login_timeout=int(fields.get("login_timeout", 15) or 15),
                    stmt_timeout=int(fields.get("stmt_timeout", 0) or 0),
                    read_only=bool(fields.get("read_only", True)))
                self.connections[conn_name] = conn
            except Exception as e:
                return {"error": "Could not connect: %s" % e}

        # Optional default database (same role as the Load tab Database picker).
        database = (cfg.get("database") or "").strip()
        if database:
            conn = self.connections.get(conn_name)
            if conn is not None:
                try:
                    # Bracket-quote a simple identifier; reject odd names.
                    safe = database.replace("]", "]]")
                    conn.execute("USE [%s]" % safe)
                except Exception as e:
                    return {"error":
                            "Could not USE database %r: %s" % (database, e)}

        # Use a hidden base name so the import stays off the tables list.
        import hashlib
        base = "__nbsql_" + hashlib.md5(
            str(node_id).encode("utf-8")).hexdigest()[:10]
        self._drop_hidden_source_table(node_id)
        res = self.import_from_connection(
            conn_name, query, base_name=base, destination="duckdb",
            query_id=query_id)
        if res.get("error") or res.get("cancelled"):
            return res
        tname = res.get("table")
        engine = res.get("engine") or "duckdb"
        self._api_node_tables[node_id] = {"table": tname, "engine": engine}
        cols = []
        try:
            eng = self.duckdb if engine == "duckdb" else self.db
            cols = list((getattr(eng, "table_columns", {}) or {}).get(tname) or [])
        except Exception:
            cols = []
        rows = None
        try:
            loaded = (res.get("loaded") or [{}])[0]
            tables = loaded.get("tables") or []
            if tables:
                rows = tables[0].get("rows")
        except Exception:
            rows = None
        return {
            "ok": True,
            "table": tname,
            "engine": engine,
            "columns": cols,
            "rows": rows,
        }

    def fetch_sharepoint_node(self, node_id, config, query_id=None):
        from . import sharepoint as _sp
        from . import sharepoint_auth as _spa
        cfg = dict(config or {})
        site = (cfg.get("site_url") or "").strip()
        mode = (cfg.get("mode") or "list").strip().lower()
        lst = (cfg.get("list_title") or "").strip()
        folder = (cfg.get("folder_path") or "").strip()
        try:
            auth = _spa.resolve_auth(self, cfg)
        except Exception as e:
            return {"error": str(e)}
        auth_mode = auth.get("mode") or "bearer"
        token = auth.get("token") or ""
        # Persist resolved secret_key so later runs find the OAuth blob.
        if auth.get("secret_key") and not (cfg.get("secret_key") or "").strip():
            cfg["secret_key"] = auth["secret_key"]
        own_run = bool(query_id)
        if own_run:
            self._register_run(query_id, None, kind="fetch", target="sharepoint")
        try:
            if own_run and self._run_is_cancelled(query_id):
                return {"cancelled": True}
            if mode == "drive":
                records, meta = _sp.browse_drive(
                    site, folder, token, auth_mode=auth_mode)
                label = folder or "/"
            else:
                records, meta = _sp.fetch_sharepoint_items(
                    site, lst, token, auth_mode=auth_mode)
                label = lst
            cols, rows, _ = _sp.records_to_columns_rows(records)
        except Exception as e:
            return {"error": str(e)}
        finally:
            if own_run:
                self._unregister_run(query_id)
        out = self._load_hidden_columns_rows(
            node_id, "__nbsp_", cols, rows,
            source_label="sharepoint:%s" % label, query_id=None)
        if out.get("ok"):
            out["meta"] = meta
            if auth.get("secret_key"):
                out["secret_key"] = auth["secret_key"]
        return out

    def download_sharepoint_file(self, config, query_id=None):
        """Download one SharePoint drive file into the user's Downloads folder."""
        from . import sharepoint as _sp
        from . import sharepoint_auth as _spa
        import os
        cfg = dict(config or {})
        site = (cfg.get("site_url") or "").strip()
        item_id = (cfg.get("item_id") or cfg.get("file_id") or "").strip()
        download_url = (cfg.get("download_url") or "").strip() or None
        try:
            auth = _spa.resolve_auth(self, cfg)
        except Exception as e:
            return {"error": str(e)}
        auth_mode = auth.get("mode") or "bearer"
        token = auth.get("token") or ""
        own_run = bool(query_id)
        if own_run:
            self._register_run(query_id, None, kind="fetch", target="sharepoint-dl")
        out_dir = self._resolve_export_dir("")
        # Resolve a unique destination path first so the download can stream
        # straight to disk (no whole-body buffer in RAM).
        probe_name = "sharepoint_file"
        if download_url:
            # Best-effort name from URL path; Graph Content-Disposition wins later.
            try:
                import urllib.parse as _up
                probe_name = _up.unquote(
                    _up.urlparse(download_url).path.rsplit("/", 1)[-1]
                ) or probe_name
            except Exception:
                pass
        safe = "".join(
            c if c.isalnum() or c in "._- " else "_"
            for c in probe_name
        ).strip() or "sharepoint_file"
        path = os.path.join(out_dir, safe)
        base, ext = os.path.splitext(path)
        n = 1
        while os.path.exists(path):
            path = "%s_%d%s" % (base, n, ext)
            n += 1
        try:
            if own_run and self._run_is_cancelled(query_id):
                return {"cancelled": True}
            nbytes, filename, meta = _sp.download_drive_item(
                site, item_id, token, download_url=download_url,
                dest_path=path, auth_mode=auth_mode)
        except Exception as e:
            try:
                if os.path.exists(path):
                    os.unlink(path)
            except Exception:
                pass
            return {"error": str(e)}
        finally:
            if own_run:
                self._unregister_run(query_id)
        # Rename if Graph supplied a better Content-Disposition name.
        final_path = path
        if filename and filename != os.path.basename(path):
            safe2 = "".join(
                c if c.isalnum() or c in "._- " else "_"
                for c in filename
            ).strip() or os.path.basename(path)
            cand = os.path.join(out_dir, safe2)
            b2, e2 = os.path.splitext(cand)
            k = 1
            while os.path.exists(cand) and cand != path:
                cand = "%s_%d%s" % (b2, k, e2)
                k += 1
            if cand != path:
                try:
                    os.replace(path, cand)
                    final_path = cand
                except Exception:
                    final_path = path
        try:
            nbytes = int(nbytes)
        except Exception:
            try:
                nbytes = os.path.getsize(final_path)
            except OSError:
                nbytes = 0
        return {
            "ok": True,
            "path": final_path,
            "filename": os.path.basename(final_path),
            "bytes": nbytes,
            "meta": meta,
        }

    def sharepoint_auth_capabilities(self):
        from . import sharepoint_auth as _spa
        return _spa.auth_capabilities()

    def sharepoint_auth_device_start(self, config):
        from . import sharepoint_auth as _spa
        try:
            return _spa.start_device_code(self, config or {})
        except Exception as e:
            return {"error": str(e)}

    def sharepoint_auth_device_poll(self, flow_id, block=False):
        from . import sharepoint_auth as _spa
        try:
            return _spa.poll_device_code(
                self, flow_id or "", block=bool(block))
        except Exception as e:
            return {"error": str(e)}

    def sharepoint_auth_interactive(self, config):
        from . import sharepoint_auth as _spa
        try:
            return _spa.interactive_sign_in(self, config or {})
        except Exception as e:
            return {"error": str(e)}

    def fetch_webscrape_node(self, node_id, config, query_id=None):
        from . import webscrape as _ws
        cfg = dict(config or {})
        url = (cfg.get("url") or "").strip()
        if not url:
            return {"error": "Give the Web scrape node a URL."}
        mode = (cfg.get("mode") or "tables").strip().lower()
        json_path = (cfg.get("json_path") or "").strip() or None
        own_run = bool(query_id)
        if own_run:
            self._register_run(query_id, None, kind="fetch", target="webscrape")
        try:
            if own_run and self._run_is_cancelled(query_id):
                return {"cancelled": True}
            cols, rows, meta = _ws.scrape_to_columns_rows(
                url,
                mode=mode,
                table_index=cfg.get("table_index") or 0,
                json_path=json_path,
            )
        except Exception as e:
            return {"error": str(e)}
        finally:
            if own_run:
                self._unregister_run(query_id)
        out = self._load_hidden_columns_rows(
            node_id, "__nbws_", cols, rows,
            source_label="webscrape:%s" % url, query_id=None)
        if out.get("ok"):
            out["meta"] = meta
            out["url"] = (meta or {}).get("url", url)
        return out

    def fetch_source_node(self, node_type, node_id, config, graph=None,
                          query_id=None):
        typ = (node_type or "").strip().lower()
        if typ == "apinode":
            return self.fetch_api_node(node_id, config, graph=graph,
                                       query_id=query_id)
        if typ == "sqlserver":
            return self.fetch_sqlserver_node(node_id, config, query_id=query_id)
        if typ == "sharepoint":
            return self.fetch_sharepoint_node(node_id, config, query_id=query_id)
        if typ == "webscrape":
            return self.fetch_webscrape_node(node_id, config, query_id=query_id)
        return {"error": "Unknown source node type: %s" % typ}

    def load_folder_files(self, folder, query_id=None):
        """Read every loadable file in a folder and stack them (UNION ALL,
        aligned by column name -- missing columns become NULL) into a hidden
        table the append-from-folder node references like an input. Cached by
        the folder's file set + mtimes. Returns {ok, table, engine, columns,
        rows, files} or {error}."""
        import os
        import hashlib
        from . import loaders
        folder = (folder or "").strip()
        if not folder or not os.path.isdir(folder):
            return {"error": "Pick a folder that exists."}
        exts = ("csv", "tsv", "txt", "json", "ndjson", "jsonl",
                "parquet", "pq", "xlsx", "xlsm", "xls")
        try:
            names = sorted(
                n for n in os.listdir(folder)
                if os.path.splitext(n)[1].lower().lstrip(".") in exts
                and os.path.isfile(os.path.join(folder, n)))
        except OSError as e:
            return {"error": err_str(e)}
        if not names:
            return {"error": "No readable files in that folder "
                    "(CSV, TSV, JSON, Parquet, Excel)."}
        paths = [os.path.join(folder, n) for n in names]
        try:
            sig = tuple((n, os.path.getmtime(p)) for n, p in zip(names, paths))
        except OSError as e:
            return {"error": err_str(e)}

        def _count(eng, name):
            try:
                _c, r = eng.execute('SELECT COUNT(*) FROM "%s"' % name)
                return int(r[0][0]) if r else None
            except Exception:
                return None

        cached = self._folder_files.get(folder)
        if cached:
            tname, eng_name, csig = cached
            eng = self.duckdb if eng_name == "duckdb" else self.db
            if csig == sig and eng is not None and tname in eng.table_columns:
                return {"ok": True, "table": tname, "engine": eng_name,
                        "columns": eng.table_columns.get(tname, []),
                        "rows": _count(eng, tname), "files": len(names)}
            try:
                eng.drop_table(tname)
            except Exception:
                pass
            self._folder_files.pop(folder, None)

        _oid = None
        try:
            _oid = opreg.begin(
                "load", target=os.path.basename(os.path.normpath(folder))
                or folder)
        except Exception:
            _oid = None
        # Make the read cancelable: register the engine it will use under the
        # query id, so cancel_query(query_id) -> engine.interrupt() aborts the
        # in-flight file read. destination="auto" resolves to the default
        # engine, so that is the one doing the work.
        if query_id:
            try:
                eng_c = (self.get_duckdb()
                         if self.default_destination() == "duckdb"
                         else self.db)
            except Exception:
                eng_c = self.db
            if eng_c is not None:
                with self._running_lock:
                    self._running[query_id] = eng_c
        temps = []  # (table, columns, engine)
        try:
            for i, p in enumerate(paths):
                h = hashlib.md5(p.encode("utf-8")).hexdigest()[:8]
                try:
                    descs = loaders.load_file(
                        self, p, destination="auto",
                        base_name="__nbf_%d_%s" % (i, h))
                except Exception as e:
                    # A cancel interrupts the current file's read -- stop the
                    # whole folder load, don't quietly skip to the next file.
                    if _is_interrupt(e):
                        raise
                    # A single unreadable file is skipped (best-effort stack).
                    continue
                d = (descs or [None])[0]
                if d and d.get("name"):
                    temps.append((d["name"], d.get("columns", []),
                                  d.get("engine", "sqlite")))
            if not temps:
                return {"error": "Couldn't read any files in that folder."}
            engset = set(t[2] for t in temps)
            if len(engset) > 1:
                for tn, _c, te in temps:
                    eng2 = self.duckdb if te == "duckdb" else self.db
                    try:
                        eng2.drop_table(tn)
                    except Exception:
                        pass
                return {"error": "All files in the folder must be the same "
                        "type so they can be stacked."}
            eng_name = temps[0][2]
            eng, _kind = self._engine_obj(
                DUCKDB_TARGET if eng_name == "duckdb" else LOCAL_TARGET)
            union_cols = []
            seen = set()
            for _t, cols, _e in temps:
                for c in cols:
                    if c not in seen:
                        seen.add(c)
                        union_cols.append(c)

            def row_iter():
                for tname, _cols, _e in temps:
                    _c, rows = eng.execute('SELECT * FROM "%s"' % tname)
                    for row in rows:
                        rowmap = dict(zip(_c, row))
                        yield tuple(rowmap.get(c) for c in union_cols)

            combined = "__nbfolder_" + hashlib.md5(
                ("%s|%s" % (folder, sig)).encode("utf-8")).hexdigest()[:10]
            try:
                eng.drop_table(combined)
            except Exception:
                pass
            if eng_name == "duckdb":
                # Combine natively: UNION ALL BY NAME stacks the per-file temps
                # inside DuckDB, aligning columns by name (missing -> NULL) just
                # like the manual union below. This avoids pulling every row of
                # every file back through Python -- that row-by-row round-trip
                # (eng.execute materializes each whole temp in memory, then we
                # rebuild a dict + tuple per row) is what made a folder of large
                # files crawl, balloon memory, and sit on "Reading folder…" with
                # no progress so it looked stuck.
                name = combined
                try:
                    union_sql = " UNION ALL BY NAME ".join(
                        'SELECT * FROM "%s"' % t for t, _c, _e in temps)
                    eng.execute(
                        'CREATE TABLE "%s" AS %s' % (combined, union_sql))
                    ccols, _ = eng.execute(
                        'SELECT * FROM "%s" LIMIT 0' % combined)
                    eng.table_columns[combined] = list(ccols or union_cols)
                    eng.table_sources[combined] = folder
                    n = _count(eng, combined)
                except Exception as ce:
                    if _is_interrupt(ce):
                        raise
                    # Columns/types don't line up across files for a native
                    # stack -> fall back to the lenient Python union (coerces to
                    # a common text type) so the load still succeeds.
                    try:
                        eng.drop_table(combined)
                    except Exception:
                        pass
                    name, n = eng.add_table_streaming(
                        combined, union_cols, row_iter(), source="folder")
            else:
                # SQLite can't read files and has no UNION ALL BY NAME; stack the
                # rows in Python (folders targeting SQLite are small).
                name, n = eng.add_table_streaming(
                    combined, union_cols, row_iter(), source="folder")
            for tname, _c, _e in temps:
                try:
                    eng.drop_table(tname)
                except Exception:
                    pass
            self._folder_files[folder] = (name, eng_name, sig)
            self._invalidate_counts()
            return {"ok": True, "table": name, "engine": eng_name,
                    "columns": eng.table_columns.get(name, union_cols),
                    "rows": n, "files": len(temps)}
        except Exception as e:
            for tname, _c, te in temps:
                eng2 = self.duckdb if te == "duckdb" else self.db
                try:
                    eng2.drop_table(tname)
                except Exception:
                    pass
            if _is_interrupt(e):
                return {"error": "cancelled", "cancelled": True}
            return {"error": "Couldn't combine the folder: %s: %s"
                    % (type(e).__name__, e)}
        finally:
            if query_id:
                with self._running_lock:
                    self._running.pop(query_id, None)
            try:
                opreg.end(_oid)
            except Exception:
                pass

    def create_table_from_grid(self, name, columns, rows, destination="auto"):
        """Create a local table from manually-entered / pasted grid data so it
        shows in the loaded-tables list. ``columns`` is a list of names;
        ``rows`` is a list of row-lists (strings). Empty cells become NULL; a
        cell that looks like an integer or float is stored as a number.
        DuckDB-first with a SQLite fallback. Returns a LoadResult-shaped dict
        (so the UI counts it) or {error}."""
        cols = [str(c).strip() for c in (columns or []) if str(c).strip()]
        if not cols:
            return {"error": "Add at least one column."}
        # de-duplicate column names
        seen, final = {}, []
        for c in cols:
            if c in seen:
                seen[c] += 1
                final.append("%s_%d" % (c, seen[c]))
            else:
                seen[c] = 0
                final.append(c)
        cols = final
        ncol = len(cols)

        def _coerce(v):
            if v is None:
                return None
            s = str(v)
            st = s.strip()
            if st == "":
                return None
            try:
                if re.fullmatch(r"[-+]?\d+", st):
                    return int(st)
                return float(st)
            except (ValueError, TypeError):
                return s  # keep text exactly, including any inline / edge spaces

        body = []
        for r in (rows or []):
            vals = list(r)[:ncol] + [None] * (ncol - len(r))
            if all(v is None or str(v).strip() == "" for v in vals):
                continue
            body.append(tuple(_coerce(v) for v in vals))
        # Restrict the table name to a safe identifier so a stray quote or
        # control character in the (free-text) name can't create a table that
        # later code -- which references it as "<name>" -- then fails to drop
        # or count.
        bn = sanitize_ident(name, "table")
        src = "manual"
        tname = n = engine = None
        if destination in ("duckdb", "auto") and HAS_DUCKDB:
            try:
                duck = self.get_duckdb()
                tname, n = duck.add_table_streaming(bn, cols, iter(body),
                                                    source=src)
                engine = "duckdb"
            except Exception:
                tname = engine = None
        if engine is None:
            tname, n = self.db.add_table_streaming(bn, cols, iter(body),
                                                   source=src)
            engine = "sqlite"
        self._invalidate_profiles()
        self._invalidate_counts()
        eng = self.get_duckdb() if engine == "duckdb" else self.db
        cols_now = eng.table_columns.get(tname, cols)
        return {
            "ok": True,
            "table": tname,
            "engine": engine,
            "loaded": [{"file": bn,
                        "tables": [{"name": tname, "rows": n,
                                    "columns": cols_now, "engine": engine}]}],
            "tables": [],
        }

    def load_catalog(self, name, database=None):
        """Register every table (and its columns) from a remote connection as
        a 'catalog' table: name + column headers only, no row data. Querying
        one in the editor is routed to this connection (passthrough). Replaces
        any previously-loaded catalog for the same connection. Returns
        {ok, count} or {error}."""
        conn = self.connections.get(name)
        if conn is None:
            return {"error": 'Connection "%s" is not active.' % name}
        cols_by = conn.list_columns(database or None)
        if not cols_by:
            # no INFORMATION_SCHEMA access (or empty) -> fall back to names only
            cols_by = {(sch, tbl): []
                       for sch, tbl in conn.list_tables(database or None)}
        # drop any earlier catalog from this same connection
        self.forget_catalog(name)
        used = set(self.db.table_columns)
        if self.duckdb is not None:
            used |= set(self.duckdb.table_columns)
        count = 0
        for (sch, tbl), cols in sorted(cols_by.items()):
            key = tbl
            if key in self.catalog_tables or key in used:
                base = ("%s_%s" % (sch, tbl)) if sch else tbl
                key, k = base, 2
                while key in self.catalog_tables or key in used:
                    key = "%s_%d" % (base, k)
                    k += 1
            qualified = (("[%s]." % database) if database else "") + \
                ((("[%s]." % sch)) if sch else "") + ("[%s]" % tbl)
            self.catalog_tables[key] = {
                "conn": name,
                "database": database or "",
                "schema": sch,
                "table": tbl,
                "qualified": qualified,
                "columns": [{"name": c, "type": t} for c, t in cols],
            }
            # route by the bare table name (what users type after FROM)
            self._catalog_route.setdefault(tbl.lower(), name)
            count += 1
        return {"ok": True, "count": count, "connection": name}

    def forget_catalog(self, name):
        """Remove catalog tables registered from a connection (e.g. on
        disconnect). Pass None to clear all catalogs."""
        drop = [k for k, v in self.catalog_tables.items()
                if name is None or v.get("conn") == name]
        for k in drop:
            self.catalog_tables.pop(k, None)
        self._catalog_route = {
            t: c for t, c in self._catalog_route.items()
            if name is not None and c != name
        } if name is not None else {}
        return len(drop)

    def catalog_columns(self, name):
        """Columns for one catalog table, fetched on demand (the sidebar omits
        them from the tables list and asks for them when a table is expanded).
        Returns {columns:[{name,type}]} or {error}."""
        info = self.catalog_tables.get(name)
        if info is None:
            return {"error": 'Unknown catalog table "%s".' % name}
        return {"columns": info.get("columns", []),
                "qualified": info.get("qualified", name)}

    def import_catalog_table(self, name):
        """Pull a SQL Server catalog table's data into a local table (DuckDB
        first, SQLite fallback) so it shows in the loaded-tables list and can
        be queried / joined locally. Streams to disk, so large tables stay
        memory-bounded. Returns a LoadResult-shaped dict or {error}."""
        info = self.catalog_tables.get(name)
        if info is None:
            return {"error": 'Unknown catalog table "%s".' % name}
        conn = info.get("conn")
        if not conn or conn not in self.connections:
            return {"error": "That table's SQL Server connection is no longer "
                    "active."}
        qualified = info.get("qualified") or ('"%s"' % info.get("table", name))
        base = info.get("table") or name
        res = self.import_from_connection(
            conn, "SELECT * FROM %s" % qualified, base_name=base,
            destination="duckdb")
        if isinstance(res, dict) and res.get("ok"):
            self._invalidate_profiles()
            self._invalidate_counts()
        return res

    def flatten_json_to_csv_dir(self, json_path, out_dir, base_name=None,
                                progress=None, cancel=None):
        """Flatten a JSON file into its relational tables and write one CSV per
        table into a '<base>_flattened' subfolder of out_dir. Returns
        {ok, dir, files:[...], table_count} or {error}. Streams through the
        tolerant JSON reader and the flattener's per-table spill, so large
        inputs stay memory-bounded -- no data is loaded into the app.
        ``progress`` (optional) is called with a dict describing the current
        stage ("reading" -> bytes, then "writing" -> tables done)."""
        import tempfile
        import shutil
        from .flatten import JSONFlattener, stream_json_records
        src = (json_path or "").strip()
        if not src:
            return {"error": "Pick a JSON file to flatten."}
        if not os.path.isfile(src):
            return {"error": "That JSON file was not found."}
        base = (out_dir or "").strip()
        if not base:
            return {"error": "Pick an output folder."}
        if not os.path.isdir(base):
            return {"error": "The output folder was not found."}
        bn = (base_name or os.path.splitext(os.path.basename(src))[0]
              or "json").strip()
        bn = JSONFlattener._sanitize(bn) or "json"
        out = os.path.join(base, bn + "_flattened")
        existed_before = os.path.isdir(out)
        try:
            os.makedirs(out, exist_ok=True)
        except Exception as e:
            return {"error": "Couldn't create the output folder: %s" % e}
        spill = getattr(loaders, "JSON_SPILL_ROWS", 50000)
        try:
            spill_dir = tempfile.mkdtemp(prefix="jf_export_",
                                         dir=tmputil.instance_dir())
        except Exception:
            spill_dir = tempfile.mkdtemp(prefix="jf_export_",
                                         dir=None)  # last resort: system temp

        def _cancel_cb(_n):
            # The flattener calls this every ~5000 emitted rows from inside
            # add_record, so a single top-level record that explodes into a
            # large nested subtree stays interruptible mid-flatten -- not just
            # between records (the outer loop already checks there). _emit
            # re-raises this LoadCancelled so the flatten aborts promptly.
            if cancel is not None and cancel():
                raise loaders.LoadCancelled()

        fl = JSONFlattener(root_name=bn, spill_threshold=spill,
                           spill_dir=spill_dir, progress_cb=_cancel_cb)
        written, used = [], set()
        created_paths = []   # CSVs this run wrote -- removed if cancelled
        records = 0

        def _read_prog(done, total):
            if progress:
                try:
                    progress({"stage": "reading", "bytes_done": done,
                              "bytes_total": total, "records": records})
                except Exception:
                    pass
        try:
            for rec in stream_json_records(src, progress=_read_prog):
                if cancel is not None and cancel():
                    raise loaders.LoadCancelled()
                fl.add_record(rec)
                records += 1
            tables = list(fl.table_names())
            tables_total = len(tables)
            if progress:
                try:
                    progress({"stage": "writing", "tables_done": 0,
                              "tables_total": tables_total, "records": records})
                except Exception:
                    pass
            for ti, table in enumerate(tables):
                if cancel is not None and cancel():
                    raise loaders.LoadCancelled()
                cols = fl.columns(table)
                safe = JSONFlattener._sanitize(table) or "table"
                fname, k = safe, 2
                while fname.lower() in used:
                    fname = "%s_%d" % (safe, k)
                    k += 1
                used.add(fname.lower())
                fpath = os.path.join(out, fname + ".csv")
                created_paths.append(fpath)
                n = 0
                with open(fpath, "w", newline="", encoding="utf-8-sig") as f:
                    w = _csv.writer(f)
                    w.writerow(cols)
                    for row in fl.iter_rows_aligned(table):
                        if (cancel is not None and n % 4096 == 0 and cancel()):
                            raise loaders.LoadCancelled()
                        w.writerow(["" if v is None else v for v in row])
                        n += 1
                written.append({"table": table, "file": fname + ".csv",
                                "rows": n, "columns": len(cols)})
                if progress:
                    try:
                        progress({"stage": "writing", "tables_done": ti + 1,
                                  "tables_total": tables_total,
                                  "detail": fname + ".csv"})
                    except Exception:
                        pass
        except loaders.LoadCancelled:
            # Cancelled mid-flatten: remove only the CSVs this run wrote, and
            # the _flattened folder itself if we created it fresh (leave a
            # pre-existing folder + any prior export untouched).
            for p in created_paths:
                try:
                    os.remove(p)
                except OSError:
                    pass
            if not existed_before:
                try:
                    os.rmdir(out)
                except OSError:
                    pass
            raise
        except Exception as e:
            return {"error": str(e)}
        finally:
            try:
                fl.close()
            except Exception:
                pass
            shutil.rmtree(spill_dir, ignore_errors=True)
        if not written:
            return {"error": "No tables were produced from that JSON file."}
        return {"ok": True, "dir": out, "table_count": len(written),
                "files": written}

    # ---- shutdown ---------------------------------------------------
    def shutdown(self):
        self._closed = True
        try:
            with self._cleanup_lock:
                if self._cleanup_timer is not None:
                    self._cleanup_timer.cancel()
                    self._cleanup_timer = None
        except Exception:
            pass
        # Abort any in-flight engine work first so recycle/unlink isn't blocked
        # by an open statement handle (Windows file locks).
        try:
            self.interrupt_loads()
        except Exception:
            pass
        # Close remote ODBC / SQL Server connections before engines recycle so
        # a clean exit never leaves network sessions or driver state hanging.
        try:
            for name, conn in list((self.connections or {}).items()):
                try:
                    close = getattr(conn, "close", None)
                    if callable(close):
                        close()
                except Exception:
                    pass
                try:
                    del self.connections[name]
                except Exception:
                    pass
            self.connections.clear()
        except Exception:
            pass
        # Graceful shutdown (Ctrl+C / SIGTERM / Exit → stop) is a clean exit:
        # drop every table (loaded + temp) AND clear the restore manifest, so
        # the next launch comes up empty with nothing loaded. (A hard crash
        # never reaches here, so its manifest survives for crash recovery.)
        # AppWindow window-close alone does NOT call this -- the server is
        # left running for reattach (.534); only /api/shutdown does.
        self.clear_all(clear_manifest=True)
        try:
            self.db.recycle()
        except Exception:
            pass
        if self.duckdb is not None:
            try:
                self.duckdb.recycle()
            except Exception:
                pass
        for p in self.temp_files:
            try:
                os.unlink(p)
            except Exception:
                pass
        self.temp_files.clear()
